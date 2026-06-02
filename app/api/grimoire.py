# app/api/grimoire.py

import os
import json
import re
import io
import base64
from copy import copy
from datetime import datetime
from fastapi import APIRouter, Request, HTTPException, Body
from fastapi.responses import JSONResponse, StreamingResponse
from weasyprint import HTML, CSS
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.cell.cell import MergedCell
from typing import List, Dict, Any, Optional

import psycopg2
from psycopg2.extras import RealDictCursor

from api.astrology import get_seed_from_request
# 🔑 [v26 수복]: DB_PATH 소각, get_db 소환
from core.database import get_db
from core.panopticon import get_pano_db

# ─────────────────────────────────────────────────────────────
# 📚 기존 컴파일러 임포트 유지
# ─────────────────────────────────────────────────────────────

# nigredo
from core.grimoire.compilers.n2_compiler import compile_n2_grimoire
from core.grimoire.compilers.n2_nak_compiler import compile_n2_nak_grimoire
from core.grimoire.compilers.n3_compiler import compile_n3_grimoire
from core.grimoire.compilers.n3_domain_compiler import compile_n3_domain_grimoire
from core.grimoire.compilers.n3_lagna_compiler import compile_n3_lagna_grimoire
from core.grimoire.compilers.n4_compiler import compile_n4_grimoire
from core.grimoire.compilers.n5_unus_compiler import compile_n5_unus_grimoire
from core.grimoire.compilers.n5_intersectus_compiler import compile_n5_intersectus_grimoire
from core.grimoire.compilers.n6_compiler import compile_n6_grimoire
from core.grimoire.compilers.n6_aspects_compiler import compile_n6_aspects_grimoire
from core.grimoire.compilers.n6_varga_compiler import compile_n6_varga_grimoire
from core.grimoire.compilers.n6_varga_kp_compiler import compile_n6_varga_kp_grimoire
from core.grimoire.compilers.n7_compiler import compile_n7_grimoire
from core.grimoire.compilers.n7_sabian_planets_compiler import compile_n7_sabian_planets_grimoire
from core.grimoire.compilers.n7_sabian_asteroids_compiler import compile_n7_sabian_asteroids_grimoire
from core.grimoire.compilers.n7_sabian_nodes_compiler import compile_n7_sabian_nodes_grimoire
from core.grimoire.compilers.n7_sabian_fates_compiler import compile_n7_sabian_fates_grimoire
from core.grimoire.compilers.n7_sabian_angles_compiler import compile_n7_sabian_angles_grimoire
from core.grimoire.compilers.n7_sabian_hermetic_compiler import compile_n7_sabian_hermetic_grimoire
from core.grimoire.compilers.n7_planets_en_compiler import compile_n7_planets_en_grimoire
from core.grimoire.compilers.n7_planets_ko_compiler import compile_n7_planets_ko_grimoire
from core.grimoire.compilers.n7_asteroids_en_compiler import compile_n7_asteroids_en_grimoire
from core.grimoire.compilers.n7_asteroids_ko_compiler import compile_n7_asteroids_ko_grimoire
from core.grimoire.compilers.n7_nodes_en_compiler import compile_n7_nodes_en_grimoire
from core.grimoire.compilers.n7_nodes_ko_compiler import compile_n7_nodes_ko_grimoire
from core.grimoire.compilers.n7_fates_en_compiler import compile_n7_fates_en_grimoire
from core.grimoire.compilers.n7_fates_ko_compiler import compile_n7_fates_ko_grimoire
from core.grimoire.compilers.n7_angles_en_compiler import compile_n7_angles_en_grimoire
from core.grimoire.compilers.n7_angles_ko_compiler import compile_n7_angles_ko_grimoire
from core.grimoire.compilers.n7_hermetic_en_compiler import compile_n7_hermetic_en_grimoire
from core.grimoire.compilers.n7_hermetic_ko_compiler import compile_n7_hermetic_ko_grimoire
from core.grimoire.compilers.n8_en_compiler import compile_n8_en_grimoire
from core.grimoire.compilers.n8_ko_compiler import compile_n8_ko_grimoire
from core.grimoire.compilers.n9_compiler import compile_n9_grimoire
from core.grimoire.compilers.n9_vd_compiler import compile_n9_vd_grimoire

# albedo
from core.grimoire.compilers.a2_comp_compiler import compile_a2_comp_grimoire
from core.grimoire.compilers.a2_compiler import compile_a2_grimoire
from core.grimoire.compilers.a2_nak_compiler import compile_a2_nak_grimoire
from core.grimoire.compilers.a3_comp_compiler import compile_a3_comp_grimoire
from core.grimoire.compilers.a3_comp_domain_compiler import compile_a3_comp_domain_grimoire
from core.grimoire.compilers.a3_compiler import compile_a3_grimoire
from core.grimoire.compilers.a3_domain_compiler import compile_a3_domain_grimoire
from core.grimoire.compilers.a3_lagna_compiler import compile_a3_lagna_grimoire
from core.grimoire.compilers.a4_compiler import compile_a4_grimoire
from core.grimoire.compilers.a5_comp_compiler import compile_a5_comp_grimoire
from core.grimoire.compilers.a5_unus_compiler import compile_a5_unus_grimoire
from core.grimoire.compilers.a5_intersectus_compiler import compile_a5_intersectus_grimoire
from core.grimoire.compilers.a6_aspects_compiler import compile_a6_aspects_grimoire
from core.grimoire.compilers.a6_compiler import compile_a6_grimoire
from core.grimoire.compilers.a6_varga_compiler import compile_a6_varga_grimoire
from core.grimoire.compilers.a6_varga_kp_compiler import compile_a6_varga_kp_grimoire
from core.grimoire.compilers.a7_compiler import compile_a7_grimoire
from core.grimoire.compilers.a7_sabian_planets_compiler import compile_a7_sabian_planets_grimoire
from core.grimoire.compilers.a7_sabian_asteroids_compiler import compile_a7_sabian_asteroids_grimoire
from core.grimoire.compilers.a7_sabian_nodes_compiler import compile_a7_sabian_nodes_grimoire
from core.grimoire.compilers.a7_sabian_fates_compiler import compile_a7_sabian_fates_grimoire
from core.grimoire.compilers.a7_sabian_angles_compiler import compile_a7_sabian_angles_grimoire
from core.grimoire.compilers.a7_sabian_hermetic_compiler import compile_a7_sabian_hermetic_grimoire
from core.grimoire.compilers.a7_planets_en_compiler import compile_a7_planets_en_grimoire
from core.grimoire.compilers.a7_planets_ko_compiler import compile_a7_planets_ko_grimoire
from core.grimoire.compilers.a7_asteroids_en_compiler import compile_a7_asteroids_en_grimoire
from core.grimoire.compilers.a7_asteroids_ko_compiler import compile_a7_asteroids_ko_grimoire
from core.grimoire.compilers.a7_nodes_en_compiler import compile_a7_nodes_en_grimoire
from core.grimoire.compilers.a7_nodes_ko_compiler import compile_a7_nodes_ko_grimoire
from core.grimoire.compilers.a7_fates_en_compiler import compile_a7_fates_en_grimoire
from core.grimoire.compilers.a7_fates_ko_compiler import compile_a7_fates_ko_grimoire
from core.grimoire.compilers.a7_angles_en_compiler import compile_a7_angles_en_grimoire
from core.grimoire.compilers.a7_angles_ko_compiler import compile_a7_angles_ko_grimoire
from core.grimoire.compilers.a7_hermetic_en_compiler import compile_a7_hermetic_en_grimoire
from core.grimoire.compilers.a7_hermetic_ko_compiler import compile_a7_hermetic_ko_grimoire
from core.grimoire.compilers.a8_en_compiler import compile_a8_en_grimoire
from core.grimoire.compilers.a8_ko_compiler import compile_a8_ko_grimoire
from core.grimoire.compilers.a8_sys_en_compiler import compile_a8_sys_en_grimoire
from core.grimoire.compilers.a8_sys_ko_compiler import compile_a8_sys_ko_grimoire
from core.grimoire.compilers.a10_compiler import compile_a10_grimoire
from core.grimoire.compilers.a9_compiler import compile_a9_grimoire
from core.grimoire.compilers.a9_synastry_compiler import compile_a9_synastry_grimoire
from core.grimoire.compilers.a9_vd_compiler import compile_a9_vd_grimoire

# citrinitas
from core.grimoire.compilers.c1_en_compiler import compile_c1_en_grimoire
from core.grimoire.compilers.c1_ko_compiler import compile_c1_ko_grimoire

class GrimoireEditPayload(BaseModel):
    stage: str
    sheet_data: List[Dict[str, Any]]

class GrimoirePDFPayload(BaseModel):
    target_name: str
    stage: str
    html_content: str  

class GrimoireRenamePayload(BaseModel):
    new_name: str

router = APIRouter(prefix="/api/grimoire", tags=["grimoire-api"])

NIGREDO_ORDER = [
    "n2", "n2_nak", "n3", "n3_domain", "n3_lagna", "n4", "n5_unus", "n5_intersectus",
    "n6", "n6_aspects", "n6_varga", "n7", "n7_sabian_planets", "n7_sabian_asteroids", 
    "n7_sabian_nodes", "n7_sabian_fates", "n7_sabian_angles", "n7_sabian_hermetic", "n7_planets",
    "n7_asteroids", "n7_nodes", "n7_fates", "n7_angles", "n7_hermetic", "n8", "n9", "n9_vd"
]

ALBEDO_ORDER = [
    "a2_comp", "a2", "a2_nak", "a3_comp", "a3_comp_domain", "a3", "a3_domain", "a3_lagna", "a4",
    "a5_comp", "a5_unus", "a5_intersectus", "a6", "a6_aspects", "a6_varga",
    "a7", "a7_sabian_planets", "a7_sabian_asteroids", "a7_sabian_nodes", "a7_sabian_fates", "a7_sabian_angles", 
    "a7_sabian_hermetic", "a7_planets", "a7_asteroids", "a7_nodes", "a7_fates", "a7_angles", "a7_hermetic",
    "a8", "a9", "a9_synastry", "a9_vd", "a10"
]

EXCEL_COMPILERS = {
    "n2": compile_n2_grimoire, "n2_nak": compile_n2_nak_grimoire, "n3": compile_n3_grimoire, "n3_domain": compile_n3_domain_grimoire,
    "n3_lagna": compile_n3_lagna_grimoire, "n4": compile_n4_grimoire, "n5_unus": compile_n5_unus_grimoire, "n5_intersectus": compile_n5_intersectus_grimoire,
    "n6": compile_n6_grimoire, "n6_aspects": compile_n6_aspects_grimoire, "n6_varga": compile_n6_varga_grimoire, "n6_varga_kp": compile_n6_varga_kp_grimoire,
    "n7": compile_n7_grimoire, "n7_sabian_planets": compile_n7_sabian_planets_grimoire, "n7_sabian_asteroids": compile_n7_sabian_asteroids_grimoire,
    "n7_sabian_nodes": compile_n7_sabian_nodes_grimoire, "n7_sabian_fates": compile_n7_sabian_fates_grimoire, "n7_sabian_angles": compile_n7_sabian_angles_grimoire,
    "n7_sabian_hermetic": compile_n7_sabian_hermetic_grimoire, "n7_planets_en": compile_n7_planets_en_grimoire, "n7_planets_ko": compile_n7_planets_ko_grimoire,
    "n7_asteroids_en": compile_n7_asteroids_en_grimoire, "n7_asteroids_ko": compile_n7_asteroids_ko_grimoire, "n7_nodes_en": compile_n7_nodes_en_grimoire,
    "n7_nodes_ko": compile_n7_nodes_ko_grimoire, "n7_fates_en": compile_n7_fates_en_grimoire, "n7_fates_ko": compile_n7_fates_ko_grimoire,
    "n7_angles_en": compile_n7_angles_en_grimoire, "n7_angles_ko": compile_n7_angles_ko_grimoire, "n7_hermetic_en": compile_n7_hermetic_en_grimoire,
    "n7_hermetic_ko": compile_n7_hermetic_ko_grimoire, "n8_en": compile_n8_en_grimoire, "n8_ko": compile_n8_ko_grimoire,
    "n9": compile_n9_grimoire, "n9_vd": compile_n9_vd_grimoire,

    "a2_comp": compile_a2_comp_grimoire, "a2": compile_a2_grimoire, "a2_nak": compile_a2_nak_grimoire, "a3_comp": compile_a3_comp_grimoire,
    "a3_comp_domain": compile_a3_comp_domain_grimoire, "a3": compile_a3_grimoire, "a3_domain": compile_a3_domain_grimoire, "a3_lagna": compile_a3_lagna_grimoire,
    "a4": compile_a4_grimoire, "a5_comp": compile_a5_comp_grimoire, "a5_unus": compile_a5_unus_grimoire, "a5_intersectus": compile_a5_intersectus_grimoire,
    "a6_aspects": compile_a6_aspects_grimoire, "a6": compile_a6_grimoire, "a6_varga": compile_a6_varga_grimoire, "a6_varga_kp": compile_a6_varga_kp_grimoire,
    "a7": compile_a7_grimoire, "a7_sabian_planets": compile_a7_sabian_planets_grimoire, "a7_sabian_asteroids": compile_a7_sabian_asteroids_grimoire,
    "a7_sabian_nodes": compile_a7_sabian_nodes_grimoire, "a7_sabian_fates": compile_a7_sabian_fates_grimoire, "a7_sabian_angles": compile_a7_sabian_angles_grimoire,
    "a7_sabian_hermetic": compile_a7_sabian_hermetic_grimoire, "a7_planets_en": compile_a7_planets_en_grimoire, "a7_planets_ko": compile_a7_planets_ko_grimoire,
    "a7_asteroids_en": compile_a7_asteroids_en_grimoire, "a7_asteroids_ko": compile_a7_asteroids_ko_grimoire, "a7_nodes_en": compile_a7_nodes_en_grimoire,
    "a7_nodes_ko": compile_a7_nodes_ko_grimoire, "a7_fates_en": compile_a7_fates_en_grimoire, "a7_fates_ko": compile_a7_fates_ko_grimoire,
    "a7_angles_en": compile_a7_angles_en_grimoire, "a7_angles_ko": compile_a7_angles_ko_grimoire, "a7_hermetic_en": compile_a7_hermetic_en_grimoire,
    "a7_hermetic_ko": compile_a7_hermetic_ko_grimoire, "a8_en": compile_a8_en_grimoire, "a8_ko": compile_a8_ko_grimoire,
    "a8_sys_en": compile_a8_sys_en_grimoire, "a8_sys_ko": compile_a8_sys_ko_grimoire, "a10": compile_a10_grimoire,
    "a9": compile_a9_grimoire, "a9_synastry": compile_a9_synastry_grimoire, "a9_vd": compile_a9_vd_grimoire,

    "c1_en": compile_c1_en_grimoire, "c1_ko": compile_c1_ko_grimoire
}

# ─────────────────────────────────────────────────────────────
# ⚙️ THE CORE UTILITIES (바이너리 화석 변환 도구)
# ─────────────────────────────────────────────────────────────

def workbook_to_base64(wb: Workbook) -> str:
    output = io.BytesIO()
    wb.save(output)
    return base64.b64encode(output.getvalue()).decode('utf-8')

def base64_to_workbook(b64_str: str) -> Workbook:
    excel_bin = base64.b64decode(b64_str)
    return load_workbook(io.BytesIO(excel_bin))

def clone_worksheet_perfectly(source_ws, target_ws):
    for col, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col].width = dim.width
    for row, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row].height = dim.height
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format) 
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)

# ─────────────────────────────────────────────────────────────
# 🌐 API ROUTERS (Binary Fossil Based)
# ─────────────────────────────────────────────────────────────

@router.get("/list/{stage}")
async def get_grimoire_list(stage: str, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)

    try:
        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT id, target_name, compiler_id, created_at 
            FROM grimoire_archives 
            WHERE user_id = %s AND stage = %s 
            ORDER BY created_at DESC
        """, (user_id, stage.lower()))
        
        archives = [{"idx": str(row["id"]), "name": row["target_name"], "file": f"{row['id']}.xlsx"} for row in cursor.fetchall()]
        return JSONResponse(content=archives)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@router.post("/save/excel/{module_id}")
async def save_excel_archive(module_id: str, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)
    
    try:
        chart_data = await request.json()
        target_name = chart_data.get('target_name', 'Unknown')
        stage = str(chart_data.get('stage', 'nigredo')).lower()
        
        compiler_func = EXCEL_COMPILERS.get(module_id)
        if not compiler_func: raise HTTPException(status_code=400)

        is_albedo = module_id.startswith("a")
        seed_data = get_seed_from_request(request, is_albedo=is_albedo)

        new_wb = compiler_func(chart_data, seed_data=seed_data)  
        new_ws = new_wb.active
        new_sheet_name = new_ws.title

        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("SELECT id, archive_fossil FROM grimoire_archives WHERE user_id = %s AND stage = %s AND target_name = %s", (user_id, stage, target_name))
        existing = cursor.fetchone()

        if existing:
            archive_id = existing["id"]
            main_wb = base64_to_workbook(existing["archive_fossil"])
            
            existing_names = main_wb.sheetnames
            final_name = new_sheet_name
            if final_name in existing_names:
                suffix = 2
                while f"{final_name} ({suffix})" in existing_names: suffix += 1
                final_name = f"{final_name} ({suffix})"
            
            target_ws = main_wb.create_sheet(title=final_name)
            clone_worksheet_perfectly(new_ws, target_ws)
            
            fossil = workbook_to_base64(main_wb)
            cursor.execute("UPDATE grimoire_archives SET archive_fossil = %s, compiler_id = %s WHERE id = %s", (fossil, module_id, archive_id))
        else:
            fossil = workbook_to_base64(new_wb)
            # 🚀 [PostgreSQL 특권]: 삽입과 동시에 부여받은 번호를 받아옵니다.
            cursor.execute("INSERT INTO grimoire_archives (user_id, stage, target_name, compiler_id, archive_fossil) VALUES (%s, %s, %s, %s, %s) RETURNING id",
                           (user_id, stage, target_name, module_id, fossil))
            archive_id = cursor.fetchone()["id"]
            
        conn.commit()
        cursor.close()
        conn.close()

        # 👁️ [PANOPTICON HOOK]
        try:
            pano_conn = get_pano_db()
            pano_cursor = pano_conn.cursor()
            pano_cursor.execute("INSERT INTO grimoire_logs (compiler_id) VALUES (%s)", (module_id,))
            pano_conn.commit()
            pano_cursor.close()
            pano_conn.close()
        except Exception as e:
            print(f"[PANOPTICON ERROR] Failed to record grimoire log: {e}")

        return JSONResponse(content={"status": "success", "file": f"{archive_id}.xlsx"})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/download/{stage}/{idx}")
async def download_grimoire_archive(stage: str, idx: str, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)
    
    try:
        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT archive_fossil FROM grimoire_archives WHERE id = %s AND user_id = %s", (idx, user_id))
        row = cursor.fetchone()
        
        if not row: raise HTTPException(status_code=404)

        binary_data = base64.b64decode(row["archive_fossil"])
        
        if stage.lower() == "rubedo":
            return StreamingResponse(io.BytesIO(binary_data), media_type="application/pdf")
        
        return StreamingResponse(
            io.BytesIO(binary_data), 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Binary fossil extraction failed: {str(e)}")
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@router.get("/info/{stage}/{idx}")
async def get_grimoire_info(stage: str, idx: str, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)
    
    try:
        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT target_name, compiler_id FROM grimoire_archives WHERE id = %s AND user_id = %s", (idx, user_id))
        row = cursor.fetchone()
        
        if not row: raise HTTPException(status_code=404)
        return JSONResponse(content={"target_name": row["target_name"], "compiler_id": row["compiler_id"]})
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@router.put("/edit/{archive_id}")
async def overwrite_grimoire_archive(archive_id: str, payload: GrimoireEditPayload, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)

    try:
        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT archive_fossil, compiler_id FROM grimoire_archives WHERE id = %s AND user_id = %s", (archive_id, user_id))
        row = cursor.fetchone()
        
        if not row: raise HTTPException(status_code=404)
        
        wb = base64_to_workbook(row["archive_fossil"])
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

        for sheet_json in payload.sheet_data:
            s_name = sheet_json["name"]
            if s_name in wb.sheetnames:
                ws = wb[s_name]
                for r_idx, row_data in enumerate(sheet_json.get("data", [])):
                    for c_idx, cell_obj in enumerate(row_data):
                        if not cell_obj: continue
                        target_cell = ws.cell(row=r_idx + 1, column=c_idx + 1)
                        if isinstance(target_cell, MergedCell): continue 
                        
                        val = cell_obj.get("v")
                        if val is not None:
                            val = ILLEGAL_CHARACTERS_RE.sub("", str(val))
                            target_cell.value = val

        new_fossil = workbook_to_base64(wb)
        cursor.execute("UPDATE grimoire_archives SET archive_fossil = %s WHERE id = %s", (new_fossil, archive_id))
        conn.commit()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@router.post("/rearrange/{stage}/{idx}")
async def rearrange_grimoire_sheets(stage: str, idx: str, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)

    try:
        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT archive_fossil FROM grimoire_archives WHERE id = %s AND user_id = %s", (idx, user_id))
        row = cursor.fetchone()
        if not row: raise HTTPException(status_code=404)
        
        wb = base64_to_workbook(row["archive_fossil"])
        
        if stage == "nigredo": order_list = NIGREDO_ORDER
        elif stage == "albedo": order_list = ALBEDO_ORDER
        else: order_list = []

        parsed_sheets = []
        for ws in wb.worksheets:
            title = ws.title.strip()
            match = re.match(r"^([a-zA-Z0-9_]+)(?:\s*\((\d+)\))?$", title)
            base_name = match.group(1) if match else title
            suffix_num = int(match.group(2)) if match and match.group(2) else 0
            lookup_name = re.sub(r'_(en|ko)$', '', base_name)
            
            try: 
                hierarchy_idx = order_list.index(lookup_name)
            except ValueError: 
                hierarchy_idx = 9999 

            parsed_sheets.append({
                "ws": ws, 
                "base_name": base_name, 
                "h_idx": hierarchy_idx, 
                "s_num": suffix_num
            })

        parsed_sheets.sort(key=lambda x: (x["h_idx"], x["s_num"]))

        for i, item in enumerate(parsed_sheets): 
            item["ws"].title = f"__TEMP_{i}__"
            
        wb._sheets = [item["ws"] for item in parsed_sheets]
        
        name_counter = {}
        for item in parsed_sheets:
            bn = item["base_name"]
            name_counter[bn] = name_counter.get(bn, 0) + 1
            count = name_counter[bn]
            item["ws"].title = bn if count == 1 else f"{bn} ({count})"

        new_fossil = workbook_to_base64(wb)
        cursor.execute("UPDATE grimoire_archives SET archive_fossil = %s WHERE id = %s", (new_fossil, idx))
        conn.commit()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@router.delete("/delete_sheet/{stage}/{idx}/{sheet_name}")
async def delete_grimoire_sheet(stage: str, idx: str, sheet_name: str, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)
    
    try:
        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT archive_fossil FROM grimoire_archives WHERE id = %s AND user_id = %s", (idx, user_id))
        row = cursor.fetchone()
        
        if not row: return JSONResponse(status_code=404, content={"detail": "Archive not found."})
        
        wb = base64_to_workbook(row["archive_fossil"])
        if sheet_name in wb.sheetnames:
            if len(wb.sheetnames) <= 1:
                return JSONResponse(status_code=400, content={"detail": "Cannot delete the only sheet."})
            del wb[sheet_name]
            
            new_fossil = workbook_to_base64(wb)
            cursor.execute("UPDATE grimoire_archives SET archive_fossil = %s WHERE id = %s", (new_fossil, idx))
            conn.commit()
            
        return JSONResponse(content={"status": "success"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": str(e)})
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@router.delete("/delete/{stage}/{idx}")
async def delete_grimoire_archive(stage: str, idx: str, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM grimoire_archives WHERE id = %s AND user_id = %s", (idx, user_id))
        conn.commit()
        return JSONResponse(content={"status": "success"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

# ─────────────────────────────────────────────────────────────
# 📜 THE CORE UTILITIES (PDF 연성 도구 - Rubedo 전용)
# ─────────────────────────────────────────────────────────────

def html_to_pdf_base64(html_content: str) -> str:
    pdf_css = """
    @import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700;800&display=swap');

    @page {
        size: A4 landscape; 
        margin: 15mm 15mm 20mm 15mm; 
        
        @bottom-center {
            content: "Created by Tetramegistus";
            color: rgba(60, 140, 90, 0.6); 
            font-family: 'JetBrains Mono', 'Consolas', monospace;
            font-size: 10pt;
            letter-spacing: 0.2em;
        }
    }
    
    body {
        font-family: 'JetBrains Mono', 'Consolas', monospace !important;
        color: #000000;
        background: #ffffff;
    }

    h1, h2, h3, .article-view-title {
        font-family: 'JetBrains Mono', 'Consolas', monospace !important;
        font-weight: 800 !important; 
    }

    thead th, thead td, tr:first-child th, tr:first-child td {
        color: #ffffff !important;
        font-weight: bold !important;
        background-color: #000000 !important; 
    }

    thead th *, thead td *, tr:first-child th *, tr:first-child td * {
        color: #ffffff !important;
    }

    table { 
        width: 100%;
        border-collapse: collapse;
        page-break-inside: auto; 
        table-layout: auto;
    }
    
    th, td {
        word-wrap: break-word; 
        overflow-wrap: break-word;
        padding: 8px; 
        border: 1px solid #ddd;
        font-family: 'JetBrains Mono', 'Consolas', monospace !important; 
    }
    
    tr { 
        page-break-inside: avoid; 
        page-break-after: auto; 
    }
    thead { display: table-header-group; }
    tfoot { display: table-footer-group; }
    p { orphans: 3; widows: 3; } 
    """
    
    full_html = f"""
    <!DOCTYPE html>
    <html>
    <head><meta charset="UTF-8"></head>
    <body>
        {html_content}
    </body>
    </html>
    """
    
    pdf_bytes = HTML(string=full_html).write_pdf(stylesheets=[CSS(string=pdf_css)])
    return base64.b64encode(pdf_bytes).decode('utf-8')

# ─────────────────────────────────────────────────────────────
# 🌐 API ROUTERS (Rubedo PDF & Common Utilities)
# ─────────────────────────────────────────────────────────────

@router.get("/check_name/{stage}")
async def check_grimoire_name(stage: str, name: str, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)
    
    try:
        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(
            "SELECT id FROM grimoire_archives WHERE user_id = %s AND stage = %s AND target_name = %s", 
            (user_id, stage.lower(), name)
        )
        row = cursor.fetchone()
        return JSONResponse(content={"exists": bool(row)})
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@router.put("/rename/{archive_id}")
async def rename_grimoire_archive(archive_id: str, payload: GrimoireRenamePayload, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE grimoire_archives SET target_name = %s WHERE id = %s AND user_id = %s", 
            (payload.new_name, archive_id, user_id)
        )
        conn.commit()
        return JSONResponse(content={"status": "success", "new_name": payload.new_name})
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@router.post("/save/pdf/{module_id}")
async def save_pdf_archive(module_id: str, payload: GrimoirePDFPayload, request: Request):
    user_id = request.cookies.get("session_user_id")
    if not user_id: raise HTTPException(status_code=401)
    
    try:
        fossil = html_to_pdf_base64(payload.html_content)
        
        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute(
            "SELECT id FROM grimoire_archives WHERE user_id = %s AND stage = %s AND target_name = %s", 
            (user_id, payload.stage, payload.target_name)
        )
        existing = cursor.fetchone()

        if existing:
            archive_id = existing["id"]
            cursor.execute(
                "UPDATE grimoire_archives SET archive_fossil = %s, compiler_id = %s WHERE id = %s", 
                (fossil, module_id, archive_id)
            )
        else:
            # 🚀 [PostgreSQL 특권]: RETURNING id 사용
            cursor.execute(
                "INSERT INTO grimoire_archives (user_id, stage, target_name, compiler_id, archive_fossil) VALUES (%s, %s, %s, %s, %s) RETURNING id",
                (user_id, payload.stage, payload.target_name, module_id, fossil)
            )
            archive_id = cursor.fetchone()["id"]
            
        conn.commit()
        cursor.close()
        conn.close()

        # 👁️ [PANOPTICON HOOK]
        try:
            pano_conn = get_pano_db()
            pano_cursor = pano_conn.cursor()
            pano_cursor.execute("INSERT INTO grimoire_logs (compiler_id) VALUES (%s)", (f"RUBEDO_{payload.target_name}",))
            pano_conn.commit()
            pano_cursor.close()
            pano_conn.close()
        except Exception as e:
            print(f"[PANOPTICON ERROR] Failed to record grimoire log: {e}")

        return JSONResponse(content={"status": "success", "file": f"{archive_id}.pdf"})
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF Alchemy failed: {str(e)}")