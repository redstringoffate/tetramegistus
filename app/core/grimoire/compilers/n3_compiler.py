import json
import os
import shutil
import math
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import calculate_principia, _ensure_float_tz, format_dms_pretty, TROPICAL_SIGNS, SYMBOL_MAP

from core.astrology.divisions.decan import get_decan
from core.astrology.divisions.duad import get_duad
from core.astrology.divisions.dodecatemoria import get_dodecatemoria
from core.astrology.divisions.egyptian_bounds import get_egyptian_bounds
from core.astrology.divisions.sabian_engine import get_sabian_index

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/n3_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/nigredo/n3.xlsx'))

SABIAN_PATHS = [
    os.path.abspath(os.path.join(BASE_DIR, '../../../data/render/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../../data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../static/data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../data/sabian.json')),
]

H_SYS_MAP = {
    'P': 'Placidus', 'W': 'Whole Sign', 'K': 'Koch', 'R': 'Regiomontanus',
    'C': 'Campanus', 'E': 'Equal', 'V': 'Vehlow'
}

RULER_SYMBOLS = {
    "Sun": "☉", "Moon": "☽", "Mercury": "☿", "Venus": "♀", "Mars": "♂", 
    "Jupiter": "♃", "Saturn": "♄", "Uranus": "♅", "Neptune": "♆", "Pluto": "♇"
}

LOCAL_SIGNS = ["Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo", "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces"]
TRADITIONAL_RULERS = {
    "Aries": "Mars", "Taurus": "Venus", "Gemini": "Mercury", "Cancer": "Moon",
    "Leo": "Sun", "Virgo": "Mercury", "Libra": "Venus", "Scorpio": "Mars", 
    "Sagittarius": "Jupiter", "Capricorn": "Saturn", "Aquarius": "Saturn", "Pisces": "Jupiter"
}

def get_col(cell_str):
    if not cell_str: return ""
    return "".join([c for c in str(cell_str) if c.isalpha()])

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def format_house_range(lon_start, lon_end):
    if lon_start is None or lon_end is None: return ""
    size = (lon_end - lon_start) % 360
    if size < 0: size += 360
    rounded_size = round(size)
    
    portions = []
    curr = lon_start
    rem = size
    while rem > 1e-5:
        next_boundary = (math.floor(curr / 30) + 1) * 30
        step = min(rem, next_boundary - curr)
        portions.append(step)
        curr = (curr + step) % 360
        rem -= step
        
    rounded_portions = [round(p) for p in portions]
    diff = rounded_size - sum(rounded_portions)
    if diff != 0 and rounded_portions:
        rounded_portions[rounded_portions.index(max(rounded_portions))] += diff
        
    portions_str = "/".join(map(str, rounded_portions))
    return f"{rounded_size} ({portions_str})"

def compile_n3_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "n3"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")
    
    try: fixed_star_orb = float(meta.get("fixed_star_orb", 1.0))
    except: fixed_star_orb = 1.0

    seed_data = chart_data.get("seed", {})
    if seed_data is None:
        raise ValueError("Grimoire requires active seed data. Please select a seed first.")

    def ensure_location_string(seed):
        loc_val = str(seed.get("location", seed.get("city", ""))).strip()
        if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
            lat = get_safe_float(seed.get("lat"), None)
            lng = get_safe_float(seed.get("lng"), None)
            if lat is not None and lng is not None:
                seed["city"] = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
                seed["location"] = seed["city"]
            else:
                seed["city"] = "Unknown Location"
                seed["location"] = "Unknown Location"
        else:
            seed["city"] = loc_val
            seed["location"] = loc_val
            
    ensure_location_string(seed_data)

    date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(seed_data.get('birth_time', '12:00:00'))
    lat = get_safe_float(seed_data.get("lat"), 37.5665)
    lng = get_safe_float(seed_data.get("lng"), 126.9780)
    tz_val = seed_data.get('tz') if seed_data.get('tz') is not None else seed_data.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)
    is_unk = bool(seed_data.get("is_time_unknown", 0))

    apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])

    res = calculate_principia(
        date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz,
        system=system, ayanamsa=ayanamsa, view=view, h_sys=h_sys,
        fixed_star_orb=fixed_star_orb, is_time_unknown=is_unk
    )

    planets_data = res.get('planets', {})
    angles_data = res.get('angles', {})
    houses_data = res.get('houses', {})
    # 🚀 [추가]: 엔진 결과에서 로드 데이터 변수를 확보합니다.
    lords_data = res.get('lords', {})

    lang_val = chart_data.get("language", meta.get("language", "en"))
    is_ko = "ko" in str(lang_val).lower() or "kr" in str(lang_val).lower()

    sabian_dict = {}
    for spath in SABIAN_PATHS:
        if os.path.exists(spath):
            try:
                with open(spath, 'r', encoding='utf-8') as f: sabian_dict = json.load(f)
                break
            except: pass
    
    def get_sabian_text(idx):
        if not idx: return ""
        idx_str = str(idx)
        def extract_text(entry):
            if isinstance(entry, str): return entry
            if isinstance(entry, dict):
                p_keys = ["text_ko", "ko", "desc_ko"] if is_ko else ["text_en", "en", "desc_en"]
                for k in p_keys + ["text", "desc", "meaning"]:
                    if k in entry and entry[k]: return str(entry[k])
                for k, v in entry.items():
                    if k != "index" and isinstance(v, str): return v
            return ""
        if isinstance(sabian_dict, dict):
            if idx_str in sabian_dict: return extract_text(sabian_dict[idx_str])
            for k, v in sabian_dict.items():
                if isinstance(v, dict) and str(v.get("index", "")) == idx_str: return extract_text(v)
        return ""

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    asc_data = angles_data.get("Ascendant") or planets_data.get("Ascendant") or {}
    mc_data = angles_data.get("Midheaven") or planets_data.get("Midheaven") or {}
    
    ws[mapping["metadata"]["ascendant"]] = asc_data.get("dms", "")
    ws[mapping["metadata"]["midheaven"]] = mc_data.get("dms", "")
    
    sys_str = system.capitalize()
    ws[mapping["metadata"]["sys_tab"]] = sys_str.upper()
    if system.upper() == "SIDEREAL":
        ws[mapping["metadata"]["ayanamsa"]] = ayanamsa.upper()
    
    h_sys_key = str(h_sys).upper()[0] if h_sys else 'P'
    ws[mapping["metadata"]["h_sys"]] = H_SYS_MAP.get(h_sys_key, h_sys.capitalize())
    
    orb_cell = ws[mapping["metadata"]["fixed_star_orb"]]
    orb_cell.value = f"{fixed_star_orb:.1f}"
    orb_cell.data_type = 's'
    orb_cell.number_format = '@'

    # 🚀 [수정]: 메타데이터 스타일 방역 (ASC/MC는 4원소 색상 적용, 나머지는 스킵)
    # 리스트에 "ascendant"와 "midheaven"을 추가했습니다.
    for key in ["ascendant", "midheaven", "sys_tab", "h_sys", "fixed_star_orb", "ayanamsa"]:
        c_ref = mapping["metadata"].get(key)
        if c_ref and ws[c_ref].value:
            cell = ws[c_ref]
            
            # ASC, MC는 도수 정보가 있으므로 배경색(Fire, Earth 등)을 허용합니다.
            # 나머지는 단순 텍스트 정보이므로 배경색을 끕니다.
            should_skip = key not in ["ascendant", "midheaven"]
            
            apply_grimoire_styles(cell, cell.value, is_info_col=False, skip_color=should_skip)
            
            # 맑은고딕 제거 및 Consolas 지정 (styler에서 지정한 폰트 색상을 유지합니다)
            color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
            cell.font = Font(name="Consolas", size=11, color=color_to_use, bold=False) 
            cell.alignment = Alignment(horizontal='left', vertical='center')

    houses_list = []
    for i in range(1, 13):
        h_str = str(i)
        if h_str in mapping.get("houses", {}):
            houses_list.append({
                "house": i,
                "map": mapping["houses"][h_str]
            })
            
    houses_list.sort(key=lambda x: x["map"].get("row_start", 0), reverse=True)
    # 🚀 [추가]: 판정용 로드 리스트 미리 확보
    day_lords_list = [d.strip() for d in str(lords_data.get('day', '-')).split('|')]
    hour_lord = str(lords_data.get('hour', '-')).strip()
    
    angle_map = {1: "Ascendant", 4: "Immum Coeli", 7: "Descendant", 10: "Midheaven"}

    for item in houses_list:
        h_num = item["house"]
        h_map = item["map"]
        row_idx = h_map.get("row_start", 0)
        
        if not row_idx: continue

        col_info = get_col(h_map.get("info", "B"))
        col_ruler = get_col(h_map.get("ruler", "C"))
        col_ruler_info = get_col(h_map.get("ruler_info", "D"))
        col_dignity = get_col(h_map.get("dignity", "E"))
        col_range = get_col(h_map.get("range", "F"))
        col_duad = get_col(h_map.get("duad", "G"))
        col_dodeca = get_col(h_map.get("dodeca", "H"))
        col_decan = get_col(h_map.get("decan", "I"))
        col_bounds = get_col(h_map.get("bounds", "J"))
        col_sabian = get_col(h_map.get("sabian", "K"))
        col_s_name = get_col(h_map.get("star_name", "L"))
        col_s_info = get_col(h_map.get("star_info", "M"))
        col_s_orb = get_col(h_map.get("star_orb", "N"))

        h_raw = houses_data.get(h_num) or houses_data.get(str(h_num)) or houses_data.get(f"H{h_num}") or {}
        lon = h_raw.get("longitude")
        
        if lon is None:
            for c_char in "BCDEFGHIJKLMNO":
                ws[f"{c_char}{row_idx}"] = ""
            continue

        sign_idx = int(lon / 30) % 12
        sign_name = LOCAL_SIGNS[sign_idx]
        deg_in_sign = lon % 30

        ws[f"{col_info}{row_idx}"] = format_dms_pretty(lon)
        
        raw_ruler = TRADITIONAL_RULERS.get(sign_name, "")
            
        r_data = {}
        matched_p_name = raw_ruler
        for p_key, p_val in planets_data.items():
            if p_key.lower() == raw_ruler.lower():
                r_data = p_val
                matched_p_name = p_key
                break
                
        ws[f"{col_ruler}{row_idx}"] = matched_p_name.capitalize() if matched_p_name else ""
        
        if r_data and 'longitude' in r_data:
            r_lon = float(r_data.get('longitude', 0))
            r_sign_name = LOCAL_SIGNS[int(r_lon / 30) % 12]
            
            # 🚀 [수정]: 띄어쓰기 제거 및 역행(r) 이중 표기 방지
            raw_dms_str = str(r_data.get("dms", format_dms_pretty(r_lon)))
            
            # 1. 사인 이름이 붙어있으면 떼어냄
            if raw_dms_str.lower().startswith(r_sign_name.lower()):
                raw_dms_str = raw_dms_str[len(r_sign_name):].strip(" ,")
                
            # 2. 엔진에서 이미 붙여서 보낸 역행 표기 찌꺼기 완벽 제거 (중복 방지)
            raw_dms_str = raw_dms_str.replace(", r", "").replace(",r", "").strip()
                
            # 3. 역행일 경우 띄어쓰기 없이 ',r' 추가
            r_retro = ",r" if r_data.get("is_retrograde") or r_data.get("is_retro") else ""
            
            # 4. 사인 이름 뒤 콤마(,)에 띄어쓰기 없이 결합!
            ws[f"{col_ruler_info}{row_idx}"] = f"{r_sign_name},{raw_dms_str}{r_retro}"
            raw_dignity = str(r_data.get("dignity", "")).strip()
            ws[f"{col_dignity}{row_idx}"] = raw_dignity.capitalize() if raw_dignity else "-"
        else:
            ws[f"{col_ruler_info}{row_idx}"] = "-"
            ws[f"{col_dignity}{row_idx}"] = "-"

        next_h = 1 if h_num == 12 else h_num + 1
        next_raw = houses_data.get(next_h) or houses_data.get(str(next_h)) or houses_data.get(f"H{next_h}") or {}
        next_lon = next_raw.get("longitude")
        if next_lon is not None:
            ws[f"{col_range}{row_idx}"] = format_house_range(lon, next_lon)
        else:
            ws[f"{col_range}{row_idx}"] = ""

        ws[f"{col_duad}{row_idx}"] = SYMBOL_MAP.get(get_duad(sign_name, deg_in_sign), "-")
        ws[f"{col_dodeca}{row_idx}"] = SYMBOL_MAP.get(get_dodecatemoria(deg_in_sign), "-")
        ws[f"{col_decan}{row_idx}"] = SYMBOL_MAP.get(get_decan(sign_name, deg_in_sign), "-")
        ws[f"{col_bounds}{row_idx}"] = SYMBOL_MAP.get(get_egyptian_bounds(sign_name, deg_in_sign), "-")
        
        s_idx = get_sabian_index(lon)
        ws[f"{col_sabian}{row_idx}"] = get_sabian_text(str(s_idx)) if s_idx else ""

        # 🚀 [수정]: 지배성(Ruler) 및 커스프(Cusp) 판정 로직 추가
        r_is_an = int(float(r_data.get('longitude', 0)) % 30) == 29 if r_data else False
        r_is_day = matched_p_name in day_lords_list
        r_is_hour = (matched_p_name == hour_lord) # 🚩 추가: Hour Lord 판별
        cusp_is_an = int(float(lon) % 30) == 29

        # 🚩 추가: Hour Lord일 경우 텍스트 강제 대문자화
        if r_is_hour:
            ws[f"{col_ruler}{row_idx}"] = str(ws[f"{col_ruler}{row_idx}"].value).upper()
            ws[f"{col_ruler_info}{row_idx}"] = str(ws[f"{col_ruler_info}{row_idx}"].value).upper()

        for c_char in "BCDEFGHIJK":
            cell = ws[f"{c_char}{row_idx}"]
            is_info = (c_char in [col_info, col_ruler_info])
            skip_p = (c_char == col_ruler_info) # Ruler Info 열은 행성 고유색 차단
            
            # 아나레틱 적용 대상 구분 (Cusp 또는 Ruler Info)
            current_an = r_is_an if c_char == col_ruler_info else (cusp_is_an if c_char == col_info else False)

            apply_grimoire_styles(
                cell, cell.value, is_info_col=is_info, 
                skip_color=(c_char == col_sabian),
                is_day_lord=(r_is_day if c_char in [col_ruler, col_ruler_info] else False),
                is_hour_lord=(r_is_hour if c_char in [col_ruler, col_ruler_info] else False), # 🚩 Styler에 Hour Lord 전달
                is_anaretic=current_an,
                skip_planet_color=skip_p
            )
            
            # Consolas 및 정렬 고정 (🚩 col_ruler를 좌측 정렬 그룹으로 편입)
            h_align = 'left' if c_char in [col_info, col_ruler, col_ruler_info, col_sabian] else 'center'
            cell.font = Font(name="Consolas", color=cell.font.color if cell.font else "000000", bold=cell.font.bold)
            cell.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=(c_char == col_sabian))

        stars = []
        if h_num in angle_map:
            angle_name = angle_map[h_num]
            a_raw = angles_data.get(angle_name) or planets_data.get(angle_name) or {}
            raw_stars = a_raw.get("fixed_stars", [])
            for s in raw_stars:
                if isinstance(s, dict): stars.append(s)
                elif isinstance(s, str): stars.append({"name": s, "position": "", "orb": ""})

        if not stars:
            ws[f"{col_s_name}{row_idx}"] = ""
            ws[f"{col_s_info}{row_idx}"] = ""
            ws[f"{col_s_orb}{row_idx}"] = ""
            for c in [col_s_name, col_s_info, col_s_orb]:
                apply_grimoire_styles(ws[f"{c}{row_idx}"], "", skip_color=True)
        else:
            ws[f"{col_s_name}{row_idx}"] = stars[0].get("name", "")
            ws[f"{col_s_info}{row_idx}"] = stars[0].get("position", "")
            orb_val = stars[0].get("orb", "")
            ws[f"{col_s_orb}{row_idx}"] = f"{orb_val}°" if orb_val else ""

            for c in [col_s_name, col_s_info, col_s_orb]:
                cell = ws[f"{c}{row_idx}"]
                is_info = (c == col_s_info)
                apply_grimoire_styles(cell, cell.value, is_info_col=is_info, skip_color=False)
                
                if cell.font: cell.font = Font(name="Consolas", color=cell.font.color, bold=cell.font.bold)
                else: cell.font = Font(name="Consolas")
                
                horiz_align = 'left' if c in [col_s_name, col_s_info] else 'center'
                cell.alignment = Alignment(horizontal=horiz_align, vertical='center', wrap_text=True)

            extra_rows = len(stars) - 1
            if extra_rows > 0:
                ws.insert_rows(row_idx + 1, amount=extra_rows)
                
                for i in range(1, len(stars)):
                    insert_idx = row_idx + i
                    ws.row_dimensions[insert_idx].height = 16.5
                    
                    ws[f"{col_s_name}{insert_idx}"] = stars[i].get("name", "")
                    ws[f"{col_s_info}{insert_idx}"] = stars[i].get("position", "")
                    o_val = stars[i].get("orb", "")
                    ws[f"{col_s_orb}{insert_idx}"] = f"{o_val}°" if o_val else ""

                    for col_chr in "BCDEFGHIJKLMNO":
                        src = ws[f"{col_chr}{row_idx}"]
                        tgt = ws[f"{col_chr}{insert_idx}"]
                        if src.has_style:
                            tgt.border = copy(src.border)
                            tgt.fill = copy(src.fill)
                            tgt.font = copy(src.font)
                            tgt.number_format = copy(src.number_format)

                    for c in [col_s_name, col_s_info, col_s_orb]:
                        cell = ws[f"{c}{insert_idx}"]
                        is_info = (c == col_s_info)
                        apply_grimoire_styles(cell, cell.value, is_info_col=is_info, skip_color=False)
                        
                        if cell.font: cell.font = Font(name="Consolas", color=cell.font.color, bold=cell.font.bold)
                        else: cell.font = Font(name="Consolas")
                        
                        horiz_align = 'left' if c in [col_s_name, col_s_info] else 'center'
                        cell.alignment = Alignment(horizontal=horiz_align, vertical='center', wrap_text=True)

                for col_chr in "ABCDEFGHIJK":
                    ws.merge_cells(f"{col_chr}{row_idx}:{col_chr}{row_idx + extra_rows}")
                    old_align = ws[f"{col_chr}{row_idx}"].alignment
                    
                    # 🚀 [요청 반영]: 병합 셀에서도 Sabian 열 좌측 정렬 유지
                    new_horiz = 'center' if col_chr in [col_range, col_dignity] else ('left' if col_chr in [col_info, col_ruler, col_ruler_info, col_sabian] else (old_align.horizontal if old_align else 'left'))
                    wrap_val = True if col_chr == col_sabian else False
                    ws[f"{col_chr}{row_idx}"].alignment = Alignment(horizontal=new_horiz, vertical='center', wrap_text=wrap_val)

    # 🚀 [Airtight]: 모든 셀의 폰트를 Consolas로 일괄 세탁 (맑은고딕 방역)
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name="Consolas", size=cell.font.size, bold=cell.font.bold,
                    italic=cell.font.italic, color=cell.font.color
                )

    # ====================================================================
    # 🚀 [수정]: Draconic / Ketunic 시스템일 경우 항성 구역 완전 파기 및 숨김
    # ====================================================================
    if system.upper() in ["DRACONIC", "KETUNIC"]:
        # 1. L~O 열(12~15)에 걸친 병합 셀 찾아서 모두 해제 (웹 뷰어 강제 확장 방지)
        merges_to_remove = [m for m in ws.merged_cells.ranges if m.max_col >= 12]
        for m_range in merges_to_remove:
            ws.unmerge_cells(str(m_range))
            
        # 2. 텍스트 잔해물 전부 삭제 (글자가 남아있으면 뷰어가 숨김을 거부함)
        for row in ws.iter_rows(min_col=12, max_col=15): # L(12) ~ O(15)
            for cell in row:
                cell.value = None
                
        # 3. L, M, N 그리고 여백 테두리가 있는 O열까지 싹 다 이중 잠금
        for col_letter in ['L', 'M', 'N', 'O']:
            ws.column_dimensions[col_letter].hidden = True
            ws.column_dimensions[col_letter].width = 0.1
    # ====================================================================

    return wb