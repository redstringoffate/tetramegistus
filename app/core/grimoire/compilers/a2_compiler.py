import json
import os
import shutil
import math
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import calculate_principia, _ensure_float_tz
from core.astrology.davison import calculate_davison_midpoint

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a2_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a2.xlsx'))

SABIAN_PATHS = [
    os.path.abspath(os.path.join(BASE_DIR, '../../../data/render/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../../data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../static/data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../data/sabian.json')),
]

RULER_SYMBOLS = {
    "Sun": "☉", "Moon": "☽", "Mercury": "☿", "Venus": "♀", "Mars": "♂", 
    "Jupiter": "♃", "Saturn": "♄", "Uranus": "♅", "Neptune": "♆", "Pluto": "♇",
    "Rahu": "☊", "Ketu": "☋", "Chiron": "⚷"
}

def get_col(cell_str):
    if not cell_str: return ""
    return "".join([c for c in str(cell_str) if c.isalpha()])

def compile_a2_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a2"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name
    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")
    try: fixed_star_orb = float(meta.get("fixed_star_orb", 1.0))
    except: fixed_star_orb = 1.0

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1']
    s2_raw = seed_data['seed2']
    dav_raw = calculate_davison_midpoint(s1_raw, s2_raw)
    seed_data = dav_raw
    seed_data['name'] = meta.get('target_name', 'CONIUNCTIO')

    date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(seed_data.get('birth_time', '12:00:00'))
    try: lat = float(seed_data.get("lat", 37.5665))
    except: lat = 37.5665
    try: lng = float(seed_data.get("lng", 126.9780))
    except: lng = 126.9780
    
    tz_val = seed_data.get('tz') if seed_data.get('tz') is not None else seed_data.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)
    is_unk = bool(s1_raw.get("is_time_unknown", 0)) or bool(s2_raw.get("is_time_unknown", 0))

    apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])

    lang_val = meta.get("language", chart_data.get("language", seed_data.get("language", seed_data.get("lang", "en"))))
    lang = str(lang_val).lower().strip()
    is_ko = "ko" in lang or "kr" in lang

    principia_res = calculate_principia(
        date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz,
        system=system, ayanamsa=ayanamsa, view=view, h_sys=h_sys,
        fixed_star_orb=fixed_star_orb, is_time_unknown=is_unk
    )

    planets_data = principia_res.get('planets', {})
    lords_data = principia_res.get('lords', {})

    sabian_dict = {}
    for spath in SABIAN_PATHS:
        if os.path.exists(spath):
            try:
                with open(spath, 'r', encoding='utf-8') as f:
                    sabian_dict = json.load(f)
                break
            except: pass
    
    def get_sabian_text(idx):
        if idx is None or str(idx).strip() == "": return ""
        idx_str = str(idx)
        
        def extract_text(entry):
            if isinstance(entry, str): return entry
            if isinstance(entry, dict):
                ko_keys = ["text_ko", "ko", "desc_ko", "ko_KR", "korean"]
                en_keys = ["text_en", "en", "desc_en", "en_US", "english"]
                primary_keys = ko_keys if is_ko else en_keys
                secondary_keys = en_keys if is_ko else ko_keys
                fallback_keys = ["text", "desc", "meaning", "symbol"]
                
                for k in primary_keys + secondary_keys + fallback_keys:
                    if k in entry and entry[k]: return str(entry[k])
                for k, v in entry.items():
                    if k != "index" and isinstance(v, str): return v
            return ""

        if isinstance(sabian_dict, dict):
            if idx_str in sabian_dict: return extract_text(sabian_dict[idx_str])
            for k, v in sabian_dict.items():
                if isinstance(v, dict) and str(v.get("index", "")) == idx_str: return extract_text(v)
        elif isinstance(sabian_dict, list):
            for item in sabian_dict:
                if isinstance(item, dict) and str(item.get("index", "")) == idx_str: return extract_text(item)
        return ""

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    # Day Lord & Hour Lord
    day_val = lords_data.get("day", "")
    day_cell_str = mapping["metadata"]["day_lord"]
    ws[day_cell_str] = day_val
    apply_grimoire_styles(ws[day_cell_str], day_val, skip_color=("|" in str(day_val)))
    
    hour_val = lords_data.get("hour", "")
    hour_cell_str = mapping["metadata"]["hour_lord"]
    ws[hour_cell_str] = hour_val
    apply_grimoire_styles(ws[hour_cell_str], hour_val, skip_color=False)

    ws[mapping["metadata"]["sys_tab"]] = system.upper()
    if system.upper() == "SIDEREAL": ws[mapping["metadata"]["ayanamsa"]] = ayanamsa.upper()
    
    # ─────────────────────────────────────────────────────────
    # 🚀 [시각화 패치]: 하우스 시스템(House Sys) 까만 글씨 복원 로직
    # ─────────────────────────────────────────────────────────
    house_sys_map = {
        "P": "PLACIDUS", "W": "WHOLE SIGN", "K": "KOCH",
        "WHOLE": "WHOLE SIGN", "WHOLE_SIGN": "WHOLE SIGN"
    }
    
    raw_h = str(h_sys).upper().strip()
    if raw_h in house_sys_map:
        h_sys_full = house_sys_map[raw_h].title()
    else:
        h_sys_full = raw_h.replace("_", " ").title()

    h_sys_cell = mapping.get("metadata", {}).get("house_sys", "B7")
    if h_sys_cell:
        ws[h_sys_cell] = h_sys_full
        
        # a2.xlsx 템플릿의 해당 셀에 서식이 없어서 글자가 까맣게 묻히는 것을 방지하기 위해
        # 바로 위 셀(B6: TROPICAL)의 폰트 서식(색상/크기)과 정렬을 강제로 복사해옵니다.
        sys_tab_cell = mapping["metadata"].get("sys_tab", "B6")
        if ws[sys_tab_cell].font:
            ws[h_sys_cell].font = copy(ws[sys_tab_cell].font)
            ws[h_sys_cell].alignment = copy(ws[sys_tab_cell].alignment)
            
        apply_grimoire_styles(ws[h_sys_cell], h_sys_full, skip_color=True)
    # ─────────────────────────────────────────────────────────

    # Orb 값을 문자로 강제 인식시키기
    orb_str = f"{fixed_star_orb:.1f}"
    orb_cell_ref = mapping["metadata"]["fixed_star_orb"]
    orb_cell = ws[orb_cell_ref]
    
    orb_cell.value = orb_str
    orb_cell.data_type = 's'       
    orb_cell.number_format = '@'   
    
    apply_grimoire_styles(orb_cell, orb_str, skip_color=True)

    if system.upper() in ["DRACONIC", "KETUNIC"]:
        for r_idx in range(2, 8):
            for src_col, tgt_col in [('A', 'D'), ('B', 'E')]:
                src_cell = ws[f"{src_col}{r_idx}"]
                tgt_cell = ws[f"{tgt_col}{r_idx}"]
                
                tgt_cell.value = src_cell.value
                tgt_cell.data_type = src_cell.data_type
                tgt_cell.number_format = src_cell.number_format
                
                if src_cell.has_style:
                    tgt_cell.font = copy(src_cell.font)
                    tgt_cell.border = copy(src_cell.border)
                    tgt_cell.fill = copy(src_cell.fill)
                    tgt_cell.alignment = copy(src_cell.alignment)
        
    all_bodies = []
    for category in ["planets", "asteroids", "lilith_nodes", "fates"]: 
        for body_name, map_data in mapping.get(category, {}).items():
            engine_name = body_name
            if engine_name == "North Node (t)": engine_name = "North Node (t)"
            elif engine_name == "North Node (m)": engine_name = "Rahu"
            elif engine_name == "South Node (m)": engine_name = "Ketu"
            elif engine_name == "South Node (t)": engine_name = "South Node (t)"
            
            p_data = planets_data.get(engine_name)
            if p_data and isinstance(p_data, dict): 
                all_bodies.append({
                    "name": body_name,
                    "map": map_data if isinstance(map_data, dict) else {},
                    "data": p_data
                })
    
    all_bodies.sort(key=lambda x: x["map"].get("row_start", 0), reverse=True)

    day_lords_list = [d.strip() for d in str(lords_data.get('day', '-')).split('|')]
    hour_lord = str(lords_data.get('hour', '-')).strip()

    for body in all_bodies:
        b_map = body["map"]
        p_data = body["data"] 
        body_name = body["name"]
        row_idx = b_map.get("row_start", 0)
        
        if not row_idx: continue
        
        # 🚀 빈칸과 테두리 잔해를 남기지 않고 행 자체를 물리적으로 완전 삭제
        if system.upper() in ["DRACONIC", "KETUNIC"] and "Node" in body_name:
            ws.delete_rows(row_idx)
            continue

        # 🚀 컬럼 밀림 방지 및 House 추가
        col_info = get_col(b_map.get("info", "E"))
        col_ruler = get_col(b_map.get("ruler", "F"))
        col_dignity = get_col(b_map.get("dignity", "G"))
        col_house = get_col(b_map.get("house", "H")) # 🚀 추가
        col_duad = get_col(b_map.get("duad", "I"))
        col_dodeca = get_col(b_map.get("dodeca", "J"))
        col_decan = get_col(b_map.get("decan", "K"))
        col_bounds = get_col(b_map.get("bounds", "L"))
        col_sabian = get_col(b_map.get("sabian", "M"))
        
        is_anaretic = p_data.get("is_anaretic", False)
        info_text = p_data.get("dms", "")
        
        is_day_lord = body_name in day_lords_list
        is_hour_lord = body_name == hour_lord
        
        if is_hour_lord: info_text = info_text.upper()
            
        info_cell = ws[f"{col_info}{row_idx}"]
        info_cell.value = info_text

        ws[f"{col_ruler}{row_idx}"] = str(p_data.get("ruler", ""))
        
        dignity_val = str(p_data.get("dignity", "")).strip()
        ws[f"{col_dignity}{row_idx}"] = dignity_val if dignity_val else "None"
        
        # 🚀 House 데이터 기록
        ws[f"{col_house}{row_idx}"] = str(p_data.get("house", ""))
        
        if view == 'zodiac':
            ws[f"{col_duad}{row_idx}"] = str(p_data.get("duad", ""))
            ws[f"{col_dodeca}{row_idx}"] = str(p_data.get("dodeca", ""))
            ws[f"{col_decan}{row_idx}"] = str(p_data.get("decan", ""))
            ws[f"{col_bounds}{row_idx}"] = str(p_data.get("bound", ""))
        else:
            nak_data = p_data.get("nakshatra", {})
            if isinstance(nak_data, dict):
                ws[f"{col_duad}{row_idx}"] = str(nak_data.get("name", ""))
            ws[f"{col_dodeca}{row_idx}"] = str(p_data.get("pada_lord", ""))
        
        s_idx = p_data.get("sabian_index")
        if not s_idx and "longitude" in p_data:
            try: s_idx = int(math.floor(float(p_data["longitude"]))) + 1
            except: pass
            
        s_val_str = str(s_idx).strip() if s_idx else ""
        sabian_text = ""
        
        if s_val_str.isdigit(): sabian_text = get_sabian_text(s_val_str)
        if not sabian_text:
            fallback_txt = p_data.get("sabian", p_data.get("sabian_text", p_data.get("sabian_symbol", "")))
            if isinstance(fallback_txt, dict):
                if is_ko: sabian_text = fallback_txt.get("text_ko", fallback_txt.get("ko", fallback_txt.get("text_en", fallback_txt.get("en", ""))))
                else: sabian_text = fallback_txt.get("text_en", fallback_txt.get("en", fallback_txt.get("text_ko", fallback_txt.get("ko", ""))))
            elif isinstance(fallback_txt, str): sabian_text = fallback_txt

        if not sabian_text:
            frontend_sabian = chart_data.get("bodies", {}).get(body_name, {}).get("sabian", "")
            if frontend_sabian: sabian_text = frontend_sabian

        if not sabian_text and s_val_str.isdigit():
            sabian_text = f"[{s_val_str}°] Symbol rendering fallback"

        ws[f"{col_sabian}{row_idx}"] = sabian_text

        # 🚀 M열까지 스타일링 확장 ("ABCDEFGHIJKL" -> "ABCDEFGHIJKLM")
        for c_char in "ABCDEFGHIJKLM":
            target_cell = ws[f"{c_char}{row_idx}"]
            is_info = (c_char in ["B", col_info]) 
            
            apply_grimoire_styles(
                target_cell, 
                target_cell.value, 
                is_info_col=is_info, 
                skip_color=(c_char in ["D", col_sabian]),
                is_day_lord=is_day_lord,   
                is_hour_lord=is_hour_lord, 
                is_anaretic=is_anaretic    
            )

        # 🚀 하우스 컬럼 중앙 정렬
        ws[f"{col_house}{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")

        # --- Fixed Stars ---
        stars = p_data.get("fixed_stars", [])
        valid_stars = []
        if stars and isinstance(stars, list):
            for s in stars:
                if isinstance(s, dict): valid_stars.append(s)
                elif isinstance(s, str): valid_stars.append({"name": s, "position": "", "orb": ""})
            
        if valid_stars:
            ws[f"A{row_idx}"] = valid_stars[0].get("name", "")
            ws[f"B{row_idx}"] = valid_stars[0].get("position", "")
            orb_val = valid_stars[0].get("orb", "")
            ws[f"C{row_idx}"] = f"{orb_val}°" if orb_val else ""
            
            for c_char in ["A", "B", "C"]:
                target_cell = ws[f"{c_char}{row_idx}"]
                apply_grimoire_styles(target_cell, target_cell.value, is_info_col=(c_char == "B"), skip_color=False)
            
            for i in range(1, len(valid_stars)):
                insert_idx = row_idx + i
                ws.insert_rows(insert_idx)
                
                ws[f"A{insert_idx}"] = valid_stars[i].get("name", "")
                ws[f"B{insert_idx}"] = valid_stars[i].get("position", "")
                orb_val2 = valid_stars[i].get("orb", "")
                ws[f"C{insert_idx}"] = f"{orb_val2}°" if orb_val2 else ""
                
                # 🚀 여기서도 동일하게 M열까지 확장 처리
                for col_chr in "ABCDEFGHIJKLM":
                    src = ws[f"{col_chr}{row_idx}"]
                    tgt = ws[f"{col_chr}{insert_idx}"]
                    
                    if src.has_style:
                        tgt.border = copy(src.border)
                        tgt.alignment = copy(src.alignment)
                        if col_chr == "D" and src.fill:
                            tgt.fill = copy(src.fill)
                        
                    is_info = (col_chr in ["B", col_info])
                    apply_grimoire_styles(tgt, tgt.value, is_info_col=is_info, skip_color=(col_chr == "D"))
        else:
            ws[f"A{row_idx}"] = ""
            ws[f"B{row_idx}"] = ""
            ws[f"C{row_idx}"] = ""

    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name="Consolas", 
                    size=cell.font.size, 
                    bold=cell.font.bold,
                    italic=cell.font.italic, 
                    color=cell.font.color
                )

    if system.upper() in ["DRACONIC", "KETUNIC"]:
        from openpyxl.utils import get_column_letter
        
        merge_to_remove = None
        for m_range in list(ws.merged_cells.ranges):
            if "A1" in m_range.coord:
                merge_to_remove = m_range
                break
                
        if merge_to_remove:
            max_col_letter = get_column_letter(merge_to_remove.max_col)
            ws.unmerge_cells(merge_to_remove.coord) 
            
            ws['D1'].value = ws['A1'].value
            ws['A1'].value = None
            
            if ws['A1'].has_style:
                ws['D1'].font = copy(ws['A1'].font)
                ws['D1'].border = copy(ws['A1'].border)
                ws['D1'].fill = copy(ws['A1'].fill)
                ws['D1'].alignment = copy(ws['A1'].alignment)
                
            ws.merge_cells(f"D1:{max_col_letter}1")

        for col_letter in ['A', 'B', 'C']:
            ws.column_dimensions[col_letter].hidden = True
            ws.column_dimensions[col_letter].width = 0.1

    return wb