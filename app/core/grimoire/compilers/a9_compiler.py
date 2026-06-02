import json
import os
import sys
from datetime import datetime
import openpyxl.utils
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp

from core.astrology.davison import calculate_davison_midpoint
from core.astrology.engine import calculate_n9_timeline, calculate_principia, calculate_arcana, SYMBOL_MAP, _ensure_float_tz

BASE_DIR = os.path.dirname(__file__)

# 🚀 A9 경로로 수정
MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a9_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a9.xlsx'))

def get_nested(d, path, default="-"):
    if not path: return default
    keys = path.split('.')
    val = d
    try:
        for key in keys:
            if isinstance(val, dict): val = val.get(key)
            else: return default
        return val if val is not None else default
    except: return default

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def compile_a9_grimoire(chart_data, seed_data=None):
    try:
        # 1. 템플릿 로드
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb["a9"] if "a9" in wb.sheetnames else wb.active
        ws.title = "a9"
        for s in list(wb.sheetnames):
            if s != "a9": del wb[s]
        wb.active = 0

        meta_req = chart_data.get("metadata", {})

        if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
            raise ValueError("Grimoire requires Albedo seed data.")

        s1_raw = seed_data['seed1'].copy()
        s2_raw = seed_data['seed2'].copy()

        # Davison 시드 생성
        seed_data = calculate_davison_midpoint(s1_raw, s2_raw)

        # Davison Midpoint 좌표 텍스트화
        d_lat = get_safe_float(seed_data.get("lat"), None)
        d_lng = get_safe_float(seed_data.get("lng"), None)
        
        if d_lat is not None and d_lng is not None:
            coord_str = f"{abs(d_lat):.2f}°{'N' if d_lat>=0 else 'S'}, {abs(d_lng):.2f}°{'E' if d_lng>=0 else 'W'}"
        else:
            coord_str = "Unknown Location"
            
        seed_data["city"] = coord_str
        seed_data["location"] = coord_str
        seed_data["city_name"] = coord_str
        seed_data["location_name"] = coord_str

        # 스탬퍼 적용
        apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])

        # 3. 데이터 추출 및 타임존 보정
        lat = get_safe_float(seed_data.get("lat"), 37.5665)
        lng = get_safe_float(seed_data.get("lng"), 126.9780)
        date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
        time_str = str(seed_data.get('birth_time', '12:00:00'))
        
        tz_val = seed_data.get('tz') if seed_data.get('tz') is not None else seed_data.get('timezone', 9.0)
        tz = _ensure_float_tz(tz_val, date_str)

        system = meta_req.get("sys_tab", "tropical")
        ayanamsa = meta_req.get("ayanamsa", "lahiri")
        h_sys_raw = str(meta_req.get("h_sys", "P")).strip().upper()
        h_sys = h_sys_raw[0] if h_sys_raw else "P"
        
        view_mode = meta_req.get("view_mode", "zodiac")
        seg_num = int(meta_req.get("segment", 1))
        is_unk = bool(seed_data.get("is_time_unknown", 0))

        # 4. 엔진 호출 (Davison 시드를 대상으로 N9과 동일하게 연산)
        principia = calculate_principia(
            date_str, time_str, lat, lng, tz, system, ayanamsa, h_sys=h_sys, is_time_unknown=is_unk
        )
        arc_p = calculate_arcana(
            date_str, time_str, lat, lng, timezone=tz, lot_schema='paulus', 
            system=system, ayanamsa=ayanamsa, h_sys=h_sys
        )
        
        p_data = principia.get("planets", {})
        lords_data = principia.get("lords", {}) 
        lots_data = arc_p.get("lots", {})

        n9_result = calculate_n9_timeline(seed_data, mode=view_mode, ayanamsa=ayanamsa)
        n9_timeline = n9_result.get("timeline", [])

        # 타임라인 오늘 날짜 깃발 꽂기
        today_str = datetime.now().strftime("%Y-%m-%d")
        for idx in range(len(n9_timeline)):
            r_date = str(n9_timeline[idx].get('date', '9999-12-31')).split('T')[0]
            next_date = str(n9_timeline[idx+1].get('date', '9999-12-31')).split('T')[0] if idx + 1 < len(n9_timeline) else "9999-12-31"
            if r_date <= today_str < next_date:
                n9_timeline[idx]['is_today'] = True
                break

        # 12년 구간 필터링 적용 (Zodiac 모드 한정)
        if view_mode == 'zodiac':
            start_age = (seg_num - 1) * 12
            end_age = start_age + 12
            filtered_timeline = []
            for r in n9_timeline:
                try: age_val = int(r.get("age", 0))
                except: age_val = 0
                
                if start_age <= age_val < end_age:
                    filtered_timeline.append(r)
            n9_timeline = filtered_timeline

        # 5. 상단 메타데이터 강제 주입 (JSON 무시하고 직접 제어)
        roman_map = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI"}
        duo_roman = roman_map.get(seg_num, "I")

        if view_mode == 'zodiac':
            ws['B5'].value = duo_roman
            ws['B6'].value = str(arc_p.get("meta", {}).get("sect", "-")).upper()
            ws['B7'].value = str(p_data.get("Ascendant", {}).get("dms", "-"))
            ws['B8'].value = str(lots_data.get("Spirit", {}).get("dms", "-"))
            ws['B9'].value = str(lots_data.get("Fortune", {}).get("dms", "-"))
            ws['B10'].value = str(lots_data.get("Eros", {}).get("dms", "-"))

            for coord in ["B5", "B6"]:
                cell = ws[coord]
                apply_grimoire_styles(cell, str(cell.value), is_info_col=False, skip_color=True)
                cell.font = Font(name="Consolas", size=9 if coord != "B6" else 10)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for coord in ["B7", "B8", "B9", "B10"]:
                cell = ws[coord]
                apply_grimoire_styles(cell, str(cell.value), is_info_col=False, skip_color=False)
                cell.font = Font(name="Consolas", size=9)
                cell.alignment = Alignment(horizontal='right', vertical='center')

            planet_rows = {
                "Saturn": 5, "Jupiter": 6, "Mars": 7, "Sun": 8,
                "Venus": 9, "Mercury": 10, "Moon": 11
            }
            
            for p_name, r_idx in planet_rows.items():
                p_info = p_data.get(p_name, {})
                ws[f'E{r_idx}'].value = p_info.get("dms", "-")
                ws[f'G{r_idx}'].value = p_info.get("dignity", "-")
                ws[f'H{r_idx}'].value = "TRUE" if p_name in str(lords_data.get("day", "")) else "-"
                ws[f'I{r_idx}'].value = "TRUE" if lords_data.get("hour") == p_name else "-"
                
                for c_char in ['E', 'G', 'H', 'I']:
                    cell = ws[f'{c_char}{r_idx}']
                    apply_grimoire_styles(cell, str(cell.value), is_info_col=False, skip_color=False)
                    cell.font = Font(name="Consolas", size=9)
                    cell.alignment = Alignment(horizontal='left' if c_char == 'E' else 'center', vertical='center')

        # 6. 타임라인 매핑을 위해 JSON 파일 로드
        with open(MAPPING_FILE, 'r', encoding='utf-8') as f: 
            mapping = json.load(f)

        start_row = 17 
        cols = mapping.get("timeline", {}).get("columns", {})

        for i, row_data in enumerate(n9_timeline):
            r_idx = start_row + i
            ws.row_dimensions[r_idx].height = 16.5
            is_current = row_data.get('is_today', False)

            for col_char, rule in cols.items():
                data_key = rule.get("data_key") if isinstance(rule, dict) else rule
                check_lb = rule.get("check_lb") if isinstance(rule, dict) else None
                
                raw_val = get_nested(row_data, data_key, "-")
                if raw_val == "-" and "_" in data_key:
                    parts = data_key.split("_")
                    if parts[0] in ["spirit", "fortune", "eros"]:
                        raw_val = row_data.get("zr", {}).get(parts[0], {}).get(parts[1], "-")
                
                is_lb = bool(get_nested(row_data, check_lb, False)) if check_lb else False
                if not is_lb and "_" in data_key:
                    parts = data_key.split("_")
                    if parts[0] in ["spirit", "fortune", "eros"]:
                        is_lb = bool(row_data.get("zr", {}).get(parts[0], {}).get(f"{parts[1]}_lb", False))

                val = SYMBOL_MAP.get(str(raw_val), str(raw_val)) if "date" not in data_key and raw_val != "-" else str(raw_val)
                cell = ws[f"{col_char}{r_idx}"]
                cell.value = val

                apply_grimoire_styles(cell, str(raw_val), is_info_col=False, skip_color=(col_char == 'A'))
                cell.font = Font(name="Consolas", size=10, color="FF0000" if is_lb else "000000", bold=is_lb)
                cell.alignment = Alignment(horizontal='left' if col_char == 'A' else 'center', vertical='center')

                if is_current and col_char == 'A':
                    cell.fill = PatternFill(start_color="B0ECFF", end_color="B0ECFF", fill_type="solid")

        # 7. 틀 고정 및 사후 병합(Post-Merge) 로직
        ws.sheet_view.pane = None
        ws.freeze_panes = 'B17'
        
        # E열과 F열만 병합, G, H, I는 개별 셀 보존
        ranges_to_merge = [
            "B5:C5", "B6:C6", "B7:C7", "B8:C8", "B9:C9", "B10:C10",
            "E5:F5", "E6:F6", "E7:F7", "E8:F8", "E9:F9", "E10:F10", "E11:F11"
        ]
        
        for m_range in ranges_to_merge:
            try:
                ws.merge_cells(m_range)
            except Exception:
                pass 

        return wb

    except Exception as ex:
        print(f">>> [A9 FINAL REFORGE ERROR] {ex}", file=sys.stderr)
        raise ex