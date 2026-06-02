import json
import os
import math
import re
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from core.grimoire.styler import apply_grimoire_styles, HOUSE_COLORS
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz, TROPICAL_SIGNS, calculate_all_star_positions
from core.astrology.davison import calculate_davison_midpoint
from core.astrology.composite import calculate_composite_chart, apply_anti_composite
from core.astrology.arabic_lots import calculate_arabic_lots
from core.astrology.constants import ASTEROIDS
import swisseph as swe

BASE_DIR = os.path.dirname(__file__)
MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/c1_en_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/citrinitas/c1_en.xlsx'))

HOUSE_SYS_NAMES = {
    'P': 'Placidus', 'W': 'Whole House', 'K': 'Koch', 'R': 'Regiomontanus', 
    'C': 'Campanus', 'E': 'Equal', 'O': 'Porphyry', 'V': 'Vehlow'
}

C1_COLORS = [
    "FF2602", "0AFFEF", "99FF05", "E814FF", "0ACC05", "FF0366", "FF9104", "0755FF", "FDFF08", "B76EFF",
    "544DA1", "68D4FF", "992314", "FDC1FF", "FCB8D6", "D8FCB6", "FCF7D6", "A8FFF3", "04FF6D", "FCA992",
    "996635", "B2F0AB", "4A3DB0", "EDEDED", "68FFE0", "FF64C5", "39B570", "E1FDFF", "F5A831", "C49D76",
    "B5478C", "AB88FF", "288559", "817F26", "1C8EFF", "F4EEFF", "CC4366", "BAD6D6", "795FA3", "3D83D6",
    "A33387", "5D6E24", "E8B919", "4C616E", "B1BAB5", "A8D6B3", "476DE0", "C941FF", "A6FF94", "3BD0D4"
]

# 🚀 [수정]: A8과 동일하게 소행성, 교점, 릴리트, 운명, 아라빅 랏 전체 허용!
ALLOWED_UPPER = [
    "SUN", "MOON", "MERCURY", "VENUS", "MARS", "JUPITER", "SATURN", "URANUS", "NEPTUNE", "PLUTO",
    "CHIRON", "CERES", "JUNO", "PALLAS", "VESTA", "ASTEROID EROS", "PSYCHE",
    "ASCENDANT", "MIDHEAVEN", "DESCENDANT", "IMMUM COELI",
    "NORTH NODE (T)", "SOUTH NODE (T)", "NORTH NODE (M)", "SOUTH NODE (M)", "RAHU", "KETU",
    "MEAN LILITH", "TRUE LILITH", "ASTEROID LILITH",
    "MOIRA", "KLOTHO", "LACHESIS", "ATROPOS",
    "FORTUNE", "SPIRIT", "NECESSITY", "NECESSITY (V)", "EROS", "EROS (V)", "COURAGE", "VICTORY", "NEMESIS",
    "VERTEX", "SYZYGY"
]

STANDARD_BODIES = ["Sun", "Moon", "Mercury", "Venus", "Mars", "Jupiter", "Saturn", "Uranus", "Neptune", "Pluto", 
                   "Chiron", "Ceres", "Pallas", "Juno", "Vesta", "Eros", "Psyche", "Moira", "Klotho", "Lachesis", "Atropos", "Lilith"]

# ====================================================================
# 🚀 1. 핵심 연산 도구들 (오류 박멸 방어막 적용)
# ====================================================================
def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def get_date_time_coords(seed):
    # 엔진과 프론트엔드의 이름 불일치를 모두 방어합니다 (date, birthDate 전부 대응)
    d = str(seed.get('birth_date') or seed.get('birthDate') or seed.get('date') or '2000-01-01').split('T')[0]
    t = str(seed.get('birth_time') or seed.get('birthTime') or seed.get('time') or '12:00:00').strip()
    if t.lower() in ["unknown", "none", ""]: t = "12:00:00"
    elif len(t) == 5: t += ":00"

    coords = seed.get("coordinates", {})
    if isinstance(coords, dict) and "lat" in coords:
        lat = get_safe_float(coords.get("lat"), 37.5665)
        lng = get_safe_float(coords.get("lng"), 126.9780)
    else:
        lat = get_safe_float(seed.get("lat"), 37.5665)
        lng = get_safe_float(seed.get("lng"), 126.9780)
        
    if lat == 0.0 and lng == 0.0: lat, lng = 37.5665, 126.9780
    tz = _ensure_float_tz(seed.get('timezone', seed.get('tz', 9.0)), d)
    return d, t, lat, lng, tz

# 🚀 [핵심]: 씨앗 데이터를 계산기에 넣기 전, 무조건 표준 포맷으로 세탁합니다!
def normalize_seed(s):
    if not s: s = {}
    d, t, lat, lng, tz = get_date_time_coords(s)
    return {
        "birth_date": d,
        "birth_time": t,
        "lat": lat,
        "lng": lng,
        "timezone": tz,
        "location": s.get("location", s.get("city", "Unknown Location")),
        "is_time_unknown": s.get("is_time_unknown", 0)
    }

def get_res(seed, sys_type, ayan, h_sys):
    d, t, lat, lng, tz = get_date_time_coords(seed)
    return calculate_principia(d, t, lat, lng, tz, sys_type, ayan, h_sys=h_sys)

def get_arc(seed, sys_type, ayan, h_sys, lot_schema='traditional'):
    d, t, lat, lng, tz = get_date_time_coords(seed)
    return calculate_arcana(d, t, lat, lng, tz, lot_schema=lot_schema, system=sys_type, ayanamsa=ayan, h_sys=h_sys)

# 🚀 [추가됨]: Paulus + Valens + Vertex 통합 추출기
def get_hermetic_lots(seed, sys_type, ayan, h_sys):
    d, t, lat, lng, tz = get_date_time_coords(seed)
    try: arc_p = calculate_arcana(d, t, lat, lng, tz, lot_schema='paulus', system=sys_type, ayanamsa=ayan, h_sys=h_sys)
    except: arc_p = {}
    try: arc_v = calculate_arcana(d, t, lat, lng, tz, lot_schema='valens', system=sys_type, ayanamsa=ayan, h_sys=h_sys)
    except: arc_v = {}
    
    res = {}
    if 'lots' in arc_p:
        for k, v in arc_p['lots'].items(): res[k] = {"longitude": v.get('value', v.get('longitude', 0.0))}
    if 'lots' in arc_v:
        for k in ['Eros', 'Necessity']:
            if k in arc_v['lots']: res[f"{k} (v)"] = {"longitude": arc_v['lots'][k].get('value', arc_v['lots'][k].get('longitude', 0.0))}
    if 'vertex' in arc_p:
        for k, v in arc_p['vertex'].items(): res[k] = {"longitude": v.get('value', v.get('longitude', 0.0))}
    if 'syzygy' in arc_p and arc_p['syzygy'].get('data'):
        res['Syzygy'] = {"longitude": arc_p['syzygy']['data'].get('value', arc_p['syzygy']['data'].get('longitude', 0.0))}
    return res

def get_jd(seed):
    d, t, lat, lng, tz = get_date_time_coords(seed)
    try: dt_obj = datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M:%S")
    except: dt_obj = datetime.strptime(f"{d} 12:00:00", "%Y-%m-%d %H:%M:%S")
    return swe.julday(dt_obj.year, dt_obj.month, dt_obj.day, dt_obj.hour + dt_obj.minute/60.0 - tz)

def create_house_bg_map(cusps_dict):
    bg_map = {deg: "H1" for deg in range(360)}
    if not cusps_dict: return bg_map
    sorted_cusps = []
    for h, lon in cusps_dict.items():
        lon_val = float(lon['longitude'] if isinstance(lon, dict) else lon)
        sabian_idx = int(math.floor(lon_val)) % 360
        sorted_cusps.append((sabian_idx, h))
    sorted_cusps.sort(key=lambda x: x[0])
    count = len(sorted_cusps)
    for i in range(count):
        curr_start, curr_h = sorted_cusps[i]
        next_start, _ = sorted_cusps[(i + 1) % count]
        if curr_start < next_start:
            for idx in range(curr_start, next_start): bg_map[idx] = f"H{curr_h}"
        else:
            for idx in range(curr_start, 360): bg_map[idx] = f"H{curr_h}"
            for idx in range(0, next_start): bg_map[idx] = f"H{curr_h}"
    return bg_map

def get_lords(res):
    dl_raw = str(res.get('lords', {}).get('day', '-'))
    hl_raw = str(res.get('lords', {}).get('hour', '-'))
    return [d.strip() for d in dl_raw.split('|') if d.strip() != '-'], [h.strip() for h in hl_raw.split('|') if h.strip() != '-']

def get_all_bodies(res_obj):
    merged = {}
    for g in ['planets', 'asteroids', 'lilith_nodes', 'angles']:
        for k, v in res_obj.get(g, {}).items():
            d_name = k.replace(" (Natal)", "")
            if d_name == "Eros": d_name = "Asteroid Eros"
            merged[d_name] = v
    return merged

def calc_minor_asteroids(jd):
    swe.set_sid_mode(0, 0, 0)
    res_dict = {}
    for ast_name, ast_num in ASTEROIDS.items():
        if ast_name in STANDARD_BODIES: continue
        try:
            res = swe.calc_ut(jd, ast_num + 10000, swe.FLG_SWIEPH | swe.FLG_SPEED)
            res_dict[ast_name] = {"longitude": res[0][0]}
        except: pass
    return res_dict

def calc_arabic_lots_full(res_trop, arc_trop, arabic_ruler):
    points = {k: v['longitude'] for k, v in res_trop.get('planets', {}).items()}
    points.update({k: v['longitude'] for k, v in res_trop.get('angles', {}).items()})
    house_cusps = {i: res_trop.get('houses', {}).get(i, {}).get('longitude', 0.0) for i in range(1, 13)}
    
    rulers_trad = ["Mars", "Venus", "Mercury", "Moon", "Sun", "Mercury", "Venus", "Mars", "Jupiter", "Saturn", "Saturn", "Jupiter"]
    rulers_mod = ["Mars", "Venus", "Mercury", "Moon", "Sun", "Mercury", "Venus", "Pluto", "Jupiter", "Saturn", "Uranus", "Neptune"]
    r_list = rulers_trad if arabic_ruler.lower() == 'traditional' else rulers_mod
    rulers = {i: r_list[int(house_cusps[i] / 30) % 12] for i in range(1, 13)}
    
    hour_lord_name = str(res_trop.get('lords', {}).get('hour', 'Sun')).split('|')[0].strip()
    lord_of_hour_lon = points.get(hour_lord_name, 0.0)
    
    p_new_moon, p_full_moon = 0.0, 0.0
    if 'syzygy' in arc_trop and 'data' in arc_trop['syzygy']:
        sz = arc_trop['syzygy']['data']
        if sz.get('phase') == 'New Moon': p_new_moon = sz.get('value', 0.0)
        else: p_full_moon = sz.get('value', 0.0)
        
    fortune_lon = arc_trop.get('lots', {}).get('Fortune', {}).get('value', 0.0)
    spirit_lon = arc_trop.get('lots', {}).get('Spirit', {}).get('value', 0.0)
    
    lots = calculate_arabic_lots(points, house_cusps, rulers, lord_of_hour_lon, p_new_moon, p_full_moon, fortune_lon, spirit_lon)
    return {k: {"longitude": v} for k, v in lots.items()}

# ====================================================================
# 🚀 2. 동적 매트릭스 엔진
# ====================================================================
def compile_c1_en_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE): raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")
    
    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: mapping = json.load(f)
    row_map = mapping.get("matrix_layout", {}).get("rows", {})
    
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active
    ws.title = "c1"
    
    meta = chart_data.get("metadata", {})
    ayanamsa = meta.get("ayanamsa", "lahiri")
    h_sys = meta.get("h_sys", "P")
    arabic_ruler = meta.get("arabic_ruler", "traditional")
    
    ws["B4"].value = ayanamsa.capitalize()
    ws["B5"].value = HOUSE_SYS_NAMES.get(h_sys, "Placidus")

    c1_data = meta.get("c1_data", [])
    c1_config = meta.get("c1_config", {})
    
    columns_to_render = []
    color_idx = 0
    start_col_idx = 2 
    
    def fmt_stamp(s):
        d, t, lat, lng, tz = get_date_time_coords(s)
        loc = s.get("location", s.get("city"))
        if not loc or loc == "Unknown Location":
            loc = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
        time_val = "Unknown" if s.get("is_time_unknown") else t
        return f"{d}, {time_val};\n{loc}"

    # 2-1. 렌더링 큐 조립 및 스탬퍼 세팅
    raw_columns = []
    sort_mode = meta.get("sort_mode", "entity") # 프론트에서 보낸 정렬 모드
    
    # 🚀 [핵심]: 컬럼의 절대 정렬 위계 (Level, Order) 정의
    SUB_WEIGHTS = {
        'asteroids': (1, 1), 'davi_ast': (1, 2),
        'tropical': (2, 1), 'comp_main': (2, 2), 'comp_anti': (2, 3), 'davi_tro': (2, 4),
        'sidereal': (3, 1), 'davi_sid': (3, 2),
        'draconic': (4, 1), 'davi_dra': (4, 2),
        'ketunic': (5, 1), 'davi_ket': (5, 2),
        'arabic': (6, 1), 'davi_lot': (6, 2)
    }

    for e_idx, entity in enumerate(c1_data):
        e_id = entity.get("id")
        conf = c1_config.get(e_id, {})
        if not conf.get("active"): continue
        
        ent_color = C1_COLORS[color_idx % 50]
        color_idx += 1
        
        # 🚀 [추가]: 사용자가 요청한 깔끔한 Suffix 명단
        SUFFIX_MAP = {
            'asteroids': 'AST', 'tropical': 'TRO', 'sidereal': 'SID',
            'draconic': 'DRA', 'ketunic': 'KET', 'arabic': 'LOT',
            'comp_main': 'COMP', 'comp_anti': 'ANTI',
            'davi_ast': 'AST', 'davi_tro': 'TRO', 'davi_sid': 'SID',
            'davi_dra': 'DRA', 'davi_ket': 'KET', 'davi_lot': 'LOT'
        }

        for sub in conf.get("subs", []):
            is_conj = (entity.get("type") == "conjunction")
            
            # 🚀 [수정]: 3글자 자르기 대신 매핑된 깔끔한 꼬리표 사용
            mapped_suffix = SUFFIX_MAP.get(sub, sub.upper()[:4])
            
            col_info = {
                "e_idx": e_idx,
                "entity": entity,
                "sub": sub,
                "name": f"{entity.get('name', 'N/A')}_{mapped_suffix}",
                "color_hex": ent_color
            }
            
            if is_conj:
                s1 = normalize_seed(entity.get("p1", entity.get("seed1", {})))
                s2 = normalize_seed(entity.get("p2", entity.get("seed2", {})))
                if "davi" in sub:
                    davi = calculate_davison_midpoint(s1, s2)
                    col_info["is_conj"] = False
                    col_info["stamp"] = fmt_stamp(davi)
                else:
                    col_info["is_conj"] = True
                    col_info["stamp1"] = fmt_stamp(s1)
                    col_info["stamp2"] = fmt_stamp(s2)
            else:
                norm_ent = normalize_seed(entity)
                col_info["is_conj"] = False
                col_info["stamp"] = fmt_stamp(norm_ent)
                
            raw_columns.append(col_info)

    # 🚀 [정렬 실행]: 모드에 따라 위계 기반의 안정 정렬 수행
    def get_sort_key(col):
        level, order = SUB_WEIGHTS.get(col["sub"], (99, 99))
        if sort_mode == "system":
            return (level, col["e_idx"], order) # 1순위: 체계, 2순위: 시드, 3순위: 디테일
        return (col["e_idx"], level, order)     # 1순위: 시드, 2순위: 체계, 3순위: 디테일

    raw_columns.sort(key=get_sort_key)

    # 정렬된 순서대로 A, B, C... 열 문자를 부여
    for i, col in enumerate(raw_columns):
        col["letter"] = get_column_letter(start_col_idx + i)
        columns_to_render.append(col)

    matrix = {col["letter"]: {"items": {d: [] for d in range(360)}, "bg": {d: None for d in range(360)}} for col in columns_to_render}

    matrix["BB"] = {"items": {d: [] for d in range(360)}, "bg": {d: None for d in range(360)}}
    # 🚀 항성 위치 계산을 위한 기준일도 p1에서 추출
    jd_base = get_jd(normalize_seed(c1_data[0].get("p1", c1_data[0].get("seed1", c1_data[0])))) if c1_data else swe.julday(2000, 1, 1, 12.0)
    stars = calculate_all_star_positions(jd_base, 'tropical', swe.FLG_SWIEPH | swe.FLG_SPEED)
    
    def add_to_matrix(col_let, item_name, longitude, is_day=False, is_hour=False):
        if longitude is None: return
        deg_abs = int(math.floor(float(longitude))) % 360
        up_name = str(item_name).upper()
        if up_name in ["MEAN NODE", "NORTH NODE (M)"]: final_name = "Rahu" if "Rahu" not in str(item_name) else item_name
        elif up_name in ["SOUTH NODE", "SOUTH NODE (M)"]: final_name = "Ketu" if "Ketu" not in str(item_name) else item_name
        elif "CUSP" in up_name:
            match = re.search(r'\d+', up_name)
            final_name = f"{match.group()}h cusp" if match else up_name
        else: final_name = str(item_name)
        matrix[col_let]["items"][deg_abs].append({"name": final_name, "lon": float(longitude), "is_day": is_day, "is_hour": is_hour})

    for s in stars: add_to_matrix("BB", s['name'], s['lon'])

    def fill_col_matrix(col_let, data_dict, allowed=None, filter_nodes=False, day_lords=None, hour_lords=None):
        if day_lords is None: day_lords = []
        if hour_lords is None: hour_lords = []
        for p_name, p_data in data_dict.items():
            up_name = p_name.upper()
            if allowed and up_name not in allowed and "CUSP" not in up_name: continue
            if filter_nodes and up_name in ['NORTH NODE (T)', 'SOUTH NODE (T)', 'RAHU', 'KETU', 'MEAN LILITH', 'TRUE LILITH', 'ASTEROID LILITH']: continue
            lon = p_data if isinstance(p_data, (int, float)) else p_data.get('longitude')
            
            is_day = p_name in day_lords or up_name in [d.upper() for d in day_lords]
            is_hour = p_name in hour_lords or up_name in [h.upper() for h in hour_lords]
            add_to_matrix(col_let, p_name, lon, is_day, is_hour)

    def calculate_and_fill(let, seed_obj, sub_type):
        if sub_type in ["tropical", "davi_tro"]:
            res_t = get_res(seed_obj, "tropical", ayanamsa, h_sys)
            bg_map = create_house_bg_map(res_t.get('houses', {}))
            for d in range(360): matrix[let]["bg"][d] = bg_map[d]
            dl, hl = get_lords(res_t)
            merged = get_all_bodies(res_t)
            
            hlots = get_hermetic_lots(seed_obj, "tropical", ayanamsa, h_sys)
            merged.update(hlots)
            
            fill_col_matrix(let, merged, allowed=ALLOWED_UPPER, day_lords=dl, hour_lords=hl)
            
            # 🚀 [수정]: 0.0도(Aries 0) 증발 방지
            for i in range(1, 13):
                if h_sys != 'W' and i in [1, 4, 7, 10]: continue
                h_lon = res_t.get('houses', {}).get(i, {}).get('longitude')
                if h_lon is not None: add_to_matrix(let, f"{i}h cusp", h_lon)
                
        elif sub_type in ["draconic", "davi_dra"]:
            res_d = get_res(seed_obj, "draconic", ayanamsa, h_sys)
            bg_map = create_house_bg_map(res_d.get('houses', {}))
            for d in range(360): matrix[let]["bg"][d] = bg_map[d]
            dl, hl = get_lords(res_d) 
            fill_col_matrix(let, get_all_bodies(res_d), allowed=ALLOWED_UPPER, filter_nodes=True, day_lords=dl, hour_lords=hl)
            
            # 🚀 [수정]: 누락되었던 Draconic 커스프 루프 추가 및 0.0도 방어
            for i in range(1, 13):
                if h_sys != 'W' and i in [1, 4, 7, 10]: continue
                h_lon = res_d.get('houses', {}).get(i, {}).get('longitude')
                if h_lon is not None: add_to_matrix(let, f"{i}h cusp", h_lon)
                
        elif sub_type in ["sidereal", "davi_sid"]:
            res_s = get_res(seed_obj, "sidereal", ayanamsa, h_sys)
            bg_map = create_house_bg_map(res_s.get('houses', {}))
            for d in range(360): matrix[let]["bg"][d] = bg_map[d]
            dl, hl = get_lords(res_s) 
            fill_col_matrix(let, get_all_bodies(res_s), allowed=ALLOWED_UPPER, day_lords=dl, hour_lords=hl)
            
            # 🚀 [수정]: 0.0도(Aries 0) 증발 방지
            for i in range(1, 13):
                if h_sys != 'W' and i in [1, 4, 7, 10]: continue
                h_lon = res_s.get('houses', {}).get(i, {}).get('longitude')
                if h_lon is not None: add_to_matrix(let, f"{i}h cusp", h_lon)
                
        elif sub_type in ["ketunic", "davi_ket"]:
            res_k = get_res(seed_obj, "ketunic", ayanamsa, h_sys)
            bg_map = create_house_bg_map(res_k.get('houses', {}))
            for d in range(360): matrix[let]["bg"][d] = bg_map[d]
            dl, hl = get_lords(res_k) 
            fill_col_matrix(let, get_all_bodies(res_k), allowed=ALLOWED_UPPER, filter_nodes=True, day_lords=dl, hour_lords=hl)
            
            # 🚀 [수정]: 누락되었던 Ketunic 커스프 루프 추가 및 0.0도 방어
            for i in range(1, 13):
                if h_sys != 'W' and i in [1, 4, 7, 10]: continue
                h_lon = res_k.get('houses', {}).get(i, {}).get('longitude')
                if h_lon is not None: add_to_matrix(let, f"{i}h cusp", h_lon)
                
        elif sub_type in ["asteroids", "davi_ast"]:
            jd = get_jd(seed_obj)
            ast = calc_minor_asteroids(jd)
            fill_col_matrix(let, ast, allowed=None)
            
        elif sub_type in ["arabic", "davi_lot"]:
            res_t = get_res(seed_obj, "tropical", ayanamsa, h_sys)
            arc_t = get_arc(seed_obj, "tropical", ayanamsa, h_sys, arabic_ruler)
            lots = calc_arabic_lots_full(res_t, arc_t, arabic_ruler)
            fill_col_matrix(let, lots, allowed=None)

    for col in columns_to_render:
        let, ent, sub = col["letter"], col["entity"], col["sub"]
        
        # 🚀 p1, p2가 있을 때만 합성차트 연산 실행
        if ent.get("type") == "conjunction" and ("p1" in ent or "seed1" in ent):
            s1 = normalize_seed(ent.get("p1", ent.get("seed1", {})))
            s2 = normalize_seed(ent.get("p2", ent.get("seed2", {})))
            
            if sub in ["comp_main", "comp_anti"]:
                res_a = get_res(s1, "tropical", ayanamsa, h_sys)
                res_b = get_res(s2, "tropical", ayanamsa, h_sys)
                comp_full = calculate_composite_chart(get_all_bodies(res_a), get_all_bodies(res_b), 
                                                      {int(k): float(v['longitude']) for k, v in res_a.get('houses', {}).items()},
                                                      {int(k): float(v['longitude']) for k, v in res_b.get('houses', {}).items()})
                c_main = comp_full.get('planets', {})
                c_houses = comp_full.get('houses', {})
                for idx, name in [(1, 'Ascendant'), (4, 'Immum Coeli'), (7, 'Descendant'), (10, 'Midheaven')]:
                    if idx in c_houses: c_main[name] = {'longitude': float(c_houses[idx])}
                
                c_anti, c_anti_houses = apply_anti_composite(c_main, c_houses)
                target_data = c_main if sub == "comp_main" else c_anti
                target_houses = c_houses if sub == "comp_main" else c_anti_houses
                
                bg_map = create_house_bg_map(target_houses)
                for d in range(360): matrix[let]["bg"][d] = bg_map[d]
                fill_col_matrix(let, target_data, allowed=ALLOWED_UPPER)
                for i in range(1, 13):
                    # 🚀 [수정]: Composite에서는 하우스 시스템(W) 상관없이 무조건 1, 4, 7, 10 생략!
                    # (위에서 이미 Ascendant, MC 등으로 덮어씌웠기 때문에 중복 출력을 차단합니다.)
                    if i in [1, 4, 7, 10]: continue 
                    
                    if i in target_houses: add_to_matrix(let, f"{i}h cusp", target_houses[i])
            elif "davi" in sub:
                davison = calculate_davison_midpoint(s1, s2)
                calculate_and_fill(let, normalize_seed(davison), sub)
        else:
            calculate_and_fill(let, normalize_seed(ent), sub)

    all_render_cols = columns_to_render + [{"letter": "BB", "is_bb": True}]

    def get_row_for_lon(lon_val):
        sign_idx, deg_in_sign = int((float(lon_val) % 360) // 30), int((float(lon_val) % 360) % 30)
        return row_map.get(f"{TROPICAL_SIGNS[sign_idx]} {deg_in_sign + 1}")

    for deg_abs in range(359, -1, -1):
        r_idx = get_row_for_lon(deg_abs)
        if not r_idx: continue

        max_lines = 1
        for col in all_render_cols:
            let = col["letter"]
            raw_items = matrix[let]["items"][deg_abs]
            raw_items.sort(key=lambda x: (float(x["lon"]), 0 if "CUSP" in str(x["name"]).upper() else 1))
            
            unique_items, seen = [], set()
            for item in raw_items:
                if item["name"] not in seen:
                    unique_items.append(item)
                    seen.add(item["name"])
            matrix[let]["items"][deg_abs] = unique_items
            if len(unique_items) > max_lines: max_lines = len(unique_items)

        if max_lines > 1:
            amount = max_lines - 1
            ins_idx = r_idx + 1
            old_heights = {r: ws.row_dimensions[r].height for r in list(ws.row_dimensions.keys())}

            ws.insert_rows(ins_idx, amount=amount)

            for r in range(ws.max_row, 0, -1):
                if r >= ins_idx + amount: ws.row_dimensions[r].height = old_heights.get(r - amount, 15)
                elif ins_idx <= r < ins_idx + amount: ws.row_dimensions[r].height = old_heights.get(r_idx, 15)
                else: ws.row_dimensions[r].height = old_heights.get(r, 15)
                
            for r_offset in range(1, max_lines):
                new_r = r_idx + r_offset
                for c_idx in range(1, 55): 
                    src = ws.cell(row=r_idx, column=c_idx)
                    tgt = ws.cell(row=new_r, column=c_idx)
                    if src.has_style:
                        tgt.font = copy(src.font); tgt.border = copy(src.border); tgt.fill = copy(src.fill)
                        tgt.number_format = copy(src.number_format); tgt.alignment = copy(src.alignment)
                    tgt.value = ""

        for col in all_render_cols:
            let = col["letter"]
            is_bb = col.get("is_bb", False)
            sub = col.get("sub", "")
            
            # 🚀 [수정]: 배경색을 제거할 보조 열 (소행성, 랏, 항성 모두 포함)
            is_minor_bg = is_bb or any(x in sub.lower() for x in ["ast", "lot", "arabic", "fixed"])
            
            # 🚀 [추가]: 글자색을 검정색으로 강제할 열 (소행성과 항성만! 아라빅 랏은 제외)
            force_black_text = is_bb or any(x in sub.lower() for x in ["ast", "fixed"])
            
            bg_hex = None if is_minor_bg else HOUSE_COLORS.get(matrix[let]["bg"][deg_abs])
            
            items = matrix[let]["items"][deg_abs]

            for line_idx in range(max_lines):
                cell = ws[f"{let}{r_idx + line_idx}"]
                if line_idx < len(items):
                    item = items[line_idx]
                    name = item["name"]
                    val_upper = name.upper()
                    
                    if not is_bb:
                        is_angle = val_upper in ["ASCENDANT", "MIDHEAVEN", "DESCENDANT", "IMMUM COELI"]
                        if item.get("is_hour") or is_angle: name = val_upper
                    
                    cell.value = name
                    apply_grimoire_styles(
                        cell, name, 
                        is_day_lord=item.get("is_day", False), 
                        is_hour_lord=item.get("is_hour", False), 
                        tabula_mode=True, 
                        house_bg=bg_hex,
                        skip_color=is_minor_bg, # 배경색 스킵 여부
                        skip_planet_color=is_minor_bg # 행성색 스킵 여부
                    )
                    
                    # 🚀 [수정]: 소행성/항성 열(force_black_text)일 때만 글자색을 검정으로 강제
                    # 이렇게 하면 Arabic Lots 열은 styler가 정해준 색상이 그대로 유지됩니다!
                    if force_black_text:
                        old_f = cell.font
                        cell.font = Font(
                            name=old_f.name, size=old_f.size, 
                            bold=old_f.bold, italic=old_f.italic, 
                            color="000000"
                        )
                else:
                    cell.value = ""
                    apply_grimoire_styles(cell, "", tabula_mode=True, house_bg=bg_hex, skip_color=is_minor_bg, skip_planet_color=is_minor_bg)
                
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # 3. 헤더 및 이중 스탬퍼 
    ws.row_dimensions[6].height = 22 
    ws.row_dimensions[7].height = 22 
    
    for col in columns_to_render:
        let = col["letter"]
        sub = col.get("sub", "")
        
        # 🚀 컬럼 폭 제어 (Arabic Lot=30 / Others=17)
        if "arabic" in sub or "lot" in sub:
            ws.column_dimensions[let].width = 30.0
        else:
            ws.column_dimensions[let].width = 17.0
        
        if col.get("is_conj"):
            ws[f"{let}6"].value = col["stamp1"]
            ws[f"{let}6"].font = Font(name="Consolas", size=7, color="555555")
            ws[f"{let}6"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws[f"{let}7"].value = col["stamp2"]
            ws[f"{let}7"].font = Font(name="Consolas", size=7, color="555555")
            ws[f"{let}7"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            ws[f"{let}7"].value = col["stamp"]
            ws[f"{let}7"].font = Font(name="Consolas", size=7, color="555555")
            ws[f"{let}7"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        header_cell = ws[f"{let}8"]
        header_cell.value = col["name"]
        header_cell.font = Font(name="Consolas", size=10, bold=True, color="000000")
        header_cell.fill = PatternFill(start_color=col["color_hex"], fill_type="solid")
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

    last_col_idx = start_col_idx + len(columns_to_render) - 1
    for empty_idx in range(last_col_idx + 1, 53):
        ws.column_dimensions[get_column_letter(empty_idx)].width = 0.1

    merge_to_remove = None
    for m_range in list(ws.merged_cells.ranges):
        if "A1" in m_range.coord:
            merge_to_remove = m_range
            break
            
    if merge_to_remove:
        ws.unmerge_cells(merge_to_remove.coord)
        ws['B1'].value = ws['A1'].value
        ws['A1'].value = None

    # 🚀 [수정]: 템플릿 기본 셀을 포함해 값이 존재하는 모든 셀에 Consolas 강제 주입
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=54):
        for cell in row:
            if cell.value is not None: # 값이 하나라도 있으면 무조건 폰트 교체
                old = cell.font
                cell.font = Font(
                    name="Consolas", 
                    size=old.size if old and old.size else 10,
                    bold=old.bold if old else False,
                    italic=old.italic if old else False,
                    color=old.color if old else "000000"
                )

    ws.freeze_panes = "B9" 
    return wb