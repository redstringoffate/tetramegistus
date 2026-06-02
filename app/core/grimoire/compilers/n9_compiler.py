import json
import os
import sys
from datetime import datetime
import openpyxl.utils
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
# 🚀 N4 벤치마킹: calculate_arcana 추가 로드
from core.astrology.engine import calculate_n9_timeline, calculate_principia, calculate_arcana, SYMBOL_MAP, _ensure_float_tz

BASE_DIR = os.path.dirname(__file__)
MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/n9_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/nigredo/n9.xlsx'))

def get_nested(d, path, default="-"):
    if not path: return default
    keys = path.split('.')
    val = d
    try:
        for key in keys:
            if isinstance(val, dict): val = val.get(key)
            else: return default
        return val if val is not None else default
    except: return default

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def ensure_location_string(seed):
    loc_val = str(seed.get("location", seed.get("city", ""))).strip()
    if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
        lat, lng = get_safe_float(seed.get("lat"), None), get_safe_float(seed.get("lng"), None)
        if lat is not None and lng is not None:
            seed["city"] = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
            seed["location"] = seed["city"]
        else:
            seed["city"] = "Unknown Location"
            seed["location"] = "Unknown Location"
    else:
        seed["city"] = loc_val
        seed["location"] = loc_val

# 🚀 [FIX 1: 매 셀마다 작동하는 Merge Cell 철통 방어 헬퍼 함수]
def write_to_merge_safe(ws, coord_str, val, align_h, font_size):
    try:
        top_left = coord_str.split(':')[0]
        r, c = openpyxl.utils.coordinate_to_tuple(top_left)
        target_cell = ws[top_left]
        
        # 워크시트의 모든 병합 셀 범위를 뒤져서, 현재 좌표가 병합셀 내부인지 검사
        for m_range in ws.merged_cells.ranges:
            if m_range.min_row <= r <= m_range.max_row and m_range.min_col <= c <= m_range.max_col:
                # 병합 셀의 진짜 주인이 되는 좌측 상단 좌표로 리타겟팅
                true_top_left = ws.cell(row=m_range.min_row, column=m_range.min_col).coordinate
                target_cell = ws[true_top_left]
                break
                
        target_cell.value = val
        apply_grimoire_styles(target_cell, str(val), is_info_col=False, skip_color=True)
        
        target_cell.font = Font(name="Consolas", size=font_size)
        target_cell.alignment = Alignment(horizontal=align_h, vertical='center')
    except Exception as e:
        pass

def compile_n9_grimoire(chart_data, seed_data=None):
    try:
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb["n9"] if "n9" in wb.sheetnames else wb.active
        ws.title = "n9"
        for s in list(wb.sheetnames):
            if s != "n9": del wb[s]
        wb.active = 0

        seed_data = chart_data.get("seed", chart_data)
        meta_req = chart_data.get("metadata", {})
        
        ensure_location_string(seed_data)
        apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])

        lat = get_safe_float(seed_data.get("lat"), 37.5665)
        lng = get_safe_float(seed_data.get("lng"), 126.9780)
        date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
        time_str = str(seed_data.get('birth_time', '12:00:00'))
        
        tz_val = seed_data.get('timezone', 9.0)
        tz = _ensure_float_tz(tz_val, date_str)

        system = meta_req.get("sys_tab", "tropical")
        ayanamsa = meta_req.get("ayanamsa", "lahiri")
        h_sys_raw = str(meta_req.get("h_sys", "P")).strip().upper()
        h_sys = h_sys_raw[0] if h_sys_raw else "P"
        
        view_mode = meta_req.get("view_mode", "zodiac")
        seg_num = int(meta_req.get("segment", 1))

        # 🚀 [FIX 2: N4의 calculate_arcana를 동원한 완벽한 데이터 연산]
        is_unk = bool(seed_data.get("is_time_unknown", 0))

        principia = calculate_principia(
            date_str, time_str, lat, lng, tz, system, ayanamsa, h_sys=h_sys, is_time_unknown=is_unk
        )
        arc_p = calculate_arcana(
            date_str, time_str, lat, lng, timezone=tz, lot_schema='paulus', 
            system=system, ayanamsa=ayanamsa, h_sys=h_sys
        )
        
        p_data = principia.get("planets", {})
        lords_data = principia.get("lords", {}) 
        lots_data = arc_p.get("lots", {})

        n9_result = calculate_n9_timeline(seed_data, mode=view_mode, ayanamsa=ayanamsa)
        n9_timeline = n9_result.get("timeline", [])

        # 타임라인 오늘 날짜 깃발 꽂기
        today_str = datetime.now().strftime("%Y-%m-%d")
        for idx in range(len(n9_timeline)):
            r_date = str(n9_timeline[idx].get('date', '9999-12-31')).split('T')[0]
            next_date = str(n9_timeline[idx+1].get('date', '9999-12-31')).split('T')[0] if idx + 1 < len(n9_timeline) else "9999-12-31"
            if r_date <= today_str < next_date:
                n9_timeline[idx]['is_today'] = True
                break

        # 12년 구간 필터링 적용
        if view_mode == 'zodiac':
            start_age = (seg_num - 1) * 12
            end_age = start_age + 12
            filtered_timeline = []
            for r in n9_timeline:
                try: age_val = int(r.get("age", 0))
                except: age_val = 0
                
                if start_age <= age_val < end_age:
                    filtered_timeline.append(r)
            n9_timeline = filtered_timeline

        # ==========================================
        # 4 & 5. 상단 메타데이터 강제 주입 (JSON 무시하고 직접 제어)
        # ==========================================
        roman_map = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI"}
        duo_roman = roman_map.get(seg_num, "I")

        if view_mode == 'zodiac':
            # 1) 좌측 지표 (B5 ~ B10)
            ws['B5'].value = duo_roman
            ws['B6'].value = str(arc_p.get("meta", {}).get("sect", "-")).upper()
            ws['B7'].value = str(p_data.get("Ascendant", {}).get("dms", "-"))
            ws['B8'].value = str(lots_data.get("Spirit", {}).get("dms", "-"))
            ws['B9'].value = str(lots_data.get("Fortune", {}).get("dms", "-"))
            ws['B10'].value = str(lots_data.get("Eros", {}).get("dms", "-"))

            for coord in ["B5", "B6"]:
                cell = ws[coord]
                apply_grimoire_styles(cell, str(cell.value), is_info_col=False, skip_color=True)
                cell.font = Font(name="Consolas", size=9 if coord != "B6" else 10)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 🚀 [요청 1]: Asc/Lots 4원소 배경색 적용 (skip_color=False 적용!)
            for coord in ["B7", "B8", "B9", "B10"]:
                cell = ws[coord]
                apply_grimoire_styles(cell, str(cell.value), is_info_col=False, skip_color=False)
                cell.font = Font(name="Consolas", size=9)
                cell.alignment = Alignment(horizontal='right', vertical='center')

            # 🚀 [요청 2]: 우측 행성 지표 분할 주입 (E: Info, G: Dignity, H: DL, I: HL)
            planet_rows = {
                "Saturn": 5, "Jupiter": 6, "Mars": 7, "Sun": 8,
                "Venus": 9, "Mercury": 10, "Moon": 11
            }
            
            for p_name, r_idx in planet_rows.items():
                p_info = p_data.get(p_name, {})
                ws[f'E{r_idx}'].value = p_info.get("dms", "-")
                ws[f'G{r_idx}'].value = p_info.get("dignity", "-")
                ws[f'H{r_idx}'].value = "TRUE" if p_name in str(lords_data.get("day", "")) else "-"
                ws[f'I{r_idx}'].value = "TRUE" if lords_data.get("hour") == p_name else "-"
                
                for c_char in ['E', 'G', 'H', 'I']:
                    cell = ws[f'{c_char}{r_idx}']
                    # 🚀 [완벽 수정]: skip_color=False 로 변경하여 4원소 색상 적용!
                    apply_grimoire_styles(cell, str(cell.value), is_info_col=False, skip_color=False)
                    cell.font = Font(name="Consolas", size=9)
                    cell.alignment = Alignment(horizontal='left' if c_char == 'E' else 'center', vertical='center')

        # 타임라인 매핑을 위해 JSON 파일 로드 (타임라인 전용)
        with open(MAPPING_FILE, 'r', encoding='utf-8') as f: 
            mapping = json.load(f)
        # ==========================================

        # 6. 타임라인 매핑 (A17부터)
        start_row = 17 
        cols = mapping.get("timeline", {}).get("columns", {})

        for i, row_data in enumerate(n9_timeline):
            r_idx = start_row + i
            ws.row_dimensions[r_idx].height = 16.5
            is_current = row_data.get('is_today', False)

            for col_char, rule in cols.items():
                data_key = rule.get("data_key") if isinstance(rule, dict) else rule
                check_lb = rule.get("check_lb") if isinstance(rule, dict) else None
                
                raw_val = get_nested(row_data, data_key, "-")
                if raw_val == "-" and "_" in data_key:
                    parts = data_key.split("_")
                    if parts[0] in ["spirit", "fortune", "eros"]:
                        raw_val = row_data.get("zr", {}).get(parts[0], {}).get(parts[1], "-")
                
                is_lb = bool(get_nested(row_data, check_lb, False)) if check_lb else False
                if not is_lb and "_" in data_key:
                    parts = data_key.split("_")
                    if parts[0] in ["spirit", "fortune", "eros"]:
                        is_lb = bool(row_data.get("zr", {}).get(parts[0], {}).get(f"{parts[1]}_lb", False))

                val = SYMBOL_MAP.get(str(raw_val), str(raw_val)) if "date" not in data_key and raw_val != "-" else str(raw_val)
                cell = ws[f"{col_char}{r_idx}"]
                cell.value = val

                apply_grimoire_styles(cell, str(raw_val), is_info_col=False, skip_color=(col_char == 'A'))
                cell.font = Font(name="Consolas", size=10, color="FF0000" if is_lb else "000000", bold=is_lb)
                cell.alignment = Alignment(horizontal='left' if col_char == 'A' else 'center', vertical='center')

                if is_current and col_char == 'A':
                    cell.fill = PatternFill(start_color="B0ECFF", end_color="B0ECFF", fill_type="solid")

        # 7. 틀 고정 및 🚀 [사후 병합(Post-Merge) 로직]
        ws.sheet_view.pane = None
        ws.freeze_panes = 'B17'
        
        # 병합할 영역 리스트 (템플릿의 원래 병합 모양대로 지정)
        # B~C열: Duodecim, Sect, Asc, Lots
        # 🚀 [완벽 수정]: E열부터 F열까지만 병합하고, G, H, I는 개별 셀로 살려둠!
        ranges_to_merge = [
            "B5:C5", "B6:C6", "B7:C7", "B8:C8", "B9:C9", "B10:C10",
            "E5:F5", "E6:F6", "E7:F7", "E8:F8", "E9:F9", "E10:F10", "E11:F11"  # 🚀 K를 F로 전부 교체 완료
        ]
        
        for m_range in ranges_to_merge:
            try:
                ws.merge_cells(m_range)
            except Exception:
                pass # 이미 병합되어 있거나 에러가 나면 부드럽게 스킵

        return wb

    except Exception as ex:
        print(f">>> [N9 FINAL REFORGE ERROR] {ex}", file=sys.stderr)
        raise ex