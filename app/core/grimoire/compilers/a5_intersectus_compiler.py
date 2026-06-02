import json
import os
import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz, format_dms_pretty
from core.astrology.aspects import calculate_all_aspects
from core.astrology.patterns import find_patterns

from api.astrology import get_seed_from_request
from core.astrology.davison import calculate_davison_midpoint
from core.astrology.composite import generate_composite_data

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a5_intersectus_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a5_intersectus.xlsx'))

ASPECT_ORDER = {
    "Conjunction": 1, "Opposition": 2, "Trine": 3, "Square": 4, "Sextile": 5,
    "Quintile": 6, "Septile": 7, "Quincunx": 8, "Octile": 9, "Novile": 10, 
    "Decile": 11, "Undecile": 12, "Semi-sextile": 13
}

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def get_body_info(pt_lon):
    if pt_lon is None: return "-"
    return format_dms_pretty(float(pt_lon))

def sanitize_unicode(text):
    if isinstance(text, str):
        return text.replace("\ufe0e", "").replace("\ufe0f", "")
    return text

def compile_a5_intersectus_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a5_intersectus"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()
    
    def resolve_seed(d):
        b_date = d.get('birth_date') or d.get('date', '2000-01-01')
        b_time = d.get('birth_time') or d.get('time', '12:00:00')
        lat = get_safe_float(d.get('lat'), 37.5665)
        lng = get_safe_float(d.get('lng'), 126.9780)
        date_str = str(b_date).split('T')[0]
        tz = _ensure_float_tz(d.get('tz') if d.get('tz') is not None else d.get('timezone', 9.0), date_str)
        return {
            'birth_date': date_str,
            'birth_time': str(b_time),
            'lat': lat,
            'lng': lng,
            'timezone': tz,
            'is_time_unknown': bool(d.get('is_time_unknown', 0))
        }

    h_a = resolve_seed(s1_raw)
    h_b = resolve_seed(s2_raw)
    
    dav_raw = calculate_davison_midpoint(s1_raw, s2_raw)
    h_dav = resolve_seed(dav_raw)

    d_lat = get_safe_float(dav_raw.get("lat"), None)
    d_lng = get_safe_float(dav_raw.get("lng"), None)
    if d_lat is not None and d_lng is not None:
        coord_str = f"{abs(d_lat):.2f}°{'N' if d_lat>=0 else 'S'}, {abs(d_lng):.2f}°{'E' if d_lng>=0 else 'W'}"
    else:
        coord_str = "Unknown Location"
    
    stamp_data = dav_raw.copy()
    stamp_data["city"] = stamp_data["location"] = stamp_data["city_name"] = stamp_data["location_name"] = coord_str
    apply_natal_stamp(ws, stamp_data, method="single", cells=["A2"])

    meta = chart_data.get("metadata", {})
    sys_a = meta.get("sys_a", "composite")
    ayan_a = meta.get("ayan_a", "lahiri")
    sys_b = meta.get("sys_b", "d_tropical")
    ayan_b = meta.get("ayan_b", "lahiri")
    
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")
    orb_factor = get_safe_float(meta.get("fixed_star_orb", 1.0))

    def get_system_points(sys_str, ayan_val):
        sys_str = str(sys_str).lower()
        parts = sys_str.split('_')
        prefix = parts[0]
        sys_name = parts[1] if len(parts) > 1 else 'tropical'
        
        if sys_str == 'composite':
            res_a = calculate_principia(date_str=h_a['birth_date'], time_str=h_a['birth_time'], lat=h_a['lat'], lng=h_a['lng'], timezone=h_a['timezone'], system='tropical', ayanamsa=ayan_val, view=view, h_sys=h_sys, fixed_star_orb=orb_factor, is_time_unknown=h_a['is_time_unknown'])
            res_b = calculate_principia(date_str=h_b['birth_date'], time_str=h_b['birth_time'], lat=h_b['lat'], lng=h_b['lng'], timezone=h_b['timezone'], system='tropical', ayanamsa=ayan_val, view=view, h_sys=h_sys, fixed_star_orb=orb_factor, is_time_unknown=h_b['is_time_unknown'])
            comp_map = generate_composite_data(res_a, res_b)
            
            h1_cusp = comp_map.get("1h cusp", comp_map.get("H1"))
            h10_cusp = comp_map.get("10h cusp", comp_map.get("H10"))
            if h1_cusp is not None:
                comp_map["Ascendant"] = float(h1_cusp)
                comp_map["Descendant"] = (float(h1_cusp) + 180) % 360
            if h10_cusp is not None:
                comp_map["Midheaven"] = float(h10_cusp)
                comp_map["Immum Coeli"] = (float(h10_cusp) + 180) % 360

            pts = {}
            for k, lon in comp_map.items():
                # 🚀 동의어 스킵
                if k in ["Asc.", "I.C.", "Dsc.", "M.C.", "Asc", "IC", "Dsc", "MC"]: continue
                
                name = k
                if k in ["Mean Node", "North Node (m)", "North Node (t)"]: name = "Rahu"
                elif k in ["South Node", "South Node (m)", "South Node (t)"]: name = "Ketu"
                elif k == "Lilith": name = "Asteroid Lilith"
                elif k == "Lilith (mean)": name = "Mean Lilith"
                elif k == "Lilith (true)": name = "True Lilith"
                elif k == "Eros": name = "Asteroid Eros"
                pts[name] = float(lon)
            return pts

        else:
            if prefix == 'd': target = h_dav
            elif prefix == 'a': target = h_a
            elif prefix == 'b': target = h_b
            else: target = h_a; sys_name = sys_str
            
            res_p = calculate_principia(date_str=target['birth_date'], time_str=target['birth_time'], lat=target['lat'], lng=target['lng'], timezone=target['timezone'], system=sys_name, ayanamsa=ayan_val, view=view, h_sys=h_sys, fixed_star_orb=orb_factor, is_time_unknown=target['is_time_unknown'])
            arc_p = calculate_arcana(date_str=target['birth_date'], time_str=target['birth_time'], lat=target['lat'], lng=target['lng'], timezone=target['timezone'], lot_schema='paulus', system=sys_name, ayanamsa=ayan_val, h_sys=h_sys)
            try: arc_v = calculate_arcana(date_str=target['birth_date'], time_str=target['birth_time'], lat=target['lat'], lng=target['lng'], timezone=target['timezone'], lot_schema='valens', system=sys_name, ayanamsa=ayan_val, h_sys=h_sys)
            except: arc_v = {'lots': {}}
            
            pts = {}
            for group in ['planets', 'asteroids', 'lilith_nodes']:
                for k, v in res_p.get(group, {}).items():
                    # 🚀 동의어 스킵
                    if k in ["Asc.", "I.C.", "Dsc.", "M.C.", "Asc", "IC", "Dsc", "MC"]: continue
                    
                    name = k
                    if k in ["Mean Node", "North Node (m)", "North Node (t)"]: name = "Rahu"
                    elif k in ["South Node", "South Node (m)", "South Node (t)"]: name = "Ketu"
                    elif k == "Lilith": name = "Asteroid Lilith"
                    elif k == "Lilith (mean)": name = "Mean Lilith"
                    elif k == "Lilith (true)": name = "True Lilith"
                    elif k == "Eros": name = "Asteroid Eros"
                    
                    lon = v.get('longitude', v.get('value'))
                    if lon is not None: pts[name] = float(lon)

            for k, v in arc_p.get('lots', {}).items(): 
                lon = v.get('longitude', v.get('value'))
                if lon is not None: pts[k] = float(lon)

            for k, v in arc_v.get('lots', {}).items():
                if k in ['Necessity', 'Eros']: 
                    lon = v.get('longitude', v.get('value'))
                    if lon is not None: pts[f"{k} (v)"] = float(lon)
            
            vx = arc_p.get('vertex', {}).get('Vertex')
            if vx:
                v_lon = vx.get('longitude', vx.get('value'))
                if v_lon is not None:
                    pts['Vertex'] = float(v_lon)
                    pts['Anti-Vertex'] = (float(v_lon) + 180) % 360
            
            syz = arc_p.get('syzygy', {}).get('data')
            if syz: 
                s_lon = syz.get('longitude', syz.get('value'))
                if s_lon is not None: pts['Syzygy'] = float(s_lon)
                
            return pts

    flat_pts_a = get_system_points(sys_a, ayan_a)
    flat_pts_b = get_system_points(sys_b, ayan_b)

    pts_merged = {f"{k}_1": v for k, v in flat_pts_a.items()}
    pts_merged.update({f"{k}_2": v for k, v in flat_pts_b.items()})

    try: all_aspects = calculate_all_aspects(pts_merged, orb_tightness=orb_factor, mode='intersectus')
    except Exception as e: all_aspects = []
    
    try: patterns = find_patterns(all_aspects, mode='intersectus')
    except Exception as e: patterns = []

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: mapping = json.load(f)

    meta_map = mapping.get("metadata", {})
    if meta_map.get("sys_a"): 
        ws[meta_map["sys_a"]] = sys_a.replace('_', ' ').upper()
        apply_grimoire_styles(ws[meta_map["sys_a"]], sys_a, skip_color=True)
        ws[meta_map["sys_a"]].font = Font(name="Consolas", bold=False)
    if meta_map.get("sys_b"): 
        ws[meta_map["sys_b"]] = sys_b.replace('_', ' ').upper()
        apply_grimoire_styles(ws[meta_map["sys_b"]], sys_b, skip_color=True)
        ws[meta_map["sys_b"]].font = Font(name="Consolas", bold=False)

    aspect_cols = mapping.get("aspects_layout", {}).get("cols", {})
    col_a, col_b, col_c = aspect_cols.get("body", "A"), aspect_cols.get("aspect", "B"), aspect_cols.get("body_info", "C")
    col_d, col_e, col_f = aspect_cols.get("object", "D"), aspect_cols.get("object_info", "E"), aspect_cols.get("orb", "F")
    aspect_rows_map = mapping.get("aspects_layout", {}).get("rows", {})
    
    sorted_bodies = sorted(aspect_rows_map.items(), key=lambda x: x[1], reverse=True)

    # 🚀 [요청 사항 완벽 구현]: System A가 Composite일 때 Atropos 하단 엑셀 행을 통째로 날려버림
    HERMETIC_BODIES = [
        "Fortune", "Spirit", "Necessity", "Necessity (v)", "Eros", "Eros (v)", 
        "Courage", "Victory", "Nemesis", "Vertex", "Anti-Vertex", "Syzygy"
    ]

    if sys_a == 'composite':
        atropos_row = aspect_rows_map.get("Atropos")
        if atropos_row:
            # Atropos 아래의 엑셀 행을 물리적으로 완전 삭제 (여백, 테두리 포맷팅 영구 제거)
            amount_to_delete = ws.max_row - atropos_row
            if amount_to_delete > 0:
                ws.delete_rows(atropos_row + 1, amount_to_delete + 5)
        
        # 반복문에서 Hermetic 항목들 제외시켜서 렌더링 스킵
        sorted_bodies = [(b, r) for b, r in sorted_bodies if b not in HERMETIC_BODIES]

    for body_name, row_idx in sorted_bodies:
        body_key = f"{body_name}_1"
        body_asps = []
        
        for a in all_aspects:
            p1, p2 = a['p1'], a['p2']
            if p1 == body_key and p2.endswith('_2'):
                body_asps.append({'asp': a['aspect'], 'obj': p2.replace('_2', ''), 'orb': a['orb']})
            elif p2 == body_key and p1.endswith('_2'):
                body_asps.append({'asp': a['aspect'], 'obj': p1.replace('_2', ''), 'orb': a['orb']})
                
        body_asps.sort(key=lambda x: (ASPECT_ORDER.get(x['asp'], 99), abs(x.get('orb', 0))))
        
        b_info = get_body_info(flat_pts_a.get(body_name))
        num_asps = len(body_asps)
        
        if num_asps == 0:
            ws[f"{col_c}{row_idx}"] = b_info
            for c_char in [col_b, col_d, col_e, col_f]: ws[f"{c_char}{row_idx}"] = ""
            for c_char in [col_b, col_c, col_d, col_e, col_f]: 
                cell = ws[f"{c_char}{row_idx}"]
                is_info = (c_char in [col_c, col_e])
                
                safe_val = sanitize_unicode(cell.value)
                apply_grimoire_styles(cell, safe_val, is_info_col=is_info, skip_color=False)
                
                color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
                cell.font = Font(name="Consolas", size=9 if is_info else 11, color=color_to_use, bold=False)
                align = 'center' if c_char == col_f else 'left'
                cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
            continue
            
        if num_asps > 1:
            ws.insert_rows(row_idx + 1, amount=num_asps - 1)
            for r in range(1, num_asps):
                insert_idx = row_idx + r
                ws.row_dimensions[insert_idx].height = 16.5
                for c_char in [col_a, col_b, col_c, col_d, col_e, col_f]:
                    src = ws[f"{c_char}{row_idx}"]
                    tgt = ws[f"{c_char}{insert_idx}"]
                    if src.has_style:
                        tgt.border, tgt.fill, tgt.font, tgt.alignment, tgt.number_format = copy(src.border), copy(src.fill), copy(src.font), copy(src.alignment), copy(src.number_format)
                    if c_char == col_a: tgt.value = "" 
        
        for i, asp_data in enumerate(body_asps):
            r_idx = row_idx + i
            asp_name = asp_data['asp']
            obj_name = asp_data['obj']
            obj_info = get_body_info(flat_pts_b.get(obj_name))
            orb_val = f"{abs(asp_data.get('orb', 0)):.2f}°"
            
            ws[f"{col_b}{r_idx}"] = asp_name
            ws[f"{col_c}{r_idx}"] = b_info if i == 0 else "" 
            ws[f"{col_d}{r_idx}"] = obj_name
            ws[f"{col_e}{r_idx}"] = obj_info
            ws[f"{col_f}{r_idx}"] = orb_val
            
            for c_char in [col_b, col_c, col_d, col_e, col_f]:
                cell = ws[f"{c_char}{r_idx}"]
                is_info = (c_char in [col_c, col_e])
                
                safe_val = sanitize_unicode(cell.value)
                skip_c = True if c_char == col_d else False
                apply_grimoire_styles(cell, safe_val, is_info_col=is_info, skip_color=skip_c)
                
                color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
                cell.font = Font(name="Consolas", size=9 if is_info else 11, color=color_to_use, bold=False)
                align = 'center' if c_char == col_f else 'left'
                cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)

    pat_map = mapping.get("patterns", {})
    pat_start = pat_map.get("row_start", 10)
    pat_cols = pat_map.get("cols", {})
    col_shape = pat_cols.get("shape", "A")
    
    if not patterns:
        cell = ws[f"{col_shape}{pat_start}"]
        cell.value = "No Shape Detected"
        cell.font = Font(name="Consolas", color="000000", bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for i in range(1, 7):
            ws[f"{pat_cols.get(f'p{i}')}{pat_start}"] = ""
            ws[f"{pat_cols.get(f'i{i}')}{pat_start}"] = ""
    else:
        patterns.sort(key=lambda x: x.get('shape', 'Unknown'))

        num_pats = len(patterns)
        if num_pats > 1:
            ws.insert_rows(pat_start + 1, amount=num_pats - 1)
            for r in range(1, num_pats):
                insert_idx = pat_start + r
                ws.row_dimensions[insert_idx].height = 16.5
                for c_char in "ABCDEFGHIJKLM":
                    src = ws[f"{c_char}{pat_start}"]
                    tgt = ws[f"{c_char}{insert_idx}"]
                    if src.has_style:
                        tgt.border, tgt.fill, tgt.font, tgt.alignment, tgt.number_format = copy(src.border), copy(src.fill), copy(src.font), copy(src.alignment), copy(src.number_format)
        
        for i, pat in enumerate(patterns):
            r_idx = pat_start + i
            
            shape_name = pat.get('shape', 'Unknown')
            if i > 0 and patterns[i-1].get('shape') == shape_name:
                ws[f"{col_shape}{r_idx}"] = ""
            else:
                ws[f"{col_shape}{r_idx}"] = shape_name
            
            for pt_i in range(1, 7):
                c_name = pat_cols.get(f"p{pt_i}")
                c_info = pat_cols.get(f"i{pt_i}")
                raw_p_name = pat.get(f'p{pt_i}', '-')
                
                if raw_p_name and raw_p_name != '-':
                    base_name = raw_p_name.split('_')[0]
                    
                    if raw_p_name.endswith('_1'):
                        prefix_char = sys_a.split('_')[0].upper()[0] 
                        display_name = f"{prefix_char}_{base_name}"
                        info_str = get_body_info(flat_pts_a.get(base_name))
                    else:
                        prefix_char = sys_b.split('_')[0].upper()[0]
                        display_name = f"{prefix_char}_{base_name}"
                        info_str = get_body_info(flat_pts_b.get(base_name))
                        
                    ws[f"{c_name}{r_idx}"] = display_name
                    ws[f"{c_info}{r_idx}"] = info_str
                else:
                    ws[f"{c_name}{r_idx}"] = ""
                    ws[f"{c_info}{r_idx}"] = ""
                    
            for c_char in "ABCDEFGHIJKLM":
                cell = ws[f"{c_char}{r_idx}"]
                is_info = c_char in ["C", "E", "G", "I", "K", "M"]
                is_name = c_char in ["B", "D", "F", "H", "J", "L"]
                
                safe_val = sanitize_unicode(cell.value)
                skip_c = True if (is_name or c_char == "A") else False
                apply_grimoire_styles(cell, safe_val, is_info_col=is_info, skip_color=skip_c)
                
                color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
                cell.font = Font(name="Consolas", size=9 if is_info else 11, color=color_to_use, bold=False)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    for r_idx in range(1, ws.max_row + 1):
        cell_val = ws[f"A{r_idx}"].value
        if str(cell_val).strip() == ".":
            ws.row_dimensions[r_idx].height = 4.5
            ws[f"A{r_idx}"].value = ""
        elif r_idx >= 10:
            if ws.row_dimensions[r_idx].height != 4.5:
                ws.row_dimensions[r_idx].height = 16.5

    return wb