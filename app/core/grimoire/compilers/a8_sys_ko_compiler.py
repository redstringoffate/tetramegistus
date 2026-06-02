import json
import os
import shutil
import math
import re
import openpyxl.utils
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import swisseph as swe

from core.grimoire.styler import apply_grimoire_styles, HOUSE_COLORS
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz, TROPICAL_SIGNS, calculate_all_star_positions
from core.astrology.davison import calculate_davison_midpoint
from core.astrology.composite import calculate_composite_chart, apply_anti_composite
from core.astrology.constants import ASTEROIDS
from core.astrology.arabic_lots import calculate_arabic_lots

BASE_DIR = os.path.dirname(__file__)

# 🚀 [절대 수복]: 어설픈 변수 조작 제거. 오직 순수한 영어 매핑/템플릿만 절대적으로 참조
MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a8_sys_ko_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a8_sys_ko.xlsx'))

ALLOWED_UPPER = [
    "SUN", "MOON", "MERCURY", "VENUS", "MARS", "JUPITER", "SATURN", "URANUS", "NEPTUNE", "PLUTO",
    "CHIRON", "CERES", "JUNO", "PALLAS", "VESTA", "ASTEROID EROS", "PSYCHE",
    "ASCENDANT", "MIDHEAVEN", "DESCENDANT", "IMMUM COELI",
    "NORTH NODE (T)", "SOUTH NODE (T)", "NORTH NODE (M)", "SOUTH NODE (M)", "RAHU", "KETU",
    "MEAN LILITH", "TRUE LILITH", "ASTEROID LILITH",
    "MOIRA", "KLOTHO", "LACHESIS", "ATROPOS",
    "FORTUNE", "SPIRIT", "NECESSITY", "NECESSITY (V)", "EROS", "EROS (V)", "COURAGE", "VICTORY", "NEMESIS",
    "VERTEX", "SYZYGY"
]

HOUSE_SYS_NAMES = {
    'P': 'Placidus', 'W': 'Whole Sign', 'K': 'Koch', 'R': 'Regiomontanus', 
    'C': 'Campanus', 'E': 'Equal', 'O': 'Porphyry', 'V': 'Vehlow'
}

STANDARD_BODIES = ["Sun", "Moon", "Mercury", "Venus", "Mars", "Jupiter", "Saturn", "Uranus", "Neptune", "Pluto", 
                   "Chiron", "Ceres", "Pallas", "Juno", "Vesta", "Eros", "Psyche", "Moira", "Klotho", "Lachesis", "Atropos", "Lilith"]

# 🚀 용어 대통합 (a/b -> seed/partner)
FALLBACK_COLS = {
    "composite_main": "B", "composite_anti": "C", 
    "davison_minor_asteroids": "D", "davison_tropical": "E", "davison_sidereal": "F", "davison_draconic": "G", "davison_ketunic": "H", "davison_arabic_lots": "I",
    "seed_minor_asteroids": "J", "seed_tropical": "K", "seed_sidereal": "L", "seed_draconic": "M", "seed_ketunic": "N", "seed_arabic_lots": "O",
    "partner_minor_asteroids": "P", "partner_tropical": "Q", "partner_sidereal": "R", "partner_draconic": "S", "partner_ketunic": "T", "partner_arabic_lots": "U",
    "fixed_stars": "W"
}

def normalize_entity(raw_e):
    e = str(raw_e).lower().strip().replace(" ", "_")
    if e in ["partner", "seed2", "seed_b", "parent_b", "b", "2", "p2"]: return "partner"
    if e in ["seed", "seed1", "seed_a", "parent_a", "a", "1", "seed_1", "p1"]: return "seed"
    if "davison" in e: return "davison"
    if "composite" in e: return "composite"
    return e

def normalize_col_key(raw_k):
    k = str(raw_k).lower().strip().replace(" ", "_")
    for p in ["partner_", "seed2_", "seed_b_", "parent_b_", "b_", "p2_"]:
        if k.startswith(p): return f"partner_{k[len(p):]}"
    for p in ["seed1_", "seed_a_", "parent_a_", "seed_", "a_", "p1_"]:
        if k.startswith(p): return f"seed_{k[len(p):]}"
    for s in ["_partner", "_seed2", "_b", "_2", "_p2"]:
        if k.endswith(s): return f"partner_{k[:-len(s)]}"
    for s in ["_seed1", "_seed", "_a", "_1", "_p1"]:
        if k.endswith(s): return f"seed_{k[:-len(s)]}"
    return k

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def get_date_time_coords(seed):
    d = str(seed.get('birth_date') or seed.get('birthDate') or '2000-01-01').split('T')[0]
    t = str(seed.get('birth_time') or seed.get('birthTime') or '12:00:00').strip()
    
    if t.lower() in ["unknown", "none", ""]: t = "12:00:00"
    elif len(t) == 5: t += ":00"

    coords = seed.get("coordinates", {})
    if isinstance(coords, dict) and "lat" in coords:
        lat = get_safe_float(coords.get("lat"), 37.5665)
        lng = get_safe_float(coords.get("lng"), 126.9780)
    else:
        lat = get_safe_float(seed.get("lat"), 37.5665)
        lng = get_safe_float(seed.get("lng"), 126.9780)
        
    if lat == 0.0 and lng == 0.0:
        lat, lng = 37.5665, 126.9780
        
    tz = _ensure_float_tz(seed.get('timezone', seed.get('tz', 9.0)), d)
    return d, t, lat, lng, tz

def get_res(seed, sys_type, ayan, h_sys):
    d, t, lat, lng, tz = get_date_time_coords(seed)
    return calculate_principia(d, t, lat, lng, tz, sys_type, ayan, h_sys=h_sys)

def get_arc(seed, sys_type, ayan, h_sys, lot_schema='paulus'):
    d, t, lat, lng, tz = get_date_time_coords(seed)
    
    # 1. Paulus 기본 랏 계산
    try: arc_p = calculate_arcana(d, t, lat, lng, tz, lot_schema=lot_schema, system=sys_type, ayanamsa=ayan, h_sys=h_sys)
    except: arc_p = {'lots': {}}
    
    # 🚀 2. Valens 랏 추가 계산 및 병합 (Eros v, Necessity v)
    try:
        arc_v = calculate_arcana(d, t, lat, lng, tz, lot_schema='valens', system=sys_type, ayanamsa=ayan, h_sys=h_sys)
        if 'lots' not in arc_p: arc_p['lots'] = {}
        for k in ['Eros', 'Necessity']:
            if k in arc_v.get('lots', {}):
                arc_p['lots'][f"{k} (v)"] = arc_v['lots'][k]
    except:
        pass
        
    return arc_p

def get_jd(seed):
    d, t, lat, lng, tz = get_date_time_coords(seed)
    try: dt_obj = datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M:%S")
    except: dt_obj = datetime.strptime(f"{d} 12:00:00", "%Y-%m-%d %H:%M:%S")
    return swe.julday(dt_obj.year, dt_obj.month, dt_obj.day, dt_obj.hour + dt_obj.minute/60.0 - tz)

def calc_minor_asteroids(jd):
    swe.set_sid_mode(0, 0, 0)
    res_dict = {}
    for ast_name, ast_num in ASTEROIDS.items():
        if ast_name in STANDARD_BODIES: continue
        try:
            res = swe.calc_ut(jd, ast_num + 10000, swe.FLG_SWIEPH | swe.FLG_SPEED)
            res_dict[ast_name] = {"longitude": res[0][0]}
        except: pass
    return res_dict

def calc_arabic_lots_full(res_trop, arc_trop, arabic_ruler):
    points = {k: v['longitude'] for k, v in res_trop.get('planets', {}).items()}
    points.update({k: v['longitude'] for k, v in res_trop.get('angles', {}).items()})
    house_cusps = {i: res_trop.get('houses', {}).get(i, {}).get('longitude', 0.0) for i in range(1, 13)}
    
    rulers_trad = ["Mars", "Venus", "Mercury", "Moon", "Sun", "Mercury", "Venus", "Mars", "Jupiter", "Saturn", "Saturn", "Jupiter"]
    rulers_mod = ["Mars", "Venus", "Mercury", "Moon", "Sun", "Mercury", "Venus", "Pluto", "Jupiter", "Saturn", "Uranus", "Neptune"]
    r_list = rulers_trad if arabic_ruler.lower() == 'traditional' else rulers_mod
    rulers = {i: r_list[int(house_cusps[i] / 30) % 12] for i in range(1, 13)}
    
    hour_lord_name = str(res_trop.get('lords', {}).get('hour', 'Sun')).split('|')[0].strip()
    lord_of_hour_lon = points.get(hour_lord_name, 0.0)
    
    p_new_moon, p_full_moon = 0.0, 0.0
    if 'syzygy' in arc_trop and 'data' in arc_trop['syzygy']:
        sz = arc_trop['syzygy']['data']
        if sz.get('phase') == 'New Moon': p_new_moon = sz.get('value', 0.0)
        else: p_full_moon = sz.get('value', 0.0)
        
    fortune_lon = arc_trop.get('lots', {}).get('Fortune', {}).get('value', 0.0)
    spirit_lon = arc_trop.get('lots', {}).get('Spirit', {}).get('value', 0.0)
    
    lots = calculate_arabic_lots(points, house_cusps, rulers, lord_of_hour_lon, p_new_moon, p_full_moon, fortune_lon, spirit_lon)
    return {k: {"longitude": v} for k, v in lots.items()}

def create_house_bg_map(cusps_dict):
    bg_map = {deg: "H1" for deg in range(360)}
    if not cusps_dict: return bg_map
    
    sorted_cusps = []
    for h, lon in cusps_dict.items():
        lon_val = float(lon['longitude'] if isinstance(lon, dict) else lon)
        sabian_idx = int(math.floor(lon_val)) % 360
        sorted_cusps.append((sabian_idx, h))
        
    sorted_cusps.sort(key=lambda x: x[0])
    count = len(sorted_cusps)
    
    for i in range(count):
        curr_start, curr_h = sorted_cusps[i]
        next_start, _ = sorted_cusps[(i + 1) % count]
        
        if curr_start < next_start:
            for idx in range(curr_start, next_start): bg_map[idx] = f"H{curr_h}"
        else:
            for idx in range(curr_start, 360): bg_map[idx] = f"H{curr_h}"
            for idx in range(0, next_start): bg_map[idx] = f"H{curr_h}"
            
    return bg_map

def ensure_location_string(seed):
    loc_val = str(seed.get("location", seed.get("city", seed.get("location_name", "")))).strip()
    if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
        d, t, lat, lng, tz = get_date_time_coords(seed)
        if lat != 0.0 and lng != 0.0:
            coord_str = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
            seed["city"] = coord_str
        else:
            seed["city"] = "Unknown Location"
    else:
        seed["city"] = loc_val

def get_lords(res):
    dl_raw = str(res.get('lords', {}).get('day', '-'))
    hl_raw = str(res.get('lords', {}).get('hour', '-'))
    return [d.strip() for d in dl_raw.split('|') if d.strip() != '-'], [h.strip() for h in hl_raw.split('|') if h.strip() != '-']

def compile_a8_sys_ko_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a8"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    meta = chart_data.get("metadata", {})
    settings = meta.get("settings", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    h_sys_raw = str(meta.get("h_sys", "P")).strip().upper()
    h_sys = h_sys_raw[0] if h_sys_raw else "P"
    arabic_ruler = meta.get("arabic_ruler", "traditional")

    albedo_data = chart_data.get("seed", {})
    if not albedo_data or 'seed1' not in albedo_data or 'seed2' not in albedo_data:
        raise ValueError("Albedo Station is vacant.")

    seed_a = albedo_data['seed1'].copy()
    seed_b = albedo_data['seed2'].copy()
    
    ensure_location_string(seed_a)
    ensure_location_string(seed_b)

    davison = calculate_davison_midpoint(seed_a, seed_b)
    ensure_location_string(davison)

    meta_map = mapping.get("metadata", {})
    if "davison_stamp" in meta_map: apply_natal_stamp(ws, davison, method="single", cells=[meta_map["davison_stamp"]])
    seed_a_stamp = meta_map.get("seed_a_stamp", meta_map.get("seed1_stamp", meta_map.get("seed_stamp")))
    if seed_a_stamp: apply_natal_stamp(ws, seed_a, method="single", cells=[seed_a_stamp])
    seed_b_stamp = meta_map.get("seed_b_stamp", meta_map.get("seed2_stamp", meta_map.get("partner_stamp")))
    if seed_b_stamp: apply_natal_stamp(ws, seed_b, method="single", cells=[seed_b_stamp])

    active_cols_set = set()
    for ent, conf in settings.items():
        if isinstance(conf, dict) and not conf.get("active", False):
            continue 
        ent_norm = normalize_entity(ent)
        subs = []
        if isinstance(conf, dict): subs = conf.get("subs", [])
        elif isinstance(conf, list): subs = conf
        for sub in subs:
            active_cols_set.add(f"{ent_norm}_{str(sub).lower().strip()}")

    raw_col_map = mapping.get("matrix_layout", {}).get("cols", {})
    col_map = {}
    for k, v in raw_col_map.items():
        norm_k = normalize_col_key(k)
        col_map[norm_k] = str(v).strip().upper()
        
    for k, v in FALLBACK_COLS.items():
        if k not in col_map: col_map[k] = v
            
    all_col_chars = list(col_map.values())
    
    if "ayanamsa" in meta_map: ws[meta_map["ayanamsa"]] = ayanamsa.upper()
    if "h_sys" in meta_map: ws[meta_map["h_sys"]] = HOUSE_SYS_NAMES.get(h_sys, "Placidus")
    if "arabic_ruler" in meta_map: ws[meta_map["arabic_ruler"]] = arabic_ruler.upper()

    for k in ["ayanamsa", "h_sys", "arabic_ruler"]:
        if k in meta_map:
            c_ref = meta_map[k]
            if ws[c_ref].value:
                ws[c_ref].font = Font(name="Consolas", size=11, bold=False)
                ws[c_ref].alignment = Alignment(horizontal='left', vertical='center')

    # 엑셀 열 덮어쓰기 원천 차단
    active_chars = {col_map[k] for k in active_cols_set if k in col_map}
    if "fixed_stars" in col_map: 
        active_chars.add(col_map["fixed_stars"])

    for col_char in set(col_map.values()):
        if col_char in active_chars:
            ws.column_dimensions[col_char].hidden = False
            if ws.column_dimensions[col_char].width is None or ws.column_dimensions[col_char].width < 5:
                ws.column_dimensions[col_char].width = 15.0
        else:
            ws.column_dimensions[col_char].hidden = True

    visible_cols = [char for key, char in col_map.items() if key in active_cols_set and key != "fixed_stars"]
    if visible_cols:
        first_col_char = sorted(visible_cols, key=lambda x: (len(x), x))[0]
        ws[f"{first_col_char}5"] = ayanamsa.upper()
        ws[f"{first_col_char}6"] = HOUSE_SYS_NAMES.get(h_sys, "Placidus")
        ws[f"{first_col_char}7"] = arabic_ruler.upper()
        
        for r_idx in [5, 6, 7]:
            c_ref = f"{first_col_char}{r_idx}"
            ws[c_ref].font = Font(name="Consolas", size=11, bold=False)
            ws[c_ref].alignment = Alignment(horizontal='left', vertical='center')

    matrix = {
        key: { "items": {deg: [] for deg in range(360)}, "bg": {deg: None for deg in range(360)} }
        for key in col_map.keys()
    }

    # ====================================================================
    # 🚀 [에러 해결]: 누락되었던 필수 데이터 정제 함수들 선언
    # ====================================================================
    def get_all_bodies(res_obj):
        merged = {}
        for g in ['planets', 'asteroids', 'lilith_nodes', 'angles']:
            for k, v in res_obj.get(g, {}).items():
                d_name = k.replace(" (Natal)", "")
                if d_name == "Eros": d_name = "Asteroid Eros"
                merged[d_name] = v
        return merged

    def add_to_matrix(col_key, item_name, longitude, is_day=False, is_hour=False):
        if longitude is None: return
        deg_abs = int(math.floor(float(longitude))) % 360
        
        up_name = str(item_name).upper()
        if up_name in ["MEAN NODE", "NORTH NODE (M)"]: final_name = "Rahu" if "Rahu" not in str(item_name) else item_name
        elif up_name in ["SOUTH NODE", "SOUTH NODE (M)"]: final_name = "Ketu" if "Ketu" not in str(item_name) else item_name
        elif "CUSP" in up_name:
            match = re.search(r'\d+', up_name)
            if match: final_name = f"{match.group()}h cusp" 
            else: final_name = up_name
        else: final_name = str(item_name)
            
        if col_key in matrix:
            matrix[col_key]["items"][deg_abs].append({
                "name": final_name,
                "lon": float(longitude),
                "is_day": is_day,
                "is_hour": is_hour
            })

    def fill_matrix(col_key, data_dict, allowed=ALLOWED_UPPER, filter_nodes=False, day_lords=None, hour_lords=None):
        if day_lords is None: day_lords = []
        if hour_lords is None: hour_lords = []
        
        if col_key not in active_cols_set and col_key != "fixed_stars": return
        for p_name, p_data in data_dict.items():
            up_name = p_name.upper()
            if allowed and up_name not in allowed and "CUSP" not in up_name: continue
            if filter_nodes and up_name in ['NORTH NODE (T)', 'SOUTH NODE (T)', 'RAHU', 'KETU', 'MEAN LILITH', 'TRUE LILITH', 'ASTEROID LILITH']: continue
            lon = p_data if isinstance(p_data, (int, float)) else p_data.get('longitude')
            
            is_day = p_name in day_lords or up_name in [d.upper() for d in day_lords]
            is_hour = p_name in hour_lords or up_name in [h.upper() for h in hour_lords]
            
            add_to_matrix(col_key, p_name, lon, is_day, is_hour)

    def fill_arcana_to_tropical(col_key, arc_obj):
        if col_key not in active_cols_set: return
        for k, v in arc_obj.get('lots', {}).items(): add_to_matrix(col_key, k, v.get('value', v.get('longitude')))
        for k, v in arc_obj.get('vertex', {}).items(): add_to_matrix(col_key, k, v.get('value', v.get('longitude')))
        if 'syzygy' in arc_obj and arc_obj['syzygy'].get('data'):
            add_to_matrix(col_key, "Syzygy", arc_obj['syzygy']['data'].get('value', arc_obj['syzygy']['data'].get('longitude')))

    # ====================================================================
    # 🚀 엔진 연산 시작
    # ====================================================================
    res_a_trop = get_res(seed_a, "tropical", ayanamsa, h_sys)
    arc_a_trop = get_arc(seed_a, "tropical", ayanamsa, h_sys, arabic_ruler.lower())
    
    res_b_trop = get_res(seed_b, "tropical", ayanamsa, h_sys)
    arc_b_trop = get_arc(seed_b, "tropical", ayanamsa, h_sys, arabic_ruler.lower())
    
    res_dav_trop = get_res(davison, "tropical", ayanamsa, h_sys)
    arc_dav_trop = get_arc(davison, "tropical", ayanamsa, h_sys, arabic_ruler.lower())
    
    jd_a = get_jd(seed_a)
    jd_b = get_jd(seed_b)
    jd_dav = get_jd(davison)

    cusps_a = {int(k): float(v['longitude'] if isinstance(v, dict) else v) for k, v in res_a_trop.get('houses', {}).items()}
    cusps_b = {int(k): float(v['longitude'] if isinstance(v, dict) else v) for k, v in res_b_trop.get('houses', {}).items()}

    # 🚀 Composite용 데이터에 소행성/교점까지 전부 합쳐서 정밀 계산
    merged_a_trop = get_all_bodies(res_a_trop)
    merged_b_trop = get_all_bodies(res_b_trop)
    
    comp_res_full = calculate_composite_chart(merged_a_trop, merged_b_trop, cusps_a, cusps_b)
    comp_main = comp_res_full.get('planets', {}) 
    comp_main_houses = comp_res_full.get('houses', {})
    
    if 1 in comp_main_houses: comp_main['Ascendant'] = {'longitude': float(comp_main_houses[1])}
    if 4 in comp_main_houses: comp_main['Immum Coeli'] = {'longitude': float(comp_main_houses[4])}
    if 7 in comp_main_houses: comp_main['Descendant'] = {'longitude': float(comp_main_houses[7])}
    if 10 in comp_main_houses: comp_main['Midheaven'] = {'longitude': float(comp_main_houses[10])}

    comp_anti, comp_anti_houses = apply_anti_composite(comp_main, comp_main_houses)

    # 🚀 각 System별로 하우스 배경색을 독립적으로 계산 (Tropical 복붙 버그 해결)
    bg_map_comp_main = create_house_bg_map(comp_main_houses)
    bg_map_comp_anti = create_house_bg_map(comp_anti_houses)

    for deg_abs in range(360):
        if "composite_main" in matrix: matrix["composite_main"]["bg"][deg_abs] = bg_map_comp_main[deg_abs]
        if "composite_anti" in matrix: matrix["composite_anti"]["bg"][deg_abs] = bg_map_comp_anti[deg_abs]

    # ====================================================================
    # 🚀 [1. 복구]: 뼈빠지게 계산한 Composite 데이터를 드디어 매트릭스에 붓습니다.
    # ====================================================================
    fill_matrix("composite_main", comp_main, allowed=ALLOWED_UPPER, filter_nodes=False)
    fill_matrix("composite_anti", comp_anti, allowed=ALLOWED_UPPER, filter_nodes=False)

    for i in range(1, 13):
        if h_sys != 'W' and i in [1, 4, 7, 10]: continue
        if "composite_main" in active_cols_set and i in comp_main_houses:
            add_to_matrix("composite_main", f"{i}h cusp", comp_main_houses[i])
        if "composite_anti" in active_cols_set and i in comp_anti_houses:
            add_to_matrix("composite_anti", f"{i}h cusp", comp_anti_houses[i])
    # ====================================================================

    def fill_system_data(prefix, sys_name, res_obj, arc_obj=None, filter_nodes=False):
        col_key = f"{prefix}_{sys_name}"
        if col_key not in active_cols_set: return
        
        bg_map = create_house_bg_map(res_obj.get('houses', {}))
        for deg_abs in range(360):
            matrix[col_key]["bg"][deg_abs] = bg_map[deg_abs]
            
        dl, hl = get_lords(res_obj)
        merged = get_all_bodies(res_obj)
        fill_matrix(col_key, merged, allowed=ALLOWED_UPPER, filter_nodes=filter_nodes, day_lords=dl, hour_lords=hl)
        
        for i in range(1, 13):
            if h_sys != 'W' and i in [1, 4, 7, 10]: continue
            c_lon = res_obj.get('houses', {}).get(i, {}).get('longitude')
            if c_lon is not None: add_to_matrix(col_key, f"{i}h cusp", c_lon)
            
        if arc_obj:
            fill_arcana_to_tropical(col_key, arc_obj)

    prefix_map = {
        "davison": (davison, res_dav_trop, arc_dav_trop, jd_dav),
        "seed": (seed_a, res_a_trop, arc_a_trop, jd_a),
        "partner": (seed_b, res_b_trop, arc_b_trop, jd_b)
    }

    for prefix, (seed_obj, res_t, arc_t, jd_val) in prefix_map.items():
        fill_system_data(prefix, "tropical", res_t, arc_obj=arc_t)
        
        if f"{prefix}_sidereal" in active_cols_set:
            res_s = get_res(seed_obj, "sidereal", ayanamsa, h_sys)
            fill_system_data(prefix, "sidereal", res_s)
            
        if f"{prefix}_draconic" in active_cols_set:
            res_d = get_res(seed_obj, "draconic", ayanamsa, h_sys)
            fill_system_data(prefix, "draconic", res_d, filter_nodes=True)
            
        if f"{prefix}_ketunic" in active_cols_set:
            res_k = get_res(seed_obj, "ketunic", ayanamsa, h_sys)
            fill_system_data(prefix, "ketunic", res_k, filter_nodes=True)
            
        if f"{prefix}_arabic_lots" in active_cols_set:
            lots = calc_arabic_lots_full(res_t, arc_t, arabic_ruler)
            fill_matrix(f"{prefix}_arabic_lots", lots, allowed=None)
            for deg_abs in range(360): matrix[f"{prefix}_arabic_lots"]["bg"][deg_abs] = None
            
        if f"{prefix}_minor_asteroids" in active_cols_set:
            ast = calc_minor_asteroids(jd_val)
            fill_matrix(f"{prefix}_minor_asteroids", ast, allowed=None)
            for deg_abs in range(360): matrix[f"{prefix}_minor_asteroids"]["bg"][deg_abs] = None

    if "fixed_stars" in col_map:
        active_stars = calculate_all_star_positions(jd_dav, 'tropical', swe.FLG_SWIEPH | swe.FLG_SPEED)
        star_dict = {s['name']: {"longitude": s['lon']} for s in active_stars}
        fill_matrix("fixed_stars", star_dict, allowed=None)

    row_map = mapping.get("matrix_layout", {}).get("rows", {})
    
    def get_row_for_lon(lon_val):
        sign_idx = int((float(lon_val) % 360) // 30)
        deg_in_sign = int((float(lon_val) % 360) % 30)
        sign_name = TROPICAL_SIGNS[sign_idx]
        row_key = f"{sign_name} {deg_in_sign + 1}"
        return row_map.get(row_key)

    for deg_abs in range(359, -1, -1):
        r_idx = get_row_for_lon(deg_abs)
        if not r_idx: continue

        max_lines = 1
        for col_key, col_char in col_map.items():
            if col_key not in active_cols_set and col_key != "fixed_stars": continue

            raw_items = matrix[col_key]["items"][deg_abs]
            raw_items.sort(key=lambda x: (float(x["lon"]), 0 if "CUSP" in str(x["name"]).upper() else 1))

            unique_items = []
            seen = set()
            for item in raw_items:
                if item["name"] not in seen:
                    unique_items.append(item)
                    seen.add(item["name"])
            matrix[col_key]["items"][deg_abs] = unique_items

            if len(unique_items) > max_lines: max_lines = len(unique_items)

        if max_lines > 1:
            amount = max_lines - 1
            ins_idx = r_idx + 1
            old_heights = {r: ws.row_dimensions[r].height for r in list(ws.row_dimensions.keys())}

            ws.insert_rows(ins_idx, amount=amount)

            for r in range(ws.max_row, 0, -1):
                if r >= ins_idx + amount: ws.row_dimensions[r].height = old_heights.get(r - amount, 15)
                elif ins_idx <= r < ins_idx + amount: ws.row_dimensions[r].height = old_heights.get(r_idx, 15)
                else: ws.row_dimensions[r].height = old_heights.get(r, 15)

            MAX_COL_IDX = max(ws.max_column, 24)
            
            for r_offset in range(1, max_lines):
                new_r = r_idx + r_offset
                for c_idx in range(1, MAX_COL_IDX + 1):
                    src = ws.cell(row=r_idx, column=c_idx)
                    tgt = ws.cell(row=new_r, column=c_idx)
                    if src.has_style:
                        tgt.font = copy(src.font)
                        tgt.border = copy(src.border)
                        tgt.fill = copy(src.fill)
                        tgt.number_format = copy(src.number_format)
                        tgt.alignment = copy(src.alignment)

                    col_letter = tgt.column_letter
                    if col_letter not in all_col_chars: 
                        tgt.value = ""
                    if c_idx == 1:
                        tgt.value = ""

        for col_key, col_char in col_map.items():
            if col_key not in active_cols_set and col_key != "fixed_stars": continue

            h_str = matrix[col_key]["bg"][deg_abs]
            is_minor = ("arabic" in col_key or "asteroids" in col_key or "fixed_stars" in col_key)
            bg_hex = None if is_minor else HOUSE_COLORS.get(h_str)
                
            items = matrix[col_key]["items"][deg_abs]

            for line_idx in range(max_lines):
                target_row = r_idx + line_idx
                cell = ws[f"{col_char}{target_row}"]

                if line_idx < len(items):
                    item = items[line_idx]
                    name = item["name"]
                    val_upper = name.upper()

                    is_angle = val_upper in ["ASCENDANT", "MIDHEAVEN", "DESCENDANT", "IMMUM COELI"]
                    if item.get("is_hour") or is_angle: name = val_upper

                    cell.value = name
                    # 🚀 [2. 오염 방지]: Minor 컬럼에서는 폰트 컬러(skip_planet_color)도 강제로 꺼버립니다.
                    apply_grimoire_styles(
                        cell, name, 
                        is_day_lord=item.get("is_day", False), 
                        is_hour_lord=item.get("is_hour", False), 
                        tabula_mode=True, 
                        house_bg=bg_hex, 
                        skip_color=is_minor,
                        skip_planet_color=is_minor  # 🚩 이 옵션이 Nemesis의 핑크색 오염을 막아줍니다!
                    )
                else:
                    cell.value = ""
                    apply_grimoire_styles(
                        cell, "", 
                        tabula_mode=True, 
                        house_bg=bg_hex, 
                        skip_color=is_minor,
                        skip_planet_color=is_minor
                    )

                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # 🚀 Day/Hour 스타일(bold 등)을 날려먹지 않도록 원본을 보존하며 Consolas 적용
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name="Consolas", size=cell.font.size, bold=cell.font.bold,
                    italic=cell.font.italic, color=cell.font.color
                )

    from openpyxl.utils import get_column_letter
    
    temp_wb_for_width = load_workbook(TEMPLATE_FILE)
    temp_ws_for_width = temp_wb_for_width[base_name] if base_name in temp_wb_for_width.sheetnames else temp_wb_for_width.active
    mapped_cols = list(all_col_chars) 
    
    # 🚩 [추가]: A8은 여러 사람의 Arabic Lots 컬럼이 있으므로 미리 찾아둡니다.
    arabic_cols = [char for key, char in col_map.items() if "arabic_lots" in key]
    
    for i in range(1, ws.max_column + 1):
        col_letter = get_column_letter(i)
        
        if ws.column_dimensions[col_letter].hidden:
            ws.column_dimensions[col_letter].width = 0.1
            continue
            
        temp_width = temp_ws_for_width.column_dimensions[col_letter].width if col_letter in temp_ws_for_width.column_dimensions else None
        
        if col_letter == 'A':
            ws.column_dimensions[col_letter].width = 15.0  
        elif col_letter in arabic_cols:                    # 🚩 [추가]: Arabic Lots 컬럼들 전부 예외 처리!
            ws.column_dimensions[col_letter].width = 30.0  # 👈 원하는 폭으로 조절하세요
        elif col_letter in mapped_cols:
            ws.column_dimensions[col_letter].width = 16.0  
        else:
            ws.column_dimensions[col_letter].width = temp_width if temp_width is not None else 40.0

    merge_to_remove = None
    for m_range in list(ws.merged_cells.ranges):
        if "A1" in m_range.coord:
            merge_to_remove = m_range
            break
            
    if merge_to_remove:
        max_col_letter = get_column_letter(merge_to_remove.max_col)
        ws.unmerge_cells(merge_to_remove.coord)
        ws['B1'].value = ws['A1'].value
        ws['A1'].value = None
        if ws['A1'].has_style:
            ws['B1'].font = copy(ws['A1'].font)
            ws['B1'].border = copy(ws['A1'].border)
            ws['B1'].fill = copy(ws['A1'].fill)
            ws['B1'].alignment = copy(ws['A1'].alignment)
        ws.merge_cells(f"B1:{max_col_letter}1")

    ws.freeze_panes = "C11"

    return wb