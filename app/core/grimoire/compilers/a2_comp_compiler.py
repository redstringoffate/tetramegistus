import json
import os
import shutil
import math
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import (
    calculate_principia, _ensure_float_tz, format_dms_pretty, 
    TROPICAL_SIGNS, SYMBOL_MAP
)
from core.astrology.composite import generate_composite_data

# Divisions (For composite re-calculation)
from core.astrology.divisions.decan import get_decan
from core.astrology.divisions.duad import get_duad
from core.astrology.divisions.dodecatemoria import get_dodecatemoria
from core.astrology.divisions.egyptian_bounds import get_egyptian_bounds
from core.astrology.divisions.sabian_engine import get_sabian_index

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a2_comp_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a2_comp.xlsx'))

SABIAN_PATHS = [
    os.path.abspath(os.path.join(BASE_DIR, '../../../data/render/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../../data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../static/data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../data/sabian.json')),
]

RULER_SYMBOLS = {
    "Sun": "☉", "Moon": "☽", "Mercury": "☿", "Venus": "♀", "Mars": "♂", 
    "Jupiter": "♃", "Saturn": "♄", "Uranus": "♅", "Neptune": "♆", "Pluto": "♇",
    "Rahu": "☊", "Ketu": "☋", "Chiron": "⚷"
}

def get_col(cell_str):
    if not cell_str: return ""
    return "".join([c for c in str(cell_str) if c.isalpha()])

def get_safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default

def compile_a2_comp_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a2_comp"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")
    mode = meta.get("mode", "normal") 

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()

    def ensure_location_string(seed):
        loc_val = str(seed.get("location", seed.get("city", ""))).strip()
        if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
            lat = get_safe_float(seed.get("lat"), None)
            lng = get_safe_float(seed.get("lng"), None)
            if lat is not None and lng is not None:
                seed["city"] = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
                seed["location"] = seed["city"]
            else:
                seed["city"] = "Unknown Location"
                seed["location"] = "Unknown Location"
        else:
            seed["city"] = loc_val
            seed["location"] = loc_val
    
    ensure_location_string(s1_raw)
    ensure_location_string(s2_raw)

    apply_natal_stamp(ws, s1_raw, method="single", cells=["A2"])
    apply_natal_stamp(ws, s2_raw, method="single", cells=["A3"])

    def map_for_engine(seed):
        date_str = str(seed.get('birth_date', '2000-01-01')).split('T')[0]
        time_str = str(seed.get('birth_time', '12:00:00'))
        tz_val = seed.get('tz') if seed.get('tz') is not None else seed.get('timezone', 9.0)
        return {
            "date_str": date_str, "time_str": time_str,
            "lat": get_safe_float(seed.get("lat"), 37.5665),
            "lng": get_safe_float(seed.get("lng"), 126.9780),
            "timezone": _ensure_float_tz(tz_val, date_str),
            "system": system, "ayanamsa": ayanamsa, "view": view, 
            "h_sys": "P",  
            "fixed_star_orb": 1.0, "is_time_unknown": bool(seed.get("is_time_unknown", 0))
        }

    res1 = calculate_principia(**map_for_engine(s1_raw))
    res2 = calculate_principia(**map_for_engine(s2_raw))
    
    comp_coords = generate_composite_data(res1, res2, mode=mode)

    def clean_cusps(res):
        c_out = {}
        for k, v in res.get('houses', {}).items():
            val = float(v['longitude']) if isinstance(v, dict) else float(v)
            c_out[int(k)] = val
        return c_out

    cusps_a = clean_cusps(res1)
    cusps_b = clean_cusps(res2)

    comp_cusps = {}
    if cusps_a and cusps_b:
        for i in range(1, 13):
            if i in cusps_a and i in cusps_b:
                c1, c2 = cusps_a[i], cusps_b[i]
                diff = abs(c1 - c2)
                mid = (c1 + c2) / 2.0
                if diff > 180:
                    mid = (mid + 180.0) % 360.0
                if mode == 'anti':
                    mid = (mid + 180.0) % 360.0
                comp_cusps[i] = mid

    s1_all = {**res1.get('planets', {}), **res1.get('angles', {}), **res1.get('asteroids', {}), **res1.get('lilith_nodes', {})}
    s2_all = {**res2.get('planets', {}), **res2.get('angles', {}), **res2.get('asteroids', {}), **res2.get('lilith_nodes', {})}
    
    for key in set(s1_all.keys()) & set(s2_all.keys()):
        if key not in comp_coords:
            lon1, lon2 = s1_all[key].get('longitude'), s2_all[key].get('longitude')
            if lon1 is not None and lon2 is not None:
                l1, l2 = float(lon1) % 360, float(lon2) % 360
                diff = abs(l1 - l2)
                mid = (l1 + l2) / 2.0 if diff <= 180 else ((l1 + l2 + 360) / 2.0) % 360.0
                if mode == 'anti': mid = (mid + 180) % 360
                comp_coords[key] = mid

    all_engine_data = {}
    for key, lon in comp_coords.items():
        if key in s1_all:
            sign_idx = int(lon / 30) % 12
            deg_in_sign = lon % 30
            sign_name = TROPICAL_SIGNS[sign_idx]
            
            found_house = "-"
            if comp_cusps:
                for h_num in range(1, 13):
                    cur = comp_cusps[h_num]
                    nxt = comp_cusps[h_num + 1] if h_num < 12 else comp_cusps[1]
                    if cur < nxt:
                        if cur <= lon < nxt: 
                            found_house = str(h_num)
                            break
                    else: 
                        if cur <= lon < 360 or 0 <= lon < nxt: 
                            found_house = str(h_num)
                            break

            p_data = s1_all[key].copy()
            p_data.update({
                "longitude": lon,
                "dms": format_dms_pretty(lon),
                "is_anaretic": deg_in_sign >= 29.0,
                "house": found_house,
                "duad": SYMBOL_MAP.get(get_duad(sign_name, deg_in_sign), "-"),
                "dodeca": SYMBOL_MAP.get(get_dodecatemoria(deg_in_sign), "-"),
                "decan": SYMBOL_MAP.get(get_decan(sign_name, deg_in_sign), "-"),
                "bound": SYMBOL_MAP.get(get_egyptian_bounds(sign_name, deg_in_sign), "-"),
                "sabian_index": get_sabian_index(lon)
            })
            all_engine_data[key] = p_data

    lang_val = chart_data.get("language", meta.get("language", "en"))
    is_ko = "ko" in str(lang_val).lower() or "kr" in str(lang_val).lower()

    sabian_dict = {}
    for spath in SABIAN_PATHS:
        if os.path.exists(spath):
            try:
                with open(spath, 'r', encoding='utf-8') as f: sabian_dict = json.load(f)
                break
            except: pass
    
    def get_sabian_text(idx):
        if idx is None or str(idx).strip() == "": return ""
        idx_str = str(idx)
        def extract_text(entry):
            if isinstance(entry, str): return entry
            if isinstance(entry, dict):
                primary_keys = ["text_ko", "ko", "desc_ko"] if is_ko else ["text_en", "en", "desc_en"]
                fallback_keys = ["text", "desc", "meaning"]
                for k in primary_keys + fallback_keys:
                    if k in entry and entry[k]: return str(entry[k])
                for k, v in entry.items():
                    if k != "index" and isinstance(v, str): return v
            return ""
        if isinstance(sabian_dict, dict):
            if idx_str in sabian_dict: return extract_text(sabian_dict[idx_str])
            for k, v in sabian_dict.items():
                if isinstance(v, dict) and str(v.get("index", "")) == idx_str: return extract_text(v)
        return ""

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    comp_type_str = "ANTI-COMPOSITE" if mode == "anti" else "COMPOSITE"
    comp_type_cell = mapping.get("metadata", {}).get("comp_type")
    if comp_type_cell:
        ws[comp_type_cell] = comp_type_str
        apply_grimoire_styles(ws[comp_type_cell], comp_type_str, skip_color=True)

    sys_str = system.upper()
    sys_tab_cell = mapping.get("metadata", {}).get("sys_tab")
    if sys_tab_cell:
        ws[sys_tab_cell] = sys_str
        apply_grimoire_styles(ws[sys_tab_cell], sys_str, skip_color=True)
    
    ayan_cell = mapping.get("metadata", {}).get("ayanamsa")
    if system.upper() == "SIDEREAL" and ayan_cell:
        ws[ayan_cell] = ayanamsa.upper()
        apply_grimoire_styles(ws[ayan_cell], ayanamsa.upper(), skip_color=True)

    # 🚀 [수복 1]: 명단 작성 시 Node들을 무조건 포함 (데이터 유무에 상관없이 물리적 삭제 대상 확보)
    all_bodies = []
    for category in ["planets", "asteroids", "lilith_nodes", "fates", "angles"]: 
        for body_name, map_data in mapping.get(category, {}).items():
            # 엔진용 키 변환
            engine_key = body_name
            if engine_key == "North Node (m)": engine_key = "Rahu"
            elif engine_key == "South Node (m)": engine_name = "Ketu"
            
            p_data = all_engine_data.get(engine_key)
            
            # Node 여부 판별 (이름에 Node, Rahu, Ketu가 포함되면 무조건 삭제 대상 명단에 올림)
            is_node_candidate = any(x in body_name for x in ["Node", "Rahu", "Ketu"])
            
            if (p_data and isinstance(p_data, dict)) or is_node_candidate: 
                all_bodies.append({
                    "name": body_name,
                    "map": map_data if isinstance(map_data, dict) else {},
                    "data": p_data or {}
                })

    # 🚀 [수복 2]: 아래쪽 행부터 지워야 인덱스가 안 꼬이므로 역순 정렬
    all_bodies.sort(key=lambda x: x["map"].get("row_start", 0), reverse=True)

    for body in all_bodies:
        b_map = body["map"]
        p_data = body["data"] 
        body_name = body["name"]
        row_idx = b_map.get("row_start", 0)
        
        if not row_idx: continue

        # 🚀 [수복 3]: Draconic/Ketunic 시 확실한 물리적 행 삭제
        is_node = any(x in body_name for x in ["Node", "Rahu", "Ketu"])
        if system.upper() in ["DRACONIC", "KETUNIC"] and is_node:
            ws.delete_rows(row_idx)
            continue
            
        # 데이터가 없는 경우(노드가 아닌데 연산에서 빠진 경우 등) 안전하게 스킵
        if not p_data: continue
        
        col_info = get_col(b_map.get("info", "B"))
        col_ruler = get_col(b_map.get("ruler", "C"))
        col_dignity = get_col(b_map.get("dignity", "D"))
        col_house = get_col(b_map.get("house", "E"))
        col_duad = get_col(b_map.get("duad", "F"))
        col_dodeca = get_col(b_map.get("dodeca", "G"))
        col_decan = get_col(b_map.get("decan", "H"))
        col_bounds = get_col(b_map.get("bounds", "I"))
        col_sabian = get_col(b_map.get("sabian", "J"))
        
        is_anaretic = p_data.get("is_anaretic", False)
        info_text = p_data.get("dms", "")
            
        info_cell = ws[f"{col_info}{row_idx}"]
        info_cell.value = info_text

        raw_ruler = str(p_data.get("ruler", ""))
        ws[f"{col_ruler}{row_idx}"] = RULER_SYMBOLS.get(raw_ruler.capitalize(), raw_ruler)
        
        raw_dignity = str(p_data.get("dignity", ""))
        if not raw_dignity and body_name in ["Ketu", "South Node (t)", "South Node (m)"]:
            raw_dignity = "None"
        ws[f"{col_dignity}{row_idx}"] = raw_dignity
        
        # House 중앙 정렬 수복
        ws[f"{col_house}{row_idx}"] = str(p_data.get("house", ""))
        ws[f"{col_house}{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")

        ws[f"{col_duad}{row_idx}"] = str(p_data.get("duad", ""))
        ws[f"{col_dodeca}{row_idx}"] = str(p_data.get("dodeca", ""))
        ws[f"{col_decan}{row_idx}"] = str(p_data.get("decan", ""))
        ws[f"{col_bounds}{row_idx}"] = str(p_data.get("bound", ""))
        
        s_idx = p_data.get("sabian_index")
        if not s_idx and "longitude" in p_data and p_data["longitude"] is not None:
            try: s_idx = int(math.floor(float(p_data["longitude"]))) + 1
            except: pass
            
        s_val_str = str(s_idx).strip() if s_idx else ""
        sabian_text = ""
        if s_val_str.isdigit(): sabian_text = get_sabian_text(s_val_str)
        if not sabian_text and s_val_str.isdigit():
            sabian_text = f"[{s_val_str}°] Symbol rendering fallback"

        ws[f"{col_sabian}{row_idx}"] = sabian_text

        # B열부터 J열까지 스타일링 확장 적용
        for c_char in "BCDEFGHIJ":
            target_cell = ws[f"{c_char}{row_idx}"]
            is_info = (c_char == col_info) 
            apply_grimoire_styles(
                target_cell, 
                target_cell.value, 
                is_info_col=is_info, 
                skip_color=(c_char == col_sabian),
                is_anaretic=is_anaretic 
            )

    # 🚀 최종 폰트 수복 (Consolas 강제)
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name="Consolas", 
                    size=cell.font.size, 
                    bold=cell.font.bold,
                    italic=cell.font.italic, 
                    color=cell.font.color
                )

    return wb