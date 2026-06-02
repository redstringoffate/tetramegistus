import json
import os
import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz, format_dms_pretty

from core.astrology.aspects import calculate_all_aspects
from core.astrology.patterns import find_patterns

# 🚀 A5(Albedo) 전용 추가 임포트
from api.astrology import get_seed_from_request
from core.astrology.davison import calculate_davison_midpoint

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a5_unus_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a5_unus.xlsx'))

ASPECT_ORDER = {
    "Conjunction": 1, "Opposition": 2, "Trine": 3, "Square": 4, "Sextile": 5,
    "Quintile": 6, "Septile": 7, "Quincunx": 8, "Octile": 9, "Novile": 10, 
    "Decile": 11, "Undecile": 12, "Semi-sextile": 13
}

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def get_body_info(pt_lon):
    if pt_lon is None: return "-"
    return format_dms_pretty(float(pt_lon))

def sanitize_unicode(text):
    if isinstance(text, str):
        return text.replace("\ufe0e", "").replace("\ufe0f", "")
    return text

def compile_a5_unus_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a5_unus"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()

    dav_hydrated = calculate_davison_midpoint(s1_raw, s2_raw)

    # "Davison Midpoint" 문자열 대신 정확한 좌표(Lat/Lon) 포맷으로 강제 덮어쓰기
    d_lat = get_safe_float(dav_hydrated.get("lat"), None)
    d_lng = get_safe_float(dav_hydrated.get("lng"), None)
    
    if d_lat is not None and d_lng is not None:
        coord_str = f"{abs(d_lat):.2f}°{'N' if d_lat>=0 else 'S'}, {abs(d_lng):.2f}°{'E' if d_lng>=0 else 'W'}"
    else:
        coord_str = "Unknown Location"
        
    dav_hydrated["city"] = coord_str
    dav_hydrated["location"] = coord_str
    dav_hydrated["city_name"] = coord_str
    dav_hydrated["location_name"] = coord_str

    apply_natal_stamp(ws, dav_hydrated, method="single", cells=["A2"])

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")

    date_str = str(dav_hydrated.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(dav_hydrated.get('birth_time', '12:00:00'))
    lat = get_safe_float(dav_hydrated.get("lat"), 37.5665)
    lng = get_safe_float(dav_hydrated.get("lng"), 126.9780)
    tz_val = dav_hydrated.get('tz') if dav_hydrated.get('tz') is not None else dav_hydrated.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)
    is_unk = bool(dav_hydrated.get("is_time_unknown", 0))

    res_p = calculate_principia(date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz, system=system, ayanamsa=ayanamsa, view=view, h_sys=h_sys, fixed_star_orb=1.0, is_time_unknown=is_unk)
    arc_p = calculate_arcana(date_str, time_str, lat, lng, timezone=tz, lot_schema='paulus', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    try: arc_v = calculate_arcana(date_str, time_str, lat, lng, timezone=tz, lot_schema='valens', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    except: arc_v = {'lots': {}}

    points_to_plot = {}
    for group in ['planets', 'asteroids', 'lilith_nodes']:
        for k, v in res_p.get(group, {}).items():
            name = k
            if k in ["Mean Node", "North Node (m)", "North Node (t)"]: name = "Rahu"
            elif k in ["South Node", "South Node (m)", "South Node (t)"]: name = "Ketu"
            elif k == "Lilith": name = "Asteroid Lilith"
            elif k == "Lilith (mean)": name = "Mean Lilith"
            elif k == "Lilith (true)": name = "True Lilith"
            elif k == "Eros": name = "Asteroid Eros"
            points_to_plot[name] = v

    for k, v in arc_p.get('lots', {}).items(): points_to_plot[k] = v
    for k, v in arc_v.get('lots', {}).items():
        if k in ['Necessity', 'Eros']: points_to_plot[f"{k} (v)"] = v
    
    vx = arc_p.get('vertex', {}).get('Vertex')
    if vx:
        points_to_plot['Vertex'] = vx
        points_to_plot['Anti-Vertex'] = {'longitude': (vx.get('value', vx.get('longitude', 0)) + 180) % 360}
    
    syz = arc_p.get('syzygy', {}).get('data')
    if syz: points_to_plot['Syzygy'] = syz

    flat_points = {}
    for k, v in points_to_plot.items():
        # 🚀 [여기에 추가] 약어 스킵 (혹시 모를 점 없는 버전도 포함)
        if k in ["Asc.", "I.C.", "Dsc.", "M.C.", "Asc", "IC", "Dsc", "MC"]:
            continue
            
        lon = v.get('longitude', v.get('value')) if isinstance(v, dict) else v
        if lon is not None:
            flat_points[k] = float(lon)

    try: all_aspects = calculate_all_aspects(flat_points, mode='unus')
    except Exception as e: all_aspects = []
        
    try: patterns = find_patterns(all_aspects, mode='unus')
    except Exception as e: patterns = []

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: mapping = json.load(f)

    aspect_cols = mapping.get("aspects_layout", {}).get("cols", {})
    col_a, col_b, col_c = aspect_cols.get("body", "A"), aspect_cols.get("aspect", "B"), aspect_cols.get("body_info", "C")
    col_d, col_e, col_f = aspect_cols.get("object", "D"), aspect_cols.get("object_info", "E"), aspect_cols.get("orb", "F")
    aspect_rows_map = mapping.get("aspects_layout", {}).get("rows", {})
    
    sorted_bodies = sorted(aspect_rows_map.items(), key=lambda x: x[1], reverse=True)

    for body_name, row_idx in sorted_bodies:
        body_asps = []
        for a in all_aspects:
            if a['p1'] == body_name and a['p2'] != body_name:
                body_asps.append({'asp': a['aspect'], 'obj': a['p2'], 'orb': a['orb']})
            elif a['p2'] == body_name and a['p1'] != body_name:
                body_asps.append({'asp': a['aspect'], 'obj': a['p1'], 'orb': a['orb']})
                
        body_asps.sort(key=lambda x: (ASPECT_ORDER.get(x['asp'], 99), abs(x.get('orb', 0))))
        
        b_info = get_body_info(flat_points.get(body_name))
        num_asps = len(body_asps)
        
        if num_asps == 0:
            ws[f"{col_c}{row_idx}"] = b_info
            for c_char in [col_b, col_d, col_e, col_f]: ws[f"{c_char}{row_idx}"] = ""
            for c_char in [col_b, col_c, col_d, col_e, col_f]:
                cell = ws[f"{c_char}{row_idx}"]
                is_info = (c_char in [col_c, col_e])
                
                safe_val = sanitize_unicode(cell.value)
                apply_grimoire_styles(cell, safe_val, is_info_col=is_info, skip_color=False)
                
                color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
                cell.font = Font(name="Consolas", size=9 if is_info else 11, color=color_to_use, bold=False)
                align = 'center' if c_char == col_f else 'left'
                cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
            continue
            
        if num_asps > 1:
            ws.insert_rows(row_idx + 1, amount=num_asps - 1)
            for r in range(1, num_asps):
                insert_idx = row_idx + r
                ws.row_dimensions[insert_idx].height = 16.5
                for c_char in [col_a, col_b, col_c, col_d, col_e, col_f]:
                    src = ws[f"{c_char}{row_idx}"]
                    tgt = ws[f"{c_char}{insert_idx}"]
                    if src.has_style:
                        tgt.border, tgt.fill, tgt.font, tgt.alignment, tgt.number_format = copy(src.border), copy(src.fill), copy(src.font), copy(src.alignment), copy(src.number_format)
                    if c_char == col_a: tgt.value = "" 
        
        for i, asp_data in enumerate(body_asps):
            r_idx = row_idx + i
            asp_name = asp_data['asp']
            obj_name = asp_data['obj']
            obj_info = get_body_info(flat_points.get(obj_name))
            orb_val = f"{abs(asp_data.get('orb', 0)):.2f}°"
            
            ws[f"{col_b}{r_idx}"] = asp_name
            ws[f"{col_c}{r_idx}"] = b_info if i == 0 else "" 
            ws[f"{col_d}{r_idx}"] = obj_name
            ws[f"{col_e}{r_idx}"] = obj_info
            ws[f"{col_f}{r_idx}"] = orb_val
            
            for c_char in [col_b, col_c, col_d, col_e, col_f]:
                cell = ws[f"{c_char}{r_idx}"]
                is_info = (c_char in [col_c, col_e])
                
                safe_val = sanitize_unicode(cell.value)
                skip_c = True if c_char == col_d else False
                apply_grimoire_styles(cell, safe_val, is_info_col=is_info, skip_color=skip_c)
                
                color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
                cell.font = Font(name="Consolas", size=9 if is_info else 11, color=color_to_use, bold=False)
                align = 'center' if c_char == col_f else 'left'
                cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)

    pat_map = mapping.get("patterns", {})
    pat_start = pat_map.get("row_start", 7)
    pat_cols = pat_map.get("cols", {})
    col_shape = pat_cols.get("shape", "A")
    
    if not patterns:
        cell = ws[f"{col_shape}{pat_start}"]
        cell.value = "No Shape Detected"
        cell.font = Font(name="Consolas", color="000000", bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for i in range(1, 7):
            ws[f"{pat_cols.get(f'p{i}')}{pat_start}"] = ""
            ws[f"{pat_cols.get(f'i{i}')}{pat_start}"] = ""
    else:
        patterns.sort(key=lambda x: x.get('shape', 'Unknown'))

        num_pats = len(patterns)
        if num_pats > 1:
            ws.insert_rows(pat_start + 1, amount=num_pats - 1)
            for r in range(1, num_pats):
                insert_idx = pat_start + r
                ws.row_dimensions[insert_idx].height = 16.5
                for c_char in "ABCDEFGHIJKLM":
                    src = ws[f"{c_char}{pat_start}"]
                    tgt = ws[f"{c_char}{insert_idx}"]
                    if src.has_style:
                        tgt.border, tgt.fill, tgt.font, tgt.alignment, tgt.number_format = copy(src.border), copy(src.fill), copy(src.font), copy(src.alignment), copy(src.number_format)
        
        for i, pat in enumerate(patterns):
            r_idx = pat_start + i
            
            shape_name = pat.get('shape', 'Unknown')
            if i > 0 and patterns[i-1].get('shape') == shape_name:
                ws[f"{col_shape}{r_idx}"] = ""
            else:
                ws[f"{col_shape}{r_idx}"] = shape_name
            
            for pt_i in range(1, 7):
                c_name = pat_cols.get(f"p{pt_i}")
                c_info = pat_cols.get(f"i{pt_i}")
                p_name = pat.get(f'p{pt_i}', '-')
                
                if p_name and p_name != '-':
                    ws[f"{c_name}{r_idx}"] = p_name
                    ws[f"{c_info}{r_idx}"] = get_body_info(flat_points.get(p_name))
                else:
                    ws[f"{c_name}{r_idx}"] = ""
                    ws[f"{c_info}{r_idx}"] = ""
                    
            for c_char in "ABCDEFGHIJKLM":
                cell = ws[f"{c_char}{r_idx}"]
                is_info = c_char in ["C", "E", "G", "I", "K", "M"]
                is_name = c_char in ["B", "D", "F", "H", "J", "L"]
                
                safe_val = sanitize_unicode(cell.value)
                skip_c = True if (is_name or c_char == "A") else False
                apply_grimoire_styles(cell, safe_val, is_info_col=is_info, skip_color=skip_c)
                
                color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
                cell.font = Font(name="Consolas", size=9 if is_info else 11, color=color_to_use, bold=False)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # Visual Border 고정
    for r_idx in range(1, ws.max_row + 1):
        cell_val = ws[f"A{r_idx}"].value
        if str(cell_val).strip() == ".":
            ws.row_dimensions[r_idx].height = 4.5
            ws[f"A{r_idx}"].value = ""
        elif r_idx >= 7:
            if ws.row_dimensions[r_idx].height != 4.5:
                ws.row_dimensions[r_idx].height = 16.5

    return wb