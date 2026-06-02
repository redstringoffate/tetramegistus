import json
import os
import shutil
import math
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import calculate_principia, _ensure_float_tz, format_dms_pretty

# 🚀 작성자님이 만들어둔 완벽한 Composite 엔진!
from core.astrology.composite import calculate_composite_chart, apply_anti_composite

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a3_comp_domain_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a3_comp_domain.xlsx'))

PLANETS_LIST = ["Sun", "Moon", "Mercury", "Venus", "Mars", "Jupiter", "Saturn", "Uranus", "Neptune", "Pluto", "Mean Lilith", "True Lilith", "Rahu", "Ketu", "North Node (t)", "South Node (t)"]
ASTEROIDS_LIST = ["Chiron", "Ceres", "Juno", "Pallas", "Vesta", "Asteroid Eros", "Psyche", "Asteroid Lilith"]

def get_col(cell_str):
    if not cell_str: return ""
    return "".join([c for c in str(cell_str) if c.isalpha()])

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def format_house_range(lon_start, lon_end):
    if lon_start is None or lon_end is None: return ""
    size = (lon_end - lon_start) % 360
    if size < 0: size += 360
    rounded_size = round(size)
    
    portions = []
    curr = lon_start
    rem = size
    while rem > 1e-5:
        next_boundary = (math.floor(curr / 30) + 1) * 30
        step = min(rem, next_boundary - curr)
        portions.append(step)
        curr = (curr + step) % 360
        rem -= step
        
    rounded_portions = [round(p) for p in portions]
    diff = rounded_size - sum(rounded_portions)
    if diff != 0 and rounded_portions:
        rounded_portions[rounded_portions.index(max(rounded_portions))] += diff
        
    portions_str = "/".join(map(str, rounded_portions))
    return f"{rounded_size} ({portions_str})"

def format_body_string(item):
    lon = float(item.get('longitude', 0))
    dms_str = format_dms_pretty(lon, is_retro=False)
    return f"{item['name']} - {dms_str}"

def compile_a3_comp_domain_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 🚀 [수복]: 메모리로 즉시 로드 (새 아키텍처 이식)
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a3_comp_domain"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name
    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")
    mode = meta.get("mode", "normal") 

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()

    def ensure_location_string(seed):
        keys = ["location_name", "city_name", "location", "city", "place"]
        loc_val = ""
        for k in keys:
            if seed.get(k):
                loc_val = str(seed.get(k)).strip()
                break
        
        if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
            lat = get_safe_float(seed.get("lat"), None)
            lng = get_safe_float(seed.get("lng"), None)
            if lat is not None and lng is not None:
                loc_val = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
            else:
                loc_val = "Unknown Location"
                
        seed["city"] = loc_val
        seed["location"] = loc_val
        seed["city_name"] = loc_val
        seed["location_name"] = loc_val

    ensure_location_string(s1_raw)
    ensure_location_string(s2_raw)

    apply_natal_stamp(ws, s1_raw, method="single", cells=["A2"])
    apply_natal_stamp(ws, s2_raw, method="single", cells=["A3"])

    def map_for_engine(seed):
        date_str = str(seed.get('birth_date', '2000-01-01')).split('T')[0]
        time_str = str(seed.get('birth_time', '12:00:00'))
        tz_val = seed.get('tz') if seed.get('tz') is not None else seed.get('timezone', 9.0)
        return {
            "date_str": date_str, "time_str": time_str,
            "lat": get_safe_float(seed.get("lat"), 37.5665),
            "lng": get_safe_float(seed.get("lng"), 126.9780),
            "timezone": _ensure_float_tz(tz_val, date_str),
            "system": system, "ayanamsa": ayanamsa, "view": view, "h_sys": h_sys,
            "fixed_star_orb": 1.0, "is_time_unknown": bool(seed.get("is_time_unknown", 0))
        }

    res1 = calculate_principia(**map_for_engine(s1_raw))
    res2 = calculate_principia(**map_for_engine(s2_raw))

    # 🚀 엔진에서 플립 방어 로직 그대로 불러오기
    p_a = {**res1.get('planets', {}), **res1.get('angles', {}), **res1.get('asteroids', {}), **res1.get('lilith_nodes', {})}
    p_b = {**res2.get('planets', {}), **res2.get('angles', {}), **res2.get('asteroids', {}), **res2.get('lilith_nodes', {})}
    
    c_a = res1.get('houses', {})
    c_b = res2.get('houses', {})

    comp_chart = calculate_composite_chart(p_a, p_b, c_a, c_b)
    comp_planets = comp_chart.get('planets', {})
    comp_cusps = comp_chart.get('houses', {})

    if mode == 'anti':
        comp_planets, comp_cusps = apply_anti_composite(comp_planets, comp_cusps)

    # 🚀 [여기 추가]: Draconic / Ketunic 시스템일 경우 교점(Nodes) 엑셀에서 원천 삭제
    if system in ['draconic', 'ketunic']:
        for n_key in ['Rahu', 'Ketu', 'North Node (m)', 'South Node (m)', 'North Node (t)', 'South Node (t)']:
            comp_planets.pop(n_key, None)

    is_unk = bool(s1_raw.get("is_time_unknown", 0)) or bool(s2_raw.get("is_time_unknown", 0))
    cusps_simple = {}
    if not is_unk:
        for i in range(1, 13):
            val = comp_cusps.get(i) or comp_cusps.get(str(i))
            if val is not None:
                cusps_simple[i] = float(val)

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    comp_type_str = "ANTI-COMPOSITE" if mode == "anti" else "COMPOSITE"
    if mapping["metadata"].get("comp_type"): ws[mapping["metadata"]["comp_type"]] = comp_type_str
    if mapping["metadata"].get("sys_tab"): ws[mapping["metadata"]["sys_tab"]] = system.upper()
    if system.upper() == "SIDEREAL" and mapping["metadata"].get("ayanamsa"):
        ayan_cell = ws[mapping["metadata"]["ayanamsa"]]
        ayan_cell.value = ayanamsa.upper()
        apply_grimoire_styles(ayan_cell, ayanamsa.upper(), skip_color=True)
        # 🚀 우측 정렬 강제 적용
        ayan_cell.alignment = Alignment(horizontal='right', vertical='center')

    comp_asc_lon = cusps_simple.get(1)
    comp_mc_lon = cusps_simple.get(10)
    
    if mapping["metadata"].get("ascendant"): ws[mapping["metadata"]["ascendant"]] = format_dms_pretty(comp_asc_lon) if comp_asc_lon is not None else "-"
    if mapping["metadata"].get("midheaven"): ws[mapping["metadata"]["midheaven"]] = format_dms_pretty(comp_mc_lon) if comp_mc_lon is not None else "-"

    for key in ["comp_type", "ascendant", "midheaven", "sys_tab", "ayanamsa"]:
        c_ref = mapping["metadata"].get(key)
        if c_ref and ws[c_ref].value:
            cell = ws[c_ref]
            should_skip_color = True if key in ["comp_type", "sys_tab", "ayanamsa"] else False
            apply_grimoire_styles(cell, cell.value, is_info_col=False, skip_color=should_skip_color)
            cell.font = Font(name="Consolas", color=cell.font.color if cell.font else "000000", bold=False, italic=False)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    def get_house_of_point(lon):
        if is_unk or not cusps_simple: return None
        lon = float(lon) % 360
        for i in range(1, 13):
            curr_c = cusps_simple[i]
            next_c = cusps_simple[i+1] if i < 12 else cusps_simple[1]
            if curr_c < next_c:
                if curr_c <= lon < next_c: return i
            else:
                if curr_c <= lon < 360 or 0 <= lon < next_c: return i
        return 1

    h_items = {i: {"P": [], "A": []} for i in range(1, 13)}
    seen_in_house = {i: set() for i in range(1, 13)}

    for p_key, p_data in comp_planets.items():
        norm_name = p_data.get('name', p_key)
        if norm_name in ["Mean Node", "North Node (m)"]: norm_name = "Rahu"
        elif norm_name in ["South Node", "South Node (m)"]: norm_name = "Ketu"
        elif norm_name == "Eros": norm_name = "Asteroid Eros"
        elif norm_name == "Lilith": norm_name = "Asteroid Lilith"
        elif norm_name == "Lilith (mean)": norm_name = "Mean Lilith"
        elif norm_name == "Lilith (true)": norm_name = "True Lilith"
        
        lon = float(p_data.get('longitude', 0))
        h_num = get_house_of_point(lon)
        if not h_num: continue
        
        # 🚀 [방역]: 이미 해당 하우스에 추가된 이름이면 스킵 (중복 차단)
        if norm_name in seen_in_house[h_num]:
            continue
        seen_in_house[h_num].add(norm_name)
        
        # 아나레틱 판정 저장
        is_an = int(float(lon) % 30) == 29
        
        d_copy = {"name": norm_name, "longitude": lon}
        if norm_name in PLANETS_LIST: h_items[h_num]["P"].append(d_copy)
        elif norm_name in ASTEROIDS_LIST: h_items[h_num]["A"].append(d_copy)

    houses_list = [{"house": i, "map": mapping["houses"][str(i)]} for i in range(1, 13) if str(i) in mapping.get("houses", {})]
    houses_list.sort(key=lambda x: x["map"].get("row_start", 0), reverse=True)

    for item in houses_list:
        h_num, h_map = item["house"], item["map"]
        row_idx = h_map.get("row_start", 0)
        if not row_idx: continue

        col_info = get_col(h_map.get("info", "B"))
        col_range = get_col(h_map.get("range", "C"))
        col_p = get_col(h_map.get("planets", "D"))
        col_a = get_col(h_map.get("asteroids", "E"))

        lon = cusps_simple.get(h_num)
        if lon is None:
            for c_char in [col_info, col_range, col_p, col_a]: 
                if c_char: ws[f"{c_char}{row_idx}"] = ""
            continue

       # 🚀 [수복 완료]: 하우스 커스프(lon) 기준 상대 거리로 정렬
        p_items = sorted(h_items[h_num]["P"], key=lambda x: (float(x.get('longitude', 0)) - lon) % 360)
        a_items = sorted(h_items[h_num]["A"], key=lambda x: (float(x.get('longitude', 0)) - lon) % 360)
        
        p_strs = [format_body_string(x) for x in p_items]
        a_strs = [format_body_string(x) for x in a_items]
        
        max_rows = max(1, len(p_strs), len(a_strs))

        if max_rows > 1:
            ws.insert_rows(row_idx + 1, amount=max_rows - 1)
            for r in range(1, max_rows):
                insert_idx = row_idx + r
                ws.row_dimensions[insert_idx].height = 16.5
                for c_char in [col_info, col_range, col_p, col_a]:
                    src = ws[f"{c_char}{row_idx}"]
                    tgt = ws[f"{c_char}{insert_idx}"]
                    if src.has_style:
                        tgt.border = copy(src.border)
                        tgt.fill = copy(src.fill)
                        tgt.font = copy(src.font)
                        tgt.alignment = copy(src.alignment)
                        tgt.number_format = copy(src.number_format)

        for i in range(max_rows):
            r_idx = row_idx + i
            ws[f"{col_p}{r_idx}"] = p_strs[i] if i < len(p_strs) else ""
            ws[f"{col_a}{r_idx}"] = a_strs[i] if i < len(a_strs) else ""

            # 🚀 [수정]: 행성(col_p) 및 소행성(col_a) 스타일링 로직
            col_targets = [(col_p, p_items), (col_a, a_items)]
            for c_char, item_list in col_targets:
                cell = ws[f"{c_char}{r_idx}"]
                text_val = str(cell.value) if cell.value else ""
                
                if text_val and i < len(item_list):
                    item_data = item_list[i]
                    # 스타일러가 문자열에서 원소명(Taurus 등)만 찾아내도록 skip_planet_color 적용
                    apply_grimoire_styles(
                        cell, 
                        text_val, 
                        is_info_col=True, 
                        skip_color=False,
                        is_anaretic=item_data.get("is_anaretic", False),
                        skip_planet_color=True # 🚩 행성색(분홍, 파랑 등) 무시 옵션
                    )
                
                # Consolas 폰트 및 정렬 고정
                cell.font = Font(name="Consolas", color=cell.font.color if cell.font else "000000", bold=False)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            
            if i == 0:
                ws[f"{col_info}{r_idx}"] = format_dms_pretty(lon)
                next_h = 1 if h_num == 12 else h_num + 1
                next_lon = cusps_simple.get(next_h)
                ws[f"{col_range}{r_idx}"] = format_house_range(lon, next_lon) if next_lon is not None else ""
                
                c_info_cell, c_range_cell = ws[f"{col_info}{r_idx}"], ws[f"{col_range}{r_idx}"]
                
                apply_grimoire_styles(c_info_cell, c_info_cell.value, is_info_col=True, skip_color=False)
                color_info = c_info_cell.font.color if c_info_cell.font and c_info_cell.font.color else "000000"
                c_info_cell.font = Font(name="Consolas", color=color_info, bold=False, italic=False)
                c_info_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                
                apply_grimoire_styles(c_range_cell, c_range_cell.value, is_info_col=True, skip_color=False)
                color_range = c_range_cell.font.color if c_range_cell.font and c_range_cell.font.color else "000000"
                c_range_cell.font = Font(name="Consolas", color=color_range, bold=False, italic=False)
                c_range_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            else:
                ws[f"{col_info}{r_idx}"] = ""
                ws[f"{col_range}{r_idx}"] = ""

    ws.column_dimensions[col_p].width = 38.0
    ws.column_dimensions[col_a].width = 38.0

    # 🚀 [Airtight]: 모든 셀의 폰트를 Consolas로 일괄 세탁하여 맑은고딕의 흔적을 지움
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name="Consolas", size=cell.font.size, bold=cell.font.bold,
                    italic=cell.font.italic, color=cell.font.color
                )

    return wb