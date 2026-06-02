import json
import os
import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import Color

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.engine import calculate_divisio, _ensure_float_tz

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/n6_varga_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/nigredo/n6_varga.xlsx'))

# 🚀 프론트엔드 영문명 -> 산스크리트어(Graha) 번역 맵
REVERSE_GRAHA_MAP = {
    'Ascendant': 'LAGNA', 'Sun': 'SURYA', 'Moon': 'CHANDRA', 
    'Mercury': 'BUDHA', 'Venus': 'SHUKRA', 'Mars': 'MANGALA', 
    'Jupiter': 'BRIHASPATI', 'Saturn': 'SHANI',
    'South Node (t)': 'KETU', 'North Node (t)': 'RAHU',
    'Ketu': 'KETU', 'Rahu': 'RAHU'
}

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def sanitize_unicode(text):
    if isinstance(text, str):
        return text.replace("\ufe0e", "").replace("\ufe0f", "")
    return text

def compile_n6_varga_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "n6_varga"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    seed_data = chart_data.get("seed", chart_data)

    def ensure_location_string(seed):
        keys = ["location_name", "city_name", "location", "city", "place"]
        loc_val = ""
        for k in keys:
            if seed.get(k):
                loc_val = str(seed.get(k)).strip()
                break
        if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
            lat, lng = get_safe_float(seed.get("lat"), None), get_safe_float(seed.get("lng"), None)
            if lat is not None and lng is not None: loc_val = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
            else: loc_val = "Unknown Location"
        seed["city"] = loc_val
        seed["location"] = loc_val

    ensure_location_string(seed_data)
    apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])

    meta = chart_data.get("metadata", {})
    ayanamsa = meta.get("ayanamsa", "lahiri")
    target_body = meta.get("target_body", "Sun")

    date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(seed_data.get('birth_time', '12:00:00'))
    lat = get_safe_float(seed_data.get("lat"), 37.5665)
    lng = get_safe_float(seed_data.get("lng"), 126.9780)
    tz_val = seed_data.get('tz') if seed_data.get('tz') is not None else seed_data.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)

    # Divisio Engine Call
    divisio_res = calculate_divisio(
        date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz, ayanamsa=ayanamsa
    )
    varga_data = divisio_res.get("varga", {}).get(target_body, {})

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: mapping = json.load(f)

    layout = mapping.get("layout", {})
    cols = layout.get("cols", {})
    rows = layout.get("rows", {})
    
    col_info = cols.get("info", "B")
    col_nak = cols.get("nakshatra", "C")
    col_pur = cols.get("purushartha", cols.get("sublord", "D"))

    for div_key, r_idx in rows.items():
        d_data = varga_data.get(div_key, {})
        
        info_val = str(d_data.get("formatted", "-"))
        nak_val = str(d_data.get("nakshatra", "-"))
        pur_val = str(d_data.get("purushartha", "-"))
        
        c_info = ws[f"{col_info}{r_idx}"]
        c_nak = ws[f"{col_nak}{r_idx}"]
        c_pur = ws[f"{col_pur}{r_idx}"]
        
        # 1. Info (Sign colored)
        c_info.value = info_val
        c_info.data_type = 's'
        apply_grimoire_styles(c_info, sanitize_unicode(info_val), is_info_col=True, skip_color=False)
        c_info.alignment = Alignment(horizontal='left', vertical='center')
        
        # 2. Nakshatra (Base colored)
        base_nak = nak_val.split('-')[0] if "-" in nak_val else nak_val
        c_nak.value = base_nak
        c_nak.data_type = 's'
        apply_grimoire_styles(c_nak, sanitize_unicode(base_nak), is_info_col=False, skip_color=False)
        c_nak.value = nak_val 
        c_nak.alignment = Alignment(horizontal='left', vertical='center')
        
        # 3. Purushartha (Prefix custom colored)
        c_pur.value = pur_val
        c_pur.data_type = 's'
        pur_base = pur_val.split('-')[0] if "-" in pur_val else pur_val
        pur_color = "000000"
        
        if pur_base == "Dharma": pur_color = "FFCC00"
        elif pur_base == "Artha": pur_color = "00CC00"
        elif pur_base == "Kama": pur_color = "FF4444"
        elif pur_base == "Moksha": pur_color = "4488FF"
        
        old_size = c_pur.font.size if c_pur.font else 11
        c_pur.font = Font(name="Consolas", size=old_size, color=pur_color)
        c_pur.alignment = Alignment(horizontal='left', vertical='center')

    # 🚀 4. Graha Text (B5: 둘 다 우측정렬 & 볼드 해제)
    graha_ref = mapping.get("metadata", {}).get("graha")
    if graha_ref:
        g_cell = ws[graha_ref]
        sanskrit_name = REVERSE_GRAHA_MAP.get(target_body, target_body).upper()
        g_cell.value = str(sanskrit_name)
        g_cell.data_type = 's'
        g_cell.number_format = '@'
        apply_grimoire_styles(g_cell, sanskrit_name, skip_color=True)
        
        color_to_use = copy(g_cell.font.color) if g_cell.font and g_cell.font.color else "000000"
        g_cell.font = Font(name="Consolas", size=11, bold=False, color=color_to_use)
        g_cell.alignment = Alignment(horizontal='right', vertical='center')

    # 🚀 5. Ayanamsa Text (B4: 둘 다 우측정렬 & 볼드 해제)
    ayan_ref = mapping.get("metadata", {}).get("ayanamsa")
    if ayan_ref:
        a_cell = ws[ayan_ref]
        a_val = str(ayanamsa).upper()
        a_cell.value = a_val
        a_cell.data_type = 's'
        a_cell.number_format = '@'
        apply_grimoire_styles(a_cell, a_val, skip_color=True)
        
        color_to_use = copy(a_cell.font.color) if a_cell.font and a_cell.font.color else "000000"
        a_cell.font = Font(name="Consolas", size=11, bold=False, color=color_to_use)
        a_cell.alignment = Alignment(horizontal='right', vertical='center')

    # Shrink Logic
    for r_idx in range(1, ws.max_row + 1):
        cell_val = ws[f"A{r_idx}"].value
        if str(cell_val).strip() == ".":
            ws.row_dimensions[r_idx].height = 4.5
            for col_char in ["A", "B", "C", "D"]:
                ws[f"{col_char}{r_idx}"].value = ""
        elif r_idx >= 7:
            if ws.row_dimensions[r_idx].height != 4.5:
                ws.row_dimensions[r_idx].height = 16.5

    # 🚀 [최종 해결책]: 엑셀의 강제 서식 복구(Calibri 회귀)를 막기 위한 '문자열 타입 강제 고정 + Consolas 복제' 콤보
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.data_type = 's' # 엑셀에게 "이건 무조건 텍스트니까 니 맘대로 서식 바꾸지 마" 라고 명령
                
            if cell.font:
                # 안전한 복제(copy)를 통해 템플릿의 색상 XML이 깨지는 것을 막습니다.
                safe_color = copy(cell.font.color) if cell.font.color else None
                cell.font = Font(name="Consolas", 
                                 size=cell.font.size, 
                                 bold=cell.font.bold, 
                                 italic=cell.font.italic, 
                                 color=safe_color)
            else:
                cell.font = Font(name="Consolas")

    # ====================================================================
    # 🚀 [수정]: 열 너비 축소(Shrink) 완벽 방어 (물리적 수치 강제 주입 + A1 해제)
    # ====================================================================
    # 1. 템플릿 의존 없이, 글자가 절대 잘리지 않을 넉넉한 고정 수치로 쐐기를 박습니다.
    forced_widths = {
        "A": 21.0, 
        "B": 20.0,  # Information (가장 긴 텍스트도 넉넉하게)
        "C": 20.0,  # Nakshatra
        "D": 20.0   # Purushartha
    }
    for col_char, target_width in forced_widths.items():
        ws.column_dimensions[col_char].width = target_width

    # 2. A1 병합 해제 및 B1으로 이사 (뷰어의 렌더링 꼬임 원천 차단)
    from openpyxl.utils import get_column_letter
    merge_to_remove = None
    for m_range in list(ws.merged_cells.ranges):
        if "A1" in m_range.coord:
            merge_to_remove = m_range
            break
            
    if merge_to_remove:
        max_col_letter = get_column_letter(merge_to_remove.max_col)
        ws.unmerge_cells(merge_to_remove.coord)
        
        ws['B1'].value = ws['A1'].value
        ws['A1'].value = None
        
        if ws['A1'].has_style:
            ws['B1'].font = copy(ws['A1'].font)
            ws['B1'].border = copy(ws['A1'].border)
            ws['B1'].fill = copy(ws['A1'].fill)
            ws['B1'].alignment = copy(ws['A1'].alignment)
            
        ws.merge_cells(f"B1:{max_col_letter}1")

    return wb