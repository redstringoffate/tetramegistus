import json
import os
import shutil
import re
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.davison import calculate_davison_midpoint # 🚀 Davison 연산 엔진 추가
from core.astrology.engine import calculate_principia, _ensure_float_tz, format_dms_pretty
from core.astrology.aspects import calculate_all_aspects

BASE_DIR = os.path.dirname(__file__)

# 🚀 A6 Aspects 경로로 수정
MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a6_aspects_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a6_aspects.xlsx'))

ASPECT_ORDER = {
    "Conjunction": 1, "Opposition": 2, "Trine": 3, "Square": 4, "Sextile": 5,
    "Quintile": 6, "Septile": 7, "Quincunx": 8, "Octile": 9, "Novile": 10, 
    "Decile": 11, "Undecile": 12, "Semi-sextile": 13
}

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def get_body_info(pt_lon):
    if pt_lon is None: return "-"
    return format_dms_pretty(float(pt_lon))

def sanitize_unicode(text):
    if isinstance(text, str):
        return text.replace("\ufe0e", "").replace("\ufe0f", "")
    return text

def compile_a6_aspects_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a6_aspects"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()

    # Davison 시드 생성
    seed_data = calculate_davison_midpoint(s1_raw, s2_raw)

    # 🚀 "Davison Midpoint" 문자열 대신 정확한 좌표(Lat/Lon) 포맷으로 강제 덮어쓰기
    d_lat = get_safe_float(seed_data.get("lat"), None)
    d_lng = get_safe_float(seed_data.get("lng"), None)
    
    if d_lat is not None and d_lng is not None:
        coord_str = f"{abs(d_lat):.2f}°{'N' if d_lat>=0 else 'S'}, {abs(d_lng):.2f}°{'E' if d_lng>=0 else 'W'}"
    else:
        coord_str = "Unknown Location"
        
    seed_data["city"] = coord_str
    seed_data["location"] = coord_str
    seed_data["city_name"] = coord_str
    seed_data["location_name"] = coord_str

    apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])

    # Harmonic 배수 추출
    raw_h_level = str(meta.get("harmonic", meta.get("h_level", "H1"))).upper()
    h_multiplier = 1
    match = re.search(r'\d+', raw_h_level)
    if match:
        h_multiplier = int(match.group())
        h_level_str = f"H{h_multiplier}"
    else:
        h_level_str = "H1"

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: mapping = json.load(f)

    date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(seed_data.get('birth_time', '12:00:00'))
    lat = get_safe_float(seed_data.get("lat"), 37.5665)
    lng = get_safe_float(seed_data.get("lng"), 126.9780)
    tz_val = seed_data.get('tz') if seed_data.get('tz') is not None else seed_data.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)
    is_unk = bool(seed_data.get("is_time_unknown", 0))

    res_p = calculate_principia(date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz, system=system, ayanamsa=ayanamsa, view=view, h_sys=h_sys, fixed_star_orb=1.0, is_time_unknown=is_unk)

    # 🚀 수정된 필터링 로직: 정확히 지정된 14개의 천체만 허용
    ALLOWED_BODIES = [
        "Sun", "Moon", "Mercury", "Venus", "Mars", 
        "Jupiter", "Saturn", "Uranus", "Neptune", "Pluto", 
        "Chiron", "Mean Lilith", "True Lilith", "North Node (t)"
    ]

    flat_points = {}
    for group in ['planets', 'asteroids', 'lilith_nodes']:
        for k, v in res_p.get(group, {}).items():
            # 1. 예외 케이스 처리 (명칭 통일)
            name = k
            if k in ["North Node", "North Node (t)", "Mean Node", "Rahu"]: 
                name = "North Node (t)"
            elif k in ["Lilith", "Asteroid Lilith"]: 
                name = "Asteroid Lilith" # 필요 시 제외 혹은 포함
            
            # 2. 화이트리스트 검증: ALLOWED_BODIES에 없는 rahu, ketu 등은 여기서 탈락
            if name in ALLOWED_BODIES:
                lon = v.get('longitude', v.get('value')) if isinstance(v, dict) else v
                if lon is not None:
                    flat_points[name] = float(lon)

    harmonic_points = {k: (v * h_multiplier) % 360.0 for k, v in flat_points.items()}

    try: all_aspects = calculate_all_aspects(harmonic_points, mode='unus')
    except: all_aspects = []

    aspect_cols = mapping.get("aspects_layout", {}).get("cols", {})
    col_a, col_b, col_c = aspect_cols.get("body", "A"), aspect_cols.get("aspect", "B"), aspect_cols.get("body_info", "C")
    col_d, col_e, col_f = aspect_cols.get("object", "D"), aspect_cols.get("object_info", "E"), aspect_cols.get("orb", "F")
    aspect_rows_map = mapping.get("aspects_layout", {}).get("rows", {})
    
    sorted_bodies = sorted(aspect_rows_map.items(), key=lambda x: x[1], reverse=True)
    dynamic_shrink_rows = list(mapping.get("aspects_layout", {}).get("shrink_rows", []))

    for body_name, row_idx in sorted_bodies:
        body_asps = []
        for a in all_aspects:
            if a['p1'] == body_name and a['p2'] != body_name:
                body_asps.append({'asp': a['aspect'], 'obj': a['p2'], 'orb': a['orb']})
            elif a['p2'] == body_name and a['p1'] != body_name:
                body_asps.append({'asp': a['aspect'], 'obj': a['p1'], 'orb': a['orb']})
        
        body_asps.sort(key=lambda x: (ASPECT_ORDER.get(x['asp'], 99), abs(x.get('orb', 0))))
        b_info = get_body_info(harmonic_points.get(body_name))
        num_asps = len(body_asps)
        
        if num_asps == 0:
            ws[f"{col_c}{row_idx}"] = b_info
            for c_char in [col_b, col_d, col_e, col_f]: ws[f"{c_char}{row_idx}"] = ""
            
            for c_char in [col_b, col_c, col_d, col_e, col_f]:
                cell = ws[f"{c_char}{row_idx}"]
                is_info = (c_char in [col_c, col_e])
                safe_val = sanitize_unicode(cell.value)
                apply_grimoire_styles(cell, safe_val, is_info_col=is_info, skip_color=False)
                
                align_h = 'center' if c_char == col_f else 'left'
                cell.alignment = Alignment(horizontal=align_h, vertical='center', wrap_text=False)
            continue
            
        if num_asps > 1:
            ws.insert_rows(row_idx + 1, amount=num_asps - 1)
            for i_sr in range(len(dynamic_shrink_rows)):
                if dynamic_shrink_rows[i_sr] > row_idx: dynamic_shrink_rows[i_sr] += (num_asps - 1)
            for r in range(1, num_asps):
                insert_idx = row_idx + r
                ws.row_dimensions[insert_idx].height = 16.5
                for c_char in [col_a, col_b, col_c, col_d, col_e, col_f]:
                    src, tgt = ws[f"{c_char}{row_idx}"], ws[f"{c_char}{insert_idx}"]
                    if src.has_style:
                        tgt.border, tgt.fill, tgt.font, tgt.alignment, tgt.number_format = copy(src.border), copy(src.fill), copy(src.font), copy(src.alignment), copy(src.number_format)
        
        for i, asp_data in enumerate(body_asps):
            r_idx = row_idx + i
            ws[f"{col_b}{r_idx}"] = asp_data['asp']
            ws[f"{col_c}{r_idx}"] = b_info if i == 0 else ""
            ws[f"{col_d}{r_idx}"] = asp_data['obj']
            ws[f"{col_e}{r_idx}"] = get_body_info(harmonic_points.get(asp_data['obj']))
            ws[f"{col_f}{r_idx}"] = f"{abs(asp_data.get('orb', 0)):.2f}°"
            
            for c_char in [col_b, col_c, col_d, col_e, col_f]:
                cell = ws[f"{c_char}{r_idx}"]
                is_info = (c_char in [col_c, col_e])
                safe_val = sanitize_unicode(cell.value)
                skip_c = (c_char == col_d)
                apply_grimoire_styles(cell, safe_val, is_info_col=is_info, skip_color=skip_c)
                
                align_h = 'center' if c_char == col_f else 'left'
                cell.alignment = Alignment(horizontal=align_h, vertical='center', wrap_text=False)

    for r_idx in dynamic_shrink_rows:
        ws.row_dimensions[r_idx].height = 4.5
        for c_char in ["A", "B", "C", "D", "E", "F"]: ws[f"{c_char}{r_idx}"].value = ""

    b4_ref = mapping.get("metadata", {}).get("h_level", "B4")
    if b4_ref:
        ws[b4_ref].value = h_level_str
        ws[b4_ref].data_type = 's'
        ws[b4_ref].number_format = '@'
        ws[b4_ref].alignment = Alignment(horizontal='center', vertical='center')

    # 🚀 [시트 전체 Consolas 강제 적용]
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(name="Consolas", 
                                 size=cell.font.size, 
                                 bold=cell.font.bold, 
                                 italic=cell.font.italic, 
                                 color=cell.font.color)
            else:
                cell.font = Font(name="Consolas")

    return wb