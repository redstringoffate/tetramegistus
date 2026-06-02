import json
import os
import shutil
import math
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz, format_dms_pretty, SYMBOL_MAP
from core.astrology.davison import calculate_davison_midpoint

from core.astrology.divisions.decan import get_decan
from core.astrology.divisions.duad import get_duad
from core.astrology.divisions.dodecatemoria import get_dodecatemoria
from core.astrology.divisions.egyptian_bounds import get_egyptian_bounds
from core.astrology.divisions.sabian_engine import get_sabian_index

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a4_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a4.xlsx'))

SABIAN_PATHS = [
    os.path.abspath(os.path.join(BASE_DIR, '../../../data/render/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../../data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../static/data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../data/sabian.json')),
]

H_SYS_MAP = {
    'P': 'Placidus', 'W': 'Whole Sign', 'K': 'Koch', 'R': 'Regiomontanus',
    'C': 'Campanus', 'E': 'Equal', 'V': 'Vehlow'
}

LOCAL_SIGNS = ["Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo", "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces"]

def get_col(cell_str):
    if not cell_str: return ""
    return "".join([c for c in str(cell_str) if c.isalpha()])

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def compile_a4_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a4"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()

    # Davison 시드 생성
    dav_hydrated = calculate_davison_midpoint(s1_raw, s2_raw)

    # 🚀 "Davison Midpoint" 문자열 대신 정확한 좌표(Lat/Lon) 포맷으로 강제 덮어쓰기 (A3 픽스 동일 적용)
    d_lat = get_safe_float(dav_hydrated.get("lat"), None)
    d_lng = get_safe_float(dav_hydrated.get("lng"), None)
    
    if d_lat is not None and d_lng is not None:
        coord_str = f"{abs(d_lat):.2f}°{'N' if d_lat>=0 else 'S'}, {abs(d_lng):.2f}°{'E' if d_lng>=0 else 'W'}"
    else:
        coord_str = "Unknown Location"
        
    dav_hydrated["city"] = coord_str
    dav_hydrated["location"] = coord_str
    dav_hydrated["city_name"] = coord_str
    dav_hydrated["location_name"] = coord_str

    # 단일 스탬프 적용
    apply_natal_stamp(ws, dav_hydrated, method="single", cells=["A2"])

    date_str = str(dav_hydrated.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(dav_hydrated.get('birth_time', '12:00:00'))
    lat = get_safe_float(dav_hydrated.get("lat"), 37.5665)
    lng = get_safe_float(dav_hydrated.get("lng"), 126.9780)
    tz_val = dav_hydrated.get('tz') if dav_hydrated.get('tz') is not None else dav_hydrated.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)
    is_unk = bool(dav_hydrated.get("is_time_unknown", 0))

    # 데이터 연산 추출
    res = calculate_principia(
        date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz,
        system=system, ayanamsa=ayanamsa, view=view, h_sys=h_sys,
        fixed_star_orb=1.0, is_time_unknown=is_unk
    )
    arc_p = calculate_arcana(date_str, time_str, lat, lng, timezone=tz, lot_schema='paulus', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    try:
        arc_v = calculate_arcana(date_str, time_str, lat, lng, timezone=tz, lot_schema='valens', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    except:
        arc_v = {'lots': {}}

    houses_data = res.get('houses', {})
    cusps_simple = {}
    if not is_unk:
        for k, v in houses_data.items():
            cusps_simple[int(k)] = float(v['longitude']) if isinstance(v, dict) else float(v)

    def get_house_of_point(lon):
        if is_unk or not cusps_simple: return None
        lon = float(lon) % 360
        for i in range(1, 13):
            curr_c = cusps_simple[i]
            next_c = cusps_simple[i+1] if i < 12 else cusps_simple[1]
            if curr_c < next_c:
                if curr_c <= lon < next_c: return i
            else:
                if curr_c <= lon < 360 or 0 <= lon < next_c: return i
        return 1

    # 🚀 Sect 계산 (Davison Sun 기준)
    sun_lon = res.get('planets', {}).get('Sun', {}).get('longitude')
    sun_house = get_house_of_point(sun_lon) if sun_lon is not None else 1
    sect_str = "DAY" if sun_house in [7, 8, 9, 10, 11, 12] else "NIGHT"

    # 아르카나 포인트 추출 맵핑
    points_to_plot = {}
    lots_p = arc_p.get('lots', {})
    lots_v = arc_v.get('lots', {})
    
    for name in ["Fortune", "Spirit", "Necessity", "Eros", "Courage", "Victory", "Nemesis"]:
        if name in lots_p: 
            points_to_plot[name] = lots_p[name].get('value', lots_p[name].get('longitude'))
            
    for name in ["Necessity", "Eros"]:
        if name in lots_v: 
            points_to_plot[f"{name} (v)"] = lots_v[name].get('value', lots_v[name].get('longitude'))
            
    vx_data = arc_p.get('vertex', {}).get('Vertex', {})
    vx_lon = vx_data.get('value', vx_data.get('longitude'))
    if vx_lon is not None:
        points_to_plot["Vertex"] = vx_lon
        points_to_plot["Anti-Vertex"] = (vx_lon + 180) % 360
        
    syz_data = arc_p.get('syzygy', {}).get('data', {})
    syz_lon = syz_data.get('value', syz_data.get('longitude'))
    if syz_lon is not None:
        points_to_plot["Syzygy"] = syz_lon

    sabian_dict = {}
    for spath in SABIAN_PATHS:
        if os.path.exists(spath):
            try:
                with open(spath, 'r', encoding='utf-8') as f: sabian_dict = json.load(f)
                break
            except: pass

    lang_val = chart_data.get("language", meta.get("language", "en"))
    is_ko = "ko" in str(lang_val).lower() or "kr" in str(lang_val).lower()
    
    def get_sabian_text(idx):
        if not idx: return ""
        idx_str = str(idx)
        def extract_text(entry):
            if isinstance(entry, str): return entry
            if isinstance(entry, dict):
                p_keys = ["text_ko", "ko", "desc_ko"] if is_ko else ["text_en", "en", "desc_en"]
                for k in p_keys + ["text", "desc", "meaning"]:
                    if k in entry and entry[k]: return str(entry[k])
                for k, v in entry.items():
                    if k != "index" and isinstance(v, str): return v
            return ""
        if isinstance(sabian_dict, dict):
            if idx_str in sabian_dict: return extract_text(sabian_dict[idx_str])
            for k, v in sabian_dict.items():
                if isinstance(v, dict) and str(v.get("index", "")) == idx_str: return extract_text(v)
        return ""

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    # 메타데이터 할당
    meta_map = mapping.get("metadata", {})
    if meta_map.get("sect"): ws[meta_map["sect"]] = sect_str
    
    h_sys_key = str(h_sys).upper()[0] if h_sys else 'P'
    if meta_map.get("h_sys"): ws[meta_map["h_sys"]] = H_SYS_MAP.get(h_sys_key, h_sys.capitalize())
    
    sz_type_raw = str(arc_p.get('syzygy', {}).get('type', 'NEW')).upper()
    sz_type_str = "FULL MOON" if "FULL" in sz_type_raw else "NEW MOON"
    if meta_map.get("syzygy_type"): ws[meta_map["syzygy_type"]] = sz_type_str

    for key in ["sect", "h_sys", "syzygy_type"]:
        c_ref = meta_map.get(key)
        if c_ref and ws[c_ref].value:
            cell = ws[c_ref]
            apply_grimoire_styles(cell, cell.value, is_info_col=False, skip_color=True)
            color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
            cell.font = Font(name="Consolas", color=color_to_use, bold=False, italic=False)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # Arcana 맵핑
    for key, map_data in mapping.get("arcana", {}).items():
        row_idx = map_data.get("row")
        if not row_idx: continue
        
        col_info = map_data.get("info")
        col_house = map_data.get("house")
        col_duad = map_data.get("duad")
        col_dodeca = map_data.get("dodeca")
        col_decan = map_data.get("decan")
        col_bounds = map_data.get("bounds")
        col_sabian = map_data.get("sabian")
        
        lon = points_to_plot.get(key)
        
        if lon is None:
            for c_char in [col_info, col_house, col_duad, col_dodeca, col_decan, col_bounds, col_sabian]:
                if c_char: ws[f"{c_char}{row_idx}"] = ""
            continue
            
        sign_idx = int(lon / 30) % 12
        sign_name = LOCAL_SIGNS[sign_idx]
        deg_in_sign = lon % 30
        
        ws[f"{col_info}{row_idx}"] = format_dms_pretty(lon)
        
        h_num = get_house_of_point(lon)
        ws[f"{col_house}{row_idx}"] = f"H{h_num}" if h_num else ""
        
        ws[f"{col_duad}{row_idx}"] = SYMBOL_MAP.get(get_duad(sign_name, deg_in_sign), "-")
        ws[f"{col_dodeca}{row_idx}"] = SYMBOL_MAP.get(get_dodecatemoria(deg_in_sign), "-")
        ws[f"{col_decan}{row_idx}"] = SYMBOL_MAP.get(get_decan(sign_name, deg_in_sign), "-")
        ws[f"{col_bounds}{row_idx}"] = SYMBOL_MAP.get(get_egyptian_bounds(sign_name, deg_in_sign), "-")
        
        s_idx = get_sabian_index(lon)
        ws[f"{col_sabian}{row_idx}"] = get_sabian_text(str(s_idx)) if s_idx else ""
        
        # 🚀 [추가됨]: 현재 행의 도수(DMS) 텍스트를 스캔하여 29도(Anaretic) 여부 판별
        dms_val = str(ws[f"{col_info}{row_idx}"].value) if ws[f"{col_info}{row_idx}"].value else ""
        is_anaretic_flag = "29°" in dms_val
        
        for c_char in [col_info, col_house, col_duad, col_dodeca, col_decan, col_bounds, col_sabian]:
            if not c_char: continue
            cell = ws[f"{c_char}{row_idx}"]
            
            # 🚀 [수정됨]: 현재 순회 중인 열(c_char)이 하우스 열이면 배경색 스킵
            should_skip_color = True if c_char in [col_house, col_sabian] else False
            
            # 🚀 [핵심 수복]: 현재 순회 중인 열이 정확히 '도수 정보 열(col_info)'일 때만 29도 여부를 검사해서 붉은색 적용!
            current_is_anaretic = True if (c_char == col_info and "29°" in str(cell.value)) else False
            
            # 스타일러 호출
            apply_grimoire_styles(
                cell, 
                cell.value, 
                is_info_col=True, 
                skip_color=should_skip_color, 
                is_anaretic=current_is_anaretic
            )
            color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
            cell.font = Font(name="Consolas", color=color_to_use, bold=False, italic=False)
            
            horiz = 'center' if c_char == col_house else ('left' if c_char in [col_info, col_sabian] else 'center')
            cell.alignment = Alignment(horizontal=horiz, vertical='center', wrap_text=False)

    return wb