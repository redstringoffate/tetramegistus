import json
import os
import shutil
import math
from copy import copy
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz, format_dms_pretty
from core.astrology.hypostases import find_persona_transit
import swisseph as swe

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/n7_sabian_planets_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/nigredo/n7_sabian_planets.xlsx'))

SABIAN_PATHS = [
    os.path.abspath(os.path.join(BASE_DIR, '../../../data/render/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../../data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../static/data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../data/sabian.json')),
]

# 🚀 하우스 시스템 Full Name 매핑 딕셔너리
HOUSE_SYS_NAMES = {
    'P': 'Placidus', 'W': 'Whole Sign', 'K': 'Koch', 'R': 'Regiomontanus', 
    'C': 'Campanus', 'E': 'Equal', 'O': 'Porphyry'
}

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def ensure_location_string(seed):
    loc_val = str(seed.get("location", seed.get("city", ""))).strip()
    if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
        lat, lng = seed.get("lat"), seed.get("lng")
        if lat is not None and lng is not None:
            seed["city"] = f"{abs(float(lat)):.2f}°{'N' if float(lat)>=0 else 'S'}, {abs(float(lng)):.2f}°{'E' if float(lng)>=0 else 'W'}"
            seed["location"] = seed["city"]
        else:
            seed["city"] = "Unknown Location"
            seed["location"] = "Unknown Location"
    else:
        seed["city"] = loc_val
        seed["location"] = loc_val

def get_house_of_point(lon, houses_dict):
    if not houses_dict: return "-"
    for i in range(1, 13):
        c1 = houses_dict.get(i, {}).get('longitude', 0.0) if isinstance(houses_dict.get(i), dict) else houses_dict.get(i, 0.0)
        c2 = houses_dict.get(i+1 if i < 12 else 1, {}).get('longitude', 0.0) if isinstance(houses_dict.get(i+1 if i < 12 else 1), dict) else houses_dict.get(i+1 if i < 12 else 1, 0.0)
        if c1 < c2:
            if c1 <= lon < c2: return f"H{i}"
        else:
            if c1 <= lon < 360 or 0 <= lon < c2: return f"H{i}"
    return "H1"

def compile_n7_sabian_planets_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "n7_sabian_planets"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)
    
    seed_data = chart_data.get("seed", chart_data)
    meta = chart_data.get("metadata", {})
    ensure_location_string(seed_data)
    is_unk = bool(seed_data.get("is_time_unknown", 0))

    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    
    # 🚀 하우스 시스템 안전 파싱
    h_sys_raw = str(meta.get("h_sys", "P")).strip().upper()
    h_sys = h_sys_raw[0] if h_sys_raw else "P"

    lang_val = meta.get("language", chart_data.get("language", seed_data.get("language", "en")))
    is_ko = "ko" in str(lang_val).lower()
    
    sabian_dict = {}
    for spath in SABIAN_PATHS:
        if os.path.exists(spath):
            try:
                with open(spath, 'r', encoding='utf-8') as f: sabian_dict = json.load(f)
                break
            except: pass

    def get_sabian_text(idx):
        idx_str = str(idx)
        def extract_text(entry):
            if isinstance(entry, str): return entry
            if isinstance(entry, dict):
                primary = ["text_ko", "ko", "desc_ko"] if is_ko else ["text_en", "en", "desc_en"]
                for k in primary + ["text", "desc"]:
                    if k in entry and entry[k]: return str(entry[k])
            return ""
        if isinstance(sabian_dict, dict):
            if idx_str in sabian_dict: return extract_text(sabian_dict[idx_str])
            for k, v in sabian_dict.items():
                if isinstance(v, dict) and str(v.get("index", "")) == idx_str: return extract_text(v)
        return ""

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: 
        mapping = json.load(f)

    apply_natal_stamp(ws, seed_data, method="single", cells=[mapping["metadata"]["stamper"]])
    ws[mapping["metadata"]["sys_tab"]] = system.upper()
    if system.lower() == "sidereal": 
        ws[mapping["metadata"]["ayanamsa"]] = ayanamsa.upper()
        
    # 🚀 [추가]: 하우스 시스템 B5 위치에 Full Name으로 출력
    if "h_sys" in mapping["metadata"]:
        ws[mapping["metadata"]["h_sys"]] = HOUSE_SYS_NAMES.get(h_sys, "Placidus")
    
    lat = get_safe_float(seed_data.get("lat"), 37.5665)
    lng = get_safe_float(seed_data.get("lng"), 126.9780)

    date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(seed_data.get('birth_time', '12:00:00'))
    tz = _ensure_float_tz(seed_data.get('timezone', 9.0), date_str)
    
    dt_obj = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
    root_jd = swe.julday(dt_obj.year, dt_obj.month, dt_obj.day, dt_obj.hour + dt_obj.minute/60.0 - tz)

    # 1. 렌더링용 루트 차트 (사용자가 선택한 system 반영)
    root_res = calculate_principia(date_str, time_str, lat, lng, tz, system, ayanamsa, h_sys=h_sys)
    root_arc_p = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='paulus', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    try: root_arc_v = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='valens', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    except: root_arc_v = {'lots': {}}

    # 2. 🚀 [추가됨]: 페르소나 역산 전용 타겟 차트 (시간 고정을 위해 무조건 "tropical" 강제 적용)
    trop_res = calculate_principia(date_str, time_str, lat, lng, tz, "tropical", ayanamsa, h_sys=h_sys)
    trop_arc_p = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='paulus', system="tropical", ayanamsa=ayanamsa, h_sys=h_sys)
    try: trop_arc_v = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='valens', system="tropical", ayanamsa=ayanamsa, h_sys=h_sys)
    except: trop_arc_v = {'lots': {}}

    target_pool = {}
    
    # 3. 🚀 [수정됨]: 타겟 풀을 담을 때는 root_res가 아닌 trop_res를 사용!
    for k, v in trop_res.get('planets', {}).items():
        target_pool[k] = v.get('longitude')
        if k == "Eros": target_pool["Asteroid Eros"] = v.get('longitude')
        elif k == "Lilith": target_pool["Asteroid Lilith"] = v.get('longitude')
        elif k == "North Node (m)": target_pool["Rahu"] = v.get('longitude')
        elif k == "South Node (m)": target_pool["Ketu"] = v.get('longitude')
        elif k == "True Node": target_pool["North Node (t)"] = v.get('longitude')
        elif k == "South Node": target_pool["South Node (t)"] = v.get('longitude')

    for k, v in trop_arc_p.get('lots', {}).items(): target_pool[k] = v.get('longitude', v.get('value'))
    for k, v in trop_arc_v.get('lots', {}).items():
        if k in ['Eros', 'Necessity']: target_pool[f"{k} (v)"] = v.get('longitude', v.get('value'))
    for k, v in trop_arc_p.get('vertex', {}).items(): target_pool[k] = v.get('longitude', v.get('value'))
    if 'syzygy' in trop_arc_p and trop_arc_p['syzygy'].get('data'):
        target_pool['Syzygy'] = trop_arc_p['syzygy']['data'].get('longitude', trop_arc_p['syzygy']['data'].get('value'))

    ts_row = mapping.get("timestamps_row", 7)

    for persona_name, col_dict in mapping["persona_columns"].items():
        c_info = col_dict["info"]
        c_house = col_dict["house"]
        
        target_lon = target_pool.get(persona_name)
        if target_lon is None: continue

        if persona_name == "Sun":
            p_res, p_arc_p, p_arc_v = root_res, root_arc_p, root_arc_v
            y, m, d = dt_obj.year, dt_obj.month, dt_obj.day
            p_hour, p_min = dt_obj.hour, dt_obj.minute
        else:
            p_jd = find_persona_transit(root_jd, target_lon)
            y_u, m_u, d_u, h_dec_u = swe.revjul(p_jd, swe.GREG_CAL)
            total_sec = int(round(h_dec_u * 3600))
            utc_dt = datetime(y_u, m_u, d_u) + timedelta(seconds=total_sec)
            local_dt = utc_dt + timedelta(hours=tz)
            
            y, m, d = local_dt.year, local_dt.month, local_dt.day
            p_hour, p_min = local_dt.hour, local_dt.minute
            
            p_res = calculate_principia(f"{y:04d}-{m:02d}-{d:02d}", f"{p_hour:02d}:{p_min:02d}:00", lat, lng, tz, system, ayanamsa, h_sys=h_sys)
            p_arc_p = calculate_arcana(f"{y:04d}-{m:02d}-{d:02d}", f"{p_hour:02d}:{p_min:02d}:00", lat, lng, tz, lot_schema='paulus', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
            try: p_arc_v = calculate_arcana(f"{y:04d}-{m:02d}-{d:02d}", f"{p_hour:02d}:{p_min:02d}:00", lat, lng, tz, lot_schema='valens', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
            except: p_arc_v = {'lots': {}}

        ts_cell = ws[f"{c_info}{ts_row}"]
        ts_cell.value = f"{y:04d}-{m:02d}-{d:02d}; {p_hour:02d}:{p_min:02d}"
        apply_grimoire_styles(ts_cell, ts_cell.value, skip_color=True)
        ts_cell.alignment = Alignment(horizontal='center', vertical='center')

        for body_name, row_dict in mapping["body_rows"].items():
            info_row = row_dict["info"]
            sabian_row = row_dict["sabian"]
            
            if is_unk and info_row > 70:
                continue

            p_data = {}
            if body_name in p_res.get('planets', {}): p_data = p_res['planets'][body_name]
            elif body_name == "Asteroid Eros" and "Eros" in p_res.get('planets', {}): p_data = p_res['planets']["Eros"]
            elif body_name == "Asteroid Lilith" and "Lilith" in p_res.get('planets', {}): p_data = p_res['planets']["Lilith"]
            elif body_name == "Rahu" and "North Node (m)" in p_res.get('planets', {}): p_data = p_res['planets']["North Node (m)"]
            elif body_name == "Ketu" and "South Node (m)" in p_res.get('planets', {}): p_data = p_res['planets']["South Node (m)"]
            elif body_name == "North Node (t)" and "True Node" in p_res.get('planets', {}): p_data = p_res['planets']["True Node"]
            elif body_name == "South Node (t)" and "South Node" in p_res.get('planets', {}): p_data = p_res['planets']["South Node"]
            elif body_name in p_arc_p.get('lots', {}): p_data = p_arc_p['lots'][body_name]
            elif "(v)" in body_name:
                clean_v = body_name.replace(" (v)", "")
                if clean_v in p_arc_v.get('lots', {}): p_data = p_arc_v['lots'][clean_v]
            elif body_name in p_arc_p.get('vertex', {}): p_data = p_arc_p['vertex'][body_name]
            elif body_name == "Syzygy" and 'syzygy' in p_arc_p: p_data = p_arc_p['syzygy'].get('data', {})

            if p_data:
                p_lon = p_data.get('longitude', p_data.get('value', 0.0))
                info_val = format_dms_pretty(p_lon)
                house_val = get_house_of_point(p_lon, p_res.get('houses', {}))
                
                s_idx = p_data.get('sabian_index', p_data.get('sabian'))
                if not s_idx: s_idx = int(math.floor(p_lon)) + 1
                sabian_val = get_sabian_text(s_idx)

                cell_info = ws[f"{c_info}{info_row}"]
                cell_house = ws[f"{c_house}{info_row}"]
                cell_sabian = ws[f"{c_info}{sabian_row}"]

                cell_info.value = info_val
                cell_house.value = house_val
                cell_sabian.value = sabian_val
                
                # 🚀 [추가됨]: 도수(info_val) 문자열에 29도가 포함되어 있는지 판별
                current_is_anaretic = "29°" in str(info_val)
                
                # 🚀 [수정됨]: cell_info(도수 셀)에만 아나레틱 깃발을 전달하여 붉은색 폰트 적용
                apply_grimoire_styles(cell_info, info_val, is_info_col=True, is_anaretic=current_is_anaretic)
                
                # 하우스와 사비안 셀은 기존처럼 배경색만 스킵하고 아나레틱 깃발은 넘기지 않음 (기본 검은 폰트 유지)
                apply_grimoire_styles(cell_house, house_val, skip_color=True) 
                apply_grimoire_styles(cell_sabian, sabian_val, is_info_col=True, skip_color=True)
                
                cell_info.alignment = Alignment(horizontal='center', vertical='center')
                cell_house.alignment = Alignment(horizontal='center', vertical='center')
                
                cell_sabian.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
                if cell_sabian.font:
                    cell_sabian.font = Font(name="Consolas", size=8, color=cell_sabian.font.color, bold=cell_sabian.font.bold)

    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name="Consolas", size=cell.font.size, bold=cell.font.bold,
                    italic=cell.font.italic, vertAlign=cell.font.vertAlign,
                    underline=cell.font.underline, strike=cell.font.strike, color=cell.font.color
                )

    return wb