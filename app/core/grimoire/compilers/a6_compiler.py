import json
import os
import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.davison import calculate_davison_midpoint
from core.astrology.engine import calculate_principia, _ensure_float_tz, TROPICAL_SIGNS, SYMBOL_MAP, format_dms_pretty

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a6_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a6.xlsx'))

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def sanitize_unicode(text):
    if isinstance(text, str):
        return text.replace("\ufe0e", "").replace("\ufe0f", "")
    return text

def get_body_info(pt_lon):
    if pt_lon is None: return "-"
    return format_dms_pretty(float(pt_lon))

def compile_a6_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a6"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()

    # Davison 시드 생성
    seed_data = calculate_davison_midpoint(s1_raw, s2_raw)

    # 🚀 "Davison Midpoint" 문자열 대신 정확한 좌표(Lat/Lon) 포맷으로 강제 덮어쓰기
    d_lat = get_safe_float(seed_data.get("lat"), None)
    d_lng = get_safe_float(seed_data.get("lng"), None)
    
    if d_lat is not None and d_lng is not None:
        coord_str = f"{abs(d_lat):.2f}°{'N' if d_lat>=0 else 'S'}, {abs(d_lng):.2f}°{'E' if d_lng>=0 else 'W'}"
    else:
        coord_str = "Unknown Location"
        
    seed_data["city"] = coord_str
    seed_data["location"] = coord_str
    seed_data["city_name"] = coord_str
    seed_data["location_name"] = coord_str

    apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])

    date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(seed_data.get('birth_time', '12:00:00'))
    lat = get_safe_float(seed_data.get("lat"), 37.5665)
    lng = get_safe_float(seed_data.get("lng"), 126.9780)
    tz_val = seed_data.get('tz') if seed_data.get('tz') is not None else seed_data.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)
    is_unk = bool(seed_data.get("is_time_unknown", 0))

    res_p = calculate_principia(date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz, system=system, ayanamsa=ayanamsa, view=view, h_sys=h_sys, fixed_star_orb=1.0, is_time_unknown=is_unk)

    flat_points = {}
    for group in ['planets', 'asteroids', 'lilith_nodes']:
        for k, v in res_p.get(group, {}).items():
            if k in ["Asc.", "I.C.", "Dsc.", "M.C.", "Asc", "IC", "Dsc", "MC"]: continue
            
            lon = v.get('longitude', v.get('value')) if isinstance(v, dict) else v
            if lon is not None:
                name = k
                if k in ["Mean Node", "North Node (m)", "North Node (t)", "North Node"]: name = "North Node (t)"
                elif k == "Lilith": name = "Asteroid Lilith"
                
                flat_points[name] = float(lon)

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: mapping = json.load(f)

    layout = mapping.get("layout", {})
    cols = layout.get("cols", {})
    rows = layout.get("rows", {})
    
    harmonics_map = {
        "H1": 1, "H2": 2, "H3": 3, "H4": 4, "H5": 5, "H6": 6, "H7": 7, "H8": 8,
        "H9": 9, "H10": 10, "H11": 11, "H12": 12, "H16": 16, "H20": 20, "H24": 24, "H32": 32
    }

    for body_name, r_idx in rows.items():
        base_lon = flat_points.get(body_name)
        
        cell_a = ws[f"A{r_idx}"]
        # 🚀 [수정]: A열은 템플릿 고유의 4원소 배경색을 보존합니다.
        apply_grimoire_styles(cell_a, cell_a.value, is_info_col=False, skip_color=True)
        
        for h_key, col_char in cols.items():
            h_multiplier = harmonics_map.get(h_key, 1)
            cell = ws[f"{col_char}{r_idx}"]
            
            if base_lon is not None:
                h_lon = (base_lon * h_multiplier) % 360.0
                info_str = get_body_info(h_lon)
                cell.value = info_str
                apply_grimoire_styles(cell, info_str, is_info_col=True, skip_color=False)
            else:
                cell.value = "-"
                
            color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
            cell.font = Font(name="Consolas", size=9, color=color_to_use, bold=False)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    shrink_rows = layout.get("shrink_rows", [])
    for r_idx in shrink_rows:
        ws.row_dimensions[r_idx].height = 4.5
        for col_char in ["A"] + list(cols.values()):
            ws[f"{col_char}{r_idx}"].value = ""

    for r_idx in range(1, ws.max_row + 1):
        cell_val = ws[f"A{r_idx}"].value
        if str(cell_val).strip() == ".":
            ws.row_dimensions[r_idx].height = 4.5
            ws[f"A{r_idx}"].value = ""
        elif r_idx >= 5 and r_idx not in shrink_rows:
            if ws.row_dimensions[r_idx].height != 4.5:
                ws.row_dimensions[r_idx].height = 16.5

    # ====================================================================
    # 1. 템플릿의 너비 속성 강제 복사 (작성자님의 똥꼬쇼 부활! 이거 필수였습니다 ㅠㅠ)
    # ====================================================================
    temp_wb_for_width = load_workbook(TEMPLATE_FILE)
    temp_ws_for_width = temp_wb_for_width[base_name] if base_name in temp_wb_for_width.sheetnames else temp_wb_for_width.active
    
    for col_char in ["A"] + list(cols.values()):
        target_width = None
        if col_char in temp_ws_for_width.column_dimensions:
            target_width = temp_ws_for_width.column_dimensions[col_char].width
        
        if target_width is None:
            target_width = 20.5 if col_char != "A" else 22.0
            
        ws.column_dimensions[col_char].width = target_width


    # ====================================================================
    # 2. A1 병합 해제 및 B1 피신 (틀 고정 버그 원천 차단)
    # ====================================================================
    from openpyxl.utils import get_column_letter
    
    merge_to_remove = None
    for m_range in list(ws.merged_cells.ranges):
        if "A1" in m_range.coord:
            merge_to_remove = m_range
            break
            
    if merge_to_remove:
        max_col_letter = get_column_letter(merge_to_remove.max_col)
        ws.unmerge_cells(merge_to_remove.coord)
        
        # A1의 타이틀과 스타일을 B1으로 피신
        ws['B1'].value = ws['A1'].value
        ws['A1'].value = None
        
        if ws['A1'].has_style:
            ws['B1'].font = copy(ws['A1'].font)
            ws['B1'].border = copy(ws['A1'].border)
            ws['B1'].fill = copy(ws['A1'].fill)
            ws['B1'].alignment = copy(ws['A1'].alignment)
            
        # B1부터 Z열(또는 끝열)까지 새롭게 병합
        ws.merge_cells(f"B1:{max_col_letter}1")


    # ====================================================================
    # 3. 완벽한 틀 고정 적용
    # ====================================================================
    ws.sheet_view.pane = None
    ws.freeze_panes = None
    # 🚀 B5: A열(행성명) 좌측 고정 + 상단 1~4행 고정
    # 만약 상단은 고정 안 하고 딱 A열만 좌측 고정하고 싶다면 'B1'으로 바꿔주세요!
    ws.freeze_panes = 'B5' 
    
    return wb