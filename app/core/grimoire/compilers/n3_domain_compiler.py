import json
import os
import shutil
import math
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz, format_dms_pretty

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/n3_domain_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/nigredo/n3_domain.xlsx'))

H_SYS_MAP = {
    'P': 'Placidus', 'W': 'Whole Sign', 'K': 'Koch', 'R': 'Regiomontanus',
    'C': 'Campanus', 'E': 'Equal', 'V': 'Vehlow'
}

PLANETS_LIST = ["Sun", "Moon", "Mercury", "Venus", "Mars", "Jupiter", "Saturn", "Uranus", "Neptune", "Pluto", "Mean Lilith", "True Lilith", "Rahu", "Ketu", "North Node (t)", "South Node (t)"]
ASTEROIDS_LIST = ["Chiron", "Ceres", "Juno", "Pallas", "Vesta", "Asteroid Eros", "Psyche", "Asteroid Lilith"]

def get_col(cell_str):
    if not cell_str: return ""
    return "".join([c for c in str(cell_str) if c.isalpha()])

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except: return default

def format_house_range(lon_start, lon_end):
    if lon_start is None or lon_end is None: return ""
    size = (lon_end - lon_start) % 360
    if size < 0: size += 360
    rounded_size = round(size)
    
    portions = []
    curr = lon_start
    rem = size
    while rem > 1e-5:
        next_boundary = (math.floor(curr / 30) + 1) * 30
        step = min(rem, next_boundary - curr)
        portions.append(step)
        curr = (curr + step) % 360
        rem -= step
        
    rounded_portions = [round(p) for p in portions]
    diff = rounded_size - sum(rounded_portions)
    if diff != 0 and rounded_portions:
        rounded_portions[rounded_portions.index(max(rounded_portions))] += diff
        
    portions_str = "/".join(map(str, rounded_portions))
    return f"{rounded_size} ({portions_str})"

def format_body_string(item):
    lon = float(item.get('longitude', 0))
    is_retro = item.get('is_retrograde', False)
    dms_str = format_dms_pretty(lon, is_retro)
    return f"{item['name']} - {dms_str}"

def compile_n3_domain_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "n3_domain"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")
    
    seed_data = chart_data.get("seed", {})
    if seed_data is None:
        raise ValueError("Grimoire requires active seed data. Please select a seed first.")

    def ensure_location_string(seed):
        loc_val = str(seed.get("location", seed.get("city", ""))).strip()
        if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
            lat, lng = get_safe_float(seed.get("lat"), None), get_safe_float(seed.get("lng"), None)
            if lat is not None and lng is not None:
                seed["city"] = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
                seed["location"] = seed["city"]
            else:
                seed["city"] = "Unknown Location"
                seed["location"] = "Unknown Location"
        else:
            seed["city"] = loc_val
            seed["location"] = loc_val
            
    ensure_location_string(seed_data)

    date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(seed_data.get('birth_time', '12:00:00'))
    lat, lng = get_safe_float(seed_data.get("lat"), 37.5665), get_safe_float(seed_data.get("lng"), 126.9780)
    tz = _ensure_float_tz(seed_data.get('tz') if seed_data.get('tz') is not None else seed_data.get('timezone', 9.0), date_str)
    is_unk = bool(seed_data.get("is_time_unknown", 0))

    apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])

    res = calculate_principia(date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz, system=system, ayanamsa=ayanamsa, view=view, h_sys=h_sys, fixed_star_orb=1.0, is_time_unknown=is_unk)
    arc_p = calculate_arcana(date_str, time_str, lat, lng, timezone=tz, lot_schema='paulus', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    try: arc_v = calculate_arcana(date_str, time_str, lat, lng, timezone=tz, lot_schema='valens', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    except: arc_v = {'lots': {}}

    planets_data = res.get('planets', {})
    angles_data = res.get('angles', {})
    houses_data = res.get('houses', {})

    # 🚀 [여기 추가]: Draconic / Ketunic 시스템일 경우 교점(Nodes) 엑셀에서 원천 삭제
    if system in ['draconic', 'ketunic']:
        for n_key in ['Rahu', 'Ketu', 'North Node (m)', 'South Node (m)', 'North Node (t)', 'South Node (t)']:
            planets_data.pop(n_key, None)

    cusps_simple = {}
    if not is_unk:
        for k, v in houses_data.items():
            cusps_simple[int(k)] = float(v['longitude']) if isinstance(v, dict) else float(v)

    def get_house_of_point(lon):
        if is_unk or not cusps_simple: return None
        lon = float(lon) % 360
        for i in range(1, 13):
            curr_c = cusps_simple[i]
            next_c = cusps_simple[i+1] if i < 12 else cusps_simple[1]
            if curr_c < next_c:
                if curr_c <= lon < next_c: return i
            else:
                if curr_c <= lon < 360 or 0 <= lon < next_c: return i
        return 1

    h_items = {i: {"P": [], "A": [], "H": []} for i in range(1, 13)}
    
    # 🚀 [추가됨]: 이미 엑셀에 들어간 천체 이름을 추적하는 중복 방지 세트
    processed_names = set()

    for name, data in planets_data.items():
        norm_name = name
        if name in ["Mean Node", "North Node (m)"]: norm_name = "Rahu"
        elif name in ["South Node", "South Node (m)"]: norm_name = "Ketu"
        elif name == "Eros": norm_name = "Asteroid Eros"
        elif name == "Lilith": norm_name = "Asteroid Lilith"
        elif name == "Lilith (mean)": norm_name = "Mean Lilith"
        elif name == "Lilith (true)": norm_name = "True Lilith"
        
        # 🚀 [추가됨]: 이미 담은 행성(예: Rahu가 이미 담겼는데 Mean Node가 또 들어온 경우)은 무시!
        if norm_name in processed_names: 
            continue
        processed_names.add(norm_name)
        
        lon = data.get('longitude')
        if lon is None: continue
        
        h_num = get_house_of_point(lon)
        if not h_num: continue
        
        d_copy = data.copy()
        d_copy['name'] = norm_name
        
        if norm_name in PLANETS_LIST: h_items[h_num]["P"].append(d_copy)
        elif norm_name in ASTEROIDS_LIST: h_items[h_num]["A"].append(d_copy)

    def add_hermetic(name_fmt, data_dict):
        lon = data_dict.get('value', data_dict.get('longitude'))
        if lon is not None:
            h_num = get_house_of_point(lon)
            if h_num: h_items[h_num]["H"].append({"name": name_fmt, "longitude": lon, "is_retrograde": False})

    for name, data in arc_p.get('lots', {}).items(): add_hermetic(name, data)
    for name, data in arc_v.get('lots', {}).items():
        if name in ['Eros', 'Necessity']: add_hermetic(f"{name} (v)", data)
    for name, data in arc_p.get('vertex', {}).items(): add_hermetic(name, data)
    if 'syzygy' in arc_p and arc_p['syzygy'].get('data'): add_hermetic("Syzygy", arc_p['syzygy']['data'])

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    asc_data = angles_data.get("Ascendant") or planets_data.get("Ascendant") or {}
    mc_data = angles_data.get("Midheaven") or planets_data.get("Midheaven") or {}
    
    ws[mapping["metadata"]["ascendant"]] = asc_data.get("dms", "")
    ws[mapping["metadata"]["midheaven"]] = mc_data.get("dms", "")
    ws[mapping["metadata"]["sys_tab"]] = system.upper()
    
    # 🚀 [Fix 1]: h_sys "Whole" -> "Whole Sign" (첫 글자 추출 후 매핑)
    h_sys_key = str(h_sys).upper()[0] if h_sys else 'P'
    ws[mapping["metadata"]["h_sys"]] = H_SYS_MAP.get(h_sys_key, h_sys.capitalize())
    
    if system.upper() == "SIDEREAL": ws[mapping["metadata"]["ayanamsa"]] = ayanamsa.upper()

    for key in ["ascendant", "midheaven", "sys_tab", "h_sys", "ayanamsa"]:
        c_ref = mapping["metadata"].get(key)
        if c_ref and ws[c_ref].value:
            cell = ws[c_ref]
            should_skip_color = True if key in ["sys_tab", "h_sys", "ayanamsa"] else False
            apply_grimoire_styles(cell, cell.value, is_info_col=False, skip_color=should_skip_color)
            
            # 🚀 [Fix 2]: Asc/MC 볼드체 완전 해제
            is_bold = True if key in [] else False
            color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
            cell.font = Font(name="Consolas", color=color_to_use, bold=is_bold, italic=False)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    houses_list = [{"house": i, "map": mapping["houses"][str(i)]} for i in range(1, 13) if str(i) in mapping.get("houses", {})]
    houses_list.sort(key=lambda x: x["map"].get("row_start", 0), reverse=True)

    for item in houses_list:
        h_num, h_map = item["house"], item["map"]
        row_idx = h_map.get("row_start", 0)
        if not row_idx: continue

        col_info, col_range = get_col(h_map.get("info", "B")), get_col(h_map.get("range", "C"))
        col_p, col_a, col_h = get_col(h_map.get("planets", "D")), get_col(h_map.get("asteroids", "E")), get_col(h_map.get("hermetic", "F"))

        lon = cusps_simple.get(h_num)
        if lon is None:
            for c_char in "BCDEF": ws[f"{c_char}{row_idx}"] = ""
            continue

        # 🚀 [수복 완료]: 하우스 커스프(lon) 기준 상대 거리로 정렬
        p_items = sorted(h_items[h_num]["P"], key=lambda x: (float(x.get('longitude', 0)) - lon) % 360)
        a_items = sorted(h_items[h_num]["A"], key=lambda x: (float(x.get('longitude', 0)) - lon) % 360)
        h_items_sorted = sorted(h_items[h_num]["H"], key=lambda x: (float(x.get('longitude', 0)) - lon) % 360)

        p_strs = [format_body_string(x) for x in p_items]
        a_strs = [format_body_string(x) for x in a_items]
        h_strs = [format_body_string(x) for x in h_items_sorted]

        max_rows = max(1, len(p_strs), len(a_strs), len(h_strs))

        if max_rows > 1:
            ws.insert_rows(row_idx + 1, amount=max_rows - 1)
            for r in range(1, max_rows):
                insert_idx = row_idx + r
                ws.row_dimensions[insert_idx].height = 16.5
                for c_char in [col_info, col_range, col_p, col_a, col_h]:
                    src, tgt = ws[f"{c_char}{row_idx}"], ws[f"{c_char}{insert_idx}"]
                    if src.has_style:
                        tgt.border, tgt.fill = copy(src.border), copy(src.fill)
                        tgt.font, tgt.alignment = copy(src.font), copy(src.alignment)
                        tgt.number_format = copy(src.number_format)

        for i in range(max_rows):
            r_idx = row_idx + i
            ws[f"{col_p}{r_idx}"] = p_strs[i] if i < len(p_strs) else ""
            ws[f"{col_a}{r_idx}"] = a_strs[i] if i < len(a_strs) else ""
            ws[f"{col_h}{r_idx}"] = h_strs[i] if i < len(h_strs) else ""

            for c_char in [col_p, col_a, col_h]:
                cell = ws[f"{c_char}{r_idx}"]
                text_val = str(cell.value) if cell.value else ""
                
                if text_val:
                    styler_target = text_val.split('-')[-1] if '-' in text_val else text_val
                    
                    # 🚀 [추가됨]: 출력될 텍스트에 "29°"가 포함되어 있는지 확인하여 Anaretic 판별
                    is_anaretic_flag = "29°" in text_val 
                    
                    # 🚀 [수정됨]: apply_grimoire_styles 호출 시 is_anaretic 인자 넘기기
                    apply_grimoire_styles(cell, styler_target, is_info_col=True, skip_color=False, is_anaretic=is_anaretic_flag)
                    
                    cell.value = text_val
                
                if cell.font: cell.font = Font(name="Consolas", color=cell.font.color, bold=cell.font.bold)
                else: cell.font = Font(name="Consolas")
                    
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            
            # 🚀 [Fix 3]: Information과 Range의 내용을 첫 줄에만 표시하고 볼드 해제
            # Information과 Range의 내용을 첫 줄에만 표시하고 볼드 해제
            if i == 0:
                ws[f"{col_info}{r_idx}"] = format_dms_pretty(lon)
                next_h = 1 if h_num == 12 else h_num + 1
                next_lon = cusps_simple.get(next_h)
                ws[f"{col_range}{r_idx}"] = format_house_range(lon, next_lon) if next_lon is not None else ""
                
                c_info, c_range = ws[f"{col_info}{r_idx}"], ws[f"{col_range}{r_idx}"]
                
                # 🚀 [추가됨]: Cusp 문자열에도 "29°"가 포함되어 있는지 확인
                is_cusp_anaretic = "29°" in str(c_info.value)
                
                # 🚀 [수정됨]: apply_grimoire_styles 호출 시 is_anaretic 인자 넘기기
                apply_grimoire_styles(c_info, c_info.value, is_info_col=True, skip_color=False, is_anaretic=is_cusp_anaretic)
                
                color_info = c_info.font.color if c_info.font and c_info.font.color else "000000"
                c_info.font = Font(name="Consolas", color=color_info, bold=False, italic=False)
                c_info.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                
                apply_grimoire_styles(c_range, c_range.value, is_info_col=True, skip_color=False)
                color_range = c_range.font.color if c_range.font and c_range.font.color else "000000"
                c_range.font = Font(name="Consolas", color=color_range, bold=False, italic=False)
                c_range.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            else:
                ws[f"{col_info}{r_idx}"] = ""
                ws[f"{col_range}{r_idx}"] = ""

    ws.column_dimensions[col_p].width = 38.0
    ws.column_dimensions[col_a].width = 38.0
    ws.column_dimensions[col_h].width = 38.0

    return wb