import json
import os
import math
import re
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz, format_dms_pretty
from core.astrology.davison import calculate_davison_midpoint
from core.astrology.composite import calculate_composite_chart, apply_anti_composite
from core.astrology.constants import ASTEROIDS
import swisseph as swe

BASE_DIR = os.path.dirname(__file__)
MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a10_mapping.json'))
# 언어별 동적 템플릿 처리 (기본: 영어)
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a10.xlsx')) 

def get_date_time_coords(seed):
    d = str(seed.get('birth_date') or seed.get('birthDate') or seed.get('date') or '2000-01-01').split('T')[0]
    t = str(seed.get('birth_time') or seed.get('birthTime') or seed.get('time') or '12:00:00').strip()
    if t.lower() in ["unknown", "none", ""]: t = "12:00:00"
    elif len(t) == 5: t += ":00"

    coords = seed.get("coordinates", {})
    if isinstance(coords, dict) and "lat" in coords:
        try: lat, lng = float(coords.get("lat", 37.5665)), float(coords.get("lng", 126.9780))
        except: lat, lng = 37.5665, 126.9780
    else:
        try: lat, lng = float(seed.get("lat", 37.5665)), float(seed.get("lng", 126.9780))
        except: lat, lng = 37.5665, 126.9780
        
    tz = _ensure_float_tz(seed.get('timezone', seed.get('tz', 9.0)), d)
    return d, t, lat, lng, tz

def get_jd(seed):
    d, t, lat, lng, tz = get_date_time_coords(seed)
    try: dt_obj = datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M:%S")
    except: dt_obj = datetime.strptime(f"{d} 12:00:00", "%Y-%m-%d %H:%M:%S")
    return swe.julday(dt_obj.year, dt_obj.month, dt_obj.day, dt_obj.hour + dt_obj.minute/60.0 - tz)

def calc_minor_asteroids(jd):
    swe.set_sid_mode(0, 0, 0)
    res_dict = {}
    for ast_name, ast_num in ASTEROIDS.items():
        try:
            res = swe.calc_ut(jd, ast_num + 10000, swe.FLG_SWIEPH | swe.FLG_SPEED)
            res_dict[ast_name] = {"longitude": res[0][0]}
        except: pass
    return res_dict

# 🚀 [방역]: C1에서 개선했던 Paulus + Valens 통합 랏 추출기 (Tropical에만 적용)
def get_hermetic_lots(seed, sys_type, ayan, h_sys):
    if sys_type != 'tropical': return {} 
    
    d, t, lat, lng, tz = get_date_time_coords(seed)
    try: arc_p = calculate_arcana(d, t, lat, lng, tz, lot_schema='paulus', system=sys_type, ayanamsa=ayan, h_sys=h_sys)
    except: arc_p = {}
    try: arc_v = calculate_arcana(d, t, lat, lng, tz, lot_schema='valens', system=sys_type, ayanamsa=ayan, h_sys=h_sys)
    except: arc_v = {}
    
    res = {}
    if 'lots' in arc_p:
        for k, v in arc_p['lots'].items(): res[k] = {"longitude": v.get('value', v.get('longitude', 0.0))}
    if 'lots' in arc_v:
        for k in ['Eros', 'Necessity']:
            if k in arc_v['lots']: res[f"{k} (v)"] = {"longitude": arc_v['lots'][k].get('value', arc_v['lots'][k].get('longitude', 0.0))}
    if 'vertex' in arc_p:
        vtx = arc_p['vertex'].get('Vertex', {})
        vtx_lon = vtx.get('value', vtx.get('longitude', 0.0))
        res['Vertex'] = {"longitude": vtx_lon}
        res['Anti-Vertex'] = {"longitude": (vtx_lon + 180) % 360}
    if 'syzygy' in arc_p and arc_p['syzygy'].get('data'):
        sz = arc_p['syzygy']['data']
        res['Syzygy'] = {"longitude": sz.get('value', sz.get('longitude', 0.0))}
    return res

SABIAN_PATHS = [
    os.path.abspath(os.path.join(BASE_DIR, '../../../data/render/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../../data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../data/sabian.json'))
]

_sabian_cache = None
def get_sabian_text(deg_val, language='en'):
    global _sabian_cache
    if _sabian_cache is None:
        _sabian_cache = {}
        for sp in SABIAN_PATHS:
            if os.path.exists(sp):
                with open(sp, 'r', encoding='utf-8') as f:
                    _sabian_cache = json.load(f)
                break
    
    if deg_val is None: return ""
    try:
        deg = float(deg_val)
        idx = str(int(math.floor(deg)) % 360)
        entry = _sabian_cache.get(idx, {})
        return entry.get(language, entry.get('en', ''))
    except: return ""

def compile_a10_grimoire(chart_data, seed_data=None, language=None):
    # 🚀 JSON 데이터에서 한국어(ko) / 영어(en) 설정만 빼옵니다 (사비안 심볼 번역용)
    if language is None:
        language = chart_data.get('language', 'en')
        
    # 🚀 쓸데없는 언어별 템플릿 찾기 로직을 삭제하고 무조건 공통 a10.xlsx 템플릿을 씁니다.
    template_path = TEMPLATE_FILE
    
    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: mapping = json.load(f)
    
    wb = load_workbook(template_path)
    ws = wb.active
    ws.title = "a10"
    
    meta = chart_data.get('metadata', {})
    sys_type = meta.get('system', 'tropical').lower()
    ayanamsa = meta.get('ayanamsa', 'lahiri')
    h_sys = meta.get('h_sys', 'P')
    
    # 🚀 1. 데이비슨 여부를 판단하여 위경도 텍스트를 강제하는 로직으로 업그레이드
    def sanitize_for_stamper(seed_obj, is_davison=False):
        lat = seed_obj.get("lat")
        lng = seed_obj.get("lng")
        if lat is not None and lng is not None:
            try:
                lat_f, lng_f = float(lat), float(lng)
                coord_str = f"{abs(lat_f):.2f}°{'N' if lat_f>=0 else 'S'}, {abs(lng_f):.2f}°{'E' if lng_f>=0 else 'W'}"
            except: coord_str = "Unknown Location"
        else: coord_str = "Unknown Location"
        
        # 🚀 [수복 1]: 데이비슨 차트면 무조건 Location에 위경도를 박아넣음
        if is_davison and coord_str != "Unknown Location":
            seed_obj["city"] = coord_str
            seed_obj["location"] = coord_str
        else:
            seed_obj["city"] = seed_obj.get("city") or seed_obj.get("location") or coord_str
            seed_obj["location"] = seed_obj.get("location") or seed_obj.get("city") or coord_str
        
        seed_obj["name"] = seed_obj.get("name", "Unknown")
        return seed_obj

    # 🚀 호출부에서 is_davison=True 플래그 전달
    seed1 = sanitize_for_stamper(chart_data.get('seed', {}).get('seed1', {}).copy())
    seed2 = sanitize_for_stamper(chart_data.get('seed', {}).get('seed2', {}).copy())
    davison = sanitize_for_stamper(calculate_davison_midpoint(seed1, seed2), is_davison=True)
    
    m_meta = mapping.get('metadata', {})
    
    # 🚀 prefix 파라미터 제거 (A7 완벽 동일 규격)
    if m_meta.get('davison_stamper'): 
        apply_natal_stamp(ws, davison, method="single", cells=[m_meta['davison_stamper']])
    if m_meta.get('seed_a_stamper'): 
        apply_natal_stamp(ws, seed1, method="single", cells=[m_meta['seed_a_stamper']])
    if m_meta.get('seed_b_stamper'): 
        apply_natal_stamp(ws, seed2, method="single", cells=[m_meta['seed_b_stamper']])
        
    # 하우스 시스템 풀네임 매핑 추가
    h_sys_map = {
        'P': 'Placidus',
        'W': 'Whole House',
        'K': 'Koch',
        'R': 'Regiomontanus',
        'C': 'Campanus',
        'E': 'Equal'
    }
    h_sys_full = h_sys_map.get(h_sys.upper(), h_sys.upper())

    ws[m_meta.get('system_tab', 'B6')].value = sys_type.upper()
    if sys_type == 'sidereal' and m_meta.get('ayanamsa'):
        ws[m_meta['ayanamsa']].value = ayanamsa.upper()
    ws[m_meta.get('h_sys', 'B7')].value = h_sys_full
    
    # 🚀 3. 엔진 연산 시 Time Unknown 상태를 명확히 전달하여 앵글 계산 오류를 방지
    unk_a = bool(seed1.get('is_time_unknown', 0))
    unk_b = bool(seed2.get('is_time_unknown', 0))
    unk_dav = unk_a or unk_b

    res_a = calculate_principia(*get_date_time_coords(seed1), sys_type, ayanamsa, h_sys=h_sys, is_time_unknown=unk_a)
    res_b = calculate_principia(*get_date_time_coords(seed2), sys_type, ayanamsa, h_sys=h_sys, is_time_unknown=unk_b)
    res_dav = calculate_principia(*get_date_time_coords(davison), sys_type, ayanamsa, h_sys=h_sys, is_time_unknown=unk_dav)
   
    ast_a = calc_minor_asteroids(get_jd(seed1))
    ast_b = calc_minor_asteroids(get_jd(seed2))
    ast_dav = calc_minor_asteroids(get_jd(davison))
    
    lots_a = get_hermetic_lots(seed1, sys_type, ayanamsa, h_sys)
    lots_b = get_hermetic_lots(seed2, sys_type, ayanamsa, h_sys)
    lots_dav = get_hermetic_lots(davison, sys_type, ayanamsa, h_sys)
    
    def get_all(res, ast, lots):
        merged = {}
        
        # 🚀 [수복 1]: 엔진 결과물은 그룹화된 구조가 맞습니다! 메인 행성과 앵글들을 정상적으로 꺼냅니다.
        for g in ['planets', 'asteroids', 'lilith_nodes', 'angles']:
            for k, v in res.get(g, {}).items():
                if isinstance(v, dict) and "longitude" in v:
                    d_name = k.replace(" (Natal)", "")
                    if d_name == "Eros": d_name = "Asteroid Eros"
                    
                    # 하우스 추출 (비어있으면 house_placements에서 가져옴)
                    h_val = v.get("house", "-")
                    if h_val in ["-", "", None]:
                        h_val = res.get("house_placements", {}).get(k, "-")
                        
                    merged[d_name] = {"longitude": v.get("longitude", 0.0), "house": h_val}
                    
        # 소행성 및 랏 추가
        for k, v in ast.items(): merged[k] = {"longitude": v.get("longitude", 0.0)}
        for k, v in lots.items(): merged[k] = {"longitude": v.get("longitude", 0.0)}
        
        # 동적 하우스 역산출 로직 (이전 정상화 코드 유지)
        cusps = {i: res.get('houses', {}).get(i, {}).get('longitude', 0.0) for i in range(1, 13)}
        if cusps and cusps.get(1) is not None:
            sorted_cusps = [(h, float(cusps[h])) for h in range(1, 13)]
            sorted_cusps.sort(key=lambda x: x[1])
            for name, data in merged.items():
                if "house" not in data or data["house"] in ["-", "", None]:
                    lon = float(data.get("longitude", 0.0))
                    placed = sorted_cusps[-1][0]
                    for idx in range(12):
                        curr_h, curr_lon = sorted_cusps[idx]
                        next_h, next_lon = sorted_cusps[(idx+1)%12]
                        if curr_lon < next_lon:
                            if curr_lon <= lon < next_lon: placed = curr_h
                        else:
                            if lon >= curr_lon or lon < next_lon: placed = curr_h
                    data["house"] = f"H{placed}"
        return merged

    data_a = get_all(res_a, ast_a, lots_a)
    data_b = get_all(res_b, ast_b, lots_b)
    data_dav = get_all(res_dav, ast_dav, lots_dav)
    
    # 🚀 [수복 완료]: 컴파일러에서도 Composite은 하우스 설정 무관하게 무조건 Placidus 기반으로 연산하여 동일한 값을 보장
    tmp_a = calculate_principia(*get_date_time_coords(seed1), sys_type, ayanamsa, h_sys='P', is_time_unknown=unk_a)
    tmp_b = calculate_principia(*get_date_time_coords(seed2), sys_type, ayanamsa, h_sys='P', is_time_unknown=unk_b)
    
    houses_a = {int(k): float(v['longitude']) for k, v in tmp_a.get('houses', {}).items()}
    houses_b = {int(k): float(v['longitude']) for k, v in tmp_b.get('houses', {}).items()}
    
    # 데이터는 원본(소행성/랏 포함)을 쓰되, 합성 커스프용 계산에만 Placidus 커스프를 주입
    comp_full = calculate_composite_chart(data_a, data_b, houses_a, houses_b)
    
    data_comp = comp_full.get('planets', {})
    houses_comp = comp_full.get('houses', {})
    
    for idx, name in [(1, 'Ascendant'), (4, 'Immum Coeli'), (7, 'Descendant'), (10, 'Midheaven')]:
        if idx in houses_comp: data_comp[name] = {'longitude': float(houses_comp[idx])}
        
    data_anti, houses_anti = apply_anti_composite(data_comp, houses_comp)
    for idx, name in [(1, 'Ascendant'), (4, 'Immum Coeli'), (7, 'Descendant'), (10, 'Midheaven')]:
        if idx in houses_anti: data_anti[name] = {'longitude': float(houses_anti[idx])}
        
    def assign_houses_to_comp(target_data, target_houses):
        if not target_houses: return
        sorted_cusps = [(h, float(target_houses[h])) for h in range(1, 13)]
        sorted_cusps.sort(key=lambda x: x[1])
        for name, data in target_data.items():
            lon = float(data.get("longitude", 0.0))
            placed = sorted_cusps[-1][0]
            for idx in range(12):
                curr_h, curr_lon = sorted_cusps[idx]
                next_h, next_lon = sorted_cusps[(idx+1)%12]
                if curr_lon < next_lon:
                    if curr_lon <= lon < next_lon: placed = curr_h
                else:
                    if lon >= curr_lon or lon < next_lon: placed = curr_h
            data["house"] = f"H{placed}"

    assign_houses_to_comp(data_comp, houses_comp)
    assign_houses_to_comp(data_anti, houses_anti)
    
    def norm_name(n):
        u = n.upper()
        if u in ["MEAN NODE", "NORTH NODE (M)", "NORTH NODE (T)"]: return "Rahu"
        if u in ["SOUTH NODE", "SOUTH NODE (M)", "SOUTH NODE (T)"]: return "Ketu"
        return n
        
    final_a = {norm_name(k): v for k, v in data_a.items()}
    final_b = {norm_name(k): v for k, v in data_b.items()}
    final_dav = {norm_name(k): v for k, v in data_dav.items()}
    final_comp = {norm_name(k): v for k, v in data_comp.items()}
    final_anti = {norm_name(k): v for k, v in data_anti.items()}
    
    col_map = mapping.get('columns', {})
    row_map = mapping.get('body_rows', {})
    
    datasets = {
        "davison": (final_dav, col_map.get("davison", {})),
        "seed_a": (final_a, col_map.get("seed_a", {})),
        "seed_b": (final_b, col_map.get("seed_b", {})),
        "composite": (final_comp, col_map.get("composite", {})),
        "anti_composite": (final_anti, col_map.get("anti_composite", {}))
    }
    
    hermetic_keys = ["Fortune", "Spirit", "Necessity", "Necessity (v)", "Eros", "Eros (v)", "Courage", "Victory", "Nemesis", "Vertex", "Anti-Vertex", "Syzygy"]
    
    for body_name, r_info in row_map.items():
        r_info_idx = r_info["info"]
        r_sab_idx = r_info["sabian"]
        
        is_hermetic = body_name in hermetic_keys

        # 🚀 [수복 2]: Tropical 시스템이 아닐 경우, Midheaven 이후의 Hermetic 로직을 완전히 패스합니다.
        if is_hermetic and sys_type != 'tropical':
            continue
        
        for ds_key, (ds_data, c_info) in datasets.items():
            if is_hermetic and ds_key in ["composite", "anti_composite"]:
                continue
                
            c_info_col = c_info.get("info")
            c_house_col = c_info.get("house")
            
            # 🚀 [방어 코드]: JSON에서 매핑 좌표가 없으면(None) 안전하게 스킵
            if not c_info_col or not c_house_col:
                continue
            
            if body_name in ds_data:
                lon = ds_data[body_name]["longitude"]
                
                # 'H'를 강제로 지우는 replace 코드를 제거하고 원형 보존
                raw_h = ds_data[body_name].get("house", "")
                if raw_h and str(raw_h).strip() != "-":
                    h_val = str(raw_h).strip()
                    if h_val.isdigit(): 
                        h_val = f"H{h_val}"
                else:
                    h_val = "-"
                
                dms_str = format_dms_pretty(lon)
                sabian_str = get_sabian_text(lon, language)
                
                is_anaretic = "29°" in dms_str
                
                cell_info = ws[f"{c_info_col}{r_info_idx}"]
                cell_house = ws[f"{c_house_col}{r_info_idx}"]
                cell_sabian = ws[f"{c_info_col}{r_sab_idx}"]
                
                cell_info.value = dms_str
                cell_house.value = h_val
                cell_sabian.value = sabian_str
                
                apply_grimoire_styles(cell_info, dms_str, is_info_col=True, is_anaretic=is_anaretic)
                apply_grimoire_styles(cell_house, h_val, skip_color=True)
                apply_grimoire_styles(cell_sabian, sabian_str, is_info_col=True, skip_color=True)
                
                cell_info.alignment = Alignment(horizontal='center', vertical='center')
                cell_house.alignment = Alignment(horizontal='center', vertical='center')
                cell_sabian.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
                
                if cell_sabian.font:
                    cell_sabian.font = Font(name="Consolas", size=8, color=cell_sabian.font.color, bold=cell_sabian.font.bold)
                    
                # 엑셀에서 사비안 도수가 텍스트를 넘어가지 않게 두 열(Info, House)을 병합해 줍니다
                try: ws.merge_cells(f"{c_info_col}{r_sab_idx}:{c_house_col}{r_sab_idx}")
                except: pass

    # 전체 폰트 Consolas 통일
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=15):
        for cell in row:
            if cell.value is not None:
                old_f = cell.font
                if old_f and old_f.name != "Consolas":
                    cell.font = Font(
                        name="Consolas", 
                        size=old_f.size if old_f and old_f.size else 10,
                        bold=old_f.bold if old_f else False,
                        italic=old_f.italic if old_f else False,
                        color=old_f.color if old_f else "000000"
                    )

    return wb