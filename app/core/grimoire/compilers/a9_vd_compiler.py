import json
import os
import sys
from datetime import datetime
import openpyxl.utils
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.davison import calculate_davison_midpoint
from core.astrology.engine import calculate_a9_synchronicum, calculate_principia, SYMBOL_MAP, _ensure_float_tz

BASE_DIR = os.path.dirname(__file__)
MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a9_vd_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a9_vd.xlsx'))

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

# 🚀 [Merge Cell 방어 헬퍼]
def write_to_merge_safe(ws, coord_str, val, align_h, font_size, skip_color=True):
    try:
        top_left = coord_str.split(':')[0]
        r, c = openpyxl.utils.coordinate_to_tuple(top_left)
        target_cell = ws[top_left]
        
        for m_range in ws.merged_cells.ranges:
            if m_range.min_row <= r <= m_range.max_row and m_range.min_col <= c <= m_range.max_col:
                true_top_left = ws.cell(row=m_range.min_row, column=m_range.min_col).coordinate
                target_cell = ws[true_top_left]
                break
                
        target_cell.value = val
        apply_grimoire_styles(target_cell, str(val), is_info_col=False, skip_color=skip_color)
        target_cell.font = Font(name="Consolas", size=font_size)
        target_cell.alignment = Alignment(horizontal=align_h, vertical='center', shrink_to_fit=False, wrap_text=False)
    except Exception as e:
        pass

def get_planet_data(p_dict, p_name):
    if p_name in p_dict: return p_dict[p_name]
    if p_name == "Rahu": return p_dict.get("North Node (m)", {})
    if p_name == "Ketu": return p_dict.get("South Node (m)", {})
    return {}

def compile_a9_vd_grimoire(chart_data, seed_data=None):
    try:
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb["a9_vd"] if "a9_vd" in wb.sheetnames else wb.active
        ws.title = "a9_vd"
        for s in list(wb.sheetnames):
            if s != "a9_vd": del wb[s]
        wb.active = 0

        if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
            raise ValueError("Grimoire requires Albedo seed data.")

        seed_a, seed_b = seed_data['seed1'], seed_data['seed2']
        seed_dav = calculate_davison_midpoint(seed_a, seed_b)
        
        # Davison 좌표 텍스트화
        d_lat, d_lng = get_safe_float(seed_dav.get("lat")), get_safe_float(seed_dav.get("lng"))
        coord_str = f"{abs(d_lat):.2f}°{'N' if d_lat>=0 else 'S'}, {abs(d_lng):.2f}°{'E' if d_lng>=0 else 'W'}"
        seed_dav["city"] = seed_dav["location"] = seed_dav["city_name"] = coord_str

        # 2. 스탬퍼 적용 (A2: Dav, A3: A, A4: B)
        # 🚀 [a10 벤치마킹]: 기존 텍스트(Seoul, Korea 등)를 우선 보존하고 Name 누락을 방지
        def sanitize_for_stamper(seed_obj):
            s_copy = seed_obj.copy()
            lat = s_copy.get("lat")
            lng = s_copy.get("lng")
            
            if lat is not None and lng is not None:
                try:
                    lat_f, lng_f = float(lat), float(lng)
                    coord_str = f"{abs(lat_f):.2f}°{'N' if lat_f>=0 else 'S'}, {abs(lng_f):.2f}°{'E' if lng_f>=0 else 'W'}"
                except: coord_str = "Unknown Location"
            else: coord_str = "Unknown Location"
            
            # 기존 city/location 문자열이 있으면 무조건 최우선 적용
            s_copy["city"] = s_copy.get("city") or s_copy.get("location") or coord_str
            s_copy["location"] = s_copy.get("location") or s_copy.get("city") or coord_str
            s_copy["name"] = s_copy.get("name", "Unknown")
            return s_copy

        apply_natal_stamp(ws, sanitize_for_stamper(seed_dav), method="single", cells=["A2"])
        apply_natal_stamp(ws, sanitize_for_stamper(seed_a), method="single", cells=["A3"])
        apply_natal_stamp(ws, sanitize_for_stamper(seed_b), method="single", cells=["A4"])

        # =========================================================
        # 🚀 [복구]: 실수로 지워졌던 Ayanamsa 변수 정의 및 엑셀 출력 부분
        meta_req = chart_data.get("metadata", {})
        ayanamsa = meta_req.get("ayanamsa", "lahiri")
        is_kp = (str(ayanamsa).lower() == 'kp')

        # Ayanamsa 명칭 주입
        ayan_map = {'lahiri': 'Lahiri', 'raman': 'Raman', 'kp': 'KP', 'fagan-bradley': 'Fagan-Bradley', 'yukteswar': 'Yukteswar'}
        display_ayan = ayan_map.get(str(ayanamsa).lower(), str(ayanamsa).capitalize())
        write_to_merge_safe(ws, "B6", display_ayan, 'right', 10, skip_color=True)
        # =========================================================

        # 3. 각 차트별 행성 지표 강제 주입 (Hardcoded)
        blocks = [
            (seed_dav, 9),  # Davison (Row 9~17)
            (seed_a, 20),   # Seed A (Row 20~28)
            (seed_b, 31)    # Seed B (Row 31~39)
        ]

        planet_order = ["Ketu", "Venus", "Sun", "Moon", "Mars", "Rahu", "Jupiter", "Saturn", "Mercury"]

        for s_data, start_r in blocks:
            d_str = str(s_data.get('birth_date')).split('T')[0]
            t_str = str(s_data.get('birth_time', '12:00:00'))
            lat, lng = get_safe_float(s_data.get('lat')), get_safe_float(s_data.get('lng'))
            tz = _ensure_float_tz(s_data.get('timezone', 9.0), d_str)
            is_unk = bool(s_data.get("is_time_unknown", 0))

            trop_p = calculate_principia(d_str, t_str, lat, lng, timezone=tz, system='tropical', ayanamsa=ayanamsa, is_time_unknown=is_unk)
            sid_p = calculate_principia(d_str, t_str, lat, lng, timezone=tz, system='sidereal', ayanamsa=ayanamsa, is_time_unknown=is_unk)

            for idx, p_name in enumerate(planet_order):
                r_idx = start_r + idx
                tp = get_planet_data(trop_p['planets'], p_name)
                sp = get_planet_data(sid_p['planets'], p_name)

                t_info, s_info = tp.get("dms", "-"), sp.get("dms", "-")
                
                if p_name in ["Rahu", "Ketu"]:
                    t_dig, s_dig = "-", "-"
                else:
                    t_raw, s_raw = tp.get("dignity", ""), sp.get("dignity", "")
                    t_dig = str(t_raw).strip() if t_raw and str(t_raw).strip() != "None" else "-"
                    s_dig = str(s_raw).strip() if s_raw and str(s_raw).strip() != "None" else "-"

                nak_name = sp.get("nakshatra", {}).get("name", "-")
                pada_sub_str = sp.get("sub_lord", "-") if is_kp else sp.get("pada_lord", "-")

                write_to_merge_safe(ws, f"B{r_idx}", t_info, 'left', 9, skip_color=False)    # Tropical Info (B:C)
                write_to_merge_safe(ws, f"D{r_idx}", s_info, 'left', 9, skip_color=False)    # Sidereal Info (D:E)
                write_to_merge_safe(ws, f"F{r_idx}", t_dig, 'center', 9, skip_color=True)    # Trop Dignity
                write_to_merge_safe(ws, f"G{r_idx}", s_dig, 'center', 9, skip_color=True)    # Sid Dignity
                write_to_merge_safe(ws, f"H{r_idx}", nak_name, 'right', 9, skip_color=False) # Nakshatra (H:I)
                write_to_merge_safe(ws, f"J{r_idx}", pada_sub_str, 'center', 9, skip_color=False) # Pada/Sub

        # 4. Shrink Rows (높이 최소화)
        for r_idx in [7, 18, 29, 40]:
            ws.row_dimensions[r_idx].height = 4.5

        # 5. A9_VD 타임라인 (Jyotish)
        a9_res = calculate_a9_synchronicum(seed_data, mode='jyotish', subMode='davison', ayanamsa=ayanamsa)
        timeline = a9_res['data']['timeline']

        start_row = 45
        today_str = datetime.now().strftime("%Y-%m-%d")

        def render_veil_block(row_dict, r_idx, prefix, cols, veil_key, veil_text_key, font_color):
            """장막(Veil) 여부를 판별하여 다크모드 및 텍스트를 연출하는 헬퍼"""
            is_veil = row_dict.get(veil_key, False)
            v_text = row_dict.get(veil_text_key)

            for idx, c_char in enumerate(cols):
                cell = ws[f'{c_char}{r_idx}']
                if is_veil:
                    cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid") # Dark Mode
                    if idx == 0 and v_text: # 첫 번째 L1 열에만 텍스트 주입
                        cell.value = v_text
                        cell.font = Font(name="Consolas", size=10, italic=True, color=font_color)
                        cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=False, wrap_text=False)
                    else:
                        cell.value = ""
                else:
                    # 정상 출력 (장막이 거친 후)
                    raw_val = row_dict.get(prefix, {}).get(f'l{idx+1}', '-')
                    val = SYMBOL_MAP.get(str(raw_val), str(raw_val))
                    cell.value = val
                    apply_grimoire_styles(cell, str(raw_val), is_info_col=False, skip_color=False)
                    cell.font = Font(name="Consolas", size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=False, wrap_text=False)

        for i, row_data in enumerate(timeline):
            r_idx = start_row + i
            ws.row_dimensions[r_idx].height = 16.5
            
            # 오늘 구간 스캔
            curr_d = row_data['date']
            next_d = timeline[i+1]['date'] if i+1 < len(timeline) else "9999-12-31"
            is_today = (curr_d <= today_str < next_d)

            # Column A (Date)
            c_a = ws[f'A{r_idx}']
            c_a.value = curr_d
            c_a.font = Font(name="Consolas", size=10)
            c_a.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=False, wrap_text=False)
            if is_today:
                c_a.fill = PatternFill(start_color="B0ECFF", end_color="B0ECFF", fill_type="solid")

            # 🚀 [The Veil Effects 연출]
            # Person A (B, C, D) -> 다 이루어졌다 (Red)
            render_veil_block(row_data, r_idx, 'a', ['B', 'C', 'D'], 'veil_a', 'veil_text_a', "FF0000")
            # Davison (E, F, G) -> 빛이 있으라 (White)
            render_veil_block(row_data, r_idx, 'dav', ['E', 'F', 'G'], 'veil_d', 'veil_text_dav', "FFFFFF")
            # Person B (H, I, J) -> 다 이루어졌다 (Red)
            render_veil_block(row_data, r_idx, 'b', ['H', 'I', 'J'], 'veil_b', 'veil_text_b', "FF0000")

        # 6. 사후 병합(Post-Merge) 및 틀 고정
        ws.sheet_view.pane = None
        ws.freeze_panes = 'B45'

        # 각 시드(Davison, A, B)의 행성 테이블 컬럼 병합
        all_p_rows = list(range(9, 18)) + list(range(20, 29)) + list(range(31, 40))
        ranges_to_merge = []
        for r in all_p_rows:
            ranges_to_merge.extend([f"B{r}:C{r}", f"D{r}:E{r}", f"H{r}:I{r}"])
        
        for m_range in ranges_to_merge:
            try: ws.merge_cells(m_range)
            except: pass

        return wb

    except Exception as ex:
        print(f">>> [A9_VD FINAL REFORGE ERROR] {ex}", file=sys.stderr)
        raise ex