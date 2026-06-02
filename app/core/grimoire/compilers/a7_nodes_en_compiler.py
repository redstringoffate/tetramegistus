import json
import os
import shutil
import math
from copy import copy
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from core.grimoire.styler import apply_grimoire_styles, HOUSE_COLORS, TABULA_FONT_COLORS 
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz
from core.astrology.hypostases import find_persona_transit
from core.astrology.davison import calculate_davison_midpoint
import swisseph as swe

BASE_DIR = os.path.dirname(__file__)

# 🚀 템플릿 및 매핑 경로 설정 (A7 Planets EN)
MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a7_nodes_en_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a7_nodes_en.xlsx'))

ALLOWED_UPPER = []
for cat in TABULA_FONT_COLORS.values():
    ALLOWED_UPPER.extend(cat["names"])

HOUSE_SYS_NAMES = {
    'P': 'Placidus', 'W': 'Whole Sign', 'K': 'Koch', 'R': 'Regiomontanus', 
    'C': 'Campanus', 'E': 'Equal', 'O': 'Porphyry'
}

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def get_house_of_point(lon, houses_dict):
    if not houses_dict: return "H1"
    for i in range(1, 13):
        c1 = houses_dict.get(i, {}).get('longitude', 0.0) if isinstance(houses_dict.get(i), dict) else houses_dict.get(i, 0.0)
        c2 = houses_dict.get(i+1 if i < 12 else 1, {}).get('longitude', 0.0) if isinstance(houses_dict.get(i+1 if i < 12 else 1), dict) else houses_dict.get(i+1 if i < 12 else 1, 0.0)
        if c1 < c2:
            if c1 <= lon < c2: return f"H{i}"
        else:
            if c1 <= lon < 360 or 0 <= lon < c2: return f"H{i}"
    return "H1"

def get_row_for_lon(lon_val, mapping_data):
    lon_val = float(lon_val) % 360
    sign_idx = int(lon_val // 30)
    deg = int(lon_val % 30)
    signs = ["Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo", "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces"]
    sign_name = signs[sign_idx]
    return mapping_data["degree_rows"][sign_name]["start"] + deg

def compile_a7_nodes_en_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a7_nodes"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()

    # Davison 시드 생성
    dav_hydrated = calculate_davison_midpoint(s1_raw, s2_raw)

    # Location 정규화
    d_lat = get_safe_float(dav_hydrated.get("lat"), None)
    d_lng = get_safe_float(dav_hydrated.get("lng"), None)
    
    if d_lat is not None and d_lng is not None:
        coord_str = f"{abs(d_lat):.2f}°{'N' if d_lat>=0 else 'S'}, {abs(d_lng):.2f}°{'E' if d_lng>=0 else 'W'}"
    else:
        coord_str = "Unknown Location"
        
    dav_hydrated["city"] = coord_str
    dav_hydrated["location"] = coord_str
    dav_hydrated["city_name"] = coord_str
    dav_hydrated["location_name"] = coord_str

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    h_sys_raw = str(meta.get("h_sys", "P")).strip().upper()
    h_sys = h_sys_raw[0] if h_sys_raw else "P"
    is_unk = bool(dav_hydrated.get("is_time_unknown", 0))

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: 
        mapping = json.load(f)

    # 스탬프 적용
    apply_natal_stamp(ws, dav_hydrated, method="single", cells=[mapping.get("metadata", {}).get("stamper", "A2")])
    
    ws["B4"] = system.upper()
    ws["B5"] = HOUSE_SYS_NAMES.get(h_sys, "Placidus")
    if system.lower() == "sidereal": 
        ws["C4"] = ayanamsa.upper()
        
    for c_ref in ["A2", "B4", "B5", "C4"]:
        if ws[c_ref].value:
            ws[c_ref].font = Font(name="Consolas", size=11)
            ws[c_ref].alignment = Alignment(horizontal='left', vertical='center')
    
    lat = get_safe_float(dav_hydrated.get("lat"), 37.5665)
    lng = get_safe_float(dav_hydrated.get("lng"), 126.9780)
    date_str = str(dav_hydrated.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(dav_hydrated.get('birth_time', '12:00:00'))
    tz_val = dav_hydrated.get('tz') if dav_hydrated.get('tz') is not None else dav_hydrated.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)
    
    dt_obj = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
    root_jd = swe.julday(dt_obj.year, dt_obj.month, dt_obj.day, dt_obj.hour + dt_obj.minute/60.0 - tz)

    # 1. 렌더링용 루트 차트
    root_res = calculate_principia(date_str, time_str, lat, lng, tz, system, ayanamsa, h_sys=h_sys)
    root_arc_p = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='paulus', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    try: root_arc_v = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='valens', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    except: root_arc_v = {'lots': {}}

    # 2. 페르소나 역산 전용 타겟 차트 (시간 고정을 위해 무조건 "tropical" 강제 적용)
    trop_res = calculate_principia(date_str, time_str, lat, lng, tz, "tropical", ayanamsa, h_sys=h_sys)
    trop_arc_p = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='paulus', system="tropical", ayanamsa=ayanamsa, h_sys=h_sys)
    try: trop_arc_v = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='valens', system="tropical", ayanamsa=ayanamsa, h_sys=h_sys)
    except: trop_arc_v = {'lots': {}}

    target_pool = {}
    for k, v in trop_res.get('planets', {}).items():
        target_pool[k] = v.get('longitude')
        if k == "Eros": target_pool["Asteroid Eros"] = v.get('longitude')
        elif k == "Lilith": target_pool["Asteroid Lilith"] = v.get('longitude')
        elif k == "North Node (m)": target_pool["Rahu"] = v.get('longitude')
        elif k == "South Node (m)": target_pool["Ketu"] = v.get('longitude')
        elif k == "True Node": target_pool["North Node (t)"] = v.get('longitude')
        elif k == "South Node": target_pool["South Node (t)"] = v.get('longitude')

    for k, v in trop_arc_p.get('lots', {}).items(): target_pool[k] = v.get('longitude', v.get('value'))
    for k, v in trop_arc_v.get('lots', {}).items():
        if k in ['Eros', 'Necessity']: target_pool[f"{k} (v)"] = v.get('longitude', v.get('value'))
    for k, v in trop_arc_p.get('vertex', {}).items(): target_pool[k] = v.get('longitude', v.get('value'))
    if 'syzygy' in trop_arc_p and trop_arc_p['syzygy'].get('data'):
        target_pool['Syzygy'] = trop_arc_p['syzygy']['data'].get('longitude', trop_arc_p['syzygy']['data'].get('value'))

    ts_row = mapping.get("timestamps_row", 7)

    matrix = {
        col_char: {
            "items": {deg: [] for deg in range(360)},
            "bg": {deg: "H1" for deg in range(360)}
        }
        for col_char in mapping.get("columns", {}).values()
    }

    for persona_name, col_char in mapping.get("columns", {}).items():
        target_lon = target_pool.get(persona_name)
        if target_lon is None: 
            for deg_abs in range(360):
                h_str = None
                for i in range(1, 13):
                    c_lon = root_res.get('houses', {}).get(i, {}).get('longitude')
                    if c_lon is not None and int(math.floor(c_lon)) == deg_abs:
                        h_str = f"H{i}"
                        break
                if not h_str:
                    h_str = get_house_of_point(deg_abs + 0.5, root_res.get('houses', {}))
                matrix[col_char]["bg"][deg_abs] = h_str
            continue

        if persona_name == "Sun":
            p_res, p_arc_p, p_arc_v = root_res, root_arc_p, root_arc_v
            y, m, d = dt_obj.year, dt_obj.month, dt_obj.day
            p_hour, p_min = dt_obj.hour, dt_obj.minute
        else:
            p_jd = find_persona_transit(root_jd, target_lon)
            y_u, m_u, d_u, h_dec_u = swe.revjul(p_jd, swe.GREG_CAL)
            total_sec = int(round(h_dec_u * 3600))
            utc_dt = datetime(y_u, m_u, d_u) + timedelta(seconds=total_sec)
            local_dt = utc_dt + timedelta(hours=tz)
            
            y, m, d = local_dt.year, local_dt.month, local_dt.day
            p_hour, p_min = local_dt.hour, local_dt.minute
            
            p_res = calculate_principia(f"{y:04d}-{m:02d}-{d:02d}", f"{p_hour:02d}:{p_min:02d}:00", lat, lng, tz, system, ayanamsa, h_sys=h_sys)
            p_arc_p = calculate_arcana(f"{y:04d}-{m:02d}-{d:02d}", f"{p_hour:02d}:{p_min:02d}:00", lat, lng, tz, lot_schema='paulus', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
            try: p_arc_v = calculate_arcana(f"{y:04d}-{m:02d}-{d:02d}", f"{p_hour:02d}:{p_min:02d}:00", lat, lng, tz, lot_schema='valens', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
            except: p_arc_v = {'lots': {}}

        ts_cell = ws[f"{col_char}{ts_row}"]
        ts_cell.value = f"{y:04d}-{m:02d}-{d:02d}; {p_hour:02d}:{p_min:02d}"
        ts_cell.font = Font(name="Consolas", size=8, color="000000")
        ts_cell.alignment = Alignment(horizontal='left', vertical='center')

        for deg_abs in range(360):
            h_str = None
            for i in range(1, 13):
                c_lon = p_res.get('houses', {}).get(i, {}).get('longitude')
                if c_lon is not None and int(math.floor(c_lon)) == deg_abs:
                    h_str = f"H{i}"
                    break
            if not h_str:
                h_str = get_house_of_point(deg_abs + 0.5, p_res.get('houses', {}))
            matrix[col_char]["bg"][deg_abs] = h_str

        day_lords = [x.strip() for x in str(p_res.get('lords', {}).get('day', '-')).split('|')]
        hour_lord = str(p_res.get('lords', {}).get('hour', '-')).strip()

        def add_item(lon, name, is_day=False, is_hour=False):
            if lon is None: return
            deg_abs = int(math.floor(lon)) % 360
            matrix[col_char]["items"][deg_abs].append({
                "name": name, "lon": lon, "is_day": is_day, "is_hour": is_hour
            })

        for g in ['planets', 'asteroids', 'lilith_nodes', 'angles']:
            for k, v in p_res.get(g, {}).items():
                if is_unk and k in ["Ascendant", "Midheaven", "Descendant", "Immum Coeli"]: continue
                
                # 🚀 [노드 차단 방역]: Draconic / Ketunic 모드 시 노드 계열 출력 스킵
                if system.lower() in ['draconic', 'ketunic']:
                    if k in ['North Node (t)', 'South Node (t)', 'Rahu', 'Ketu', 'North Node (m)', 'South Node (m)', 'True Node', 'South Node']:
                        continue

                d_name = k.replace(" (Natal)", "")
                if d_name.upper() not in ALLOWED_UPPER: continue
                add_item(v.get('longitude'), d_name, is_day=(k in day_lords), is_hour=(k == hour_lord))

        if not is_unk:
            for k, v in p_arc_p.get('lots', {}).items(): add_item(v.get('value', v.get('longitude')), k)
            for k, v in p_arc_v.get('lots', {}).items():
                if k in ['Eros', 'Necessity']: add_item(v.get('value', v.get('longitude')), f"{k} (v)")
            for k, v in p_arc_p.get('vertex', {}).items(): add_item(v.get('value', v.get('longitude')), k)
            if 'syzygy' in p_arc_p and p_arc_p['syzygy'].get('data'):
                add_item(p_arc_p['syzygy']['data'].get('value', p_arc_p['syzygy']['data'].get('longitude')), "Syzygy")

        if not is_unk:
            is_whole_sign = (h_sys == 'W')
            for i in range(1, 13):
                if not is_whole_sign and i in [1, 4, 7, 10]: continue
                c_lon = p_res.get('houses', {}).get(i, {}).get('longitude')
                if c_lon is not None: add_item(c_lon, f"{i}H CUSP")

    all_col_chars = mapping.get("columns", {}).values()
    
    # 🚀 [핵심 방역 4]: 파이썬이 행 높이(Row Dimensions)를 통제하며 줄 삽입! (Shrink Row 에러 원천 차단)
    for deg_abs in range(359, -1, -1):
        r_idx = get_row_for_lon(deg_abs, mapping)
        
        max_lines = 1
        for col_char in all_col_chars:
            raw_items = matrix[col_char]["items"][deg_abs]
            raw_items.sort(key=lambda x: x["lon"])
            unique_items = []
            seen_names = set()
            for item in raw_items:
                if item["name"] not in seen_names:
                    unique_items.append(item)
                    seen_names.add(item["name"])
            matrix[col_char]["items"][deg_abs] = unique_items
            if len(unique_items) > max_lines:
                max_lines = len(unique_items)
                
        if max_lines > 1:
            amount = max_lines - 1
            ins_idx = r_idx + 1
            
            # 현재 시트의 모든 행 높이 안전하게 백업
            old_heights = {r: ws.row_dimensions[r].height for r in list(ws.row_dimensions.keys())}
                    
            ws.insert_rows(ins_idx, amount=amount)
            
            # 행 높이 매핑 (Shrink Row 보존)
            for r in range(ws.max_row, 0, -1):
                if r >= ins_idx + amount:
                    orig_r = r - amount
                    ws.row_dimensions[r].height = old_heights.get(orig_r, 15)
                elif ins_idx <= r < ins_idx + amount:
                    ws.row_dimensions[r].height = old_heights.get(r_idx, 15)
                else:
                    ws.row_dimensions[r].height = old_heights.get(r, 15)
            
            # 새로 생긴 줄에 서식 복사 (A열 텍스트 중복 방지)
            for r_offset in range(1, max_lines):
                new_r = r_idx + r_offset
                for c_idx in range(1, ws.max_column + 1):
                    src = ws.cell(row=r_idx, column=c_idx)
                    tgt = ws.cell(row=new_r, column=c_idx)
                    if src.has_style:
                        tgt.font = copy(src.font)
                        tgt.border = copy(src.border)
                        tgt.fill = copy(src.fill)
                        tgt.number_format = copy(src.number_format)
                        tgt.alignment = copy(src.alignment)
                    
                    if c_idx == 1: 
                        tgt.value = "" 

        # 텍스트 및 색상 주입
        for col_char in all_col_chars:
            h_str = matrix[col_char]["bg"][deg_abs]
            bg_hex = HOUSE_COLORS.get(h_str)
            items = matrix[col_char]["items"][deg_abs]
            
            for line_idx in range(max_lines):
                target_row = r_idx + line_idx
                cell = ws[f"{col_char}{target_row}"]
                
                if line_idx < len(items):
                    item = items[line_idx]
                    name = item["name"]
                    val_upper = name.upper()
                    
                    is_angle = val_upper in ["ASCENDANT", "MIDHEAVEN", "DESCENDANT", "IMMUM COELI"]
                    if item["is_hour"] or is_angle:
                        name = val_upper
                        
                    cell.value = name
                    apply_grimoire_styles(cell, name, is_day_lord=item["is_day"], is_hour_lord=item["is_hour"], tabula_mode=True, house_bg=bg_hex)
                else:
                    cell.value = ""
                    apply_grimoire_styles(cell, "", tabula_mode=True, house_bg=bg_hex)
                    
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # 🚀 시트 전체 폰트를 Consolas로 통일
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name="Consolas", 
                    size=cell.font.size, 
                    bold=cell.font.bold,
                    italic=cell.font.italic, 
                    color=cell.font.color
                )

    # 🚀 [추가된 과잉친절]: A열(Sabian Number)과 상단 헤더(1~8행) 완벽 고정
    ws.freeze_panes = "B9"

    # ... (기존 폰트 세탁 반복문 끝) ...

    # ====================================================================
    # 🚀 [추가]: Tabula 컬럼 Shrink 완벽 방어 및 너비 균일화
    # ====================================================================
    from openpyxl.utils import get_column_letter
    
    # 1. 중간 열(행성 데이터)의 폭을 일정하게 고정하고, 나머지는 넉넉하게 설정
    temp_wb_for_width = load_workbook(TEMPLATE_FILE)
    temp_ws_for_width = temp_wb_for_width[base_name] if base_name in temp_wb_for_width.sheetnames else temp_wb_for_width.active
    
    # mapping에서 가져온 데이터 열 문자들 (B, C, D...)
    mapped_cols = list(all_col_chars) 
    
    for i in range(1, ws.max_column + 1):
        col_letter = get_column_letter(i)
        temp_width = temp_ws_for_width.column_dimensions[col_letter].width
        
        if col_letter == 'A':
            ws.column_dimensions[col_letter].width = 15.0  # Sabian Number 열 고정
        elif col_letter in mapped_cols:
            ws.column_dimensions[col_letter].width = 15.0  # 🚀 작성자님 요청: 중간 행성/소행성 열 폭 균일화!
        else:
            # Sabian Symbols(해석) 등 나머지 열은 템플릿 값 복사 또는 넉넉하게 40.0
            ws.column_dimensions[col_letter].width = temp_width if temp_width is not None else 40.0

    # 2. A1 병합 해제 및 B1으로 피신 (뷰어의 Shrink 꼬임 원천 차단)
    merge_to_remove = None
    for m_range in list(ws.merged_cells.ranges):
        if "A1" in m_range.coord:
            merge_to_remove = m_range
            break
            
    if merge_to_remove:
        max_col_letter = get_column_letter(merge_to_remove.max_col)
        ws.unmerge_cells(merge_to_remove.coord)
        
        ws['B1'].value = ws['A1'].value
        ws['A1'].value = None
        
        if ws['A1'].has_style:
            ws['B1'].font = copy(ws['A1'].font)
            ws['B1'].border = copy(ws['A1'].border)
            ws['B1'].fill = copy(ws['A1'].fill)
            ws['B1'].alignment = copy(ws['A1'].alignment)
            
        ws.merge_cells(f"B1:{max_col_letter}1")
    # ====================================================================

    # (기존에 넣었던) A열과 상단 헤더 완벽 고정
    ws.freeze_panes = "B9"

    return wb