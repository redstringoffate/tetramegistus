import json
import os
import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import calculate_principia, _ensure_float_tz, format_dms_pretty
from core.astrology.davison import calculate_davison_midpoint
from core.astrology.lagna import calculate_all_jaimini_padas

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a3_lagna_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a3_lagna.xlsx'))

LOCAL_SIGNS = ["Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo", "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces"]

def get_safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default

def compile_a3_lagna_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 🚀 [수복]: 메모리로 즉시 로드 (새 아키텍처 이식)
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a3_lagna"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name
    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)
    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()

    dav_hydrated = calculate_davison_midpoint(s1_raw, s2_raw)

    # 🚀 "Davison Midpoint" 문자열 대신 정확한 좌표(Lat/Lon) 포맷으로 강제 덮어쓰기
    d_lat = get_safe_float(dav_hydrated.get("lat"), None)
    d_lng = get_safe_float(dav_hydrated.get("lng"), None)
    
    if d_lat is not None and d_lng is not None:
        coord_str = f"{abs(d_lat):.2f}°{'N' if d_lat>=0 else 'S'}, {abs(d_lng):.2f}°{'E' if d_lng>=0 else 'W'}"
    else:
        coord_str = "Unknown Location"
        
    dav_hydrated["city"] = coord_str
    dav_hydrated["location"] = coord_str
    dav_hydrated["city_name"] = coord_str
    dav_hydrated["location_name"] = coord_str

    # 단일 스탬프 적용
    apply_natal_stamp(ws, dav_hydrated, method="single", cells=["A2"])

    date_str = str(dav_hydrated.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(dav_hydrated.get('birth_time', '12:00:00'))
    lat = get_safe_float(dav_hydrated.get("lat"), 37.5665)
    lng = get_safe_float(dav_hydrated.get("lng"), 126.9780)
    tz_val = dav_hydrated.get('tz') if dav_hydrated.get('tz') is not None else dav_hydrated.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)
    is_unk = bool(dav_hydrated.get("is_time_unknown", 0))

    res = calculate_principia(
        date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz,
        system=system, ayanamsa=ayanamsa, view=view, h_sys=h_sys,
        fixed_star_orb=1.0, is_time_unknown=is_unk
    )

    planets_data = res.get('planets', {})
    angles_data = res.get('angles', {})
    houses_data = res.get('houses', {})

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    # 🚀 KeyError를 원천 차단하는 안전한 메타데이터 할당 (sys_tab 에러 방지)
    meta_map = mapping.get("metadata", {})
    asc_data = angles_data.get("Ascendant") or planets_data.get("Ascendant") or {}
    mc_data = angles_data.get("Midheaven") or planets_data.get("Midheaven") or {}
    
    if meta_map.get("ascendant"): 
        ws[meta_map["ascendant"]] = asc_data.get("dms", "")
    if meta_map.get("midheaven"): 
        ws[meta_map["midheaven"]] = mc_data.get("dms", "")
    if system.upper() == "SIDEREAL" and meta_map.get("ayanamsa"): 
        ws[meta_map["ayanamsa"]] = ayanamsa.upper()

    for key in ["ascendant", "midheaven", "ayanamsa", "sys_tab", "h_sys"]:
        c_ref = meta_map.get(key)
        if c_ref and ws[c_ref].value:
            cell = ws[c_ref]
            should_skip_color = True if key in ["ayanamsa", "sys_tab", "h_sys"] else False
            apply_grimoire_styles(cell, cell.value, is_info_col=False, skip_color=should_skip_color)
            
            color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
            cell.font = Font(name="Consolas", color=color_to_use, bold=False, italic=False)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    norm_planets = copy(planets_data)
    for k, v in planets_data.items():
        if k in ["Mean Node", "North Node (t)", "North Node (m)", "True Node"]: 
            norm_planets["Rahu"] = v
        if k in ["South Node", "South Node (t)", "South Node (m)"]: 
            norm_planets["Ketu"] = v

    cusps_simple = {}
    angle_map = {1: "Ascendant", 4: "Immum Coeli", 7: "Descendant", 10: "Midheaven"}
    
    for i in range(1, 13):
        if i in angle_map:
            ang_data = angles_data.get(angle_map[i]) or planets_data.get(angle_map[i]) or {}
            cusps_simple[i] = float(ang_data.get('longitude', 0.0))
        else:
            h_data = houses_data.get(i) or houses_data.get(str(i)) or houses_data.get(f"H{i}") or {}
            val = h_data.get('longitude', 0.0) if isinstance(h_data, dict) else h_data
            if val == {} or val == '' or val is None: val = 0.0
            cusps_simple[i] = float(val)
        
    asc_lon = float(asc_data.get('longitude', 0.0))
    asc_sign_idx = int(asc_lon / 30) % 12

    try:
        raw_padas = calculate_all_jaimini_padas(norm_planets, cusps_simple, asc_sign_idx)
    except Exception as e:
        print(f"[LAGNA ERROR] {e}")
        raw_padas = {}

    padas = {
        "AL": raw_padas.get(1, {}),
        "A2": raw_padas.get(2, {}),
        "A3": raw_padas.get(3, {}),
        "A4": raw_padas.get(4, {}),
        "A5": raw_padas.get(5, {}),
        "A6": raw_padas.get(6, {}),
        "A7": raw_padas.get(7, {}),
        "A8": raw_padas.get(8, {}),
        "A9": raw_padas.get(9, {}),
        "A10": raw_padas.get(10, {}),
        "A11": raw_padas.get(11, {}),
        "UL": raw_padas.get(12, {})
    }

    for lagna_key, map_data in mapping.get("lagnas", {}).items():
        cell_ref = map_data["info"]
        pada_info = padas.get(lagna_key)
        
        cell_val = ""
        if pada_info is not None and pada_info != {}:
            if isinstance(pada_info, dict):
                # 🚀 position_str 우선순위 (Pisces 9H 형태 출력)
                if "position_str" in pada_info:
                    cell_val = pada_info["position_str"]
                else:
                    lon = pada_info.get("longitude", pada_info.get("lon", pada_info.get("value")))
                    if lon is not None:
                        lon = float(lon)
                        sign_val = pada_info.get("sign")
                        
                        if sign_val is not None:
                            sign_name = LOCAL_SIGNS[int(sign_val)] if isinstance(sign_val, int) else str(sign_val)
                        else:
                            sign_name = LOCAL_SIGNS[int(lon / 30) % 12]
                            
                        raw_dms = str(pada_info.get("dms", format_dms_pretty(lon)))
                        if raw_dms.lower().startswith(sign_name.lower()):
                            raw_dms = raw_dms[len(sign_name):].strip(" ,")
                            
                        cell_val = f"{sign_name}, {raw_dms}"
            elif isinstance(pada_info, (int, float)):
                lon = float(pada_info)
                sign_name = LOCAL_SIGNS[int(lon / 30) % 12]
                raw_dms = format_dms_pretty(lon)
                cell_val = f"{sign_name}, {raw_dms}"
            elif isinstance(pada_info, str):
                cell_val = pada_info

        ws[cell_ref] = cell_val

        # 스타일 및 우그러짐 방지 적용
        cell = ws[cell_ref]
        apply_grimoire_styles(cell, cell.value, is_info_col=True, skip_color=False)
        
        # 🚀 볼드체 해제
        color_info = cell.font.color if cell.font and cell.font.color else "000000"
        cell.font = Font(name="Consolas", color=color_info, bold=False, italic=False)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    return wb