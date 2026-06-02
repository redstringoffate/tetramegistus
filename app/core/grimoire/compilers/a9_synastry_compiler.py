import json
import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.engine import calculate_a9_synchronicum, calculate_principia, calculate_arcana, SYMBOL_MAP, _ensure_float_tz

BASE_DIR = os.path.dirname(__file__)
MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a9_synastry_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a9_synastry.xlsx'))

def get_nested(d, path, default="-"):
    if not path: return default
    keys = path.split('.')
    val = d
    try:
        for key in keys:
            if isinstance(val, dict): val = val.get(key)
            else: return default
        return val if val is not None else default
    except: return default

# 🚀 [FIX 1]: NoneType 에러를 막아주는 강력한 안전장치 추가!
def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def compile_a9_synastry_grimoire(chart_data, seed_data=None):
    try:
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb["a9_synastry"] if "a9_synastry" in wb.sheetnames else wb.active
        ws.title = "a9_synastry"
        for s in list(wb.sheetnames):
            if s != "a9_synastry": del wb[s]
        wb.active = 0

        if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
            raise ValueError("Grimoire requires Albedo seed data.")

        seed_a, seed_b = seed_data['seed1'], seed_data['seed2']
        meta_req = chart_data.get("metadata", {})
        ayanamsa = meta_req.get("ayanamsa", "lahiri")
        seg_num = int(meta_req.get("segment", 1))

        # 2. 스탬퍼 적용 (A2: Seed A, A3: Seed B)
        # 🚀 [a10 벤치마킹]: 기존 텍스트(Seoul, Korea 등)를 우선 보존하고 Name 누락을 방지
        def sanitize_for_stamper(seed_obj):
            s_copy = seed_obj.copy()
            lat = s_copy.get("lat")
            lng = s_copy.get("lng")
            
            if lat is not None and lng is not None:
                try:
                    lat_f, lng_f = float(lat), float(lng)
                    coord_str = f"{abs(lat_f):.2f}°{'N' if lat_f>=0 else 'S'}, {abs(lng_f):.2f}°{'E' if lng_f>=0 else 'W'}"
                except: coord_str = "Unknown Location"
            else: coord_str = "Unknown Location"
            
            # 기존 city/location 문자열이 있으면 무조건 최우선 적용
            s_copy["city"] = s_copy.get("city") or s_copy.get("location") or coord_str
            s_copy["location"] = s_copy.get("location") or s_copy.get("city") or coord_str
            s_copy["name"] = s_copy.get("name", "Unknown")
            return s_copy

        apply_natal_stamp(ws, sanitize_for_stamper(seed_a), method="single", cells=["A2"])
        apply_natal_stamp(ws, sanitize_for_stamper(seed_b), method="single", cells=["A3"])

        # 🚀 [잘림 방지]: 스탬퍼 글자가 템플릿 가운데 정렬에 막혀 잘리지 않게 강제 좌측 정렬
        for coord in ["A2", "A3"]:
            ws[coord].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            
        # 👇 여기서부터 새로 추가! (SheetJS 웹 뷰어 오버플로우 렌더링 방어)
        try:
            ws.merge_cells('A2:F2')
            ws.merge_cells('A3:F3')
        except:
            pass

        # 3. 개별 Principia/Arcana 연산 (상단 정보용)
        def get_full_data(sd):
            d_str = str(sd.get('birth_date')).split('T')[0]
            t_str = str(sd.get('birth_time', '12:00:00'))
            # 🚀 [FIX 2]: float() 대신 get_safe_float()를 사용하여 에러 원천 차단
            lat = get_safe_float(sd.get('lat'), 0.0)
            lng = get_safe_float(sd.get('lng'), 0.0)
            tz = _ensure_float_tz(sd.get('timezone', 9.0), d_str)
            
            p = calculate_principia(d_str, t_str, lat, lng, timezone=tz, ayanamsa=ayanamsa)
            a = calculate_arcana(d_str, t_str, lat, lng, timezone=tz, ayanamsa=ayanamsa)
            return p, a

        data_a_p, data_a_arc = get_full_data(seed_a)
        data_b_p, data_b_arc = get_full_data(seed_b)

        # 4. A9 Synchronicum 연산 (타임라인용)
        # Synastry 모드이므로 나중에 태어난 사람 기준 시작
        a9_res = calculate_a9_synchronicum(seed_data, mode='zodiac', subMode='synastry', ayanamsa=ayanamsa)
        timeline = a9_res['data']['timeline']

        # 5. 상단 정보 직접 주입
        roman_map = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI"}
        ws['B5'].value = roman_map.get(seg_num, "I")
        ws['B5'].alignment = Alignment(horizontal='center')

        def inject_person_meta(p_data, arc_data, start_row):
            # Sect, Asc, Lots
            ws[f'B{start_row}'].value = str(arc_data.get("meta", {}).get("sect", "-")).upper()
            ws[f'B{start_row+1}'].value = str(p_data['planets'].get("Ascendant", {}).get("dms", "-"))
            ws[f'B{start_row+2}'].value = str(arc_data['lots'].get("Spirit", {}).get("dms", "-"))
            ws[f'B{start_row+3}'].value = str(arc_data['lots'].get("Fortune", {}).get("dms", "-"))
            ws[f'B{start_row+4}'].value = str(arc_data['lots'].get("Eros", {}).get("dms", "-"))

            # Planets
            p_order = ["Saturn", "Jupiter", "Mars", "Sun", "Venus", "Mercury", "Moon"]
            lords = p_data.get('lords', {})
            for i, p_name in enumerate(p_order):
                r = start_row + i
                p_info = p_data['planets'].get(p_name, {})
                ws[f'E{r}'].value = p_info.get("dms", "-")
                ws[f'G{r}'].value = p_info.get("dignity", "-")
                ws[f'H{r}'].value = "TRUE" if p_name in str(lords.get("day", "")) else "-"
                ws[f'I{r}'].value = "TRUE" if p_name in str(lords.get("hour", "")) else "-"
                
                # Styling
                for c in ['E', 'G', 'H', 'I']:
                    cell = ws[f'{c}{r}']
                    apply_grimoire_styles(cell, str(cell.value), is_info_col=(c=='E'), skip_color=False)
                    cell.font = Font(name="Consolas", size=9)
                    cell.alignment = Alignment(horizontal='left' if c=='E' else 'center', vertical='center')

            # Left side Styling
            for r in range(start_row, start_row + 5):
                cell = ws[f'B{r}']
                is_sect = (r == start_row)
                apply_grimoire_styles(cell, str(cell.value), is_info_col=not is_sect, skip_color=is_sect)
                cell.font = Font(name="Consolas", size=10 if is_sect else 9)
                cell.alignment = Alignment(horizontal='center' if is_sect else 'right', vertical='center')

        inject_person_meta(data_a_p, data_a_arc, 8)  # Person A (Row 8~14)
        inject_person_meta(data_b_p, data_b_arc, 17) # Person B (Row 17~23)

        # 6. 타임라인 매핑 (12년 필터링 포함)
        start_age = (seg_num - 1) * 12
        end_age = start_age + 12
        filtered_timeline = [r for r in timeline if start_age <= r.get('age', 0) < end_age]

        with open(MAPPING_FILE, 'r', encoding='utf-8') as f: mapping = json.load(f)
        cols = mapping['timeline']['columns']
        today_str = datetime.now().strftime("%Y-%m-%d")

        for i, row_data in enumerate(filtered_timeline):
            r_idx = 29 + i
            is_today_row = False
            
            # 오늘 날짜 판별 로직
            curr_d = row_data['date']
            next_d = filtered_timeline[i+1]['date'] if i+1 < len(filtered_timeline) else "9999-12-31"
            if curr_d <= today_str < next_d: is_today_row = True

            for col_char, rule in cols.items():
                val_raw = get_nested(row_data, rule['data_key'], "-")
                val = SYMBOL_MAP.get(str(val_raw), str(val_raw))
                cell = ws[f"{col_char}{r_idx}"]
                cell.value = val
                
                # 스타일 적용
                apply_grimoire_styles(cell, str(val_raw), is_info_col=False, skip_color=(col_char == 'A'))
                cell.font = Font(name="Consolas", size=10)
                cell.alignment = Alignment(horizontal='left' if col_char == 'A' else 'center', vertical='center')
                
                if is_today_row and col_char == 'A':
                    cell.fill = PatternFill(start_color="B0ECFF", end_color="B0ECFF", fill_type="solid")

        # 7. 사후 병합 처리
        ranges = [f"{c}{r}:{c_next}{r}" for r in list(range(8,13)) + list(range(17,22)) for c, c_next in [('B','C')]]
        ranges += [f"E{r}:F{r}" for r in list(range(8,15)) + list(range(17,24))]
        for r in ranges:
            try: ws.merge_cells(r)
            except: pass

        ws.freeze_panes = 'B29'
        return wb

    except Exception as ex:
        print(f">>> [A9 SYNASTRY ERROR] {ex}", file=sys.stderr)
        raise ex