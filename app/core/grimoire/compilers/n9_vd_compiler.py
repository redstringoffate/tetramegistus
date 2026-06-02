import json
import os
import sys
from datetime import datetime
import openpyxl.utils
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.engine import calculate_n9_timeline, calculate_principia, SYMBOL_MAP, _ensure_float_tz

BASE_DIR = os.path.dirname(__file__)
MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/n9_vd_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/nigredo/n9_vd.xlsx'))

def get_nested(d, path, default="-"):
    if not path: return default
    keys = path.split('.')
    val = d
    try:
        for key in keys:
            if isinstance(val, dict): val = val.get(key)
            else: return default
        return val if val is not None else default
    except: return default

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def ensure_location_string(seed):
    loc_val = str(seed.get("location", seed.get("city", ""))).strip()
    if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
        lat, lng = get_safe_float(seed.get("lat"), None), get_safe_float(seed.get("lng"), None)
        if lat is not None and lng is not None:
            seed["city"] = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
            seed["location"] = seed["city"]
        else:
            seed["city"] = "Unknown Location"
            seed["location"] = "Unknown Location"
    else:
        seed["city"] = loc_val
        seed["location"] = loc_val

# 🚀 [Merge Cell 철통 방어 헬퍼 함수]
def write_to_merge_safe(ws, coord_str, val, align_h, font_size, skip_color=True):
    try:
        top_left = coord_str.split(':')[0]
        r, c = openpyxl.utils.coordinate_to_tuple(top_left)
        target_cell = ws[top_left]
        
        for m_range in ws.merged_cells.ranges:
            if m_range.min_row <= r <= m_range.max_row and m_range.min_col <= c <= m_range.max_col:
                true_top_left = ws.cell(row=m_range.min_row, column=m_range.min_col).coordinate
                target_cell = ws[true_top_left]
                break
                
        target_cell.value = val
        apply_grimoire_styles(target_cell, str(val), is_info_col=False, skip_color=skip_color)
        
        target_cell.font = Font(name="Consolas", size=font_size)
        # 🚀 [FIX 1]: shrink_to_fit=False, wrap_text=False 를 강제 적용하여 글씨가 쪼그라드는 현상 원천 차단
        target_cell.alignment = Alignment(horizontal=align_h, vertical='center', shrink_to_fit=False, wrap_text=False)
    except Exception as e:
        pass

# 엔진에서 반환하는 노드 이름 대응 헬퍼
def get_planet_data(p_dict, p_name):
    if p_name in p_dict: return p_dict[p_name]
    if p_name == "Rahu": return p_dict.get("North Node (m)", {})
    if p_name == "Ketu": return p_dict.get("South Node (m)", {})
    return {}

def compile_n9_vd_grimoire(chart_data, seed_data=None):
    try:
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb["n9_vd"] if "n9_vd" in wb.sheetnames else wb.active
        ws.title = "n9_vd"
        for s in list(wb.sheetnames):
            if s != "n9_vd": del wb[s]
        wb.active = 0

        # 1. 시드 & 메타데이터 파싱
        seed_data = chart_data.get("seed", chart_data)
        meta_req = chart_data.get("metadata", {})
        
        ensure_location_string(seed_data)
        apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])

        lat = get_safe_float(seed_data.get("lat"), 37.5665)
        lng = get_safe_float(seed_data.get("lng"), 126.9780)
        date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
        time_str = str(seed_data.get('birth_time', '12:00:00'))
        
        tz_val = seed_data.get('timezone', 9.0)
        tz = _ensure_float_tz(tz_val, date_str)

        ayanamsa = meta_req.get("ayanamsa", "lahiri")
        is_unk = bool(seed_data.get("is_time_unknown", 0))

        # 2. 엔진 듀얼 연산 (Tropical & Sidereal 동시 추출)
        trop_principia = calculate_principia(
            date_str, time_str, lat, lng, tz, system='tropical', ayanamsa=ayanamsa, is_time_unknown=is_unk
        )
        sid_principia = calculate_principia(
            date_str, time_str, lat, lng, tz, system='sidereal', ayanamsa=ayanamsa, is_time_unknown=is_unk
        )

        trop_p = trop_principia.get("planets", {})
        sid_p = sid_principia.get("planets", {})

        n9_result = calculate_n9_timeline(seed_data, mode='jyotish', ayanamsa=ayanamsa)
        n9_timeline = n9_result.get("timeline", [])

        # 3. 상단 메타데이터 & 행성 테이블 강제 주입 (Hardcoded)
        # B5: Ayanamsa 
        # 🚀 [요청 1 반영]: 아야남사 대소문자 예쁘게 맵핑 (KP, Fagan-Bradley 등)
        ayan_map = {
            'lahiri': 'Lahiri', 'raman': 'Raman', 'kp': 'KP', 
            'fagan-bradley': 'Fagan-Bradley', 'yukteswar': 'Yukteswar'
        }
        display_ayan = ayan_map.get(str(ayanamsa).lower(), str(ayanamsa).capitalize())
        write_to_merge_safe(ws, "B5", display_ayan, 'right', 10)

        planet_rows = {
            "Ketu": 7, "Venus": 8, "Sun": 9, "Moon": 10, "Mars": 11,
            "Rahu": 12, "Jupiter": 13, "Saturn": 14, "Mercury": 15
        }
        
        is_kp = (str(ayanamsa).lower() == 'kp')

        for p_name, r_idx in planet_rows.items():
            tp = get_planet_data(trop_p, p_name)
            sp = get_planet_data(sid_p, p_name)

            # 데이터 추출
            t_info = tp.get("dms", "-")
            s_info = sp.get("dms", "-")
            
            nak_data = sp.get("nakshatra", {})
            nak_name = nak_data.get("name", "-")
            
            if is_kp:
                pada_sub_str = sp.get("sub_lord", "-")
            else:
                pada_sub_str = sp.get("pada_lord", "-")

            # 🚀 [요청 2 반영]: Rahu, Ketu는 위계가 없으므로 "-"로 강제 통일, 빈칸도 방어
            if p_name in ["Rahu", "Ketu"]:
                t_dig, s_dig = "-", "-"
            else:
                t_raw = tp.get("dignity", "")
                s_raw = sp.get("dignity", "")
                t_dig = str(t_raw).strip() if t_raw and str(t_raw).strip() != "None" else "-"
                s_dig = str(s_raw).strip() if s_raw and str(s_raw).strip() != "None" else "-"

            # 셀에 주입 및 정렬 적용
            # B열 (Info - Left), C열 (Sidereal Info - Left)
            write_to_merge_safe(ws, f"B{r_idx}", t_info, 'left', 9, skip_color=False)
            write_to_merge_safe(ws, f"C{r_idx}", s_info, 'left', 9, skip_color=False)
            
            # 🚀 [FIX 2]: D열 (Nakshatra) skip_color=False 적용하여 색상 입히기
            write_to_merge_safe(ws, f"D{r_idx}", nak_name, 'right', 9, skip_color=False)
            
            # 🚀 [FIX 3-3]: E열 (Pada/Sub) skip_color=False 적용하여 행성 기호 색상 입히기
            write_to_merge_safe(ws, f"E{r_idx}", pada_sub_str, 'center', 9, skip_color=False)
            
            write_to_merge_safe(ws, f"F{r_idx}", t_dig, 'center', 9, skip_color=True)
            write_to_merge_safe(ws, f"G{r_idx}", s_dig, 'center', 9, skip_color=True)

        # 4. 타임라인 매핑 (Row 19~)
        with open(MAPPING_FILE, 'r', encoding='utf-8') as f: 
            mapping = json.load(f)

        start_row = mapping.get("timeline", {}).get("start_row", 19)
        cols = mapping.get("timeline", {}).get("columns", {})

        # 오늘 날짜 정확하게 스캔
        today_str = datetime.now().strftime("%Y-%m-%d")
        for idx in range(len(n9_timeline)):
            r_date = str(n9_timeline[idx].get('date', '9999-12-31')).split('T')[0]
            next_date = str(n9_timeline[idx+1].get('date', '9999-12-31')).split('T')[0] if idx + 1 < len(n9_timeline) else "9999-12-31"
            if r_date <= today_str < next_date:
                n9_timeline[idx]['is_today'] = True
                break

        for i, row_data in enumerate(n9_timeline):
            r_idx = start_row + i
            ws.row_dimensions[r_idx].height = 16.5
            is_current = row_data.get('is_today', False)

            # 🚀 [추가됨]: 엔진에 나이(age) 데이터가 없으면 생일 기준으로 즉석에서 만 나이 계산
            if "age" not in row_data:
                try:
                    r_date_str = str(row_data.get('date', '')).split('T')[0]
                    r_date_obj = datetime.strptime(r_date_str, "%Y-%m-%d")
                    b_date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    # 만 나이 계산 공식
                    age = r_date_obj.year - b_date_obj.year - ((r_date_obj.month, r_date_obj.day) < (b_date_obj.month, b_date_obj.day))
                    row_data["age"] = max(0, age) # 태어나기 전이면 0살 처리
                except:
                    pass

            for col_char, rule in cols.items():
                data_key = rule.get("data_key")
                raw_val = get_nested(row_data, data_key, "-")

                # Jyotish Dasa 기호 변환 적용
                val = SYMBOL_MAP.get(str(raw_val), str(raw_val)) if "date" not in data_key and raw_val != "-" else str(raw_val)
                
                cell = ws[f"{col_char}{r_idx}"]
                cell.value = val

                apply_grimoire_styles(cell, str(raw_val), is_info_col=False, skip_color=(col_char == 'A'))
                
                # B, C, D 기호 열은 중앙 정렬, A 날짜 열은 좌측 정렬
                align_h = 'left' if col_char == 'A' else 'center'
                cell.font = Font(name="Consolas", size=10)
                # 🚀 [FIX 1]: 타임라인 셀에도 shrink 방지 쐐기 박기
                cell.alignment = Alignment(horizontal=align_h, vertical='center', shrink_to_fit=False, wrap_text=False)

                # 오늘 구간 하늘색 형광펜!
                if is_current and col_char == 'A':
                    cell.fill = PatternFill(start_color="B0ECFF", end_color="B0ECFF", fill_type="solid")

        # 틀 고정
        ws.sheet_view.pane = None
        ws.freeze_panes = f'B{start_row}'
        
        # 🚀 [최종 마법]: 파이썬이 엑셀 열 너비(Column Width)를 강제로 시원하게 넓혀버립니다!
        # (템플릿에서 마우스로 드래그할 필요 없음)
        ws.column_dimensions['B'].width = 20.0  # Tropical Info
        ws.column_dimensions['C'].width = 20.0  # Sidereal Info
        ws.column_dimensions['D'].width = 20.0  # Nakshatra
        ws.column_dimensions['E'].width = 13.0  # Pada/Sub-Lord

        return wb

    except Exception as ex:
        print(f">>> [N9_VD FINAL REFORGE ERROR] {ex}", file=sys.stderr)
        raise ex