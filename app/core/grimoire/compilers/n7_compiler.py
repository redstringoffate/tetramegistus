import json
import os
import shutil
import math
from copy import copy
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz, format_dms_pretty, SYMBOL_MAP
from core.astrology.hypostases import find_persona_transit
import swisseph as swe

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/n7_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/nigredo/n7.xlsx'))

SABIAN_PATHS = [
    os.path.abspath(os.path.join(BASE_DIR, '../../../data/render/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../../data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../static/data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../data/sabian.json')),
]

HOUSE_SYS_NAMES = {
    'P': 'Placidus', 'W': 'Whole Sign', 'K': 'Koch', 'R': 'Regiomontanus', 
    'C': 'Campanus', 'E': 'Equal', 'O': 'Porphyry'
}

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def ensure_location_string(seed):
    loc_val = str(seed.get("location", seed.get("city", ""))).strip()
    if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
        lat, lng = seed.get("lat"), seed.get("lng")
        if lat is not None and lng is not None:
            seed["city"] = f"{abs(float(lat)):.2f}°{'N' if float(lat)>=0 else 'S'}, {abs(float(lng)):.2f}°{'E' if float(lng)>=0 else 'W'}"
            seed["location"] = seed["city"]
        else:
            seed["city"] = "Unknown Location"
            seed["location"] = "Unknown Location"
    else:
        seed["city"] = loc_val
        seed["location"] = loc_val

def _lords_to_symbols(lord_str):
    if not lord_str or lord_str == "-": return "-"
    syms = {"Sun": "☉", "Moon": "☽", "Mercury": "☿", "Venus": "♀", "Mars": "♂", "Jupiter": "♃", "Saturn": "♄"}
    parts = [p.strip() for p in str(lord_str).split('|')]
    return " | ".join([syms.get(p, p) for p in parts])

def get_house_of_point(lon, houses_dict):
    if not houses_dict: return "-"
    for i in range(1, 13):
        c1 = houses_dict.get(i, {}).get('longitude', 0.0) if isinstance(houses_dict.get(i), dict) else houses_dict.get(i, 0.0)
        c2 = houses_dict.get(i+1 if i < 12 else 1, {}).get('longitude', 0.0) if isinstance(houses_dict.get(i+1 if i < 12 else 1), dict) else houses_dict.get(i+1 if i < 12 else 1, 0.0)
        if c1 < c2:
            if c1 <= lon < c2: return f"H{i}"
        else:
            if c1 <= lon < 360 or 0 <= lon < c2: return f"H{i}"
    return "H1"

def compile_n7_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "n7"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    seed_data = chart_data.get("seed", chart_data)
    meta = chart_data.get("metadata", {})
    ensure_location_string(seed_data) 
    
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    
    # 🚀 [FIX 2]: 하우스 시스템 파싱 강화 (W -> Whole Sign 추출)
    h_sys_raw = str(meta.get("h_sys", "P")).strip().upper()
    h_sys = h_sys_raw[0] if h_sys_raw else "P"
    is_unk = bool(seed_data.get("is_time_unknown", 0))

    lang_val = meta.get("language", chart_data.get("language", seed_data.get("language", "en")))
    is_ko = "ko" in str(lang_val).lower()
    
    sabian_dict = {}
    for spath in SABIAN_PATHS:
        if os.path.exists(spath):
            try:
                with open(spath, 'r', encoding='utf-8') as f: sabian_dict = json.load(f)
                break
            except: pass

    def get_sabian_text(idx):
        idx_str = str(idx)
        def extract_text(entry):
            if isinstance(entry, str): return entry
            if isinstance(entry, dict):
                primary = ["text_ko", "ko", "desc_ko"] if is_ko else ["text_en", "en", "desc_en"]
                for k in primary + ["text", "desc"]:
                    if k in entry and entry[k]: return str(entry[k])
            return ""
        if isinstance(sabian_dict, dict):
            if idx_str in sabian_dict: return extract_text(sabian_dict[idx_str])
            for k, v in sabian_dict.items():
                if isinstance(v, dict) and str(v.get("index", "")) == idx_str: return extract_text(v)
        return ""

    apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])
    ws["B4"] = system.upper()
    if system.lower() == "sidereal": ws["C4"] = ayanamsa.upper()
    
    # 하우스 시스템 Full Name 출력
    ws["B5"] = HOUSE_SYS_NAMES.get(h_sys, "Placidus")

    lat = get_safe_float(seed_data.get("lat"), 37.5665)
    lng = get_safe_float(seed_data.get("lng"), 126.9780)
    date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(seed_data.get('birth_time', '12:00:00'))
    tz = _ensure_float_tz(seed_data.get('timezone', 9.0), date_str)
    
    dt_obj = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
    root_jd = swe.julday(dt_obj.year, dt_obj.month, dt_obj.day, dt_obj.hour + dt_obj.minute/60.0 - tz)

    # 1. 렌더링용 루트 차트 (사용자가 선택한 system 반영)
    root_res = calculate_principia(date_str, time_str, lat, lng, tz, system, ayanamsa, h_sys=h_sys)
    root_arc_p = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='paulus', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    try: root_arc_v = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='valens', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
    except: root_arc_v = {'lots': {}}

    # 2. 🚀 [추가됨]: 페르소나 역산 전용 타겟 차트 (시간 고정을 위해 무조건 "tropical" 강제 적용)
    trop_res = calculate_principia(date_str, time_str, lat, lng, tz, "tropical", ayanamsa, h_sys=h_sys)
    trop_arc_p = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='paulus', system="tropical", ayanamsa=ayanamsa, h_sys=h_sys)
    try: trop_arc_v = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='valens', system="tropical", ayanamsa=ayanamsa, h_sys=h_sys)
    except: trop_arc_v = {'lots': {}}

    target_pool = {}
    
    # 3. 🚀 [수정됨]: 타겟 풀을 담을 때는 root_res가 아닌 trop_res를 사용!
    for k, v in trop_res.get('planets', {}).items():
        target_pool[k] = v.get('longitude')
        if k == "Eros": target_pool["Asteroid Eros"] = v.get('longitude')
        elif k == "Lilith": target_pool["Asteroid Lilith"] = v.get('longitude')
        elif k == "North Node (m)": target_pool["Rahu"] = v.get('longitude')
        elif k == "South Node (m)": target_pool["Ketu"] = v.get('longitude')
        elif k == "True Node": target_pool["North Node (t)"] = v.get('longitude')
        elif k == "South Node": target_pool["South Node (t)"] = v.get('longitude')

    for k, v in trop_arc_p.get('lots', {}).items(): target_pool[k] = v.get('longitude', v.get('value'))
    for k, v in trop_arc_v.get('lots', {}).items():
        if k in ['Eros', 'Necessity']: target_pool[f"{k} (v)"] = v.get('longitude', v.get('value'))
    for k, v in trop_arc_p.get('vertex', {}).items(): target_pool[k] = v.get('longitude', v.get('value'))
    if 'syzygy' in trop_arc_p and trop_arc_p['syzygy'].get('data'):
        target_pool['Syzygy'] = trop_arc_p['syzygy']['data'].get('longitude', trop_arc_p['syzygy']['data'].get('value'))
        
    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)
    row_map = mapping["matrix_layout"]["rows"]
    col_map = mapping["matrix_layout"]["cols"]

    for body_name, r_idx in row_map.items():
        if is_unk and r_idx > 42: break
        
        target_lon = target_pool.get(body_name)
        if target_lon is None: continue

        if body_name == "Sun":
            p_res, p_arc_p, p_arc_v = root_res, root_arc_p, root_arc_v
            y, m, d = dt_obj.year, dt_obj.month, dt_obj.day
            p_hour, p_min = dt_obj.hour, dt_obj.minute
        else:
            p_jd = find_persona_transit(root_jd, target_lon) 
            y_u, m_u, d_u, h_dec_u = swe.revjul(p_jd, swe.GREG_CAL)
            total_sec = int(round(h_dec_u * 3600))
            utc_dt = datetime(y_u, m_u, d_u) + timedelta(seconds=total_sec)
            local_dt = utc_dt + timedelta(hours=tz)
            
            y, m, d = local_dt.year, local_dt.month, local_dt.day
            p_hour, p_min = local_dt.hour, local_dt.minute
            
            p_res = calculate_principia(f"{y:04d}-{m:02d}-{d:02d}", f"{p_hour:02d}:{p_min:02d}:00", lat, lng, tz, system, ayanamsa, h_sys=h_sys)
            p_arc_p = calculate_arcana(f"{y:04d}-{m:02d}-{d:02d}", f"{p_hour:02d}:{p_min:02d}:00", lat, lng, tz, lot_schema='paulus', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
            try: p_arc_v = calculate_arcana(f"{y:04d}-{m:02d}-{d:02d}", f"{p_hour:02d}:{p_min:02d}:00", lat, lng, tz, lot_schema='valens', system=system, ayanamsa=ayanamsa, h_sys=h_sys)
            except: p_arc_v = {'lots': {}}

        ws[f"{col_map['birth_info']}{r_idx}"] = f"{y:04d}-{m:02d}-{d:02d}, {p_hour:02d}:{p_min:02d}"

        p_data = {}
        if body_name in p_res.get('planets', {}): p_data = p_res['planets'][body_name]
        elif body_name == "Asteroid Eros" and "Eros" in p_res.get('planets', {}): p_data = p_res['planets']["Eros"]
        elif body_name == "Asteroid Lilith" and "Lilith" in p_res.get('planets', {}): p_data = p_res['planets']["Lilith"]
        elif body_name == "Rahu" and "North Node (m)" in p_res.get('planets', {}): p_data = p_res['planets']["North Node (m)"]
        elif body_name == "Ketu" and "South Node (m)" in p_res.get('planets', {}): p_data = p_res['planets']["South Node (m)"]
        elif body_name == "North Node (t)" and "True Node" in p_res.get('planets', {}): p_data = p_res['planets']["True Node"]
        elif body_name == "South Node (t)" and "South Node" in p_res.get('planets', {}): p_data = p_res['planets']["South Node"]
        elif body_name in p_arc_p.get('lots', {}): p_data = p_arc_p['lots'][body_name]
        elif "(v)" in body_name:
            clean_v = body_name.replace(" (v)", "")
            if clean_v in p_arc_v.get('lots', {}): p_data = p_arc_v['lots'][clean_v]
        elif body_name in p_arc_p.get('vertex', {}): p_data = p_arc_p['vertex'][body_name]
        elif body_name == "Syzygy" and 'syzygy' in p_arc_p: p_data = p_arc_p['syzygy'].get('data', {})
        
        if p_data:
            p_lon = p_data.get('longitude', p_data.get('value', 0.0))
            ws[f"{col_map['info']}{r_idx}"] = format_dms_pretty(p_lon)
            
            ws[f"{col_map['house']}{r_idx}"] = get_house_of_point(p_lon, p_res.get('houses', {}))
            ws[f"{col_map['aries_0']}{r_idx}"] = get_house_of_point(0.0, p_res.get('houses', {}))
            
            dl_sym = _lords_to_symbols(p_res.get('lords', {}).get('day', '-'))
            hl_sym = _lords_to_symbols(p_res.get('lords', {}).get('hour', '-'))
            ws[f"{col_map['day_lord']}{r_idx}"] = dl_sym
            ws[f"{col_map['hour_lord']}{r_idx}"] = hl_sym
            
            ws[f"{col_map['duad']}{r_idx}"] = p_data.get('duad', '-')
            ws[f"{col_map['dodeca']}{r_idx}"] = p_data.get('dodeca', '-')
            ws[f"{col_map['decan']}{r_idx}"] = p_data.get('decan', '-')
            ws[f"{col_map['bounds']}{r_idx}"] = p_data.get('bound', '-')
            
            s_idx = p_data.get('sabian_index', p_data.get('sabian'))
            if not s_idx: s_idx = int(math.floor(p_lon)) + 1
            ws[f"{col_map['sabian']}{r_idx}"] = get_sabian_text(s_idx)

        # 🚀 [수정됨]: 스타일러 적용 루프
        for key, c_char in col_map.items():
            cell = ws[f"{c_char}{r_idx}"]
            is_info = (key in ['info', 'sabian'])
            
            # 🚀 [추가됨]: House, Aries 0, 그리고 Sabian 컬럼은 배경색 스킵! 
            # (4원소 및 행성 배경색 컬러링은 듀아드/데칸/바운드 등의 기호 셀에만 정상 적용됨)
            skip_c = True if key in ['house', 'aries_0', 'sabian'] else False
            
            # Lord 기호가 2개 이상(| 포함)이면 컬러 스킵
            if key in ['day_lord', 'hour_lord'] and "|" in str(cell.value):
                skip_c = True
                
            # 🚀 [추가됨]: Info(도수) 열에 29도가 있을 때만 핀포인트로 아나레틱 붉은색 폰트 적용
            current_is_anaretic = True if (key == 'info' and "29°" in str(cell.value)) else False
                
            # Day/Hour Lord 인자를 아예 넘기지 않으므로(기본값 False), 폰트가 굵어지는 현상은 완벽히 차단됨
            apply_grimoire_styles(
                cell, 
                cell.value, 
                is_info_col=is_info, 
                skip_color=skip_c,
                is_anaretic=current_is_anaretic
            )
            
            if c_char in ["D", "E", "F", "G", "H", "I", "J", "K"]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 🚀 [FIX 3]: 시트 전체 폰트를 Consolas로 통일 (기존 색상, 굵기 등은 완벽 보존)
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name="Consolas",
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    vertAlign=cell.font.vertAlign,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color=cell.font.color
                )

    return wb