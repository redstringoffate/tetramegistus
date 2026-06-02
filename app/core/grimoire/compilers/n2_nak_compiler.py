import json
import os
import shutil
import math
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import calculate_principia, _ensure_float_tz

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/n2_nak_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/nigredo/n2_nak.xlsx'))

SABIAN_PATHS = [
    os.path.abspath(os.path.join(BASE_DIR, '../../../data/render/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../../data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../static/data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../data/sabian.json')),
]

H_SYS_MAP = {
    "P": "Placidus", "W": "Whole House", "K": "Koch"
}

RULER_SYMBOLS = {
    "Sun": "☉", "Moon": "☽", "Mercury": "☿", "Venus": "♀", "Mars": "♂", 
    "Jupiter": "♃", "Saturn": "♄", "Uranus": "♅", "Neptune": "♆", "Pluto": "♇",
    "Rahu": "☊", "Ketu": "☋", "Chiron": "⚷"
}

def get_col(cell_str):
    if not cell_str: return ""
    return "".join([c for c in str(cell_str) if c.isalpha()])

def compile_n2_nak_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 🚀 [수복]: 메모리로 즉시 로드 (새 아키텍처 이식)
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "n2_nak"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name
    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")
    try: fixed_star_orb = float(meta.get("fixed_star_orb", 1.0))
    except: fixed_star_orb = 1.0

    if not seed_data:
        raise ValueError("Station is vacant. Cannot calculate Principia for Grimoire.")

    # 🚀 [여기에 삽입!]: N3 컴파일러들의 Location 우선순위 정규화 로직
    def ensure_location_string(seed):
        loc_val = str(seed.get("location", seed.get("city", ""))).strip()
        if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
            lat, lng = seed.get("lat"), seed.get("lng")
            if lat is not None and lng is not None:
                seed["city"] = f"{abs(float(lat)):.2f}°{'N' if float(lat)>=0 else 'S'}, {abs(float(lng)):.2f}°{'E' if float(lng)>=0 else 'W'}"
                seed["location"] = seed["city"]
            else:
                seed["city"] = "Unknown Location"
                seed["location"] = "Unknown Location"
        else:
            seed["city"] = loc_val
            seed["location"] = loc_val
            
    ensure_location_string(seed_data)
    # -------------------------------------------------------------

    date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(seed_data.get('birth_time', '12:00:00'))
    try: lat = float(seed_data.get("lat", 37.5665))
    except: lat = 37.5665
    try: lng = float(seed_data.get("lng", 126.9780))
    except: lng = 126.9780
    
    tz_val = seed_data.get('tz') if seed_data.get('tz') is not None else seed_data.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)
    is_unk = bool(seed_data.get("is_time_unknown", 0))

    apply_natal_stamp(ws, seed_data, method="single", cells=["A2"])

    lang_val = chart_data.get("language", meta.get("language", seed_data.get("language", seed_data.get("lang", "en"))))
    lang = str(lang_val).lower()
    is_ko = "ko" in lang or "kr" in lang

    principia_res = calculate_principia(
        date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz,
        system=system, ayanamsa=ayanamsa, view=view, h_sys=h_sys,
        fixed_star_orb=fixed_star_orb, is_time_unknown=is_unk
    )

    all_engine_data = {}
    for key in ['planets', 'angles', 'fates', 'houses']:
        if key in principia_res and isinstance(principia_res[key], dict):
            all_engine_data.update(principia_res[key])
            
    lords_data = principia_res.get('lords', {})

    sabian_dict = {}
    for spath in SABIAN_PATHS:
        if os.path.exists(spath):
            try:
                with open(spath, 'r', encoding='utf-8') as f:
                    sabian_dict = json.load(f)
                break
            except: pass
    
    def get_sabian_text(idx):
        if idx is None or str(idx).strip() == "": return ""
        idx_str = str(idx)
        
        def extract_text(entry):
            if isinstance(entry, str): return entry
            if isinstance(entry, dict):
                ko_keys = ["text_ko", "ko", "desc_ko", "ko_KR", "korean"]
                en_keys = ["text_en", "en", "desc_en", "en_US", "english"]
                primary_keys = ko_keys if is_ko else en_keys
                secondary_keys = en_keys if is_ko else ko_keys
                fallback_keys = ["text", "desc", "meaning", "symbol"]
                
                for k in primary_keys + secondary_keys + fallback_keys:
                    if k in entry and entry[k]: return str(entry[k])
                for k, v in entry.items():
                    if k != "index" and isinstance(v, str): return v
            return ""

        if isinstance(sabian_dict, dict):
            if idx_str in sabian_dict: return extract_text(sabian_dict[idx_str])
            for k, v in sabian_dict.items():
                if isinstance(v, dict) and str(v.get("index", "")) == idx_str: return extract_text(v)
        elif isinstance(sabian_dict, list):
            for item in sabian_dict:
                if isinstance(item, dict) and str(item.get("index", "")) == idx_str: return extract_text(item)
        return ""

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    day_val = lords_data.get("day", "")
    day_cell_str = mapping.get("metadata", {}).get("day_lord")
    if day_cell_str:
        ws[day_cell_str] = day_val
        apply_grimoire_styles(ws[day_cell_str], day_val, skip_color=("|" in str(day_val)))
        ws[day_cell_str].alignment = Alignment(horizontal='right', vertical='center') # 🚀 우측 정렬 추가

    hour_val = lords_data.get("hour", "")
    hour_cell_str = mapping.get("metadata", {}).get("hour_lord")
    if hour_cell_str:
        ws[hour_cell_str] = hour_val
        apply_grimoire_styles(ws[hour_cell_str], hour_val, skip_color=False)
        ws[hour_cell_str].alignment = Alignment(horizontal='right', vertical='center')

    ayanamsa_cell = mapping.get("metadata", {}).get("ayanamsa")
    if ayanamsa_cell and system.upper() == "SIDEREAL":
        ws[ayanamsa_cell] = ayanamsa.upper()
        apply_grimoire_styles(ws[ayanamsa_cell], ayanamsa.upper(), skip_color=True)
        ws[ayanamsa_cell].alignment = Alignment(horizontal='right', vertical='center') # 🚀 우측 정렬 추가
    
    # 🚀 [New]: Orb B7 각인 (mapping.json 기반)
    orb_cell_ref = mapping["metadata"].get("fixed_star_orb", "B7")
    ws[orb_cell_ref].value = f"{fixed_star_orb:.1f}"
    ws[orb_cell_ref].number_format = '@' # 텍스트 형식 유지
    ws[orb_cell_ref].alignment = Alignment(horizontal='right', vertical='center')
        
    h_sys_cell = mapping.get("metadata", {}).get("h_sys")
    if h_sys_cell:
        full_h_sys = H_SYS_MAP.get(h_sys.upper(), h_sys.capitalize())
        ws[h_sys_cell] = full_h_sys
        apply_grimoire_styles(ws[h_sys_cell], full_h_sys, skip_color=True)

    all_bodies = []
    for category in ["planets", "asteroids", "lilith_nodes", "fates", "angles"]: 
        for body_name, map_data in mapping.get(category, {}).items():
            engine_name = body_name
            if engine_name == "North Node (t)": engine_name = "North Node (t)"
            elif engine_name == "North Node (m)": engine_name = "Rahu"
            elif engine_name == "South Node (m)": engine_name = "Ketu"
            elif engine_name == "South Node (t)": engine_name = "South Node (t)"
            
            p_data = all_engine_data.get(engine_name)
            if p_data and isinstance(p_data, dict): 
                all_bodies.append({
                    "name": body_name,
                    "map": map_data if isinstance(map_data, dict) else {},
                    "data": p_data
                })
    
    all_bodies.sort(key=lambda x: x["map"].get("row_start", 0), reverse=True)

    day_lords_list = [d.strip() for d in str(lords_data.get('day', '-')).split('|')]
    hour_lord = str(lords_data.get('hour', '-')).strip()

    for body in all_bodies:
        b_map = body["map"]
        p_data = body["data"] 
        body_name = body["name"]
        row_idx = b_map.get("row_start", 0)
        
        if not row_idx: continue

        # 🚀 [수정 2]: 아나레틱 판정식 추가 (엔진 데이터에서 직접 계산)
        lon_val = p_data.get('longitude', 0)
        is_anaretic = int(float(lon_val) % 30) == 29
        
        col_info = get_col(b_map.get("info", "E"))
        col_ruler = get_col(b_map.get("ruler", "F"))
        col_dignity = get_col(b_map.get("dignity", "G"))
        col_nakshatra = get_col(b_map.get("nakshatra", "H"))
        col_pada = get_col(b_map.get("pada", "I"))
        col_sabian = get_col(b_map.get("sabian", "J"))
        
        info_text = p_data.get("dms", "")
        
        is_day_lord = body_name in day_lords_list
        is_hour_lord = body_name == hour_lord
        
        if is_hour_lord: info_text = info_text.upper()
            
        info_cell = ws[f"{col_info}{row_idx}"]
        info_cell.value = info_text

        raw_ruler = str(p_data.get("ruler", ""))
        ws[f"{col_ruler}{row_idx}"] = RULER_SYMBOLS.get(raw_ruler.capitalize(), raw_ruler)
        
        # 🚀 [FIX]: Ketu의 Dignity가 빈 값("")으로 오는 현상 강제 교정
        raw_dignity = str(p_data.get("dignity", "")).strip()
        empty_dignity_targets = ["Ketu", "South Node (t)", "South Node (m)", "Ascendant", "Midheaven", "Descendant", "Immum Coeli"]
        
        if not raw_dignity and body_name in empty_dignity_targets:
            raw_dignity = "None"
            
        ws[f"{col_dignity}{row_idx}"] = raw_dignity
        
        nak_data = p_data.get("nakshatra", {})
        if isinstance(nak_data, dict):
            ws[f"{col_nakshatra}{row_idx}"] = str(nak_data.get("name", ""))
            
            is_kp = (ayanamsa.lower() == "kp" or system.lower() == "kp")
            if is_kp:
                raw_lord = str(p_data.get("sub_lord", ""))
            else:
                raw_lord = str(p_data.get("pada_lord", ""))
                
            lord_sym = RULER_SYMBOLS.get(raw_lord.capitalize(), raw_lord)
            ws[f"{col_pada}{row_idx}"] = lord_sym
        
        s_idx = p_data.get("sabian_index")
        if not s_idx and "longitude" in p_data:
            try: s_idx = int(math.floor(float(p_data["longitude"]))) + 1
            except: pass
            
        s_val_str = str(s_idx).strip() if s_idx else ""
        sabian_text = ""
        
        if s_val_str.isdigit(): sabian_text = get_sabian_text(s_val_str)
        if not sabian_text:
            fallback_txt = p_data.get("sabian", p_data.get("sabian_text", p_data.get("sabian_symbol", "")))
            if isinstance(fallback_txt, dict):
                if is_ko: sabian_text = fallback_txt.get("text_ko", fallback_txt.get("ko", fallback_txt.get("text_en", fallback_txt.get("en", ""))))
                else: sabian_text = fallback_txt.get("text_en", fallback_txt.get("en", fallback_txt.get("text_ko", fallback_txt.get("ko", ""))))
            elif isinstance(fallback_txt, str): sabian_text = fallback_txt

        if not sabian_text:
            frontend_sabian = chart_data.get("bodies", {}).get(body_name, {}).get("sabian", "")
            if frontend_sabian: sabian_text = frontend_sabian

        if not sabian_text and s_val_str.isdigit():
            sabian_text = f"[{s_val_str}°] Symbol rendering fallback"

        ws[f"{col_sabian}{row_idx}"] = sabian_text

        # 🚀 [수정 3]: 컬럼 범위 및 skip_color 로직 최적화
        # n2_nak은 A~J열까지 사용하므로 범위를 좁히고, 사비안 열만 색상을 건너뜁니다.
        # 🚀 [수정 1]: A, C, D, J열은 템플릿 배경색 유지 / H열(Nakshatra)은 컬러링 허용
        for c_char in "ABCDEFGHIJ":
            target_cell = ws[f"{c_char}{row_idx}"]
            
            should_skip = (c_char in ["A", "C", "D", "J"])
            
            apply_grimoire_styles(
                target_cell, 
                target_cell.value, 
                is_info_col=(c_char in ["B", "E"]),
                skip_color=should_skip,
                is_day_lord=is_day_lord,
                is_hour_lord=is_hour_lord,
                is_anaretic=is_anaretic
            )

        # --- Fixed Stars ---
        stars = p_data.get("fixed_stars", [])
        valid_stars = []
        if stars and isinstance(stars, list):
            for s in stars:
                if isinstance(s, dict): valid_stars.append(s)
                elif isinstance(s, str): valid_stars.append({"name": s, "position": "", "orb": ""})
            
        if valid_stars:
            ws[f"A{row_idx}"] = valid_stars[0].get("name", "")
            ws[f"B{row_idx}"] = valid_stars[0].get("position", "")
            orb_val = valid_stars[0].get("orb", "")
            if orb_val:
                ws[f"C{row_idx}"].value = f"{orb_val}°"
                ws[f"C{row_idx}"].data_type = 's'
                ws[f"C{row_idx}"].number_format = '@'
            else:
                ws[f"C{row_idx}"].value = ""
            
            # 별 이름(A)은 색상 방어, 정보(B)는 4원소 색상 허용
            # 🚀 [수정 2]: 별 이름(A)과 Orb(C)의 배경색 오염 금지
            for c_char in ["A", "B", "C"]:
                target_cell = ws[f"{c_char}{row_idx}"]
                apply_grimoire_styles(target_cell, target_cell.value, is_info_col=(c_char == "B"), skip_color=(c_char in ["A", "C"]))
            
            for i in range(1, len(valid_stars)):
                insert_idx = row_idx + i
                ws.insert_rows(insert_idx)
                
                for col_chr in "ABCDEFGHIJ":
                    src = ws[f"{col_chr}{row_idx}"]
                    tgt = ws[f"{col_chr}{insert_idx}"]
                    if src.has_style:
                        if src.border: tgt.border = copy(src.border)
                        if src.alignment: tgt.alignment = copy(src.alignment)
                        if src.number_format: tgt.number_format = src.number_format
                        if col_chr == "D" and src.fill:
                            tgt.fill = copy(src.fill)
                
                ws[f"A{insert_idx}"] = valid_stars[i].get("name", "")
                ws[f"B{insert_idx}"] = valid_stars[i].get("position", "")
                orb_val2 = valid_stars[i].get("orb", "")
                if orb_val2:
                    ws[f"C{insert_idx}"].value = f"{orb_val2}°"
                    ws[f"C{insert_idx}"].data_type = 's'
                    ws[f"C{insert_idx}"].number_format = '@'
                else:
                    ws[f"C{insert_idx}"].value = ""
                
                for col_chr in "ABCDEFGHIJ":
                    tgt = ws[f"{col_chr}{insert_idx}"]
                    is_info = (col_chr in ["B", col_info])
                    apply_grimoire_styles(tgt, tgt.value, is_info_col=is_info, skip_color=(col_chr in ["D", col_sabian]))
        else:
            ws[f"A{row_idx}"] = ""
            ws[f"B{row_idx}"] = ""
            ws[f"C{row_idx}"] = ""

    # 1. 11번 행부터 넉넉하게 모든 행의 높이를 16.5(표준)로 강제 덮어쓰기
    for r in range(11, ws.max_row + 5):
        ws.row_dimensions[r].height = 16.5

    # 2. 확실한 타겟팅: A열에 작성자님이 찍어둔 마커('.')를 찾습니다.
    target_row = None
    for r in range(11, ws.max_row + 1):
        val_a = ws.cell(row=r, column=1).value
        if val_a and str(val_a).strip() == ".":
            target_row = r
            break
            
    # 3. 만약 마커를 못 찾았다면(안 찍으셨을 경우 대비) 기존처럼 Lilith 윗줄을 찾습니다.
    if not target_row:
        for r in range(11, ws.max_row + 1):
            val_d = ws.cell(row=r, column=4).value
            if val_d and "lilith" in str(val_d).lower():
                target_row = r - 1
                break

    # 4. 찾아낸 바로 그 '한 줄'만 4.5로 완벽하게 압축시킵니다.
    if target_row:
        ws.row_dimensions[target_row].height = 4.5

    # 🚀 [수정 3]: 리턴 직전, 시트 전체에 Consolas 폰트 일괄 각인
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(name="Consolas", size=cell.font.size, bold=cell.font.bold, color=cell.font.color)

    return wb