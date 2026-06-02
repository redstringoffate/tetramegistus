import json
import os
import shutil
import math
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from api.astrology import get_seed_from_request
from core.astrology.engine import calculate_principia, _ensure_float_tz, format_dms_pretty, SYMBOL_MAP

from core.astrology.divisions.decan import get_decan
from core.astrology.divisions.duad import get_duad
from core.astrology.divisions.dodecatemoria import get_dodecatemoria
from core.astrology.divisions.egyptian_bounds import get_egyptian_bounds
from core.astrology.divisions.sabian_engine import get_sabian_index

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a3_comp_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a3_comp.xlsx'))

SABIAN_PATHS = [
    os.path.abspath(os.path.join(BASE_DIR, '../../../data/render/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../../data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../../static/data/sabian.json')),
    os.path.abspath(os.path.join(BASE_DIR, '../../data/sabian.json')),
]

PLANETS_LIST = ["Sun", "Moon", "Mercury", "Venus", "Mars", "Jupiter", "Saturn", "Uranus", "Neptune", "Pluto", "Mean Lilith", "True Lilith", "Rahu", "Ketu", "North Node (t)", "South Node (t)"]
ASTEROIDS_LIST = ["Chiron", "Ceres", "Juno", "Pallas", "Vesta", "Asteroid Eros", "Psyche", "Asteroid Lilith"]
LOCAL_SIGNS = ["Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo", "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces"]

def get_col(cell_str):
    if not cell_str: return ""
    return "".join([c for c in str(cell_str) if c.isalpha()])

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def calculate_midpoint(lon1, lon2):
    """두 경도 사이의 최단 거리 중점 산출 (Circular Logic)"""
    l1, l2 = float(lon1) % 360, float(lon2) % 360
    diff = abs(l1 - l2)
    if diff <= 180:
        return (l1 + l2) / 2.0
    else:
        return ((l1 + l2 + 360) / 2.0) % 360.0

def format_house_range(lon_start, lon_end):
    if lon_start is None or lon_end is None: return ""
    size = (lon_end - lon_start) % 360
    if size < 0: size += 360
    rounded_size = round(size)
    
    portions = []
    curr = lon_start
    rem = size
    while rem > 1e-5:
        next_boundary = (math.floor(curr / 30) + 1) * 30
        step = min(rem, next_boundary - curr)
        portions.append(step)
        curr = (curr + step) % 360
        rem -= step
        
    rounded_portions = [round(p) for p in portions]
    diff = rounded_size - sum(rounded_portions)
    if diff != 0 and rounded_portions:
        rounded_portions[rounded_portions.index(max(rounded_portions))] += diff
        
    portions_str = "/".join(map(str, rounded_portions))
    return f"{rounded_size} ({portions_str})"

def format_body_string(item):
    lon = float(item.get('longitude', 0))
    dms_str = format_dms_pretty(lon, is_retro=False)
    return f"{item['name']} - {dms_str}"

def compile_a3_comp_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 🚀 [수복]: 메모리로 즉시 로드 (새 아키텍처 이식)
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a3_comp"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name
    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    meta = chart_data.get("metadata", {})
    system = meta.get("sys_tab", "tropical")
    ayanamsa = meta.get("ayanamsa", "lahiri")
    view = meta.get("view_mode", "zodiac")
    h_sys = meta.get("h_sys", "P")
    mode = meta.get("mode", "normal") 

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()

    # 🚀 [Fix 1]: 스탬퍼에 도시 정보를 완벽하게 넘기기 위한 강력한 정규화
    def ensure_location_string(seed):
        keys = ["location_name", "city_name", "location", "city", "place"]
        loc_val = ""
        for k in keys:
            if seed.get(k):
                loc_val = str(seed.get(k)).strip()
                break
        
        if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
            lat = get_safe_float(seed.get("lat"), None)
            lng = get_safe_float(seed.get("lng"), None)
            if lat is not None and lng is not None:
                loc_val = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
            else:
                loc_val = "Unknown Location"
                
        seed["city"] = loc_val
        seed["location"] = loc_val
        seed["city_name"] = loc_val
        seed["location_name"] = loc_val

    ensure_location_string(s1_raw)
    ensure_location_string(s2_raw)

    apply_natal_stamp(ws, s1_raw, method="single", cells=["A2"])
    apply_natal_stamp(ws, s2_raw, method="single", cells=["A3"])

    def map_for_engine(seed):
        date_str = str(seed.get('birth_date', '2000-01-01')).split('T')[0]
        time_str = str(seed.get('birth_time', '12:00:00'))
        tz_val = seed.get('tz') if seed.get('tz') is not None else seed.get('timezone', 9.0)
        return {
            "date_str": date_str, "time_str": time_str,
            "lat": get_safe_float(seed.get("lat"), 37.5665),
            "lng": get_safe_float(seed.get("lng"), 126.9780),
            "timezone": _ensure_float_tz(tz_val, date_str),
            "system": system, "ayanamsa": ayanamsa, "view": view, "h_sys": h_sys,
            "fixed_star_orb": 1.0, "is_time_unknown": bool(seed.get("is_time_unknown", 0))
        }

    res1 = calculate_principia(**map_for_engine(s1_raw))
    res2 = calculate_principia(**map_for_engine(s2_raw))

    def get_val(item):
        return float(item.get('longitude', 0)) if isinstance(item, dict) else float(item)

    cusps_simple = {}
    is_unk = bool(s1_raw.get("is_time_unknown", 0)) or bool(s2_raw.get("is_time_unknown", 0))
    
    # 🚀 [Fix 2]: 180도 플립을 원천봉쇄하는 "MC Anchor 로직"을 컴파일러 내부에 완벽 이식!
    if not is_unk:
        # 1. 10하우스(MC)를 기준점(Anchor)으로 먼저 잡습니다.
        mc1 = get_val(res1.get('houses', {}).get(10) or res1.get('houses', {}).get('10', 0))
        mc2 = get_val(res2.get('houses', {}).get(10) or res2.get('houses', {}).get('10', 0))
        mc_mid = calculate_midpoint(mc1, mc2)
        if mode == 'anti': mc_mid = (mc_mid + 180) % 360
        cusps_simple[10] = mc_mid
        
        # 2. 나머지 하우스들은 MC와의 상대적 거리(오차)를 계산해 뒤집힌 경우 강제 원복시킵니다.
        for i in range(1, 13):
            if i == 10: continue
            v1 = get_val(res1.get('houses', {}).get(i) or res1.get('houses', {}).get(str(i), 0))
            v2 = get_val(res2.get('houses', {}).get(i) or res2.get('houses', {}).get(str(i), 0))
            
            mid = calculate_midpoint(v1, v2)
            if mode == 'anti': mid = (mid + 180) % 360
            
            expected_offset = ((i - 10) * 30) % 360
            actual_offset = (mid - cusps_simple[10]) % 360
            delta = abs(actual_offset - expected_offset)
            if delta > 180: delta = 360 - delta
            
            if delta > 90: # 90도 이상 오차가 난다면 거꾸로 뒤집힌 것!
                mid = (mid + 180) % 360
                
            cusps_simple[i] = mid

    # 🚀 행성 및 소행성 중점 계산 (Flat Dict 의존도 제거)
    comp_coords = {}
    s1_all = {**res1.get('planets', {}), **res1.get('angles', {}), **res1.get('asteroids', {}), **res1.get('lilith_nodes', {})}
    s2_all = {**res2.get('planets', {}), **res2.get('angles', {}), **res2.get('asteroids', {}), **res2.get('lilith_nodes', {})}
    
    for key in set(s1_all.keys()) & set(s2_all.keys()):
        lon1 = s1_all[key].get('longitude')
        lon2 = s2_all[key].get('longitude')
        if lon1 is not None and lon2 is not None:
            mid = calculate_midpoint(lon1, lon2)
            if mode == 'anti': mid = (mid + 180) % 360
            comp_coords[key] = mid

    lang_val = chart_data.get("language", meta.get("language", "en"))
    is_ko = "ko" in str(lang_val).lower() or "kr" in str(lang_val).lower()

    sabian_dict = {}
    for spath in SABIAN_PATHS:
        if os.path.exists(spath):
            try:
                with open(spath, 'r', encoding='utf-8') as f: sabian_dict = json.load(f)
                break
            except: pass
    
    def get_sabian_text(idx):
        if not idx: return ""
        idx_str = str(idx)
        def extract_text(entry):
            if isinstance(entry, str): return entry
            if isinstance(entry, dict):
                p_keys = ["text_ko", "ko", "desc_ko"] if is_ko else ["text_en", "en", "desc_en"]
                for k in p_keys + ["text", "desc", "meaning"]:
                    if k in entry and entry[k]: return str(entry[k])
                for k, v in entry.items():
                    if k != "index" and isinstance(v, str): return v
            return ""
        if isinstance(sabian_dict, dict):
            if idx_str in sabian_dict: return extract_text(sabian_dict[idx_str])
            for k, v in sabian_dict.items():
                if isinstance(v, dict) and str(v.get("index", "")) == idx_str: return extract_text(v)
        return ""

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    comp_type_str = "ANTI-COMPOSITE" if mode == "anti" else "COMPOSITE"
    if mapping["metadata"].get("comp_type"): ws[mapping["metadata"]["comp_type"]] = comp_type_str
    if mapping["metadata"].get("sys_tab"): ws[mapping["metadata"]["sys_tab"]] = system.upper()
    if system.upper() == "SIDEREAL" and mapping["metadata"].get("ayanamsa"): ws[mapping["metadata"]["ayanamsa"]] = ayanamsa.upper()

    for key in ["comp_type", "sys_tab", "ayanamsa"]:
        c_ref = mapping["metadata"].get(key)
        if c_ref and ws[c_ref].value:
            cell = ws[c_ref]
            apply_grimoire_styles(cell, cell.value, is_info_col=False, skip_color=True)
            cell.font = Font(name="Consolas", bold=False, italic=False)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    def get_house_of_point(lon):
        if is_unk or not cusps_simple: return None
        lon = float(lon) % 360
        for i in range(1, 13):
            curr_c = cusps_simple[i]
            next_c = cusps_simple[i+1] if i < 12 else cusps_simple[1]
            if curr_c < next_c:
                if curr_c <= lon < next_c: return i
            else:
                if curr_c <= lon < 360 or 0 <= lon < next_c: return i
        return 1

    # 🚀 [방역]: 한 하우스 내에서 중복된 행성 이름이 나오지 않도록 관리
    h_items = {i: {"P": [], "A": []} for i in range(1, 13)}
    seen_in_house = {i: set() for i in range(1, 13)} # 하우스별 중복 체크용

    for name, lon in comp_coords.items():
        norm_name = name
        if name in ["Mean Node", "North Node (m)"]: norm_name = "Rahu"
        elif name in ["South Node", "South Node (m)"]: norm_name = "Ketu"
        elif name == "Eros": norm_name = "Asteroid Eros"
        elif name == "Lilith": norm_name = "Asteroid Lilith"
        elif name == "Lilith (mean)": norm_name = "Mean Lilith"
        elif name == "Lilith (true)": norm_name = "True Lilith"
        
        h_num = get_house_of_point(lon)
        if not h_num: continue
        
        # 🚀 [Fix]: 이미 해당 하우스에 동일한 이름(예: Rahu)이 있다면 스킵
        if norm_name in seen_in_house[h_num]:
            continue
        seen_in_house[h_num].add(norm_name)
        
        is_an = int(float(lon) % 30) == 29
        d_copy = {"name": norm_name, "longitude": lon, "is_anaretic": is_an}
        
        if norm_name in PLANETS_LIST: h_items[h_num]["P"].append(d_copy)
        elif norm_name in ASTEROIDS_LIST: h_items[h_num]["A"].append(d_copy)

    houses_list = [{"house": i, "map": mapping["houses"][str(i)]} for i in range(1, 13) if str(i) in mapping.get("houses", {})]
    houses_list.sort(key=lambda x: x["map"].get("row_start", 0), reverse=True)

    for item in houses_list:
        h_num, h_map = item["house"], item["map"]
        row_idx = h_map.get("row_start", 0)
        if not row_idx: continue

        col_info, col_p, col_a, col_range = get_col(h_map.get("info", "B")), get_col(h_map.get("planets", "C")), get_col(h_map.get("asteroids", "D")), get_col(h_map.get("range", "E"))
        col_duad, col_dodeca, col_decan, col_bounds, col_sabian = get_col(h_map.get("duad", "F")), get_col(h_map.get("dodeca", "G")), get_col(h_map.get("decan", "H")), get_col(h_map.get("bounds", "I")), get_col(h_map.get("sabian", "J"))

        lon = cusps_simple.get(h_num)
        if lon is None:
            for c_char in "BCDEFGHIJ": ws[f"{c_char}{row_idx}"] = ""
            continue

        p_items = sorted(h_items[h_num]["P"], key=lambda x: float(x.get('longitude', 0)))
        a_items = sorted(h_items[h_num]["A"], key=lambda x: float(x.get('longitude', 0)))
        p_strs, a_strs = [format_body_string(x) for x in p_items], [format_body_string(x) for x in a_items]
        max_rows = max(1, len(p_strs), len(a_strs))

        if max_rows > 1:
            ws.insert_rows(row_idx + 1, amount=max_rows - 1)
            for r in range(1, max_rows):
                insert_idx = row_idx + r
                ws.row_dimensions[insert_idx].height = 16.5
                for c_char in [col_info, col_p, col_a, col_range, col_duad, col_dodeca, col_decan, col_bounds, col_sabian]:
                    src, tgt = ws[f"{c_char}{row_idx}"], ws[f"{c_char}{insert_idx}"]
                    if src.has_style:
                        tgt.border, tgt.fill = copy(src.border), copy(src.fill)
                        tgt.font, tgt.alignment = copy(src.font), copy(src.alignment)
                        tgt.number_format = copy(src.number_format)

        for i in range(max_rows):
            r_idx = row_idx + i
            ws[f"{col_p}{r_idx}"] = p_strs[i] if i < len(p_strs) else ""
            ws[f"{col_a}{r_idx}"] = a_strs[i] if i < len(a_strs) else ""

            # 행성(C열)과 소행성(D열) 각각에 매칭되는 데이터 리스트 준비
            col_targets = [(col_p, p_items), (col_a, a_items)]

            for c_char, item_list in col_targets:
                cell = ws[f"{c_char}{r_idx}"]
                if cell.value and i < len(item_list):
                    item_data = item_list[i]
                    # 🚀 [수정]: 저장된 아나레틱 플래그를 스타일러에 전달
                    apply_grimoire_styles(
                        cell, 
                        str(cell.value), 
                        is_info_col=True, 
                        is_anaretic=item_data.get("is_anaretic", False),
                        skip_planet_color=True # 🚀 [핵심]: 행성 고유색은 무시하고 4원소색만 찾음
                    )
                
                cell.font = Font(name="Consolas", color=cell.font.color if cell.font else "000000", bold=False)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            
            if i == 0:
                sign_idx = int(lon / 30) % 12
                sign_name = LOCAL_SIGNS[sign_idx]
                deg_in_sign = lon % 30

                ws[f"{col_info}{r_idx}"] = format_dms_pretty(lon)
                next_h = 1 if h_num == 12 else h_num + 1
                next_lon = cusps_simple.get(next_h)
                ws[f"{col_range}{r_idx}"] = format_house_range(lon, next_lon) if next_lon is not None else ""
                
                ws[f"{col_duad}{r_idx}"] = SYMBOL_MAP.get(get_duad(sign_name, deg_in_sign), "-")
                ws[f"{col_dodeca}{r_idx}"] = SYMBOL_MAP.get(get_dodecatemoria(deg_in_sign), "-")
                ws[f"{col_decan}{r_idx}"] = SYMBOL_MAP.get(get_decan(sign_name, deg_in_sign), "-")
                ws[f"{col_bounds}{r_idx}"] = SYMBOL_MAP.get(get_egyptian_bounds(sign_name, deg_in_sign), "-")
                s_idx = get_sabian_index(lon)
                ws[f"{col_sabian}{r_idx}"] = get_sabian_text(str(s_idx)) if s_idx else ""
                
                # 하우스 커스프 자체가 아나레틱인지 판정
                cusp_is_an = int(float(lon) % 30) == 29
                
                for c_char in [col_info, col_range, col_duad, col_dodeca, col_decan, col_bounds, col_sabian]:
                    cell = ws[f"{c_char}{r_idx}"]
                    # 🚀 [수정]: B열(info)에서만 빨간색이 허용되도록 is_info_col 조건을 정확히 전달
                    apply_grimoire_styles(
                        cell, 
                        cell.value, 
                        is_info_col=(c_char == col_info), 
                        is_anaretic=cusp_is_an,
                        skip_color=(c_char == col_sabian)
                    )
                    color_to_use = cell.font.color if cell.font and cell.font.color else "000000"
                    cell.font = Font(name="Consolas", color=color_to_use, bold=False, italic=False)
                    horiz = 'left' if c_char in [col_info, col_sabian] else 'center'
                    cell.alignment = Alignment(horizontal=horiz, vertical='center', wrap_text=False)
            else:
                for c_char in [col_info, col_range, col_duad, col_dodeca, col_decan, col_bounds, col_sabian]:
                    ws[f"{c_char}{r_idx}"] = ""

    ws.column_dimensions[col_p].width = 38.0
    ws.column_dimensions[col_a].width = 38.0

    return wb