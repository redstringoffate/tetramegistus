import json
import os
import shutil
import math
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from core.grimoire.styler import apply_grimoire_styles, HOUSE_COLORS, TABULA_FONT_COLORS 
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.engine import calculate_principia, calculate_arcana, _ensure_float_tz, calculate_all_star_positions
from core.astrology.arabic_lots import calculate_arabic_lots
from core.astrology.constants import ASTEROIDS
import swisseph as swe

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/n8_en_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/nigredo/n8_en.xlsx'))

# 🚀 [방역]: 파이썬 엔진이 뱉어내는 중의어(Asc., I.C., Lilith (mean) 등)를 완벽 차단하기 위한 화이트리스트
ALLOWED_UPPER = [
    "SUN", "MOON", "MERCURY", "VENUS", "MARS", "JUPITER", "SATURN", "URANUS", "NEPTUNE", "PLUTO",
    "CHIRON", "CERES", "JUNO", "PALLAS", "VESTA", "ASTEROID EROS", "PSYCHE",
    "ASCENDANT", "MIDHEAVEN", "DESCENDANT", "IMMUM COELI",
    "FORTUNE", "SPIRIT", "NECESSITY", "NECESSITY (V)", "EROS", "EROS (V)", "COURAGE", "VICTORY", "NEMESIS", "VERTEX", "SYZYGY",
    "MOIRA", "KLOTHO", "LACHESIS", "ATROPOS",
    "MEAN LILITH", "TRUE LILITH", "ASTEROID LILITH", "NORTH NODE (T)", "RAHU", "SOUTH NODE (T)", "KETU"
]

HOUSE_SYS_NAMES = {
    'P': 'Placidus', 'W': 'Whole Sign', 'K': 'Koch', 'R': 'Regiomontanus', 
    'C': 'Campanus', 'E': 'Equal', 'O': 'Porphyry'
}

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def ensure_location_string(seed):
    loc_val = str(seed.get("location", seed.get("city", ""))).strip()
    if not loc_val or loc_val.lower() in ["none", "unknown", "unknown location"]:
        lat, lng = get_safe_float(seed.get("lat"), None), get_safe_float(seed.get("lng"), None)
        if lat is not None and lng is not None:
            seed["city"] = f"{abs(lat):.2f}°{'N' if lat>=0 else 'S'}, {abs(lng):.2f}°{'E' if lng>=0 else 'W'}"
            seed["location"] = seed["city"]
        else:
            seed["city"] = "Unknown Location"
            seed["location"] = "Unknown Location"
    else:
        seed["city"] = loc_val
        seed["location"] = loc_val

def get_house_of_point(lon, houses_dict):
    if not houses_dict: return "H1"
    for i in range(1, 13):
        c1 = houses_dict.get(i, {}).get('longitude', 0.0) if isinstance(houses_dict.get(i), dict) else houses_dict.get(i, 0.0)
        c2 = houses_dict.get(i+1 if i < 12 else 1, {}).get('longitude', 0.0) if isinstance(houses_dict.get(i+1 if i < 12 else 1), dict) else houses_dict.get(i+1 if i < 12 else 1, 0.0)
        if c1 < c2:
            if c1 <= lon < c2: return f"H{i}"
        else:
            if c1 <= lon < 360 or 0 <= lon < c2: return f"H{i}"
    return "H1"

def get_row_for_lon(lon_val, mapping_data):
    lon_val = float(lon_val) % 360
    sign_idx = int(lon_val // 30)
    deg = int(lon_val % 30)
    signs = ["Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo", "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces"]
    sign_name = signs[sign_idx]
    return mapping_data["degree_rows"][sign_name]["start"] + deg

def compile_n8_en_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "n8"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    seed_data = chart_data.get("seed", chart_data)
    meta = chart_data.get("metadata", {})
    ensure_location_string(seed_data)
    # 🚀 [엔진 방식 이식]: N8 본체 엔진(calculate_codex_tenebris)과 판정 로직 통일
    is_unk = str(meta.get("is_time_unknown", "0")) in ["1", "True", "true"] or \
             str(seed_data.get("is_time_unknown", "0")) in ["1", "True", "true"] or \
             str(seed_data.get("birth_time", "")).lower() in ["unknown", "none", "", "null"]

    # 🚀 [시간 강제 고정]: Unknown일 때 파싱 에러 방지용 (엔진과 동일)
    raw_birth_time = str(seed_data.get('birth_time', '12:00:00')).strip()
    time_str = "12:00:00" if is_unk else raw_birth_time

    ayanamsa = meta.get("ayanamsa", "lahiri")
    h_sys_raw = str(meta.get("h_sys", "P")).strip().upper()
    h_sys = h_sys_raw[0] if h_sys_raw else "P"
    arabic_ruler = str(meta.get("arabic_ruler", "traditional")).strip().upper()

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: 
        mapping = json.load(f)

    col_minor = mapping["columns"]["minor_asteroids"]
    col_trop = mapping["columns"]["tropical"]
    col_sid = mapping["columns"]["sidereal"]
    col_dra = mapping["columns"]["draconic"]
    col_ket = mapping["columns"]["ketunic"]
    col_arabic = mapping["columns"]["arabic_lots"]
    col_stars = mapping["columns"]["fixed_stars"]
    
    all_col_chars = mapping["columns"].values()

    # 메타데이터 출력
    apply_natal_stamp(ws, seed_data, method="single", cells=[mapping["metadata"]["stamper"]])
    ws[mapping["metadata"]["ayanamsa"]] = ayanamsa.upper()
    ws[mapping["metadata"]["h_sys"]] = HOUSE_SYS_NAMES.get(h_sys, "Placidus")
    ws[mapping["metadata"]["arabic_ruler"]] = arabic_ruler
        
    for c_ref in ["A2", "B5", "B6", "B7"]:
        if ws[c_ref].value:
            ws[c_ref].font = Font(name="Consolas", size=11)
            ws[c_ref].alignment = Alignment(horizontal='left', vertical='center')

    # 🚀 Time Unknown일 경우 G열(Arabic Lots) 완벽하게 숨김 (I열 Fixed Stars는 보호됨)
    if is_unk:
        ws.column_dimensions[col_arabic].hidden = True

    lat = get_safe_float(seed_data.get("lat"), 37.5665)
    lng = get_safe_float(seed_data.get("lng"), 126.9780)
    date_str = str(seed_data.get('birth_date', '2000-01-01')).split('T')[0]
    
    # 🚀 [방역 1]: 'Unknown' 텍스트 에러 방어 및 기본 시간(12시) 할당
    time_str = str(seed_data.get('birth_time', '12:00:00')).strip()
    if time_str.lower() in ["unknown", "none", ""]:
        time_str = "12:00:00"
        
    try:
        dt_obj = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
    except ValueError:
        time_str = "12:00:00"
        dt_obj = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
        
    tz = _ensure_float_tz(seed_data.get('timezone', 9.0), date_str)
    root_jd = swe.julday(dt_obj.year, dt_obj.month, dt_obj.day, dt_obj.hour + dt_obj.minute/60.0 - tz)

    # 1. 4대 체계 연산
    trop_res = calculate_principia(date_str, time_str, lat, lng, tz, "tropical", ayanamsa, h_sys=h_sys)
    trop_arc = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='paulus', system="tropical", ayanamsa=ayanamsa, h_sys=h_sys)
    
    # 🚀 [핵심 추가]: Valens 랏도 마저 계산해서 trop_arc에 합쳐줍니다.
    try:
        valens_arc = calculate_arcana(date_str, time_str, lat, lng, tz, lot_schema='valens', system="tropical", ayanamsa=ayanamsa, h_sys=h_sys)
        if 'lots' not in trop_arc: trop_arc['lots'] = {}
        for k in ['Eros', 'Necessity']:
            if k in valens_arc.get('lots', {}):
                trop_arc['lots'][f"{k} (v)"] = valens_arc['lots'][k]
    except:
        pass
    
    sid_res = calculate_principia(date_str, time_str, lat, lng, tz, "sidereal", ayanamsa, h_sys=h_sys)    
    dra_res = calculate_principia(date_str, time_str, lat, lng, tz, "draconic", ayanamsa, h_sys=h_sys)
    ket_res = calculate_principia(date_str, time_str, lat, lng, tz, "ketunic", ayanamsa, h_sys=h_sys)

    matrix = {
        col_char: { "items": {deg: [] for deg in range(360)}, "bg": {deg: "H1" for deg in range(360)} }
        for col_char in all_col_chars
    }

    # 🚀 하우스 배경색 계산 (Time Unknown 시 전체 컬러링 방어 + B, G, I열 컬러링 방지)
    for deg_abs in range(360):
        if is_unk:
            matrix[col_trop]["bg"][deg_abs] = None
            matrix[col_sid]["bg"][deg_abs] = None
            matrix[col_dra]["bg"][deg_abs] = None
            matrix[col_ket]["bg"][deg_abs] = None
        else:
            # 🚀 [수정됨] N7의 커스프 강제 매핑 로직을 4대 시스템에 일괄 적용
            systems = [
                (col_trop, trop_res),
                (col_sid, sid_res),
                (col_dra, dra_res),
                (col_ket, ket_res)
            ]
            
            for col_char, res_obj in systems:
                h_str = None
                # 1. 해당 도수에 커스프가 정확히 떨어지는지 확인
                for i in range(1, 13):
                    c_lon = res_obj.get('houses', {}).get(i, {}).get('longitude')
                    if c_lon is not None and int(math.floor(c_lon)) == deg_abs:
                        h_str = f"H{i}"
                        break
                
                # 2. 커스프가 없는 일반 도수라면 기본 +0.5 보정 로직 사용
                if not h_str:
                    h_str = get_house_of_point(deg_abs + 0.5, res_obj.get('houses', {}))
                    
                matrix[col_char]["bg"][deg_abs] = h_str
            
        # B, G, I 열은 언제나 하우스 배경색 적용 안 함
        matrix[col_minor]["bg"][deg_abs] = None
        matrix[col_arabic]["bg"][deg_abs] = None
        matrix[col_stars]["bg"][deg_abs] = None

    def add_item(col_char, lon, name, is_day=False, is_hour=False):
        if lon is None: return
        deg_abs = int(math.floor(lon)) % 360
        matrix[col_char]["items"][deg_abs].append({"name": name, "lon": lon, "is_day": is_day, "is_hour": is_hour})

    day_lords = [x.strip() for x in str(trop_res.get('lords', {}).get('day', '-')).split('|')]
    hour_lord = str(trop_res.get('lords', {}).get('hour', '-')).strip()
    is_whole_sign = (h_sys == 'W')

    # [1] C/D/E/F 열 (System 렌더링)
    systems_map = [
        (col_trop, trop_res, trop_arc, False),
        (col_sid, sid_res, None, False),
        (col_dra, dra_res, None, True), 
        (col_ket, ket_res, None, True)  
    ]

    for col_char, res_obj, arc_obj, filter_nodes in systems_map:
        for g in ['planets', 'asteroids', 'lilith_nodes', 'angles']:
            for k, v in res_obj.get(g, {}).items():
                d_name = k.replace(" (Natal)", "")
                
                # 🚀 [Asteroid Eros 완벽 분리] - 물리적 천체(res_obj)에서 나온 Eros는 무조건 Asteroid Eros로 강제 치환!
                if d_name == 'Eros':
                    d_name = 'Asteroid Eros'
                    
                # 🚀 [중의어/찌꺼기 필터링]: ALLOWED_UPPER에 없는 이름(Asc., I.C., True Node 등)은 자동 컷트
                if d_name.upper() not in ALLOWED_UPPER: continue
                
                if is_unk and d_name.upper() in ["ASCENDANT", "MIDHEAVEN", "DESCENDANT", "IMMUM COELI"]: continue
                if filter_nodes and d_name.upper() in ['NORTH NODE (T)', 'SOUTH NODE (T)', 'RAHU', 'KETU']: continue
                
                add_item(col_char, v.get('longitude'), d_name, is_day=(k in day_lords), is_hour=(k == hour_lord))
        
        if arc_obj and not is_unk:
            for k, v in arc_obj.get('lots', {}).items(): 
                add_item(col_char, v.get('value', v.get('longitude')), k)
            for k, v in arc_obj.get('vertex', {}).items(): 
                add_item(col_char, v.get('value', v.get('longitude')), k)
            if 'syzygy' in arc_obj and arc_obj['syzygy'].get('data'):
                add_item(col_char, arc_obj['syzygy']['data'].get('value', arc_obj['syzygy']['data'].get('longitude')), "Syzygy")

        if not is_unk:
            for i in range(1, 13):
                if not is_whole_sign and i in [1, 4, 7, 10]: continue
                c_lon = res_obj.get('houses', {}).get(i, {}).get('longitude')
                if c_lon is not None: add_item(col_char, c_lon, f"{i}H CUSP")

    # [2] B열: Minor Asteroids
    STANDARD_BODIES = ["Sun", "Moon", "Mercury", "Venus", "Mars", "Jupiter", "Saturn", "Uranus", "Neptune", "Pluto", 
                       "Chiron", "Ceres", "Pallas", "Juno", "Vesta", "Eros", "Psyche", "Moira", "Klotho", "Lachesis", "Atropos", "Lilith"]
    swe.set_sid_mode(0, 0, 0)
    for ast_name, ast_num in ASTEROIDS.items():
        if ast_name in STANDARD_BODIES: continue
        try:
            res = swe.calc_ut(root_jd, ast_num + 10000, swe.FLG_SWIEPH | swe.FLG_SPEED)
            add_item(col_minor, res[0][0], ast_name)
        except: pass

    # [3] G열: Arabic Lots
    if not is_unk:
        points = {}
        for name, p in trop_res.get('planets', {}).items(): points[name] = p['longitude']
        for name, p in trop_res.get('angles', {}).items(): points[name] = p['longitude']
        
        house_cusps = {i: trop_res.get('houses', {}).get(i, {}).get('longitude', 0.0) for i in range(1, 13)}
        
        rulers_trad = ["Mars", "Venus", "Mercury", "Moon", "Sun", "Mercury", "Venus", "Mars", "Jupiter", "Saturn", "Saturn", "Jupiter"]
        rulers_mod = ["Mars", "Venus", "Mercury", "Moon", "Sun", "Mercury", "Venus", "Pluto", "Jupiter", "Saturn", "Uranus", "Neptune"]
        r_list = rulers_trad if arabic_ruler.lower() == 'traditional' else rulers_mod
        
        rulers = {i: r_list[int(house_cusps[i] / 30) % 12] for i in range(1, 13)}
        
        hour_lord_name = str(trop_res.get('lords', {}).get('hour', 'Sun')).split('|')[0].strip()
        lord_of_hour_lon = points.get(hour_lord_name, 0.0)
        
        p_new_moon, p_full_moon = 0.0, 0.0
        if 'syzygy' in trop_arc and 'data' in trop_arc['syzygy']:
            sz = trop_arc['syzygy']['data']
            if sz.get('phase') == 'New Moon': p_new_moon = sz.get('value', 0.0)
            else: p_full_moon = sz.get('value', 0.0)
            
        fortune_lon = trop_arc.get('lots', {}).get('Fortune', {}).get('value', 0.0)
        spirit_lon = trop_arc.get('lots', {}).get('Spirit', {}).get('value', 0.0)
        
        a_lots = calculate_arabic_lots(points, house_cusps, rulers, lord_of_hour_lon, p_new_moon, p_full_moon, fortune_lon, spirit_lon)
        for lot_name, lot_lon in a_lots.items():
            add_item(col_arabic, lot_lon, lot_name)

    # [4] I열: Fixed Stars (col_stars = 'I')
    active_stars = calculate_all_star_positions(root_jd, 'tropical', swe.FLG_SWIEPH | swe.FLG_SPEED)
    for star in active_stars:
        add_item(col_stars, star['lon'], star['name'])

    # 🚀 [핵심 방역]: 파이썬이 행 높이를 통제하며 물리적 줄 삽입 (역순 스캔)
    for deg_abs in range(359, -1, -1):
        r_idx = get_row_for_lon(deg_abs, mapping)
        
        max_lines = 1
        for col_char in all_col_chars:
            if col_char == col_arabic and is_unk: continue
            
            raw_items = matrix[col_char]["items"][deg_abs]
            # 🚀 [Tie-Breaker]: 도수 오름차순, 동점일 땐 무조건 CUSP를 0순위로 강제 견인
            raw_items.sort(key=lambda x: (float(x["lon"]), 0 if "CUSP" in str(x["name"]).upper() else 1))
            
            unique_items = []
            seen = set()
            for item in raw_items:
                if item["name"] not in seen:
                    unique_items.append(item)
                    seen.add(item["name"])
            matrix[col_char]["items"][deg_abs] = unique_items
            
            if len(unique_items) > max_lines:
                max_lines = len(unique_items)
                
        if max_lines > 1:
            amount = max_lines - 1
            ins_idx = r_idx + 1
            old_heights = {r: ws.row_dimensions[r].height for r in list(ws.row_dimensions.keys())}
            
            ws.insert_rows(ins_idx, amount=amount)
            
            for r in range(ws.max_row, 0, -1):
                if r >= ins_idx + amount: ws.row_dimensions[r].height = old_heights.get(r - amount, 15)
                elif ins_idx <= r < ins_idx + amount: ws.row_dimensions[r].height = old_heights.get(r_idx, 15)
                else: ws.row_dimensions[r].height = old_heights.get(r, 15)
            
            for r_offset in range(1, max_lines):
                new_r = r_idx + r_offset
                for c_idx in range(1, ws.max_column + 1):
                    src = ws.cell(row=r_idx, column=c_idx)
                    tgt = ws.cell(row=new_r, column=c_idx)
                    if src.has_style:
                        tgt.font = copy(src.font)
                        tgt.border = copy(src.border)
                        tgt.fill = copy(src.fill)
                        tgt.number_format = copy(src.number_format)
                        tgt.alignment = copy(src.alignment)
                    
                    # 🚀 A열(도수)과 H열(사비안 심볼)은 복제된 줄에서 텍스트 중복을 피하기 위해 지움
                    if c_idx in [1, 8]: 
                        tgt.value = "" 

        # 텍스트 및 하우스 색상 주입
        for col_char in all_col_chars:
            if col_char == col_arabic and is_unk: continue
            
            h_str = matrix[col_char]["bg"][deg_abs]
            bg_hex = HOUSE_COLORS.get(h_str) if h_str else None  # B, G, I열은 bg_hex가 None으로 들어감
            items = matrix[col_char]["items"][deg_abs]
            
            # 🚀 [추가]: Asteroids, Arabic Lots, Fixed Stars 열인지 판별하여 컬러링 오염 차단
            is_minor = col_char in [col_minor, col_arabic, col_stars]
            
            for line_idx in range(max_lines):
                target_row = r_idx + line_idx
                cell = ws[f"{col_char}{target_row}"]
                
                if line_idx < len(items):
                    item = items[line_idx]
                    name = item["name"]
                    val_upper = name.upper()
                    
                    is_angle = val_upper in ["ASCENDANT", "MIDHEAVEN", "DESCENDANT", "IMMUM COELI"]
                    if item["is_hour"] or is_angle: name = val_upper
                        
                    cell.value = name
                    apply_grimoire_styles(
                        cell, name, 
                        is_day_lord=item["is_day"], 
                        is_hour_lord=item["is_hour"], 
                        tabula_mode=True, 
                        house_bg=bg_hex,
                        skip_color=is_minor,
                        skip_planet_color=is_minor  # 🚩 Nemesis 등 소행성의 폰트 컬러 강제 덮어쓰기 방지!
                    )
                else:
                    cell.value = ""
                    apply_grimoire_styles(
                        cell, "", 
                        tabula_mode=True, 
                        house_bg=bg_hex,
                        skip_color=is_minor,
                        skip_planet_color=is_minor
                    )
                    
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # 전체 폰트를 Consolas로 통일
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name="Consolas", size=cell.font.size, bold=cell.font.bold,
                    italic=cell.font.italic, color=cell.font.color
                )

    # ... (기존 폰트 세탁 반복문 끝) ...

    # ====================================================================
    # 🚀 [추가]: N8/A8 Tabula 컬럼 Shrink 완벽 방어 및 너비 균일화
    # ====================================================================
    from openpyxl.utils import get_column_letter
    
    temp_wb_for_width = load_workbook(TEMPLATE_FILE)
    temp_ws_for_width = temp_wb_for_width[base_name] if base_name in temp_wb_for_width.sheetnames else temp_wb_for_width.active
    
    mapped_cols = list(all_col_chars) 
    
    for i in range(1, ws.max_column + 1):
        col_letter = get_column_letter(i)
        
        if ws.column_dimensions[col_letter].hidden:
            ws.column_dimensions[col_letter].width = 0.1
            continue
            
        temp_width = temp_ws_for_width.column_dimensions[col_letter].width if col_letter in temp_ws_for_width.column_dimensions else None
        
        if col_letter == 'A':
            ws.column_dimensions[col_letter].width = 15.0  
        elif col_letter == col_arabic:                     # 🚩 [추가]: Arabic Lots 컬럼 예외 처리!
            ws.column_dimensions[col_letter].width = 30.0  # 👈 원하는 폭으로 조절하세요 (기존 14.0)
        elif col_letter in mapped_cols:
            ws.column_dimensions[col_letter].width = 16.0  
        else:
            ws.column_dimensions[col_letter].width = temp_width if temp_width is not None else 40.0

    # A1 병합 해제 및 B1으로 피신 (뷰어 Shrink 꼬임 원천 차단)
    merge_to_remove = None
    for m_range in list(ws.merged_cells.ranges):
        if "A1" in m_range.coord:
            merge_to_remove = m_range
            break
            
    if merge_to_remove:
        max_col_letter = get_column_letter(merge_to_remove.max_col)
        ws.unmerge_cells(merge_to_remove.coord)
        
        ws['B1'].value = ws['A1'].value
        ws['A1'].value = None
        
        if ws['A1'].has_style:
            ws['B1'].font = copy(ws['A1'].font)
            ws['B1'].border = copy(ws['A1'].border)
            ws['B1'].fill = copy(ws['A1'].fill)
            ws['B1'].alignment = copy(ws['A1'].alignment)
            
        ws.merge_cells(f"B1:{max_col_letter}1")
    # ====================================================================

    # 🚀 N8은 A, B열과 상단 뼈대를 고정해야 하므로 C11을 유지합니다!
    ws.freeze_panes = "C11"

    return wb