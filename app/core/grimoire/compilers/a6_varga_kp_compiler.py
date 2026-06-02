import json
import os
import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import Color

from core.grimoire.styler import apply_grimoire_styles
from core.grimoire.stamper import apply_natal_stamp
from core.astrology.engine import calculate_divisio, _ensure_float_tz
from core.astrology.davison import calculate_davison_midpoint
from api.astrology import get_seed_from_request

BASE_DIR = os.path.dirname(__file__)

MAPPING_FILE = os.path.abspath(os.path.join(BASE_DIR, '../mappings/a6_varga_kp_mapping.json'))
TEMPLATE_FILE = os.path.abspath(os.path.join(BASE_DIR, '../../../templates_excel/albedo/a6_varga_kp.xlsx'))

# 🚀 프론트엔드 영문명 -> 산스크리트어(Graha) 번역 맵
REVERSE_GRAHA_MAP = {
    'Ascendant': 'LAGNA', 'Sun': 'SURYA', 'Moon': 'CHANDRA', 
    'Mercury': 'BUDHA', 'Venus': 'SHUKRA', 'Mars': 'MANGALA', 
    'Jupiter': 'BRIHASPATI', 'Saturn': 'SHANI',
    'South Node (t)': 'KETU', 'North Node (t)': 'RAHU',
    'Ketu': 'KETU', 'Rahu': 'RAHU'
}

def get_safe_float(val, default=0.0):
    try: return float(val) if val is not None else default
    except (ValueError, TypeError): return default

def sanitize_unicode(text):
    if isinstance(text, str):
        return text.replace("\ufe0e", "").replace("\ufe0f", "")
    return text

def compile_a6_varga_kp_grimoire(chart_data, seed_data=None):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template missing: {TEMPLATE_FILE}")

    # 2. 상단 로직을 n2_compiler.py 스타일로 교체
    wb = load_workbook(TEMPLATE_FILE)
    base_name = "a6_varga"
    ws = wb[base_name] if base_name in wb.sheetnames else wb.active
    ws.title = base_name

    for s in list(wb.sheetnames):
        if s != base_name: del wb[s]
    wb.active = wb._sheets.index(ws)

    if not seed_data or 'seed1' not in seed_data or 'seed2' not in seed_data:
        raise ValueError("Grimoire requires Albedo seed data. Please run A1 first.")

    s1_raw = seed_data['seed1'].copy()
    s2_raw = seed_data['seed2'].copy()
    dav_hydrated = calculate_davison_midpoint(s1_raw, s2_raw)

    # 🚀 A6와 동일하게 위치 정보를 좌표 포맷으로 강제 덮어쓰기
    d_lat = get_safe_float(dav_hydrated.get("lat"), None)
    d_lng = get_safe_float(dav_hydrated.get("lng"), None)
    
    if d_lat is not None and d_lng is not None:
        coord_str = f"{abs(d_lat):.2f}°{'N' if d_lat>=0 else 'S'}, {abs(d_lng):.2f}°{'E' if d_lng>=0 else 'W'}"
    else:
        coord_str = "Unknown Location"
        
    dav_hydrated["city"] = coord_str
    dav_hydrated["location"] = coord_str
    dav_hydrated["city_name"] = coord_str
    dav_hydrated["location_name"] = coord_str

    # 🚀 chart_data 꼼수 대신 확실한 단일(single) 스탬프 방식 사용!
    apply_natal_stamp(ws, dav_hydrated, method="single", cells=["A2"])

    meta = chart_data.get("metadata", {})
    ayanamsa = meta.get("ayanamsa", "kp")
    target_body = meta.get("target_body", "Sun")

    date_str = str(dav_hydrated.get('birth_date', '2000-01-01')).split('T')[0]
    time_str = str(dav_hydrated.get('birth_time', '12:00:00'))
    lat = get_safe_float(dav_hydrated.get("lat"), 37.5665)
    lng = get_safe_float(dav_hydrated.get("lng"), 126.9780)
    tz_val = dav_hydrated.get('tz') if dav_hydrated.get('tz') is not None else dav_hydrated.get('timezone', 9.0)
    tz = _ensure_float_tz(tz_val, date_str)

    # Divisio Engine Call
    divisio_res = calculate_divisio(
        date_str=date_str, time_str=time_str, lat=lat, lng=lng, timezone=tz, ayanamsa=ayanamsa
    )
    varga_data = divisio_res.get("varga", {}).get(target_body, {})

    with open(MAPPING_FILE, 'r', encoding='utf-8') as f: mapping = json.load(f)

    layout = mapping.get("layout", {})
    cols = layout.get("cols", {})
    rows = layout.get("rows", {})
    
    col_info = cols.get("info", "B")
    col_nak = cols.get("nakshatra", "C")
    col_pur = cols.get("purushartha", cols.get("sublord", "D"))

    for div_key, r_idx in rows.items():
        d_data = varga_data.get(div_key, {})
        
        info_val = str(d_data.get("formatted", "-"))
        nak_val = str(d_data.get("nakshatra", "-"))
        pur_val = str(d_data.get("purushartha", "-"))
        
        c_info = ws[f"{col_info}{r_idx}"]
        c_nak = ws[f"{col_nak}{r_idx}"]
        c_pur = ws[f"{col_pur}{r_idx}"]
        
        # 1. Info
        c_info.value = info_val
        c_info.data_type = 's'
        apply_grimoire_styles(c_info, sanitize_unicode(info_val), is_info_col=True, skip_color=False)
        c_info.alignment = Alignment(horizontal='left', vertical='center')
        
        # 2. Nakshatra
        base_nak = nak_val.split('-')[0] if "-" in nak_val else nak_val
        c_nak.value = base_nak
        c_nak.data_type = 's'
        apply_grimoire_styles(c_nak, sanitize_unicode(base_nak), is_info_col=False, skip_color=False)
        c_nak.value = nak_val 
        c_nak.alignment = Alignment(horizontal='left', vertical='center')
        
        # 3. Purushartha (Custom colored)
        c_pur.value = pur_val
        c_pur.data_type = 's'
        pur_base = pur_val.split('-')[0] if "-" in pur_val else pur_val
        pur_color = "000000"
        
        if pur_base == "Dharma": pur_color = "FFCC00"
        elif pur_base == "Artha": pur_color = "00CC00"
        elif pur_base == "Kama": pur_color = "FF4444"
        elif pur_base == "Moksha": pur_color = "4488FF"
        
        old_size = c_pur.font.size if c_pur.font else 11
        c_pur.font = Font(name="Consolas", size=old_size, color=pur_color)
        c_pur.alignment = Alignment(horizontal='left', vertical='center')

    # 🚀 4. Graha Text (B5: 우측 정렬, 원래 사이즈 유지, 볼드 해제)
    graha_ref = mapping.get("metadata", {}).get("graha")
    if graha_ref:
        g_cell = ws[graha_ref]
        sanskrit_name = REVERSE_GRAHA_MAP.get(target_body, target_body).upper()
        g_cell.value = str(sanskrit_name)
        g_cell.data_type = 's'
        g_cell.number_format = '@'
        apply_grimoire_styles(g_cell, sanskrit_name, skip_color=True)
        
        color_to_use = copy(g_cell.font.color) if g_cell.font and g_cell.font.color else "000000"
        old_size = g_cell.font.size if g_cell.font else 11
        g_cell.font = Font(name="Consolas", size=old_size, bold=False, color=color_to_use)
        g_cell.alignment = Alignment(horizontal='right', vertical='center')

    # Shrink Logic (구분선 납작하게)
    for r_idx in range(1, ws.max_row + 1):
        cell_val = ws[f"A{r_idx}"].value
        if str(cell_val).strip() == ".":
            ws.row_dimensions[r_idx].height = 4.5
            for col_char in ["A", "B", "C", "D"]:
                ws[f"{col_char}{r_idx}"].value = ""
        elif r_idx >= 7:
            if ws.row_dimensions[r_idx].height != 4.5:
                ws.row_dimensions[r_idx].height = 16.5

    # 🚀 [최종 해결책]: 엑셀 서식 복구 방어 (문자열 타입 쐐기 + 안전한 Consolas 덮어쓰기)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.data_type = 's'
                
            if cell.font:
                safe_color = copy(cell.font.color) if cell.font.color else None
                cell.font = Font(name="Consolas", 
                                 size=cell.font.size,  # 템플릿 원본 사이즈 유지
                                 bold=cell.font.bold, 
                                 italic=cell.font.italic, 
                                 color=safe_color)
            else:
                cell.font = Font(name="Consolas")

    # ====================================================================
    # 🚀 [수정]: 열 너비 축소(Shrink) 완벽 방어 (물리적 수치 강제 주입 + A1 해제)
    # ====================================================================
    # 1. 템플릿 의존 없이, 글자가 절대 잘리지 않을 넉넉한 고정 수치로 쐐기를 박습니다.
    forced_widths = {
        "A": 22.0, 
        "B": 20.0,  # Information (가장 긴 텍스트도 넉넉하게)
        "C": 20.0,  # Nakshatra
        "D": 20.0   # Purushartha
    }
    for col_char, target_width in forced_widths.items():
        ws.column_dimensions[col_char].width = target_width

    # 2. A1 병합 해제 및 B1으로 이사 (뷰어의 렌더링 꼬임 원천 차단)
    from openpyxl.utils import get_column_letter
    merge_to_remove = None
    for m_range in list(ws.merged_cells.ranges):
        if "A1" in m_range.coord:
            merge_to_remove = m_range
            break
            
    if merge_to_remove:
        max_col_letter = get_column_letter(merge_to_remove.max_col)
        ws.unmerge_cells(merge_to_remove.coord)
        
        ws['B1'].value = ws['A1'].value
        ws['A1'].value = None
        
        if ws['A1'].has_style:
            ws['B1'].font = copy(ws['A1'].font)
            ws['B1'].border = copy(ws['A1'].border)
            ws['B1'].fill = copy(ws['A1'].fill)
            ws['B1'].alignment = copy(ws['A1'].alignment)
            
        ws.merge_cells(f"B1:{max_col_letter}1")

    return wb