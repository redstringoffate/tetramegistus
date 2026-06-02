import re
from copy import copy
from openpyxl.styles import Font, PatternFill

# ─────────────────────────────────────────────────────────────
# 🎨 1. 컬러 팔레트 (Hex Code)
# ─────────────────────────────────────────────────────────────
HOUSE_COLORS = {
    "H1": "F16F65", "H2": "FCC26C", "H3": "E2D782",
    "H4": "E7F258", "H5": "F296BD", "H6": "F1B3AD",
    "H7": "6472F2", "H8": "9872F6", "H9": "E480E4",
    "H10": "72F05A", "H11": "95E3F3", "H12": "AEB9F0"
}

PLANET_COLORS = {
    "Rahu": "595959", "☊": "595959", "Ketu": "EEEEEE", "☋": "EEEEEE",
    "Saturn": "FFFF00", "♄": "FFFF00", "Jupiter": "00B050", "♃": "00B050",
    "Mars": "FF0000", "♂": "FF0000", "Sun": "FF66CC", "☉": "FF66CC",
    "Venus": "A6A6A6", "♀": "A6A6A6", "Mercury": "00B0F0", "☿": "00B0F0",
    "Moon": "FFCCFF", "☽": "FFCCFF",
}

ASPECT_COLORS = {
    "Conjunction": "FF4343", "Opposition": "FF00FF", "Trine": "4BD0FF",
    "Square": "FFFF00", "Sextile": "66FFCC", "Quintile": "FFCC99",
    "Septile": "EFCFFD", "Quincunx": "FF8FFA", "Octile": "64B7CE",
    "Novile": "9A9DDA", "Decile": "CCECFF", "Undecile": "FFFFC5", "Semi-sextile": "C1FFEA"
}

ELEMENT_COLORS = { "Fire": "FFCCCC", "Earth": "FFFF99", "Air": "F2F2F2", "Water": "CCFFFF" }
VEDIC_COLORS = { "Dharma": "FFC000", "Artha": "66FF05", "Kama": "FF99FF", "Moksha": "CCCCFF" }

KEYWORD_MAP = {
    "Fire": "Fire", "Aries": "Fire", "Leo": "Fire", "Sagittarius": "Fire", 
    "♈︎": "Fire", "♌︎": "Fire", "♐︎": "Fire",
    "Mesha": "Fire", "Simha": "Fire", "Dhanu": "Fire",
    "Earth": "Earth", "Taurus": "Earth", "Virgo": "Earth", "Capricorn": "Earth", 
    "♉︎": "Earth", "♍︎": "Earth", "♑︎": "Earth",
    "Vrishabha": "Earth", "Kanya": "Earth", "Makara": "Earth",
    "Air": "Air", "Gemini": "Air", "Libra": "Air", "Aquarius": "Air", 
    "♊︎": "Air", "♎︎": "Air", "♒︎": "Air",
    "Mithuna": "Air", "Tula": "Air", "Kumbha": "Air",
    "Water": "Water", "Cancer": "Water", "Scorpio": "Water", "Pisces": "Water", 
    "♋︎": "Water", "♏︎": "Water", "♓︎": "Water",
    "Karka": "Water", "Vrishchika": "Water", "Meena": "Water"
}

NAKSHATRA_LORDS = {
    "Ashwini": "Ketu", "Magha": "Ketu", "Mula": "Ketu", 
    "Bharani": "Venus", "Purva Phalguni": "Venus", "Purva Ashadha": "Venus",
    "Krittika": "Sun", "Uttara Phalguni": "Sun", "Uttara Ashadha": "Sun",
    "Rohini": "Moon", "Hasta": "Moon", "Shravana": "Moon",
    "Mrigashira": "Mars", "Chitra": "Mars", "Dhanishta": "Mars",
    "Ardra": "Rahu", "Swati": "Rahu", "Shatabhisha": "Rahu",
    "Punarvasu": "Jupiter", "Vishakha": "Jupiter", "Purva Bhadrapada": "Jupiter",
    "Pushya": "Saturn", "Anuradha": "Saturn", "Uttara Bhadrapada": "Saturn",
    "Ashlesha": "Mercury", "Jyeshtha": "Mercury", "Revati": "Mercury"
}

TABULA_FONT_COLORS = {
    "angles": {"color": "FF0000", "names": ["ASCENDANT", "MIDHEAVEN", "DESCENDANT", "IMMUM COELI"]}, 
    "lots": {"color": "FF1493", "names": ["FORTUNE", "SPIRIT", "NECESSITY", "NECESSITY (V)", "EROS", "EROS (V)", "COURAGE", "VICTORY", "NEMESIS", "VERTEX", "SYZYGY"]}, 
    "planets": {"color": "0070C0", "names": ["SUN", "MOON", "MERCURY", "VENUS", "MARS", "JUPITER", "SATURN", "URANUS", "NEPTUNE", "PLUTO", "CHIRON", "CERES", "JUNO", "PALLAS", "VESTA", "ASTEROID EROS", "PSYCHE"]}, 
    "fates": {"color": "7030A0", "names": ["MOIRA", "KLOTHO", "LACHESIS", "ATROPOS"]}, 
    "nodes": {"color": "FFFFFF", "names": ["MEAN LILITH", "TRUE LILITH", "ASTEROID LILITH", "NORTH NODE (T)", "RAHU", "SOUTH NODE (T)", "KETU"]}, 
}

# 🚀 [추가됨]: Arabic Lots 폰트 컬러 (어두운 배경용 가시성 극대화)
ARABIC_LOT_COLORS = {
    # Green (00B050 or 7CFF9B)
    "ADVANCEMENT": "7CFF9B", "BUSINESS PARTNERSHIPS": "7CFF9B", "IF AN EVENT WILL COME ABOUT": "7CFF9B", 
    "IF AN OUTCOME WILL BE USEFUL": "7CFF9B", "INCREASE (POMEGRANATE)": "7CFF9B", "MIRACLES": "7CFF9B", 
    "REVELATION": "7CFF9B", "TRANSFORMATION (POINT OF SONIA)": "7CFF9B",
    
    # Purple (D391FF)
    "AUTHORITY": "D391FF", "FAME AND REPUTATION": "D391FF", "OFFICE & KING": "D391FF", "THOSE SUDDENLY ELEVATED": "D391FF",
    
    # Yellow (FFFF00)
    "COMPULSION": "FFFF00", "DELUSION": "FFFF00", "MADNESS": "FFFF00",
    
    # Red (FF4343)
    "COURAGE, VIOLENCE & COMBAT": "FF4343", "ENEMIES (OLYMPIODORUS A)": "FF4343", "ENEMIES (FIRMICUS)": "FF4343",
    "ENEMIES (HERMES)": "FF4343", "ENEMIES (LAURENTIANUS)": "FF4343", "ENEMIES (OLYMPIODORUS B)": "FF4343",
    "ENEMIES (OLYMPIODORUS C)": "FF4343", "FATALITY (HOURGLASS)": "FF4343", "MISUNDERSTANDING": "FF4343",
    "RUMORS": "FF4343", "SUDDEN PARTING": "FF4343",
    
    # Black -> 다크 그레이 (808080)
    "DEATH (AL-TABARI 1)": "808080", "DEATH POINT (EMERSON)": "808080",
    
    # Navy (4BD0FF or 87CEFA)
    "DESTINY": "87CEFA", "PROSPERITY": "87CEFA", "WORK OF TRUTH": "87CEFA",
    
    # Blue (00B0F0)
    "LIFE (MALE)": "00B0F0", "LIFE (FEMALE)": "00B0F0", "MIND AND CAPTIVITY": "00B0F0", "NEED & DESIRE": "00B0F0",
    "PREJUDICE": "00B0F0", "REPRESSION": "00B0F0", "STABILITY": "00B0F0", "TEMPERAMENT": "00B0F0", 
    "VANITY": "00B0F0", "WAR (AL-TABARI)": "00B0F0"
}

def apply_grimoire_styles(cell, text_value, is_info_col=False, skip_color=False, 
                          is_day_lord=False, is_hour_lord=False, 
                          is_anaretic=False, skip_planet_color=False,
                          tabula_mode=False, house_bg=None):
    
    current_font = copy(cell.font) if cell.font else Font()
    current_font.name = "Consolas"
    
    # 1. 🚀 [방역]: 빈 값일 경우 즉시 종료 (에러 방지)
    if not text_value or not str(text_value).strip():
        if is_info_col: current_font.size = 9
        cell.font = current_font
        if tabula_mode and house_bg:
            clean_hex = house_bg.replace("#", "")
            cell.fill = PatternFill(start_color="FF" + clean_hex, end_color="FF" + clean_hex, fill_type="solid")
        return

    # 2. 🚀 [변수 선언]: 모든 로직 이전에 변수를 먼저 정의해야 에러가 안 납니다.
    val_str = str(text_value).strip()
    val_str_lower = val_str.lower()
    val_str_upper = val_str.upper()

    # 3. 폰트 크기 및 효과 결정
    if is_info_col: current_font.size = 9
    
    # [굵기]: Day Lord이면서 정보 열(또는 N8)일 때만 볼드 적용
    current_font.bold = True if is_day_lord and (is_info_col or tabula_mode) else False
    
    # [색상]: 아나레틱이면서 정보 열(또는 N8)일 때만 빨간색 적용
    if is_anaretic and (is_info_col or tabula_mode):
        current_font.color = "FF0000"
    
    # 4. 🚀 Tabula 모드 (360도 매트릭스) 고유 스타일링
    if tabula_mode:
        if "CUSP" in val_str_upper:
            current_font.color = "808080"
            current_font.size = 8
        elif is_anaretic:
            current_font.color = "FF0000"
        else:
            lines = [line.strip() for line in val_str_upper.split('\n')]
            matched = False
            for category, style_data in TABULA_FONT_COLORS.items():
                for name in style_data["names"]:
                    if name in lines:
                        current_font.color = style_data["color"]
                        matched = True; break
                if matched: break
            if not matched:
                for line in lines:
                    if line in ARABIC_LOT_COLORS:
                        current_font.color = ARABIC_LOT_COLORS[line]
                        matched = True; break
        
        cell.font = current_font
        if house_bg:
            clean_hex = house_bg.replace("#", "")
            cell.fill = PatternFill(start_color="FF" + clean_hex, end_color="FF" + clean_hex, fill_type="solid")
        return

    # 5. 🚀 배경색 매칭 (A2/A3 등 리스트/하우스형)
    if skip_color:
        cell.font = current_font
        return

    matched_bg_color = None
    
    # [우선순위 1]: 하우스 및 아스펙트 (정확히 일치할 때)
    if val_str_upper in HOUSE_COLORS:
        matched_bg_color = HOUSE_COLORS[val_str_upper]
    
    if not matched_bg_color:
        for aspect_key, hex_color in ASPECT_COLORS.items():
            if aspect_key.lower() == val_str_lower:
                matched_bg_color = hex_color; break

    # [우선순위 2]: 🚀 4원소 및 베딕 키워드 (행성보다 먼저 검사하여 우선권 부여)
    if not matched_bg_color:
        for keyword, category in KEYWORD_MAP.items():
            if keyword.lower() in val_str_lower:
                if category in ELEMENT_COLORS: matched_bg_color = ELEMENT_COLORS[category]
                elif category in VEDIC_COLORS: matched_bg_color = VEDIC_COLORS[category]
                break 
    
    # 👉 [여기에 추가] [우선순위 2.5]: 🚀 낙샤트라 로드(행성) 색상 매칭
    if not matched_bg_color and not skip_planet_color:
        for nak, lord in NAKSHATRA_LORDS.items():
            if nak.lower() in val_str_lower:
                matched_bg_color = PLANET_COLORS.get(lord)
                break

    # [우선순위 3]: 행성 색상 (skip_planet_color가 False일 때만)
    if not matched_bg_color and not skip_planet_color:
        for p_key, hex_color in PLANET_COLORS.items():
            if p_key.lower() in val_str_lower:
                matched_bg_color = hex_color; break
            
    cell.font = current_font
    if matched_bg_color:
        clean_hex = matched_bg_color.replace("#", "")
        cell.fill = PatternFill(start_color="FF" + clean_hex, end_color="FF" + clean_hex, fill_type="solid")
    else:
        # 매칭되는 배경색이 없으면 템플릿의 색상을 지움 (Airtight)
        cell.fill = PatternFill(fill_type=None)